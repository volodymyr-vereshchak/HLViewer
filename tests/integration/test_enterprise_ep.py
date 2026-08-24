"""Enterprise (промисловість) endpoints: DB CRUD, Excel template/export/import
(workbooks built in-memory with openpyxl) and /enterprise/volumes/ with a
mocked DPDClient — no live DPD API calls."""

import io
import json

import openpyxl
import pytest_asyncio

from backend.db.engine import async_session_factory
from backend.db.models.device_catalog_model import CorectorType, Manufacturer
from backend.db.models.enterprise_model import EPOCH_INSTALLED_FROM


def tagging_get_volumes(records: list[dict]):
    """A stand-in for DPDClient.get_volumes that echoes the caller's `tag`.

    The real client copies it off each device dict onto every record it
    produces, and the pipeline attributes records by it — the identity
    quadruple no longer identifies a caller-side row, since one corrector can
    serve two metering points over time. A mock that skips this returns
    records nothing can be attributed to.
    """
    async def _get(devices, *args, **kwargs):
        out = []
        for device in devices:
            for rec in records:
                if (rec["serNum"] == device["serNum"]
                        and rec.get("chNum", 0) == device["chNum"]):
                    out.append({**rec, "tag": device.get("tag")})
        return out
    return _get


async def read_stream_events(client, params) -> list[dict]:
    """Consume the NDJSON progress stream into a list of event dicts."""
    events = []
    async with client.stream(
        "GET", "/enterprise/volumes/stream", params=params
    ) as resp:
        assert resp.status_code == 200, await resp.aread()
        async for line in resp.aiter_lines():
            if line.strip():
                events.append(json.loads(line))
    return events


@pytest_asyncio.fixture
async def device_catalog(clean_db) -> dict:
    """Manufacturer + corrector model used by import/export tests."""
    async with async_session_factory() as session:
        mfr = Manufacturer(short_name="РадмирТех", full_name="РадмирТех ТОВ СП", mf_dev=1)
        session.add(mfr)
        await session.flush()
        ct = CorectorType(manufacturer_id=mfr.id, model_name="ВЕГА-1.01", type_dev=3)
        session.add(ct)
        await session.commit()
        await session.refresh(mfr)
        await session.refresh(ct)
        return {"manufacturer": mfr.id, "corector_type": ct.id, "mf_dev": 1, "type_dev": 3}


def _enterprise_payload(seed_topology, **overrides) -> dict:
    """A point with one corrector standing there since forever — the shape
    every point migrated from the pre-history schema has."""
    device = {
        "ser_num": 123456,
        "ch_num": 0,
        "installed_from": EPOCH_INSTALLED_FROM.isoformat(),
        # Unlinked device: raw DPD codes stand in for the catalog, as they do
        # for rows the catalog backfill could not match.
        "mf_dev": 1,
        "type_dev": 3,
    }
    for key in ("ser_num", "ch_num", "corector_type_id", "mf_dev", "type_dev"):
        if key in overrides:
            device[key] = overrides.pop(key)
    payload = {
        "enterprise_name": "ТОВ Завод №1",
        "branch_id": seed_topology["branch"],
        "line_id": seed_topology["line1"],
        "active": True,
        "enabled": True,
        "devices": [device],
    }
    payload.update(overrides)
    return payload


class TestEnterpriseCrud:
    async def test_create_list_update_delete(self, admin_client, seed_topology):
        created = await admin_client.post(
            "/enterprise-mappings/", json=_enterprise_payload(seed_topology)
        )
        assert created.status_code == 201, created.text
        ent = created.json()

        listed = (await admin_client.get("/enterprise-mappings/")).json()
        assert len(listed) == 1
        assert listed[0]["enterprise_name"] == "ТОВ Завод №1"

        patched = await admin_client.patch(
            f"/enterprise-mappings/{ent['id']}", json={"enabled": False}
        )
        assert patched.status_code == 200
        assert patched.json()["enabled"] is False

        assert (
            await admin_client.delete(f"/enterprise-mappings/{ent['id']}")
        ).status_code == 204
        assert (await admin_client.get("/enterprise-mappings/")).json() == []

    async def test_viewer_list_scoped_by_branch(
        self, scoped_viewer_client, seed_two_branches, admin_client
    ):
        # one enterprise in each branch
        for n in (1, 2):
            resp = await admin_client.post(
                "/enterprise-mappings/",
                json={
                    "enterprise_name": f"Підприємство {n}",
                    "branch_id": seed_two_branches[f"branch{n}"],
                    "line_id": seed_two_branches[f"line{n}"],
                    "devices": [{
                        "ser_num": 1000 + n,
                        "ch_num": 0,
                        "mf_dev": 1,
                        "type_dev": 3,
                        "installed_from": EPOCH_INSTALLED_FROM.isoformat(),
                    }],
                },
            )
            assert resp.status_code == 201

        listed = (await scoped_viewer_client.get("/enterprise-mappings/")).json()
        assert [e["enterprise_name"] for e in listed] == ["Підприємство 1"]


class TestExcelTemplateAndExport:
    async def test_template_download(self, admin_client, device_catalog, seed_topology):
        resp = await admin_client.get("/enterprise-mappings/template")
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert wb.sheetnames == ["Дані", "Довідник"]
        ws = wb["Дані"]
        assert ws.cell(row=1, column=1).value == "Підприємство"
        # reference sheet lists the seeded corrector model and lines
        ref = wb["Довідник"]
        assert ref.cell(row=2, column=2).value == "ВЕГА-1.01"
        line_ids = {ref.cell(row=i, column=4).value for i in range(2, 4)}
        assert line_ids == {seed_topology["line1"], seed_topology["line2"]}
        # composite pickable labels carry the id
        labels = {ref.cell(row=i, column=10).value for i in range(2, 4)}
        assert any(f"[ID {seed_topology['line1']}]" in l for l in labels)

    async def test_export_contains_enterprise(
        self, admin_client, device_catalog, seed_topology
    ):
        await admin_client.post(
            "/enterprise-mappings/",
            json=_enterprise_payload(
                seed_topology, corector_type_id=device_catalog["corector_type"]
            ),
        )
        resp = await admin_client.get("/enterprise-mappings/export")
        assert resp.status_code == 200
        ws = openpyxl.load_workbook(io.BytesIO(resp.content))["Дані"]
        row = [ws.cell(row=3, column=c).value for c in range(1, 10)]
        assert row[0] == "ТОВ Завод №1"
        assert row[1] == 123456
        assert row[2] == "РадмирТех"
        assert row[3] == "ВЕГА-1.01"
        # H is a formula deriving the id from the picked label in I
        assert str(row[7]).startswith("=IFERROR(INDEX(")
        assert f"[ID {seed_topology['line1']}]" in row[8]

    @staticmethod
    def _dropdowns_by_column(ws) -> dict:
        """First data-validation per data-sheet column letter."""
        out = {}
        for dv in ws.data_validations.dataValidation:
            col = str(dv.sqref).strip()[0]
            out[col] = dv
        return out

    async def test_template_and_export_have_strict_dropdowns(
        self, admin_client, device_catalog, seed_topology
    ):
        """Manufacturer, model, line (by label) and Так/Ні columns are
        restricted by stop-style list validation in BOTH downloadable
        workbooks; the id column is formula-driven and the sheet protection
        leaves only the entry columns editable."""
        for path in ("/enterprise-mappings/template", "/enterprise-mappings/export"):
            resp = await admin_client.get(path)
            assert resp.status_code == 200, path
            wb = openpyxl.load_workbook(io.BytesIO(resp.content))
            ws = wb["Дані"]
            dvs = self._dropdowns_by_column(ws)

            assert set(dvs) >= {"C", "D", "F", "G", "I"}, path
            assert "H" not in dvs  # id is not user-selectable at all
            assert dvs["C"].formula1.startswith("'Довідник'!$I$2")
            assert dvs["D"].formula1.startswith("'Довідник'!$B$2")
            assert dvs["I"].formula1.startswith("'Довідник'!$J$2")
            assert dvs["F"].formula1 == '"Так,Ні"'
            for dv in dvs.values():
                assert dv.errorStyle == "stop"
                assert dv.showErrorMessage is True

            # The id column is a locked formula; entry columns are unlocked.
            assert ws.protection.sheet is True
            assert str(ws.cell(row=5, column=8).value).startswith("=IFERROR(INDEX(")
            assert ws.cell(row=5, column=8).protection.locked is True
            assert ws.cell(row=5, column=9).protection.locked is False
            assert ws.cell(row=5, column=1).protection.locked is False

            # Reference sheet: dropdown sources + fixed header
            ref = wb["Довідник"]
            assert ref.cell(row=1, column=9).value == "Виробники (унікальні)"
            assert ref.cell(row=2, column=9).value == "РадмирТех"
            assert ref.cell(row=1, column=10).value == "Лінія (для вибору)"
            assert ref.cell(row=1, column=6).value == "Обчислювач"

    async def test_template_spelling_fixed(
        self, admin_client, device_catalog, seed_topology
    ):
        resp = await admin_client.get("/enterprise-mappings/template")
        ref = openpyxl.load_workbook(io.BytesIO(resp.content))["Довідник"]
        headers = [ref.cell(row=1, column=c).value for c in range(1, 10)]
        assert "Вичислювач" not in headers
        assert "Обчислювач" in headers


class TestExcelImport:
    def _workbook_bytes(self, rows: list[list]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Підприємство", "Серійний номер", "Виробник", "Модель коректора",
                   "Канал (0-based)", "Активний", "Увімкнений", "ID лінії", "Назва лінії"])
        ws.append(["hint"] * 9)  # row 2 = hints, data starts at row 3
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def _upload(self, client, content: bytes, **params):
        return await client.post(
            "/enterprise-mappings/upload",
            params=params,
            files={"file": ("import.xlsx", content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    async def test_import_creates_records(
        self, admin_client, device_catalog, seed_topology
    ):
        content = self._workbook_bytes([
            ["ТОВ Імпорт", 555001, "РадмирТех", "ВЕГА-1.01", 0, "Так", "Так",
             seed_topology["line1"], "l1"],
        ])
        resp = await self._upload(
            admin_client, content, branch_id=seed_topology["branch"]
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["imported"] == 1
        assert body["errors"] == []

        listed = (await admin_client.get("/enterprise-mappings/")).json()
        assert len(listed) == 1
        ent = listed[0]
        assert ent["line_id"] == seed_topology["line1"]
        # One corrector, standing there since forever: a file without the
        # install column says "the whole archive belongs to this device",
        # which is exactly how every row behaved before the history existed.
        assert len(ent["devices"]) == 1
        device = ent["devices"][0]
        assert device["ser_num"] == 555001
        assert device["corector_type_id"] == device_catalog["corector_type"]
        assert device["installed_from"] == EPOCH_INSTALLED_FROM.isoformat()
        # effective device codes resolved through the catalog
        assert device["mf_dev"] == device_catalog["mf_dev"]
        assert device["type_dev"] == device_catalog["type_dev"]

    async def test_import_recovers_line_id_from_label(
        self, admin_client, device_catalog, seed_topology
    ):
        """When the id cell is empty (the ID formula was never recalculated
        by Excel), the id is recovered from the picked label's [ID N] tail."""
        label = f"l1 — a12 / TESTLUMG [ID {seed_topology['line1']}]"
        content = self._workbook_bytes([
            ["ТОВ Імпорт", 555002, "РадмирТех", "ВЕГА-1.01", 0, "Так", "Так",
             None, label],
        ])
        resp = await self._upload(
            admin_client, content, branch_id=seed_topology["branch"]
        )
        assert resp.status_code == 200
        assert resp.json()["errors"] == []
        listed = (await admin_client.get("/enterprise-mappings/")).json()
        assert listed[0]["line_id"] == seed_topology["line1"]

    async def test_reimport_updates_not_duplicates(
        self, admin_client, device_catalog, seed_topology
    ):
        row = ["ТОВ Імпорт", 555001, "РадмирТех", "ВЕГА-1.01", 0, "Так", "Так", None, None]
        await self._upload(admin_client, self._workbook_bytes([row]))
        row[0] = "ТОВ Імпорт (перейменовано)"
        row[6] = "Ні"
        resp = await self._upload(admin_client, self._workbook_bytes([row]))
        assert resp.json()["imported"] == 1

        listed = (await admin_client.get("/enterprise-mappings/")).json()
        assert len(listed) == 1
        assert listed[0]["enterprise_name"] == "ТОВ Імпорт (перейменовано)"
        assert listed[0]["enabled"] is False

    async def test_import_collects_row_errors(
        self, admin_client, device_catalog, seed_topology
    ):
        content = self._workbook_bytes([
            ["Невідомий виробник", 1, "НемаТакого", "ВЕГА-1.01", 0, "Так", "Так", None, None],
            ["Невідома модель", 2, "РадмирТех", "НемаМоделі", 0, "Так", "Так", None, None],
            ["Погана лінія", 3, "РадмирТех", "ВЕГА-1.01", 0, "Так", "Так", 9999, None],
        ])
        resp = await self._upload(admin_client, content)
        body = resp.json()
        # row 3 imports (bad line id → warning, line_id=null); rows 1-2 skipped
        assert body["imported"] == 1
        assert body["warnings"] == 3

    async def test_import_broken_file_400(self, admin_client, device_catalog):
        resp = await self._upload(admin_client, b"not an xlsx at all")
        assert resp.status_code == 400

    # ── Install date and hour ────────────────────────────────────────────────

    async def _one_device(self, client) -> dict:
        listed = (await client.get("/enterprise-mappings/")).json()
        assert len(listed) == 1
        assert len(listed[0]["devices"]) == 1
        return listed[0]["devices"][0]

    async def test_install_date_without_hour_uses_the_contract_hour(
        self, admin_client, device_catalog, seed_topology
    ):
        """NOT midnight: the hours before the commercial day opens belong to
        the previous day, and so to whatever corrector stood there before."""
        content = self._workbook_bytes([
            ["ТОВ Імпорт", 555001, "РадмирТех", "ВЕГА-1.01", 0, "Так", "Так",
             seed_topology["line1"], "l1", "15.03.2026", None],
        ])
        resp = await self._upload(
            admin_client, content, branch_id=seed_topology["branch"]
        )
        assert resp.status_code == 200, resp.text
        device = await self._one_device(admin_client)
        assert device["installed_from"] == "2026-03-15T07:00:00"

    async def test_install_hour_is_taken_as_given(
        self, admin_client, device_catalog, seed_topology
    ):
        content = self._workbook_bytes([
            ["ТОВ Імпорт", 555001, "РадмирТех", "ВЕГА-1.01", 0, "Так", "Так",
             seed_topology["line1"], "l1", "15.03.2026", 14],
        ])
        await self._upload(admin_client, content, branch_id=seed_topology["branch"])
        device = await self._one_device(admin_client)
        assert device["installed_from"] == "2026-03-15T14:00:00"

    async def test_bad_install_date_is_a_row_error(
        self, admin_client, device_catalog, seed_topology
    ):
        content = self._workbook_bytes([
            ["ТОВ Імпорт", 555001, "РадмирТех", "ВЕГА-1.01", 0, "Так", "Так",
             seed_topology["line1"], "l1", "позавчора", None],
        ])
        resp = await self._upload(
            admin_client, content, branch_id=seed_topology["branch"]
        )
        assert resp.json()["imported"] == 0
        assert any("дата встановлення" in e for e in resp.json()["errors"])

    async def test_reimport_keeps_a_replacement_entered_in_the_admin_panel(
        self, admin_client, device_catalog, seed_topology
    ):
        """The file only ever carries the corrector fitted now. Re-importing
        an older one must not wipe the history somebody entered by hand."""
        content = self._workbook_bytes([
            ["ТОВ Імпорт", 555001, "РадмирТех", "ВЕГА-1.01", 0, "Так", "Так",
             seed_topology["line1"], "l1", None, None],
        ])
        await self._upload(admin_client, content, branch_id=seed_topology["branch"])
        ent = (await admin_client.get("/enterprise-mappings/")).json()[0]

        # A replacement is fitted and recorded in the admin panel.
        patched = await admin_client.patch(
            f"/enterprise-mappings/{ent['id']}",
            json={"devices": [
                {"ser_num": 555001, "ch_num": 0,
                 "corector_type_id": device_catalog["corector_type"],
                 "installed_from": EPOCH_INSTALLED_FROM.isoformat()},
                {"ser_num": 555002, "ch_num": 0,
                 "corector_type_id": device_catalog["corector_type"],
                 "installed_from": "2026-03-15T07:00:00"},
            ]},
        )
        assert patched.status_code == 200, patched.text

        # The same old file is uploaded again.
        again = await self._upload(
            admin_client, content, branch_id=seed_topology["branch"]
        )
        assert again.status_code == 200
        listed = (await admin_client.get("/enterprise-mappings/")).json()
        assert len(listed) == 1, "the point must not be duplicated"
        assert [d["ser_num"] for d in listed[0]["devices"]] == [555001, 555002]

    async def test_unknown_device_on_a_known_name_is_skipped(
        self, admin_client, device_catalog, seed_topology
    ):
        """A serial the point never had means a swap, and a swap needs a date.
        Creating a second point silently would split the consumer's archive."""
        first = self._workbook_bytes([
            ["ТОВ Імпорт", 555001, "РадмирТех", "ВЕГА-1.01", 0, "Так", "Так",
             seed_topology["line1"], "l1", None, None],
        ])
        await self._upload(admin_client, first, branch_id=seed_topology["branch"])

        swapped = self._workbook_bytes([
            ["ТОВ Імпорт", 999999, "РадмирТех", "ВЕГА-1.01", 0, "Так", "Так",
             seed_topology["line1"], "l1", None, None],
        ])
        resp = await self._upload(
            admin_client, swapped, branch_id=seed_topology["branch"]
        )
        assert resp.json()["imported"] == 0
        assert any("адмінці" in e for e in resp.json()["errors"])
        listed = (await admin_client.get("/enterprise-mappings/")).json()
        assert len(listed) == 1
        assert [d["ser_num"] for d in listed[0]["devices"]] == [555001]

    async def test_export_writes_the_current_device_and_its_date(
        self, admin_client, device_catalog, seed_topology
    ):
        created = await admin_client.post(
            "/enterprise-mappings/",
            json={
                "enterprise_name": "ТОВ Експорт",
                "branch_id": seed_topology["branch"],
                "line_id": seed_topology["line1"],
                "devices": [
                    {"ser_num": 1, "ch_num": 0,
                     "corector_type_id": device_catalog["corector_type"],
                     "installed_from": EPOCH_INSTALLED_FROM.isoformat()},
                    {"ser_num": 2, "ch_num": 0,
                     "corector_type_id": device_catalog["corector_type"],
                     "installed_from": "2026-03-15T14:00:00"},
                ],
            },
        )
        assert created.status_code == 201, created.text

        resp = await admin_client.get("/enterprise-mappings/export")
        ws = openpyxl.load_workbook(io.BytesIO(resp.content))["Дані"]
        assert ws.cell(row=3, column=2).value == 2       # the one fitted now
        assert ws.cell(row=3, column=10).value == "15.03.2026"
        assert ws.cell(row=3, column=11).value == 14
        # Points with a longer history are flagged so the file is not mistaken
        # for the whole story.
        assert ws.cell(row=3, column=1).comment is not None

    async def test_export_leaves_the_date_blank_for_an_original_device(
        self, admin_client, device_catalog, seed_topology
    ):
        """«Від початку» must export as an empty cell — re-importing it has to
        mean the same thing, not invent an install date."""
        await admin_client.post(
            "/enterprise-mappings/",
            json=_enterprise_payload(
                seed_topology, corector_type_id=device_catalog["corector_type"]
            ),
        )
        resp = await admin_client.get("/enterprise-mappings/export")
        ws = openpyxl.load_workbook(io.BytesIO(resp.content))["Дані"]
        assert ws.cell(row=3, column=10).value is None
        assert ws.cell(row=3, column=11).value is None


class TestEnterpriseVolumes:
    async def test_volumes_with_mocked_dpd(
        self, admin_client, seed_topology, mocker
    ):
        await admin_client.post(
            "/enterprise-mappings/", json=_enterprise_payload(seed_topology)
        )
        await admin_client.post(
            "/enterprise-mappings/",
            json=_enterprise_payload(
                seed_topology, enterprise_name="ТОВ Завод №2", ser_num=123457
            ),
        )

        # two devices on the same line & date → volumes are summed
        mock_client = mocker.AsyncMock()
        mock_client.get_volumes = mocker.AsyncMock(side_effect=tagging_get_volumes([
            {"serNum": 123456, "mfDev": 1, "typeDev": 3, "chNum": 0,
             "date": "2024-12-25", "dvstAlwrk": 100.5, "temper": 18.0,
             "press": 101.3, "pressUnit": "kPa"},
            {"serNum": 123457, "mfDev": 1, "typeDev": 3, "chNum": 0,
             "date": "2024-12-25", "dvstAlwrk": 49.5},
        ]))
        mocker.patch(
            "backend.services.enterprise_volume_service.DPDClient.for_branch",
            mocker.AsyncMock(return_value=mock_client),
        )

        resp = await admin_client.get(
            "/enterprise/volumes/",
            params={
                "line_id": [seed_topology["line1"]],
                "from_date": "2024-12-25",
                "to_date": "2024-12-25",
            },
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert len(body) == 1
        agg = body[0]
        assert agg["line_id"] == seed_topology["line1"]
        assert agg["total_volume"] == 150.0
        assert agg["device_count"] == 2
        volumes = {d["serNum"]: d["volume"] for d in agg["devices"]}
        assert volumes == {123456: 100.5, 123457: 49.5}

    async def test_volumes_no_mappings_returns_empty(
        self, admin_client, seed_topology, mocker
    ):
        mocker.patch(
            "backend.services.enterprise_volume_service.DPDClient.for_branch",
            mocker.AsyncMock(),
        )
        resp = await admin_client.get(
            "/enterprise/volumes/",
            params={
                "line_id": [seed_topology["line1"]],
                "from_date": "2024-12-25",
                "to_date": "2024-12-25",
            },
        )
        assert resp.status_code == 200
        assert resp.json() == []

    async def test_volumes_invalid_dates_400(self, admin_client, seed_topology):
        resp = await admin_client.get(
            "/enterprise/volumes/",
            params={
                "line_id": [seed_topology["line1"]],
                "from_date": "25.12.2024",
                "to_date": "2024-12-26",
            },
        )
        assert resp.status_code == 400

    async def test_stream_emits_progress_then_result(
        self, admin_client, seed_topology, mocker
    ):
        await admin_client.post(
            "/enterprise-mappings/", json=_enterprise_payload(seed_topology)
        )
        mock_client = mocker.AsyncMock()

        async def fake_get_volumes(devices, date_from, date_to, *,
                                   type_request="daily", max_retries=3,
                                   device_ranges=None, progress_cb=None):
            if progress_cb:
                for i in range(1, len(devices) + 1):
                    progress_cb(i, len(devices))
            return [{
                "tag": d.get("tag"),
                "serNum": 123456, "mfDev": 1, "typeDev": 3, "chNum": 0,
                "date": "2024-12-25", "dvstAlwrk": 100.5,
            } for d in devices]

        mock_client.get_volumes = fake_get_volumes
        mocker.patch(
            "backend.services.enterprise_volume_service.DPDClient.for_branch",
            mocker.AsyncMock(return_value=mock_client),
        )

        events = await read_stream_events(admin_client, {
            "line_id": [seed_topology["line1"]],
            "from_date": "2024-12-25",
            "to_date": "2024-12-25",
        })

        kinds = [e["type"] for e in events]
        assert "progress" in kinds
        assert events[-1]["type"] == "result"
        result = events[-1]["data"]
        assert len(result) == 1
        assert result[0]["line_id"] == seed_topology["line1"]
        assert result[0]["total_volume"] == 100.5

    async def test_stream_include_devices_false_strips_breakdowns(
        self, admin_client, seed_topology, mocker
    ):
        """Reports only need line totals: include_devices=false drops the
        per-device arrays (a month of hourly data shrank from ~18 MB)."""
        await admin_client.post(
            "/enterprise-mappings/", json=_enterprise_payload(seed_topology)
        )
        mock_client = mocker.AsyncMock()
        mock_client.get_volumes = mocker.AsyncMock(side_effect=tagging_get_volumes([{
            "serNum": 123456, "mfDev": 1, "typeDev": 3, "chNum": 0,
            "date": "2024-12-25", "dvstAlwrk": 100.5,
        }]))
        mocker.patch(
            "backend.services.enterprise_volume_service.DPDClient.for_branch",
            mocker.AsyncMock(return_value=mock_client),
        )

        events = await read_stream_events(admin_client, {
            "line_id": [seed_topology["line1"]],
            "from_date": "2024-12-25",
            "to_date": "2024-12-25",
            "include_devices": "false",
        })

        result = events[-1]["data"]
        assert result[0]["total_volume"] == 100.5
        assert result[0]["device_count"] == 1
        assert result[0]["devices"] == []

    async def test_stream_hours_filter_keeps_only_those_hours(
        self, admin_client, seed_topology, mocker
    ):
        """The night report reads nine hours of the day; the other fifteen are
        never aggregated and never sent."""
        await admin_client.post(
            "/enterprise-mappings/", json=_enterprise_payload(seed_topology)
        )
        mock_client = mocker.AsyncMock()
        # Stamps inside the commercial day 25.12 (07:00 → 26.12 06:00): the
        # night hours of that day fall on the NEXT calendar date.
        mock_client.get_volumes = mocker.AsyncMock(side_effect=tagging_get_volumes([
            {"serNum": 123456, "mfDev": 1, "typeDev": 3, "chNum": 0,
             "date": f"{day}T{hour:02d}:00:00", "dvstAlwrk": float(hour)}
            for day, hour in (
                ("2024-12-25", 12), ("2024-12-25", 22),
                ("2024-12-26", 2), ("2024-12-26", 3),
            )
        ]))
        mocker.patch(
            "backend.services.enterprise_volume_service.DPDClient.for_branch",
            mocker.AsyncMock(return_value=mock_client),
        )
        params = {
            "line_id": [seed_topology["line1"]],
            "from_date": "2024-12-25",
            "to_date": "2024-12-25",
            "period_type": "hourly",
            "include_devices": "false",
        }

        every = (await read_stream_events(admin_client, params))[-1]["data"]
        assert sorted(r["period"][11:13] for r in every) == ["02", "03", "12", "22"]

        # Same request, hours narrowed: the totals of the kept hours are
        # untouched and nothing else comes back.
        night = (await read_stream_events(
            admin_client, {**params, "hours": [2, 3]}
        ))[-1]["data"]
        assert {(r["period"], r["total_volume"]) for r in night} == {
            (r["period"], r["total_volume"]) for r in every if r["period"][11:13] in ("02", "03")
        }

    async def test_stream_rejects_hours_out_of_range(self, admin_client, seed_topology):
        async with admin_client.stream("GET", "/enterprise/volumes/stream", params={
            "line_id": [seed_topology["line1"]],
            "from_date": "2024-12-25",
            "to_date": "2024-12-25",
            "hours": [24],
        }) as resp:
            await resp.aread()
            assert resp.status_code == 400

    async def test_stream_reports_dpd_failure_in_band(
        self, admin_client, seed_topology, mocker
    ):
        """The stream is already 200 when the poll fails — the failure arrives
        as an error event, not as an HTTP status."""
        await admin_client.post(
            "/enterprise-mappings/", json=_enterprise_payload(seed_topology)
        )
        mock_client = mocker.AsyncMock()
        mock_client.get_volumes.side_effect = ConnectionError("DPD is down")
        mocker.patch(
            "backend.services.enterprise_volume_service.DPDClient.for_branch",
            mocker.AsyncMock(return_value=mock_client),
        )

        events = await read_stream_events(admin_client, {
            "line_id": [seed_topology["line1"]],
            "from_date": "2024-12-25",
            "to_date": "2024-12-25",
        })

        assert events[-1]["type"] == "error"
        assert "DPD is down" in events[-1]["detail"]

    async def test_stream_no_devices_returns_empty_result(
        self, admin_client, seed_topology
    ):
        events = await read_stream_events(admin_client, {
            "line_id": [seed_topology["line1"]],  # no mappings created
            "from_date": "2024-12-25",
            "to_date": "2024-12-25",
        })
        assert events == [{"type": "result", "data": []}]

    async def test_stream_requires_auth(self, anon_client):
        resp = await anon_client.get("/enterprise/volumes/stream", params={
            "line_id": [1],
            "from_date": "2024-12-25",
            "to_date": "2024-12-25",
        })
        assert resp.status_code == 401

    async def test_cache_clear_admin_only(
        self, admin_client, viewer_client, seed_topology, mocker
    ):
        """DELETE /enterprise/cache/ wipes the DPD cache; the auth middleware
        rejects non-admins (any DELETE needs the admin role)."""
        await admin_client.post(
            "/enterprise-mappings/", json=_enterprise_payload(seed_topology)
        )
        mock_client = mocker.AsyncMock()
        mock_client.get_volumes = mocker.AsyncMock(side_effect=tagging_get_volumes([{
            "serNum": 123456, "mfDev": 1, "typeDev": 3, "chNum": 0,
            "date": "2024-12-25", "dvstAlwrk": 100.5,
        }]))
        mocker.patch(
            "backend.services.enterprise_volume_service.DPDClient.for_branch",
            mocker.AsyncMock(return_value=mock_client),
        )
        params = {
            "line_id": [seed_topology["line1"]],
            "from_date": "2024-12-25",
            "to_date": "2024-12-25",
        }
        await admin_client.get("/enterprise/volumes/", params=params)

        assert (await viewer_client.delete("/enterprise/cache/")).status_code == 403

        resp = await admin_client.delete("/enterprise/cache/")
        assert resp.status_code == 200, resp.text
        assert resp.json()["cleared"] is True

        # The next request finds an empty archive (coverage wiped) and
        # backfills from DPD again.
        mock_client.get_volumes.reset_mock()
        await admin_client.get("/enterprise/volumes/", params=params)
        mock_client.get_volumes.assert_awaited_once()

    async def test_volumes_dpd_down_503(self, admin_client, seed_topology, mocker):
        await admin_client.post(
            "/enterprise-mappings/", json=_enterprise_payload(seed_topology)
        )
        mock_client = mocker.AsyncMock()
        mock_client.get_volumes.side_effect = ConnectionError("DPD is down")
        mocker.patch(
            "backend.services.enterprise_volume_service.DPDClient.for_branch",
            mocker.AsyncMock(return_value=mock_client),
        )
        resp = await admin_client.get(
            "/enterprise/volumes/",
            params={
                "line_id": [seed_topology["line1"]],
                "from_date": "2024-12-25",
                "to_date": "2024-12-25",
            },
        )
        assert resp.status_code == 503
