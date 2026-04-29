"""
Parse sys.xlsx and edit.xlsx exported from metering software,
append deduplicated records to SYSNAME_new.json / EDITNAME_new.json.

Row 3 = code 0, row 4 = code 1, etc.
Only column D (index 3) is used — the type description.

Usage:
    python scripts/parse_type_excel.py --calc-type-id 33
    python scripts/parse_type_excel.py --calc-type-id 33 --sys-xlsx path/sys.xlsx --edit-xlsx path/edit.xlsx
"""

import argparse
import json
import os
import re

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROBE_DIR  = os.path.join(SCRIPT_DIR, "probe_hostlib")
PRELOAD_DIR = os.path.join(SCRIPT_DIR, "..", "backend", "db", "preload_db")

SYS_JSON  = os.path.join(PRELOAD_DIR, "SYSNAME_new.json")
EDIT_JSON = os.path.join(PRELOAD_DIR, "EDITNAME_new.json")

UNKNOWN = "Невідомий код"

# Remove trailing zero-decimal annotations in any form:
#   " (об'єм :0,0 )"  "0,0 м3"  ", 0,0000 Мпа/м2"  etc.
# Strips from the leading comma/space + optional opening paren up to end of line.
_TRAILING_RE = re.compile(r'[,\s]*(?:\([^)]*)?0+,0+.*$')


def clean_sys_name(text: str) -> str:
    return _TRAILING_RE.sub('', text).strip()


def load_json(path: str, key: str) -> list:
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)[key]
    return []


def save_json(path: str, key: str, data: list) -> None:
    with open(path, 'w', encoding='utf-8') as f:
        json.dump({key: data}, f, ensure_ascii=False, indent='\t')


def parse_sys(xlsx_path: str, calc_type_id: int) -> list:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    records = []
    seen = set()
    code = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        raw = row[3] if len(row) > 3 else None
        if raw is None:
            code += 1
            continue
        desc = str(raw).strip()
        if not desc or UNKNOWN in desc:
            code += 1
            continue
        desc = clean_sys_name(desc)
        if not desc or desc.isspace():
            code += 1
            continue
        if desc not in seen:
            seen.add(desc)
            records.append({"ID_TYPE": calc_type_id, "SYS_ID": code, "SYSNAME": desc})
        code += 1
    wb.close()
    return records


def parse_edit(xlsx_path: str, calc_type_id: int) -> list:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    records = []
    seen = set()
    code = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        raw = row[3] if len(row) > 3 else None
        if raw is None:
            code += 1
            continue
        desc = str(raw).strip()
        if not desc or UNKNOWN in desc:
            code += 1
            continue
        if desc not in seen:
            seen.add(desc)
            records.append({"ID_TYPE": calc_type_id, "EDIT_ID": code, "EDITNAME": desc})
        code += 1
    wb.close()
    return records


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--calc-type-id', type=int, required=True, help='ID_TYPE (gas_volume_calc_type_id)')
    parser.add_argument('--sys-xlsx',  default=os.path.join(PROBE_DIR, 'sys.xlsx'))
    parser.add_argument('--edit-xlsx', default=os.path.join(PROBE_DIR, 'edit.xlsx'))
    args = parser.parse_args()

    cid = args.calc_type_id

    # ── SYS ──────────────────────────────────────────────────────────────────
    sys_all = load_json(SYS_JSON, 'SYSNAME')
    sys_all = [r for r in sys_all if r['ID_TYPE'] != cid]
    new_sys = parse_sys(args.sys_xlsx, cid)
    sys_all.extend(new_sys)
    save_json(SYS_JSON, 'SYSNAME', sys_all)
    print(f"SYS : {len(new_sys):4d} records added  (ID_TYPE={cid})  ->  SYSNAME_new.json")

    # ── EDIT ─────────────────────────────────────────────────────────────────
    edit_all = load_json(EDIT_JSON, 'EDITNAME')
    edit_all = [r for r in edit_all if r['ID_TYPE'] != cid]
    new_edit = parse_edit(args.edit_xlsx, cid)
    edit_all.extend(new_edit)
    save_json(EDIT_JSON, 'EDITNAME', edit_all)
    print(f"EDIT: {len(new_edit):4d} records added  (ID_TYPE={cid})  ->  EDITNAME_new.json")


if __name__ == '__main__':
    main()
