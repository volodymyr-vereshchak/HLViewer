"""
Excel workflows for enterprise (промисловість) mappings.

Builds the import template and the DB export workbook, and parses an uploaded
workbook back into enterprise records. Moved out of enterprise_ep.py so the
router only handles HTTP concerns; the column layout, catalog lookups and
row-level validation all live here.
"""

import io
import logging
from collections import defaultdict
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import select

from backend.db.dao.device_catalog_dao import CorectorTypeDao, ManufacturerDao
from backend.db.dao.enterprise_dao import EnterpriseDao
from backend.db.dao.line_dao import LineDao
from backend.db.engine import async_session_factory
from backend.db.models.enterprise_model import Enterprise

logger = logging.getLogger(__name__)

COLUMNS = ["Підприємство", "Серійний номер", "Виробник", "Модель коректора",
           "Канал (0-based)", "Активний", "Увімкнений", "ID лінії", "Назва лінії (довідково)"]
COL_WIDTHS = [40, 18, 16, 24, 16, 12, 14, 12, 32]

_HDR_FILL = "2E7D32"
_HINT_FILL = "1B5E20"
_HINT_COLOR = "A5D6A7"
_REF_HDR_FILL = "1565C0"


class ExcelParseError(ValueError):
    """The uploaded file could not be opened as a workbook."""


def _write_header(ws, hints: list[str]) -> None:
    hdr_fill = PatternFill("solid", fgColor=_HDR_FILL)
    hint_fill = PatternFill("solid", fgColor=_HINT_FILL)
    hdr_font = Font(bold=True, color="FFFFFF")
    hint_font = Font(italic=True, color=_HINT_COLOR)

    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx, hint in enumerate(hints, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font = hint_font
        cell.fill = hint_fill

    ws.freeze_panes = "A3"


async def _lines_with_context(session):
    """Lines for the reference sheet: (line_id, line_name, calc_name, lumg_name).

    Line names are NOT globally unique (a line is unique only within its
    calculator), so the reference carries the ID plus calculator + LUMG context
    to disambiguate. The data sheet then references the line by ID."""
    from backend.db.models.line_model import Line
    from backend.db.models.gas_volume_calc_model import GasVolumeCalc
    from backend.db.models.lumg_model import Lumg

    stmt = (
        select(Line.id, Line.name, GasVolumeCalc.name, Lumg.name)
        .join(GasVolumeCalc, Line.gas_volume_calc_id == GasVolumeCalc.id)
        .join(Lumg, GasVolumeCalc.lumg_id == Lumg.id)
        .order_by(Lumg.name, GasVolumeCalc.name, Line.name)
    )
    return (await session.execute(stmt)).all()


async def build_template_workbook() -> openpyxl.Workbook:
    """Empty import template: data sheet with hints + reference sheet with the
    device catalog and all lines (ID + calculator + LUMG context)."""
    async with async_session_factory() as session:
        manufacturers = await ManufacturerDao(session).get_all()
        corector_types = await CorectorTypeDao(session).get_all()
        lines_ctx = await _lines_with_context(session)  # (id, name, calc, lumg)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дані"

    mfr_names = " / ".join(m.short_name for m in manufacturers) or "РадмирТех / ГРЕМПІС / Тандем / Укргазтех"
    _write_header(ws, [
        "Назва точки обліку",
        "Напр.: 123456 (без дефісів)",
        mfr_names,
        "Модель (напр. ВЕГА-1.01) — див. аркуш 'Довідник'",
        "0, 1, 2, …",
        "Так / Ні",
        "Так / Ні",
        "ID лінії з аркуша 'Довідник' (або порожньо)",
        "Заповнюється автоматично при експорті",
    ])

    # example row
    ex_mfr = manufacturers[0].short_name if manufacturers else "РадмирТех"
    ex_model = corector_types[0].model_name if corector_types else "ВЕГА-1.01"
    ex_lid = lines_ctx[0][0] if lines_ctx else ""
    ex_lname = lines_ctx[0][1] if lines_ctx else ""
    for col_idx, val in enumerate(
        ["ТОВ Завод №1", 123456, ex_mfr, ex_model, 0, "Так", "Так", ex_lid, ex_lname], start=1
    ):
        ws.cell(row=3, column=col_idx, value=val)

    # ── Sheet 2: Reference ────────────────────────────────────────────────────
    ref = wb.create_sheet("Довідник")
    ref_hdr_fill = PatternFill("solid", fgColor=_REF_HDR_FILL)
    ref_hdr_font = Font(bold=True, color="FFFFFF")

    # Manufacturers + models
    ref.cell(row=1, column=1, value="Виробник (скорочено)").font = ref_hdr_font
    ref.cell(row=1, column=1).fill = ref_hdr_fill
    ref.cell(row=1, column=2, value="Модель коректора").font = ref_hdr_font
    ref.cell(row=1, column=2).fill = ref_hdr_fill
    ref.column_dimensions["A"].width = 20
    ref.column_dimensions["B"].width = 28

    mfr_map = {m.id: m.short_name for m in manufacturers}
    row = 2
    for ct in corector_types:
        ref.cell(row=row, column=1, value=mfr_map.get(ct.manufacturer_id, ""))
        ref.cell(row=row, column=2, value=ct.model_name)
        row += 1

    # Lines — ID + назва + вичислювач + ЛУМГ (look up the ID by name here, then
    # put the ID into the 'ID лінії' column on the data sheet).
    line_ref_cols = [("ID лінії", 10), ("Назва лінії", 32), ("Вичислювач", 28), ("ЛУМГ", 24)]
    for j, (title, width) in enumerate(line_ref_cols, start=4):  # columns D, E, F, G
        c = ref.cell(row=1, column=j, value=title)
        c.font = ref_hdr_font
        c.fill = ref_hdr_fill
        ref.column_dimensions[get_column_letter(j)].width = width
    for i, (lid, lname, cname, lumgname) in enumerate(lines_ctx, start=2):
        ref.cell(row=i, column=4, value=lid)
        ref.cell(row=i, column=5, value=lname)
        ref.cell(row=i, column=6, value=cname)
        ref.cell(row=i, column=7, value=lumgname)

    return wb


async def build_export_workbook() -> openpyxl.Workbook:
    """Export the current enterprise table; device codes resolved through the
    corrector-type catalog when linked, legacy columns otherwise."""
    async with async_session_factory() as session:
        enterprises = await EnterpriseDao(session).get_all()
        manufacturers = await ManufacturerDao(session).get_all()
        corector_types = await CorectorTypeDao(session).get_all()
        lines = await LineDao(session).get_all()

    mfr_by_mfdev = {m.mf_dev: m.short_name for m in manufacturers}
    mfr_id_by_mfdev = {m.mf_dev: m.id for m in manufacturers}
    mfr_by_id = {m.id: m for m in manufacturers}
    ct_by_id = {ct.id: ct for ct in corector_types}
    ct_model: dict[tuple, str] = {}
    for ct in corector_types:
        key = (ct.manufacturer_id, ct.type_dev)
        if key not in ct_model:
            ct_model[key] = ct.model_name
    line_by_id = {ln.id: ln.name for ln in lines}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дані"

    _write_header(ws, [
        "Назва точки обліку", "Серійний номер", "Скорочена назва",
        "Модель з довідника", "0, 1, 2…", "Так / Ні", "Так / Ні",
        "ID лінії (ключ для імпорту)", "Назва лінії (довідково)",
    ])

    for row_idx, ent in enumerate(enterprises, start=3):
        # Prefer the corrector-type catalog (source of truth); fall back to the
        # legacy mf_dev/type_dev codes for rows that aren't linked yet.
        if ent.corector_type_id is not None and ent.corector_type_id in ct_by_id:
            ct = ct_by_id[ent.corector_type_id]
            model_name = ct.model_name
            mfr = mfr_by_id.get(ct.manufacturer_id)
            mfr_short = mfr.short_name if mfr else ""
        else:
            mfr_short = mfr_by_mfdev.get(ent.mf_dev, str(ent.mf_dev) if ent.mf_dev is not None else "")
            mfr_id = mfr_id_by_mfdev.get(ent.mf_dev)
            model_name = ct_model.get((mfr_id, ent.type_dev), "") if mfr_id else ""
        line_name = line_by_id.get(ent.line_id, "") if ent.line_id else ""
        ws.cell(row=row_idx, column=1, value=ent.enterprise_name)
        ws.cell(row=row_idx, column=2, value=ent.ser_num)
        ws.cell(row=row_idx, column=3, value=mfr_short)
        ws.cell(row=row_idx, column=4, value=model_name)
        ws.cell(row=row_idx, column=5, value=ent.ch_num)
        ws.cell(row=row_idx, column=6, value="Так" if ent.active else "Ні")
        ws.cell(row=row_idx, column=7, value="Так" if ent.enabled else "Ні")
        ws.cell(row=row_idx, column=8, value=ent.line_id if ent.line_id else None)
        ws.cell(row=row_idx, column=9, value=line_name)

    return wb


def workbook_bytes(wb: openpyxl.Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _parse_bool(val, default=True) -> bool:
    if val is None:
        return default
    return str(val).strip().lower() not in ("ні", "нет", "no", "false", "0", "н")


async def parse_upload(content: bytes, branch_id: Optional[int]) -> tuple[list[dict], list[str]]:
    """Parse an uploaded workbook into enterprise records.

    Returns (records, row_errors). Raises ExcelParseError when the file is not
    a readable workbook. Rows with unknown manufacturer/model are skipped with
    an error; an unknown line id/name only nulls the line link."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise ExcelParseError(f"Не вдалося відкрити файл: {e}")

    # Pre-load lookup tables from DB
    async with async_session_factory() as session:
        manufacturers = await ManufacturerDao(session).get_all()
        corector_types = await CorectorTypeDao(session).get_all()
        all_lines = await LineDao(session).get_all()

    mfr_by_short: dict[str, object] = {m.short_name.strip(): m for m in manufacturers}
    ct_by_mfr_model: dict[tuple, object] = {
        (ct.manufacturer_id, ct.model_name.strip()): ct for ct in corector_types
    }
    line_ids_set: set = {ln.id for ln in all_lines}
    # name → [line_id, …]; a name maps to >1 id when calculators reuse line names
    line_ids_by_name: dict[str, list] = defaultdict(list)
    for ln in all_lines:
        line_ids_by_name[ln.name.strip().lower()].append(ln.id)

    records: list[dict] = []
    errors: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or not any(row):
            continue

        row_padded = (list(row) + [None] * 9)[:9]
        (enterprise_name, ser_num, mfr_str, model_str, ch_num,
         active_str, enabled_str, line_key, _line_name_info) = row_padded

        if not enterprise_name:
            errors.append(f"Рядок {row_idx}: порожня назва — пропущено")
            continue

        # ser_num
        try:
            ser_num_int = int(str(ser_num).replace("-", "").lstrip("0") or "0")
        except (ValueError, TypeError):
            errors.append(f"Рядок {row_idx}: некоректний серійний номер '{ser_num}'")
            continue

        # manufacturer → mf_dev
        mfr_clean = str(mfr_str).strip() if mfr_str else ""
        mfr = mfr_by_short.get(mfr_clean)
        if not mfr:
            errors.append(
                f"Рядок {row_idx}: невідомий виробник '{mfr_clean}'. "
                f"Допустимі: {', '.join(mfr_by_short)}"
            )
            continue

        # model → type_dev
        model_clean = str(model_str).strip() if model_str else ""
        ct = ct_by_mfr_model.get((mfr.id, model_clean))
        if not ct:
            errors.append(
                f"Рядок {row_idx}: модель '{model_clean}' не знайдена для виробника '{mfr_clean}'"
            )
            continue

        # ch_num
        try:
            ch_num_int = int(ch_num)
        except (ValueError, TypeError):
            errors.append(f"Рядок {row_idx}: некоректний канал '{ch_num}'")
            continue

        # line (optional): column "ID лінії". Primary = numeric line ID; fall back
        # to a name match for older files — but reject ambiguous names (a name that
        # belongs to several lines across calculators), telling the user to use the ID.
        line_id = None
        if line_key is not None and str(line_key).strip():
            key = str(line_key).strip()
            if key.replace(".", "", 1).isdigit():            # numeric → line ID
                lid = int(float(key))
                if lid in line_ids_set:
                    line_id = lid
                else:
                    errors.append(f"Рядок {row_idx}: ID лінії {lid} не існує — line_id=null")
            else:                                            # text → name lookup
                ids = line_ids_by_name.get(key.lower(), [])
                if len(ids) == 1:
                    line_id = ids[0]
                elif len(ids) > 1:
                    errors.append(
                        f"Рядок {row_idx}: назва лінії '{key}' неоднозначна "
                        f"({len(ids)} збігів) — вкажіть ID лінії з аркуша 'Довідник'"
                    )
                else:
                    errors.append(f"Рядок {row_idx}: лінія '{key}' не знайдена — line_id=null")

        records.append({
            "enterprise_name": str(enterprise_name).strip(),
            "branch_id": branch_id,
            "ser_num": ser_num_int,
            # Device identity now lives in the corrector-type catalog; store the FK
            # (the resolved CorectorType) instead of the raw mf_dev/type_dev codes.
            "corector_type_id": ct.id,
            "ch_num": ch_num_int,
            "active": _parse_bool(active_str),
            "enabled": _parse_bool(enabled_str),
            "line_id": line_id,
        })

    return records, errors


async def upsert_enterprises(records: list[dict]) -> list[int]:
    """Bulk upsert parsed records by device identity (uq_enterprise_device_ct).
    Returns the ids of the affected rows."""
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    async with async_session_factory() as session:
        stmt = pg_insert(Enterprise).values(records)
        stmt = stmt.on_conflict_do_update(
            constraint='uq_enterprise_device_ct',
            set_={
                'enterprise_name': stmt.excluded.enterprise_name,
                'branch_id': stmt.excluded.branch_id,
                'line_id': stmt.excluded.line_id,
                'active': stmt.excluded.active,
                'enabled': stmt.excluded.enabled,
            }
        ).returning(Enterprise.id)

        result = await session.execute(stmt)
        await session.commit()
        return [row[0] for row in result]
