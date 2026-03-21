"""
analyze_grs_profiles.py

Діагностика per-GRS відхилень у декомпозиті.
Гіпотеза: споживачі різних ГРС мають різні сезонні профілі білінгу,
що спричиняє residual bias у поточному глобальному декомпозиті.

Виходи:
  data/analyze_grs_profiles.xlsx  — 4 листи (GRS_summary, Profile_heatmap,
                                     Comparison_top15, Profile_weights)
  Консоль — кореляція, топ покращень, загальне MAE
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from pobut_predictor import GROUP_MERGE, CT_PRIV, META_COLS

SUBS_FILE   = 'data/input/all_pobut_enriched.csv'
MODEM_FILE  = 'data/input/profile_pobut_daily_result.csv'
RESULT_FILE = 'data/annual_decompose_result.xlsx'
OUT_XLSX    = 'data/analyze_grs_profiles.xlsx'

MIN_CLEAN_GRS    = 50    # мінімальна кількість "ідеальних" абонентів для GRS-профілю
MIN_CORR_PROFILE = 0.85  # мінімальна корел. з глобальним профілем (≈ зсув ≤ 1 міс.)

MONTH_ENG = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
             'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}
MONTHS = list(range(1, 13))

# ── 1. Завантаження даних ─────────────────────────────────────────────────────
print("=== 1. Завантаження даних ===")

md_raw = pd.read_csv(MODEM_FILE, sep=';', low_memory=False)
md_raw.columns = META_COLS + list(md_raw.columns[len(META_COLS):])
md_raw = md_raw[md_raw['gas_off'].isna() & md_raw['alternative'].isna() & md_raw['dacha'].isna()]
modem_ids = set(pd.to_numeric(md_raw.iloc[:, 0], errors='coerce').dropna().astype(int))
print(f"  Modem IDs: {len(modem_ids):,}")

ap = pd.read_csv(SUBS_FILE, low_memory=False)
ap['account_id']      = pd.to_numeric(ap['account_id'], errors='coerce')
ap['appliance_group'] = ap['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
ap['consumer_type']   = ap['consumer_type'].fillna(CT_PRIV)
ap['gas_off']         = ap['gas_off'].notna()

bill_cols = {}
for eng, m in MONTH_ENG.items():
    col = f'{eng}_2025'
    if col in ap.columns:
        ap[col] = pd.to_numeric(ap[col], errors='coerce').fillna(0)
        bill_cols[m] = col

mc = [bill_cols[m] for m in MONTHS if m in bill_cols]
ap['annual_billing'] = ap[mc].sum(axis=1)
print(f"  Всього абонентів: {len(ap):,}")

# ── 2. "Чисті" споживачі (базовий фільтр) ────────────────────────────────────
print("\n=== 2. 'Чисті' споживачі (базовий фільтр) ===")

ap['max_month_bill']   = ap[mc].max(axis=1)
ap['n_months_nonzero'] = (ap[mc] > 0).sum(axis=1)
ap['max_month_frac']   = ap['max_month_bill'] / ap['annual_billing'].clip(lower=1)

clean_mask = (
    ~ap['account_id'].isin(modem_ids)
    & ~ap['gas_off']
    & ap['grs'].notna()
    & (ap['annual_billing'] > 0)
    & (ap['max_month_frac'] < 0.5)
    & (ap['n_months_nonzero'] >= 8)
)
clean_all = ap[clean_mask].copy()
print(f"  Після базового фільтру: {len(clean_all):,}  (з {len(ap):,} всього)")

# ── 3. Попередні профілі PER GROUP (еталон для кореляційного фільтру) ─────────
# Кожна (appliance_group, consumer_type) отримує власний еталон:
#   ОП-групи → зимово-піковий профіль
#   ПГ-групи → рівномірний профіль (~1/12)
# Так ПГ-абоненти не штрафуються за "неправильну" seasonality.
print("\n=== 3. Попередні профілі еталонів per-group ===")

MIN_PRELIM = 200   # мінімум абонентів у групі для власного еталону

group_prelim = {}  # (grp, ct) → np.array(12,)
for (grp, ct), sub in clean_all.groupby(['appliance_group', 'consumer_type']):
    if len(sub) < MIN_PRELIM:
        continue
    fracs = np.array([
        float((sub[bill_cols[m]] / sub['annual_billing'].clip(lower=1)).mean())
        for m in MONTHS
    ])
    total = fracs.sum()
    if total > 0:
        group_prelim[(grp, ct)] = fracs / total

# Глобальний fallback (якщо групи замало)
_gf = np.array([
    float((clean_all[bill_cols[m]] / clean_all['annual_billing'].clip(lower=1)).mean())
    for m in MONTHS
])
_global_fallback = _gf / _gf.sum()

print(f"  Груп з власним еталоном (≥{MIN_PRELIM}): {len(group_prelim)}")
for (grp, ct), ref in sorted(group_prelim.items()):
    print(f"  [{grp}, {ct[:18]}]  "
          + "  ".join(f"{UA[m]}:{ref[m-1]*100:.1f}%" for m in MONTHS))

# ── 2b. Кореляційний фільтр per-group ────────────────────────────────────────
print(f"\n=== 2b. Кореляційний фільтр per-group (поріг ≥ {MIN_CORR_PROFILE}) ===")
print("  Кожна група порівнюється з власним еталоном.")
print("  ПГ-абоненти більше не штрафуються за рівномірний профіль.\n")

def _pearson_corr_matrix(fracs_mat, ref_vec):
    """Vectorized Pearson corr of each row in fracs_mat with ref_vec."""
    ref_c   = ref_vec - ref_vec.mean()
    ref_std = np.sqrt((ref_c ** 2).sum())
    fracs_c = fracs_mat - fracs_mat.mean(axis=1, keepdims=True)
    std_f   = np.sqrt((fracs_c ** 2).sum(axis=1))
    dot     = fracs_c @ ref_c
    with np.errstate(invalid='ignore', divide='ignore'):
        return np.where(std_f > 0, dot / (std_f * ref_std), 0.0)

corr_series = pd.Series(np.nan, index=clean_all.index)

group_stats = []
for (grp, ct), sub in clean_all.groupby(['appliance_group', 'consumer_type']):
    ref_vec  = group_prelim.get((grp, ct), _global_fallback)
    fracs_m  = (
        sub[[bill_cols[m] for m in MONTHS]].values
        / sub['annual_billing'].clip(lower=1).values[:, np.newaxis]
    )
    corr_g = _pearson_corr_matrix(fracs_m, ref_vec)
    corr_series.loc[sub.index] = corr_g

    n_pass = (corr_g >= MIN_CORR_PROFILE).sum()
    group_stats.append({
        'grp': grp, 'ct': ct,
        'n_total': len(sub),
        'n_ideal': n_pass,
        'pct_ideal': n_pass / len(sub) * 100,
        'corr_avg': corr_g.mean(),
        'etalon': 'group' if (grp, ct) in group_prelim else 'global_fallback',
    })
    print(f"  [{grp:12s}, {ct[:18]:18s}]  "
          f"n={len(sub):>6,}  ідеальних={n_pass:>6,} ({n_pass/len(sub)*100:>5.1f}%)  "
          f"corr_avg={corr_g.mean():.3f}")

clean_all = clean_all.copy()
clean_all['corr_with_profile'] = corr_series

# Статистика по порогах (загалом)
corr_all = corr_series.values
print(f"\n  Розподіл кореляцій (всі групи разом):")
for thr in [0.60, 0.70, 0.80, 0.85, 0.90, 0.95]:
    n_thr  = (corr_all >= thr).sum()
    marker = "  ← поточний поріг" if thr == MIN_CORR_PROFILE else ""
    print(f"    corr ≥ {thr:.2f}: {n_thr:>7,}  ({n_thr/len(corr_all)*100:.1f}%){marker}")

# Застосовуємо поріг
ideal_all  = clean_all[clean_all['corr_with_profile'] >= MIN_CORR_PROFILE].copy()
n_rejected = len(clean_all) - len(ideal_all)
print(f"\n  'Ідеальних' абонентів: {len(ideal_all):,}  "
      f"(відфільтровано {n_rejected:,} / {n_rejected/len(clean_all)*100:.1f}%)")

# ── 3b. Фінальний глобальний профіль (з "ідеальних") ─────────────────────────
print("\n=== 3. Глобальний профіль білінгу (лише ідеальні абоненти) ===")

global_fracs = {}
for m in MONTHS:
    col = bill_cols.get(m)
    if col:
        global_fracs[m] = float(
            (ideal_all[col] / ideal_all['annual_billing'].clip(lower=1)).mean()
        )

total_g = sum(global_fracs.values())
global_profile = {m: v / total_g for m, v in global_fracs.items()}
_global_prelim = {m: float(_global_fallback[m-1]) for m in MONTHS}
print("  Базовий:   " + "  ".join(f"{UA[m]}:{_global_prelim[m]*100:.1f}%" for m in MONTHS))
print("  Ідеальний: " + "  ".join(f"{UA[m]}:{global_profile[m]*100:.1f}%" for m in MONTHS))
print("  Різниця:   " + "  ".join(
    f"{UA[m]}:{(global_profile[m]-_global_prelim[m])*100:+.1f}%"
    for m in MONTHS
))

# ── 4. GRS-специфічні профілі ─────────────────────────────────────────────────
print(f"\n=== 4. GRS-специфічні профілі (MIN_CLEAN={MIN_CLEAN_GRS}, corr≥{MIN_CORR_PROFILE}) ===")

all_grs_in_clean = ideal_all['grs'].unique()
grs_profiles     = {}   # grs → {month: fraction}
grs_clean_counts = {}   # grs → n_ideal

for grs in all_grs_in_clean:
    sub = ideal_all[ideal_all['grs'] == grs]
    n   = len(sub)
    grs_clean_counts[grs] = n
    if n < MIN_CLEAN_GRS:
        continue

    fracs = {}
    for m in MONTHS:
        col = bill_cols.get(m)
        if col:
            fracs[m] = float(
                (sub[col] / sub['annual_billing'].clip(lower=1)).mean()
            )

    total = sum(fracs.values())
    if total <= 0:
        continue

    grs_profiles[grs] = {m: v / total for m, v in fracs.items()}

n_all_grs   = len(ideal_all['grs'].unique())
n_with_prof = len(grs_profiles)
print(f"  ГРС з профілем: {n_with_prof} / {n_all_grs}  "
      f"({n_all_grs - n_with_prof} ГРС нижче порогу)")

# ── 5. Відхилення GRS-профілів від глобального ────────────────────────────────
print("\n=== 5. Відхилення GRS-профілів від глобального ===")

deviations = {}   # grs → {month: ratio}  (1.0 = рівно як global)
for grs, prof in grs_profiles.items():
    deviations[grs] = {}
    for m in MONTHS:
        g = global_profile.get(m, 1/12)
        deviations[grs][m] = (prof.get(m, 0) / g) if g > 0 else 1.0

# Найбільші відхилення
max_dev_ratio = max(d for dev in deviations.values() for d in dev.values())
min_dev_ratio = min(d for dev in deviations.values() for d in dev.values())

n_large_dev = sum(
    1 for dev in deviations.values()
    if any(abs(d - 1.0) > 0.20 for d in dev.values())
)
print(f"  Макс. відхилення (будь-який місяць): "
      f"+{(max_dev_ratio-1)*100:.1f}% / {(min_dev_ratio-1)*100:.1f}%")
print(f"  ГРС з відхиленням > ±20%: {n_large_dev}")

# Місяці з найбільшою варіативністю між ГРС
print("\n  Std відхилень по місяцях (між ГРС):")
for m in MONTHS:
    vals = [deviations[grs][m] for grs in deviations]
    print(f"    {UA[m]}: std={np.std(vals)*100:.1f}%  "
          f"min={min(vals)*100:.1f}%  max={max(vals)*100:.1f}%")

# ── 6. Поточний bias per GRS (з annual_decompose_result.xlsx) ─────────────────
print("\n=== 6. Поточний bias per GRS ===")

result_df = pd.read_excel(RESULT_FILE, sheet_name='All_GRS_monthly')
# Колонки: grs, billing, month, pred, modem_fact, total_pred, error, bias_pct, is_large

# Сума non-modem annual billing per GRS (фільтр = той самий що в annual_decompose.py)
_mc2 = [bill_cols[m] for m in MONTHS if m in bill_cols]
nm_df = ap[
    ~ap['account_id'].isin(modem_ids)
    & ap['grs'].notna()
    & (ap[_mc2].abs().sum(axis=1) > 0)
].copy()
nm_annual_grs = nm_df.groupby('grs')['annual_billing'].sum()

# ── 7. Порівняння global vs GRS-specific декомпозиту ──────────────────────────
print("\n=== 7. Порівняння global vs GRS-specific декомпозиту ===")

comp_rows = []
for _, row in result_df.iterrows():
    grs = row['grs']
    m   = int(row['month'])

    # GRS-specific: GRS-профіль якщо є, інакше global
    prof      = grs_profiles.get(grs, global_profile)
    nm_annual = float(nm_annual_grs.get(grs, 0))
    w         = prof.get(m, global_profile.get(m, 1/12))
    decompose_grs_val = nm_annual * w + float(row['modem_fact'])

    comp_rows.append({
        'grs':              grs,
        'month':            m,
        'billing':          float(row['billing']),
        'decompose_global': float(row['total_pred']),
        'modem_fact':       float(row['modem_fact']),
        'decompose_grs':    decompose_grs_val,
        'has_grs_profile':  grs in grs_profiles,
        'is_large':         bool(row.get('is_large', False)),
    })

comp_df = pd.DataFrame(comp_rows)
b_clip = comp_df['billing'].clip(lower=1)
comp_df['bias_global_pct'] = (comp_df['decompose_global'] - comp_df['billing']) / b_clip * 100
comp_df['bias_grs_pct']    = (comp_df['decompose_grs']    - comp_df['billing']) / b_clip * 100

# Per-GRS summary — використовуємо місячне MAE (не річний bias, який = 0 по конструкції)
grs_summary_rows = []
for grs, sub in comp_df.groupby('grs'):
    b = sub['billing'].sum()
    if b <= 0:
        continue

    has_p  = bool(sub['has_grs_profile'].iloc[0])
    is_lg  = bool(sub['is_large'].iloc[0])

    # Місячне MAE |bias%|: середнє з 12 місяців |pred-bill|/bill
    sub_v = sub[sub['billing'] > 0]
    mae_g = sub_v['bias_global_pct'].abs().mean() if len(sub_v) > 0 else None
    mae_r = sub_v['bias_grs_pct'].abs().mean()    if len(sub_v) > 0 else None
    delta_mae = (round(mae_g - mae_r, 2) if mae_g is not None and mae_r is not None else None)

    max_dev = (
        max(abs(deviations[grs][m] - 1.0) for m in MONTHS) * 100
        if grs in deviations else None
    )
    modem_pct = sub['modem_fact'].sum() / b * 100

    row_out = {
        'grs':             grs,
        'ann_bill_M':      round(b / 1e6, 3),
        'is_large':        is_lg,
        'n_clean':         grs_clean_counts.get(grs, 0),
        'has_grs_profile': has_p,
        'modem_%':         round(modem_pct, 1),
        'mae_global%':     round(mae_g, 2) if mae_g is not None else None,
        'mae_grs%':        round(mae_r, 2) if mae_r is not None else None,
        'delta_mae_pp':    delta_mae,   # >0 = місячне MAE покращилось
        'max_dev%':        round(max_dev, 1) if max_dev is not None else None,
    }
    # Monthly bias (global) for reference
    for m in MONTHS:
        sub_m = sub[sub['month'] == m]
        row_out[UA[m]] = (
            round(float(sub_m['bias_global_pct'].values[0]), 1)
            if len(sub_m) == 1 else None
        )
    grs_summary_rows.append(row_out)

grs_sum_df = pd.DataFrame(grs_summary_rows).sort_values('ann_bill_M', ascending=False)

# ── 8. Кореляційний аналіз ────────────────────────────────────────────────────
print("\n=== 8. Кореляційний аналіз ===")

valid_corr = grs_sum_df.dropna(subset=['max_dev%', 'mae_global%'])
valid_corr = valid_corr[valid_corr['has_grs_profile']]

if len(valid_corr) >= 5:
    corr_val = np.corrcoef(valid_corr['max_dev%'], valid_corr['mae_global%'])[0, 1]
    print(f"  Корел. (max_profile_dev vs місячне_MAE_global%): {corr_val:+.3f}  (n={len(valid_corr)})")
else:
    corr_val = None
    print(f"  Недостатньо ГРС з профілями для кореляції (n={len(valid_corr)})")

# ── 9. Консольний звіт ────────────────────────────────────────────────────────
print("\n" + "="*65)
print("=== Підсумок: Аналіз GRS-профілів ===")
print(f"  ГРС з достатньою кількістю 'чистих' абонентів (≥{MIN_CLEAN_GRS}): "
      f"{n_with_prof} / {n_all_grs}")

print(f"\n  Відхилення GRS-профілів від глобального:")
print(f"    Макс. відхилення (будь-який місяць): "
      f"+{(max_dev_ratio-1)*100:.1f}% / {(min_dev_ratio-1)*100:.1f}%")
print(f"    ГРС з відхиленням > ±20%: {n_large_dev} ГРС")

if corr_val is not None:
    print(f"\n  Корел. (max_profile_dev vs місячне_MAE_global%): {corr_val:+.3f}")

# Топ-10 покращень від GRS-профілів (за зниженням місячного MAE)
has_p_df = grs_sum_df[grs_sum_df['has_grs_profile']].dropna(subset=['delta_mae_pp']).copy()
top_improv = has_p_df.sort_values('delta_mae_pp', ascending=False).head(10)
print(f"\n  Покращення місячного MAE при GRS-профілях (топ-10):")
for _, r in top_improv.iterrows():
    print(f"    {str(r['grs'])[:40]:40s}: "
          f"MAE_global={r['mae_global%']:.1f}% → "
          f"MAE_grs={r['mae_grs%']:.1f}%  "
          f"(Δ={r['delta_mae_pp']:+.2f}pp)")

# MAE загальне (місячний рівень)
valid_all = comp_df[comp_df['billing'] > 0]
mae_global = valid_all['bias_global_pct'].abs().mean()
mae_grs    = valid_all['bias_grs_pct'].abs().mean()
print(f"\n  Місячне MAE |bias%|: global={mae_global:.2f}%  grs={mae_grs:.2f}%")

valid_hp = valid_all[valid_all['has_grs_profile']]
if len(valid_hp) > 0:
    mae_g_hp = valid_hp['bias_global_pct'].abs().mean()
    mae_r_hp = valid_hp['bias_grs_pct'].abs().mean()
    delta_mae = mae_g_hp - mae_r_hp
    verdict = "ПОКРАЩЕННЯ" if delta_mae > 0 else "ПОГІРШЕННЯ"
    print(f"  (тільки ГРС з профілями): "
          f"global={mae_g_hp:.2f}%  grs={mae_r_hp:.2f}%  "
          f"→ {verdict} {abs(delta_mae):.2f}pp")

print()
print("  ВИСНОВОК:")
if mae_grs < mae_global:
    print("  ✓ GRS-специфічні профілі ПОКРАЩУЮТЬ місячний декомпозит")
    print("    Гіпотеза підтверджена: різні ГРС мають різні сезонні профілі білінгу")
else:
    print("  ✗ GRS-специфічні профілі НЕ ПОКРАЩУЮТЬ місячний декомпозит")
    print("    Гіпотеза не підтверджена: per-GRS bias обумовлений іншими факторами,")
    print("    не різницею у сезонних профілях чистих абонентів.")
    print("    Ймовірні причини: нерівномірні перерозрахунки, квартальні платники,")
    print("    особливості дати прийому на обслуговування.")
print("="*65)

# ── 10. Збереження Excel ───────────────────────────────────────────────────────
print(f"\n=== 10. Збереження → {OUT_XLSX} ===")

# Лист 2: Profile_heatmap — рядки=ГРС, стовпці=місяці, значення=deviation ratio
heat_rows = []
for grs in sorted(deviations.keys()):
    r = {'grs': grs, 'n_clean': grs_clean_counts.get(grs, 0)}
    for m in MONTHS:
        r[UA[m]] = round(deviations[grs].get(m, 1.0), 3)
    # Додаємо max_dev та bias_global для довідки
    r['max_dev%']    = round(max(abs(deviations[grs][m] - 1.0) for m in MONTHS) * 100, 1)
    r['mae_global%'] = grs_sum_df.set_index('grs')['mae_global%'].get(grs)
    heat_rows.append(r)
heat_df = pd.DataFrame(heat_rows)

# Лист 3: Comparison_top15 — топ-15 ГРС за місячним MAE (global)
comp_top15_grs = (
    grs_sum_df.dropna(subset=['mae_global%'])
    .nlargest(15, 'mae_global%')
)

# Лист 4: Profile_weights — місячний профіль кожного ГРС (%) vs global
pw_rows = []
for m in MONTHS:
    row_pw = {'month': UA[m], 'global_%': round(global_profile[m] * 100, 2)}
    for grs, prof in grs_profiles.items():
        row_pw[grs] = round(prof.get(m, 0) * 100, 2)
    pw_rows.append(row_pw)
pw_df = pd.DataFrame(pw_rows)

with pd.ExcelWriter(OUT_XLSX, engine='openpyxl') as xw:
    # Лист 1: GRS_summary
    grs_sum_df.to_excel(xw, sheet_name='GRS_summary', index=False)

    # Лист 2: Profile_heatmap
    heat_df.to_excel(xw, sheet_name='Profile_heatmap', index=False)
    ws_heat = xw.book['Profile_heatmap']
    # Кольорова схема: > 1.2 = червоний, < 0.8 = синій
    month_col_indices = list(range(3, 3 + 12))   # cols C..N (1-based = 3..14)
    for row in ws_heat.iter_rows(min_row=2, min_col=3, max_col=14):
        for cell in row:
            try:
                v = float(cell.value)
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='center')
                if v >= 1.2:
                    cell.fill = PatternFill('solid', fgColor='FF8888')
                    cell.font = Font(bold=True)
                elif v >= 1.1:
                    cell.fill = PatternFill('solid', fgColor='FFCCCC')
                elif v <= 0.80:
                    cell.fill = PatternFill('solid', fgColor='8888FF')
                    cell.font = Font(bold=True, color='FFFFFF')
                elif v <= 0.90:
                    cell.fill = PatternFill('solid', fgColor='CCCCFF')
            except (TypeError, ValueError):
                pass

    # Лист 3: Comparison_top15
    comp_top15_grs.to_excel(xw, sheet_name='Comparison_top15', index=False)

    # Лист 4: Profile_weights
    pw_df.to_excel(xw, sheet_name='Profile_weights', index=False)

print(f"  DONE → {OUT_XLSX}")
print(f"  Листи: GRS_summary | Profile_heatmap | Comparison_top15 | Profile_weights")
