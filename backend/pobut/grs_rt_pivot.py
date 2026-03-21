"""
grs_rt_pivot.py

RT предикт: реалтайм прогноз для всіх не-модемних споживачів
на основі поточних модемних даних.

ПРИНЦИП:
  - Модемні (~1200 абон.)  : факт-дані → будуємо RT патерни (rate_sum, heat_m2_sum)
  - Не-модемні (~219K абон.): pbl = summer_baseline з білінгу (де є, ≈92%)
                               інакше → групова медіана з модемних (fallback 8%)
  - rate_sum та heat_m2_sum — виключно з модемних даних (реалтайм)
  - Ніякого порівняння з білінговими ОБСЯГАМИ

Запуск:
    cd backend/pobut
    python grs_rt_pivot.py

Вихід:
    data/grs_rt_pivot.xlsx
      "Предикт_RT"   — пивот GRS × Місяць (м³), всі доступні місяці
      "GRS_місяць"   — довгий формат (grs, рік, міс, факт, предикт, разом)
"""
import sys
import io
import os
import calendar

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pobut_predictor import PobUtPredictor, CT_PRIV

# ─── Config ───────────────────────────────────────────────────────────────────
MODEM_FILE = 'data/input/profile_pobut_daily_result.csv'
SUBS_FILE  = 'data/input/all_pobut_enriched.csv'
OUT_EXCEL  = 'data/grs_rt_pivot.xlsx'

MONTH_UA = {
    1: 'Січ', 2: 'Лют',  3: 'Бер', 4: 'Кві',
    5: 'Тра', 6: 'Чер',  7: 'Лип', 8: 'Сер',
    9: 'Вер', 10: 'Жов', 11: 'Лис', 12: 'Гру',
}


# ─── Subclass: uses individual summer_baseline as pbl ─────────────────────────
class PobUtPredictorSB(PobUtPredictor):
    """
    PobUtPredictor with individual summer_baseline as non-modem pbl.

    For each non-modem consumer where summer_baseline > 0, uses it as the
    individual baseline level instead of the group median.
    Rate patterns (rate_sum, heat_m2_sum) are still purely from modem data.
    """

    def _prepare_non_modem(self):
        super()._prepare_non_modem()

        # summer_baseline is the billing-derived average summer consumption
        # for each consumer. Use it as individual pbl where valid.
        sb_map = (
            self._all_pobut
            .assign(summer_baseline=lambda d: pd.to_numeric(
                d['summer_baseline'], errors='coerce'
            ))
            .set_index('account_id')['summer_baseline']
            .to_dict()
        )

        nm = self._non_modem
        nm['_sb'] = nm['account_id'].map(sb_map)
        valid = nm['_sb'].notna() & (nm['_sb'] > 0)
        nm.loc[valid, 'pbl'] = nm.loc[valid, '_sb']
        nm.drop(columns=['_sb'], inplace=True)

        n_ind = int(valid.sum())
        n_grp = int((~valid).sum())
        print(f"  pbl source: individual summer_baseline={n_ind:,}  "
              f"group_median_fallback={n_grp:,}")


# ─── Run prediction ────────────────────────────────────────────────────────────
print("=" * 60)
print("RT Предикт з individual summer_baseline")
print("=" * 60)

p = PobUtPredictorSB(MODEM_FILE, SUBS_FILE, mode='offline', summer_op_vpg=False)
df_all = p.predict()          # all available periods (no year filter)

periods_in_data = sorted(set(zip(df_all['year'], df_all['month'])))
first_p, last_p = periods_in_data[0], periods_in_data[-1]
print(f"\nПеріоди в даних: "
      f"{first_p[0]}-{first_p[1]:02d} … {last_p[0]}-{last_p[1]:02d} "
      f"({len(periods_in_data)} місяців)")
print(f"Рядків у combined: {len(df_all):,}")

# ─── Aggregate by GRS × year × month ─────────────────────────────────────────
agg = (
    df_all
    .groupby(['grs', 'year', 'month', 'type'])['volume']
    .sum()
    .reset_index()
)

pivot_long = agg.pivot_table(
    index=['grs', 'year', 'month'],
    columns='type',
    values='volume',
    aggfunc='sum',
    fill_value=0,
).reset_index()
pivot_long.columns.name = None

pivot_long['fact']  = pivot_long.get('fact',  pd.Series(0, index=pivot_long.index))
pivot_long['pred']  = pivot_long.get('pred',  pd.Series(0, index=pivot_long.index))
pivot_long['total'] = pivot_long['fact'] + pivot_long['pred']

# Period labels: Жов-24, Лис-24, ..., Гру-25, Січ-26
pivot_long['period'] = pivot_long.apply(
    lambda r: f"{MONTH_UA[int(r['month'])]}-{str(int(r['year']))[2:]}", axis=1
)

# Sorted period list (chronological)
period_order = (
    pivot_long[['year', 'month', 'period']]
    .drop_duplicates()
    .sort_values(['year', 'month'])
    ['period'].tolist()
)

# ─── Console summary ──────────────────────────────────────────────────────────
by_month = (
    pivot_long
    .groupby(['year', 'month', 'period'])[['fact', 'pred', 'total']]
    .sum()
    .reset_index()
    .sort_values(['year', 'month'])
)
by_month['fact_share%'] = (
    by_month['fact'] / by_month['total'].replace(0, np.nan) * 100
).round(1)

print(f"\n{'Місяць':>8}  {'Факт-модем':>12}  {'Предикт':>12}  "
      f"{'Разом':>12}  {'Факт%':>6}")
print("-" * 60)
for _, r in by_month.iterrows():
    print(f"{r['period']:>8}  {r['fact']:>12,.0f}  {r['pred']:>12,.0f}  "
          f"{r['total']:>12,.0f}  {r['fact_share%']:>5.1f}%")
print("-" * 60)
tot_f = by_month['fact'].sum()
tot_p = by_month['pred'].sum()
tot_t = by_month['total'].sum()
print(f"{'РАЗОМ':>8}  {tot_f:>12,.0f}  {tot_p:>12,.0f}  "
      f"{tot_t:>12,.0f}  {tot_f/tot_t*100:>5.1f}%")

# ─── Pivot table: GRS × period ────────────────────────────────────────────────
grs_pivot = pivot_long.pivot_table(
    index='grs',
    columns='period',
    values='total',
    aggfunc='sum',
    fill_value=0,
)[period_order].copy()

# Sort GRS by total descending
grs_pivot['_sort'] = grs_pivot.sum(axis=1)
grs_pivot = grs_pivot.sort_values('_sort', ascending=False).drop(columns=['_sort'])
grs_pivot = grs_pivot.round(0).astype(int)

# РАЗОМ row
total_row = pd.DataFrame(
    [grs_pivot.sum()],
    index=pd.Index(['РАЗОМ'], name='grs'),
)
grs_pivot = pd.concat([grs_pivot, total_row])
grs_pivot['РАЗОМ'] = grs_pivot[period_order].sum(axis=1)

# ─── Excel output with formatting ─────────────────────────────────────────────
print(f"\nЗберігаємо {OUT_EXCEL} ...")

# ── Sheet helpers ──────────────────────────────────────────────────────────────
COLOR_HEADER = 'FF1F4E79'   # dark blue
COLOR_TOTAL  = 'FFD6E4F0'   # light blue
COLOR_GRS    = 'FFE9EFF7'   # very light blue for GRS column
FONT_WHITE   = Font(name='Calibri', bold=True, color='FFFFFFFF', size=10)
FONT_BOLD    = Font(name='Calibri', bold=True, size=10)
FONT_NORMAL  = Font(name='Calibri', size=10)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal='right',  vertical='center')
ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center')

THIN = Side(style='thin', color='FFC0C0C0')
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_THICK_BOTTOM = Border(
    left=THIN, right=THIN,
    bottom=Side(style='medium', color='FF1F4E79'),
    top=THIN,
)


def fmt_num(ws, row, col, value, bold=False):
    cell = ws.cell(row=row, column=col, value=int(value))
    cell.number_format = '#,##0'
    cell.font = FONT_BOLD if bold else FONT_NORMAL
    cell.alignment = ALIGN_RIGHT
    cell.border = BORDER_THIN
    return cell


with pd.ExcelWriter(OUT_EXCEL, engine='openpyxl') as xw:

    # ══ Sheet 1: Pivot GRS × Month ════════════════════════════════════════════
    ws = xw.book.create_sheet('Предикт_RT')

    all_cols  = ['ГРС'] + period_order + ['РАЗОМ']
    n_periods = len(period_order)
    n_grs     = len(grs_pivot)

    # Header row
    for c_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font      = FONT_WHITE
        cell.fill      = PatternFill('solid', fgColor=COLOR_HEADER)
        cell.alignment = ALIGN_CENTER
        cell.border    = BORDER_THICK_BOTTOM

    # Data rows
    for r_idx, (grs_name, row_data) in enumerate(grs_pivot.iterrows(), start=2):
        is_total_row = (grs_name == 'РАЗОМ')
        row_fill = PatternFill('solid', fgColor=COLOR_TOTAL) if is_total_row else None

        # GRS name cell
        grs_cell = ws.cell(row=r_idx, column=1, value=str(grs_name))
        grs_cell.font      = FONT_BOLD if is_total_row else FONT_NORMAL
        grs_cell.fill      = (row_fill if is_total_row
                               else PatternFill('solid', fgColor=COLOR_GRS))
        grs_cell.alignment = ALIGN_LEFT
        grs_cell.border    = BORDER_THIN

        # Period + РАЗОМ value cells
        for c_idx, col_name in enumerate(period_order + ['РАЗОМ'], start=2):
            val  = row_data.get(col_name, 0)
            cell = fmt_num(ws, r_idx, c_idx, val, bold=is_total_row)
            if is_total_row:
                cell.fill = row_fill

    # Column widths
    ws.column_dimensions['A'].width = 42
    for c_idx in range(2, len(all_cols) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 9
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'B2'

    # ══ Sheet 2: Long format GRS × month ══════════════════════════════════════
    detail = (
        pivot_long[['grs', 'year', 'month', 'period', 'fact', 'pred', 'total']]
        .sort_values(['grs', 'year', 'month'])
        .copy()
    )
    detail.columns = ['ГРС', 'Рік', 'Місяць', 'Період',
                      'Факт_модем', 'Предикт_RT', 'Разом_RT']
    detail.to_excel(xw, sheet_name='GRS_місяць', index=False)

    # ══ Remove default Sheet ══════════════════════════════════════════════════
    if 'Sheet' in xw.book.sheetnames:
        del xw.book['Sheet']

print(f"[OK] {OUT_EXCEL}")
print(f"\nГРС у пивоті: {len(grs_pivot) - 1}")  # -1 for РАЗОМ row
print(f"Місяців: {len(period_order)}  ({period_order[0]} … {period_order[-1]})")
print("\n=== DONE ===")
