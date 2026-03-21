"""
analyze_billing_profiles.py

Кластеризація за формою місячного профілю (без прив'язки до груп ОП/ПГ/ВПГ).

Для кожного абонента будується 12D вектор:
    p_m = billing_m / annual_billing   (для не-модемних)
    p_m = monthly_sum_m / annual_sum_m (для модемних)

Потім KMeans (k=2..8, силует) → порівняння кластерів з існуючими групами.
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
from sklearn.decomposition import PCA
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from pobut_predictor import GROUP_MERGE, META_COLS, CT_PRIV

SUBS_FILE  = 'data/input/all_pobut_enriched.csv'
MODEM_FILE = 'data/input/profile_pobut_daily_result.csv'
OUT_XLSX   = 'data/billing_profile_clusters.xlsx'

UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}
MONTHS = list(range(1, 13))
MONTH_ENG = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
             'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

# Фільтри "чистих" абонентів
MIN_NONZERO  = 10    # мінімум ненульових місяців з 12
MAX_SPIKE    = 0.25  # макс. частка одного місяця (<25% — виключає квартальних)
ALLOW_NEG    = False # виключаємо перерахунки (від'ємні місяці)
K_RANGE      = range(2, 9)

# ── 1. Білінгові дані (не-модемні) ────────────────────────────────────────────
print("=== 1. Завантаження білінгових даних ===")

md_raw = pd.read_csv(MODEM_FILE, sep=';', low_memory=False)
md_raw.columns = META_COLS + list(md_raw.columns[len(META_COLS):])
md_raw = md_raw[md_raw['gas_off'].isna() & md_raw['alternative'].isna() & md_raw['dacha'].isna()]
modem_ids = set(pd.to_numeric(md_raw.iloc[:, 0], errors='coerce').dropna().astype(int))
print(f"  Модемних абонентів: {len(modem_ids):,}")

ap = pd.read_csv(SUBS_FILE, low_memory=False)
ap['account_id']      = pd.to_numeric(ap['account_id'], errors='coerce')
ap['appliance_group'] = ap['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
ap['consumer_type']   = ap['consumer_type'].fillna(CT_PRIV)

bill_cols = {}
for eng, m in MONTH_ENG.items():
    col = f'{eng}_2025'
    if col in ap.columns:
        ap[col] = pd.to_numeric(ap[col], errors='coerce').fillna(0)
        bill_cols[m] = col

mc = [bill_cols[m] for m in MONTHS if m in bill_cols]
ap['annual'] = ap[mc].sum(axis=1)

# Не-модемні з достатнім білінгом
nm = ap[
    ~ap['account_id'].isin(modem_ids)
    & (ap['annual'] > 0)
].copy()

# 12D профіль
prof_cols = [f'p_{m}' for m in MONTHS]
for m in MONTHS:
    nm[f'p_{m}'] = nm[bill_cols[m]] / nm['annual']

nm['n_nonzero']  = (nm[mc] > 0).sum(axis=1)
nm['n_negative'] = (nm[mc] < 0).sum(axis=1)
nm['max_frac']   = nm[prof_cols].max(axis=1)

# Фільтр чистих абонентів
clean_mask = (
    (nm['n_nonzero'] >= MIN_NONZERO)
    & (nm['max_frac'] < MAX_SPIKE)
    & nm['grs'].notna()
)
if not ALLOW_NEG:
    clean_mask &= (nm['n_negative'] == 0)

nm_clean = nm[clean_mask].copy()

# Статистика відсіювання
print(f"  Не-модемних всього: {len(nm):,}")
print(f"  Відсіяно — мало місяців (<{MIN_NONZERO}):     {(nm['n_nonzero'] < MIN_NONZERO).sum():>7,}")
print(f"  Відсіяно — стрибок місяця (≥{MAX_SPIKE:.0%}):   {(nm['n_nonzero'] >= MIN_NONZERO) & (nm['max_frac'] >= MAX_SPIKE) & (nm['n_negative'] == 0):}")

n_few   = (nm['n_nonzero'] < MIN_NONZERO).sum()
n_spike = ((nm['n_nonzero'] >= MIN_NONZERO) & (nm['max_frac'] >= MAX_SPIKE)).sum()
n_neg   = ((nm['n_nonzero'] >= MIN_NONZERO) & (nm['max_frac'] < MAX_SPIKE) & (nm['n_negative'] > 0)).sum()
print(f"  Відсіяно — мало місяців (<{MIN_NONZERO}):   {n_few:>7,}  ({n_few/len(nm)*100:.1f}%)")
print(f"  Відсіяно — стрибок (max≥{MAX_SPIKE:.0%}):     {n_spike:>7,}  ({n_spike/len(nm)*100:.1f}%)")
if not ALLOW_NEG:
    print(f"  Відсіяно — перерахунки (neg):      {n_neg:>7,}  ({n_neg/len(nm)*100:.1f}%)")
print(f"  Чистих залишилось:                 {len(nm_clean):>7,}  ({len(nm_clean)/len(nm)*100:.1f}%)")

# ── 2. Модемні дані (місячна агрегація) ───────────────────────────────────────
print("\n=== 2. Модемні дані — місячна агрегація ===")

date_cols = []
for c in md_raw.columns:
    if isinstance(c, str) and c.count('.') == 2 and len(c) == 10:
        try:
            dd, mm, yy = c.split('.')
            dt = pd.Timestamp(f'{yy}-{mm}-{dd}')
            if dt.year == 2025:
                date_cols.append((dt, c))
        except Exception:
            pass
date_cols.sort()

md_raw['account_id'] = pd.to_numeric(md_raw['account_id'], errors='coerce')
md_raw['appliance_group'] = md_raw['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
md_raw['consumer_type']   = md_raw['consumer_type'].fillna(CT_PRIV)

md_num = (md_raw[[c for _, c in date_cols]]
          .apply(pd.to_numeric, errors='coerce').fillna(0).clip(lower=0))
md_num.index = md_raw['account_id'].values

# Місячні суми
modem_monthly = {}
for dt, col in date_cols:
    m = dt.month
    if m not in modem_monthly:
        modem_monthly[m] = np.zeros(len(md_raw))
    modem_monthly[m] += md_num[col].values

modem_df = pd.DataFrame(modem_monthly, index=md_raw['account_id'].values)
modem_df.columns.name = 'month'
modem_df['annual'] = modem_df[list(modem_monthly.keys())].sum(axis=1)
modem_df = modem_df[modem_df['annual'] > 0].copy()

# 12D профіль модемних
for m in MONTHS:
    if m in modem_df.columns:
        modem_df[f'p_{m}'] = modem_df[m] / modem_df['annual']
    else:
        modem_df[f'p_{m}'] = 0.0

modem_df['n_nonzero'] = (modem_df[list(modem_monthly.keys())] > 0).sum(axis=1)
modem_df['max_frac']  = modem_df[prof_cols].max(axis=1)
modem_df = modem_df.reset_index().rename(columns={'index': 'account_id'})

# Додаємо метадані
md_meta = md_raw[['account_id', 'appliance_group', 'consumer_type']].copy()
modem_df = modem_df.merge(md_meta, on='account_id', how='left')
modem_df['grs'] = modem_df['account_id'].map(
    ap.set_index('account_id')['grs'].to_dict()
)

modem_clean = modem_df[
    (modem_df['n_nonzero'] >= MIN_NONZERO)
    & (modem_df['max_frac'] < MAX_SPIKE)
].copy()

print(f"  Модемних абонентів: {len(modem_df):,}")
print(f"  Після фільтру: {len(modem_clean):,}")

# ── 3. KMeans кластеризація ────────────────────────────────────────────────────
print("\n=== 3. KMeans по формі профілю (12D) ===")

def best_kmeans(X, k_range=K_RANGE, n_init=20, seed=42):
    """Підбір k за силуетом. Повертає (k, sil, labels)."""
    best_k, best_sil, best_labels = 2, -1.0, None
    results = []
    for k in k_range:
        km = KMeans(n_clusters=k, n_init=n_init, random_state=seed)
        lbl = km.fit_predict(X)
        sizes = np.bincount(lbl)
        if min(sizes) < max(5, int(0.02 * len(X))):
            sil = -1.0
        elif len(X) < 20:
            sil = -1.0
        else:
            try:
                sil = float(silhouette_score(X, lbl, sample_size=min(5000, len(X))))
            except Exception:
                sil = -1.0
        results.append({'k': k, 'sil': round(sil, 4),
                        'sizes': sorted(sizes.tolist(), reverse=True)})
        if sil > best_sil:
            best_sil, best_k, best_labels = sil, k, lbl.copy()
    return best_k, best_sil, best_labels, results

# 3a. Білінгові
X_nm = nm_clean[prof_cols].values
print(f"\n  [Білінг, n={len(X_nm)}]")
k_nm, sil_nm, lbl_nm, res_nm = best_kmeans(X_nm)
nm_clean = nm_clean.copy()
nm_clean['cluster'] = lbl_nm
print(f"  Оптимальний k={k_nm}, силует={sil_nm:.4f}")
for r in res_nm:
    print(f"    k={r['k']}: sil={r['sil']:.4f}  sizes={r['sizes']}")

# 3b. Модемні
X_md = modem_clean[prof_cols].values
print(f"\n  [Модеми, n={len(X_md)}]")
k_md, sil_md, lbl_md, res_md = best_kmeans(X_md)
modem_clean = modem_clean.copy()
modem_clean['cluster'] = lbl_md
print(f"  Оптимальний k={k_md}, силует={sil_md:.4f}")
for r in res_md:
    print(f"    k={r['k']}: sil={r['sil']:.4f}  sizes={r['sizes']}")

# ── 4. Аналіз кластерів ───────────────────────────────────────────────────────
print("\n=== 4. Профілі кластерів ===")

def cluster_profiles(df, k, source_label):
    """Середні профілі + склад по групах."""
    rows = []
    for c in range(k):
        sub = df[df['cluster'] == c]
        prof = {UA[m]: round(float(sub[f'p_{m}'].mean() * 100), 2) for m in MONTHS}
        grp_counts = sub['appliance_group'].value_counts()
        ct_counts  = sub['consumer_type'].value_counts()
        peak_m = max(MONTHS, key=lambda m: sub[f'p_{m}'].mean())

        # Сезонні агрегати
        heat_pct   = sum(sub[f'p_{m}'].mean() for m in [1,2,3,4,10,11,12]) * 100
        summer_pct = sum(sub[f'p_{m}'].mean() for m in [6,7,8]) * 100

        rows.append({
            'Джерело':   source_label,
            'Кластер':   c,
            'N':         len(sub),
            'N%':        round(len(sub) / len(df) * 100, 1),
            'Пік_міс':   UA[peak_m],
            'Опал_%':    round(heat_pct, 1),
            'Літо_%':    round(summer_pct, 1),
            **prof,
            'Топ_група': grp_counts.index[0] if len(grp_counts) else '?',
            'Топ_тип':   ct_counts.index[0]  if len(ct_counts)  else '?',
        })
    return pd.DataFrame(rows)

prof_nm = cluster_profiles(nm_clean, k_nm, 'Білінг')
prof_md = cluster_profiles(modem_clean, k_md, 'Модеми')

print("\n[Білінг — середні профілі кластерів, %]")
print(f"  {'Кл':>3}  {'N':>6}  {'Опал%':>6}  {'Літо%':>6}  " +
      "  ".join(f"{UA[m]:>5}" for m in MONTHS))
print("  " + "─" * 90)
for _, r in prof_nm.iterrows():
    print(f"  {r['Кластер']:>3}  {r['N']:>6,}  {r['Опал_%']:>6.1f}  {r['Літо_%']:>6.1f}  " +
          "  ".join(f"{r[UA[m]]:>5.1f}" for m in MONTHS) +
          f"  | {r['Топ_група']}")

print("\n[Модеми — середні профілі кластерів, %]")
print(f"  {'Кл':>3}  {'N':>6}  {'Опал%':>6}  {'Літо%':>6}  " +
      "  ".join(f"{UA[m]:>5}" for m in MONTHS))
print("  " + "─" * 90)
for _, r in prof_md.iterrows():
    print(f"  {r['Кластер']:>3}  {r['N']:>6,}  {r['Опал_%']:>6.1f}  {r['Літо_%']:>6.1f}  " +
          "  ".join(f"{r[UA[m]]:>5.1f}" for m in MONTHS) +
          f"  | {r['Топ_група']}")

# ── 5. Склад кластерів по групах ─────────────────────────────────────────────
print("\n=== 5. Склад кластерів по appliance_group ===")

def cluster_composition(df, k, label):
    print(f"\n  [{label}]")
    grp_order = df['appliance_group'].value_counts().index.tolist()
    header = f"  {'Кл':>3}  {'N':>6} | " + " | ".join(f"{g[:18]:>18}" for g in grp_order)
    print(header)
    print("  " + "─" * len(header))
    for c in range(k):
        sub = df[df['cluster'] == c]
        vals = " | ".join(
            f"{sub[sub['appliance_group']==g].shape[0]:>14} ({sub[sub['appliance_group']==g].shape[0]/len(sub)*100:>4.1f}%)"
            for g in grp_order
        )
        print(f"  {c:>3}  {len(sub):>6,} | {vals}")

cluster_composition(nm_clean, k_nm, 'Білінг')
cluster_composition(modem_clean, k_md, 'Модеми')

# ── 6. Excel ─────────────────────────────────────────────────────────────────
print(f"\n=== 6. Збереження → {OUT_XLSX} ===")

HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', size=9)
CLUSTER_COLORS = ['FFF2CC','DDEBF7','E2EFDA','FCE4D6','EAD1DC','D9D2E9','D0E0E3','F4CCCC']

def write_profile_sheet(ws, prof_df, title):
    ws.append([title])
    ws.cell(1, 1).font = Font(bold=True, size=11)
    # Header
    hdrs = ['Кластер','N','N%','Пік_міс','Опал_%','Літо_%'] + [UA[m] for m in MONTHS] + ['Топ_група','Топ_тип']
    ws.append(hdrs)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(2, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 10
    for ci in range(2, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.column_dimensions[get_column_letter(len(hdrs)-1)].width = 22
    ws.column_dimensions[get_column_letter(len(hdrs))].width = 22

    for ri, (_, row) in enumerate(prof_df.iterrows(), 3):
        cluster_id = int(row['Кластер'])
        fill = PatternFill('solid', fgColor=CLUSTER_COLORS[cluster_id % len(CLUSTER_COLORS)])
        vals = [cluster_id, int(row['N']), round(float(row['N%']), 1),
                row['Пік_міс'], round(float(row['Опал_%']), 1), round(float(row['Літо_%']), 1)] + \
               [round(float(row[UA[m]]), 2) for m in MONTHS] + \
               [row['Топ_група'], row['Топ_тип']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.fill = fill
            c.alignment = Alignment(horizontal='center')
            if isinstance(v, float):
                c.number_format = '0.00'

def write_composition_sheet(ws, df, k, label):
    ws.append([f'Склад кластерів ({label})'])
    ws.cell(1, 1).font = Font(bold=True, size=11)
    grp_order = df['appliance_group'].value_counts().index.tolist()
    ct_order  = df['consumer_type'].value_counts().index.tolist()

    hdrs = ['Кластер', 'N', 'N%'] + grp_order + ct_order
    ws.append(hdrs)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(2, ci, str(h)[:25])
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[2].height = 30
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 7
    for ci in range(4, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    for c in range(k):
        sub = df[df['cluster'] == c]
        fill = PatternFill('solid', fgColor=CLUSTER_COLORS[c % len(CLUSTER_COLORS)])
        row_vals = [c, len(sub), round(len(sub)/len(df)*100, 1)]
        for g in grp_order:
            n = (sub['appliance_group'] == g).sum()
            row_vals.append(f"{n} ({n/len(sub)*100:.1f}%)" if len(sub) else '0')
        for ct in ct_order:
            n = (sub['consumer_type'] == ct).sum()
            row_vals.append(f"{n} ({n/len(sub)*100:.1f}%)" if len(sub) else '0')
        for ci, v in enumerate(row_vals, 1):
            c_ = ws.cell(3 + c, ci, v)
            c_.fill = fill
            c_.alignment = Alignment(horizontal='center')

wb = Workbook()

# Лист 1: Профілі білінгу
ws1 = wb.active; ws1.title = 'Білінг_профілі'
write_profile_sheet(ws1, prof_nm, f'Кластери білінгових профілів (k={k_nm}, sil={sil_nm:.4f})')

# Лист 2: Профілі модемів
ws2 = wb.create_sheet('Модеми_профілі')
write_profile_sheet(ws2, prof_md, f'Кластери модемних профілів (k={k_md}, sil={sil_md:.4f})')

# Лист 3: Склад кластерів — білінг
ws3 = wb.create_sheet('Білінг_склад')
write_composition_sheet(ws3, nm_clean, k_nm, 'Білінг')

# Лист 4: Склад кластерів — модеми
ws4 = wb.create_sheet('Модеми_склад')
write_composition_sheet(ws4, modem_clean, k_md, 'Модеми')

# Лист 5: Силует scores
ws5 = wb.create_sheet('Силует')
ws5.append(['Джерело','k','Силует','Розміри кластерів'])
for r in [{'source':'Білінг',**x} for x in res_nm] + \
         [{'source':'Модеми',**x} for x in res_md]:
    ws5.append([r['source'], r['k'], r['sil'], str(r['sizes'])])

wb.save(OUT_XLSX)
print(f"  DONE → {OUT_XLSX}")
print("  Листи: Білінг_профілі | Модеми_профілі | Білінг_склад | Модеми_склад | Силует")
