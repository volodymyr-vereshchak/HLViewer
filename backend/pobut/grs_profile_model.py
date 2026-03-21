"""
grs_profile_model.py

Покращена модель A2 для прогнозу місячного споживання по ГРС.

Підхід:
  1. w_heat[m], w_cook[m] — нормований місячний профіль з модемів 2025
  2. hs_appl(GRS)          — частка опалення з appliance_group (≈1.0 для малих ГРС)
  3. target                — декомпозит (annual_decompose_result.xlsx, total_pred)
  4. k[m]                  — глобальна місячна корекція: median(decompose/pred) по ГРС

  pred(GRS,m) = annual_decomp(GRS) × [hs × w_heat[m] + (1-hs) × w_cook[m]] × k[m]

Результат: data/grs_profile_model.xlsx
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pobut_predictor import GROUP_MERGE, META_COLS

OP  = '\u041e\u041f'   # «ОП» (кирилиця)
UA  = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
       7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}
MONTHS = list(range(1, 13))
WINTER = [1, 2, 3, 10, 11, 12]
SUMMER = [6, 7, 8, 9]
OUT_XLSX = 'data/grs_profile_model.xlsx'

SKIP_GRS = {'С-1(Запоріж) ГРС-1,3 кільце', 'ГРС (2) м.Запоріжжя'}

# ── 1. Модемні профілі за 2025 ─────────────────────────────────────────────────
print("=== 1. Модемні профілі ===")
md = pd.read_csv('data/input/profile_pobut_daily_result.csv', sep=';', low_memory=False)
md.columns = META_COLS + list(md.columns[len(META_COLS):])
md = md[md['gas_off'].isna() & md['alternative'].isna() & md['dacha'].isna()]
md['appliance_group'] = md['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
md['has_op'] = md['appliance_group'].str.contains(OP, na=False)

day_cols = [c for c in md.columns if '.' in c and len(c) == 10]
dates = pd.to_datetime(day_cols, format='%d.%m.%Y')

heat_s = pd.Series(md[md['has_op']][day_cols].sum().values,  index=dates).sort_index()
cook_s = pd.Series(md[~md['has_op']][day_cols].sum().values, index=dates).sort_index()

heat_m25 = heat_s['2025'].resample('MS').sum()
cook_m25 = cook_s['2025'].resample('MS').sum()

w_heat = (heat_m25 / heat_m25.sum()).values   # [12]
w_cook = (cook_m25 / cook_m25.sum()).values   # [12]

print(f"  Modems: {len(md):,}  (heat={md['has_op'].sum()}, cook={(~md['has_op']).sum()})")
print(f"  Профіль heating (%): " + '  '.join(f'{v*100:.1f}' for v in w_heat))
print(f"  Профіль cooking (%): " + '  '.join(f'{v*100:.1f}' for v in w_cook))

# ── 2. Декомпозит (target) та білінг (reference) по ГРС ───────────────────────
print("\n=== 2. Декомпозит та білінг ===")
dc = pd.read_excel('data/annual_decompose_result.xlsx', sheet_name='All_GRS_monthly')
dc['grs'] = dc['grs'].astype(str)
dc['month'] = dc['month'].astype(int)

decomp_piv   = dc.pivot(index='month', columns='grs', values='total_pred')
billing_piv  = dc.pivot(index='month', columns='grs', values='billing')
annual_decomp  = dc.groupby('grs')['total_pred'].sum()
annual_billing = dc.groupby('grs')['billing'].sum()

target_grs = [g for g in annual_decomp.index if g not in SKIP_GRS]
print(f"  ГРС для моделі: {len(target_grs)}")

# ── 3. Heating share з appliance_group ────────────────────────────────────────
print("\n=== 3. Heating share (appliance_group) ===")
MONTH_ENG = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
             'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
ap = pd.read_csv('data/input/all_pobut_enriched.csv', low_memory=False)
ap['account_id']      = pd.to_numeric(ap['account_id'], errors='coerce')
ap['appliance_group'] = ap['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
ap['has_op']          = ap['appliance_group'].str.contains(OP, na=False)
for eng, m in MONTH_ENG.items():
    col = f'{eng}_2025'
    if col in ap.columns:
        ap[col] = pd.to_numeric(ap[col], errors='coerce').fillna(0)
mc = [f'{e}_2025' for e in MONTH_ENG if f'{e}_2025' in ap.columns]
ap['annual'] = ap[mc].sum(axis=1)

modem_ids = set(pd.to_numeric(md.iloc[:, 0], errors='coerce').dropna().astype(int))
nm = ap[~ap['account_id'].isin(modem_ids) & ap['grs'].notna()].copy()

hs_appl = nm.groupby('grs').apply(
    lambda g: g.loc[g['has_op'], 'annual'].sum() / g['annual'].sum()
    if g['annual'].sum() > 0 else 1.0, include_groups=False
).rename('hs')

hs = {g: float(hs_appl.get(g, 1.0)) for g in target_grs}
print(f"  heating_share: min={min(hs.values()):.2f}  max={max(hs.values()):.2f}  "
      f"mean={np.mean(list(hs.values())):.2f}")

# ── 4. Базовий прогноз (без корекції) ─────────────────────────────────────────
def base_pred(grs, m, ann_dict):
    ann = ann_dict.get(grs, 0)
    h   = hs.get(grs, 1.0)
    return ann * (h * w_heat[m-1] + (1-h) * w_cook[m-1])

# ── 5. Місячні корекційні коефіцієнти k[m] ────────────────────────────────────
print("\n=== 4. Місячні коефіцієнти корекції k[m] ===")
ann_d = dict(annual_decomp)
ann_b = dict(annual_billing)

k_dec = {}
for m in MONTHS:
    ratios = []
    for grs in target_grs:
        if grs not in decomp_piv.columns: continue
        f = decomp_piv.loc[m, grs] if m in decomp_piv.index else np.nan
        p = base_pred(grs, m, ann_d)
        if not np.isnan(f) and f != 0 and p > 0:
            ratios.append(f / p)
    k_dec[m] = float(np.median(ratios)) if ratios else 1.0

print(f"  {'Місяць':6}  " + '  '.join(f'{UA[m]:>5}' for m in MONTHS))
print(f"  {'k[m]':6}  " + '  '.join(f'{k_dec[m]:>5.3f}' for m in MONTHS))

# Нормуємо k[m] щоб Σ_m k[m]*w[m] = 1/12 * 12 = 1 (зберігаємо річний баланс)
# По суті k[m] вже близьке до 1, нормалізація не потрібна — але збережемо per-GRS баланс далі

# ── 6. Фінальні прогнози ──────────────────────────────────────────────────────
print("\n=== 5. Результати (decompose + k[m]) ===")
results = {}
for grs in target_grs:
    ann = ann_d.get(grs, 0)
    if ann <= 0: continue
    row = {}
    for m in MONTHS:
        p_raw = base_pred(grs, m, ann_d)
        p_cor = p_raw * k_dec[m]
        # Нормалізуємо щоб сума за рік = annual_decomp (зберігаємо баланс)
        row[m] = p_cor
    # Масштаб: сума = annual_decomp
    total_pred = sum(row.values())
    scale = ann / total_pred if total_pred > 0 else 1.0
    row = {m: v * scale for m, v in row.items()}
    results[grs] = row

# Метрики vs decompose
print(f"  {'ГРС':<35}  " + '  '.join(f'{UA[m]:>4}' for m in MONTHS) + '  MAPE%')
print("  " + "-"*120)
mapes_dec, mapes_bill = [], []
for grs in target_grs:
    if grs not in results or grs not in decomp_piv.columns: continue
    errs_d, errs_b = [], []
    row_str = []
    for m in MONTHS:
        p = results[grs][m]
        fd = decomp_piv.loc[m, grs] if m in decomp_piv.index else np.nan
        fb = billing_piv.loc[m, grs] if grs in billing_piv.columns and m in billing_piv.index else np.nan
        if not np.isnan(fd) and fd != 0:
            bias = (p - fd) / abs(fd) * 100
            errs_d.append(abs(bias))
            row_str.append(f'{bias:>+4.0f}%')
        else:
            row_str.append('    ')
        if not np.isnan(fb) and fb != 0:
            errs_b.append(abs(p - fb) / abs(fb) * 100)
    mape_d = np.mean(errs_d) if errs_d else np.nan
    mape_b = np.mean(errs_b) if errs_b else np.nan
    mapes_dec.append(mape_d)
    mapes_bill.append(mape_b)
    print(f"  {grs[:34]:<35}  " + '  '.join(row_str) + f'  {mape_d:>5.1f}%')

print("  " + "-"*120)
print(f"  {'Середній MAPE vs decompose':35}  {'':55}  {np.nanmean(mapes_dec):>5.1f}%")
print(f"  {'Середній MAPE vs billing':35}  {'':55}  {np.nanmean(mapes_bill):>5.1f}%")

# Зимовий MAPE
mapes_win_d, mapes_win_b = [], []
for grs in target_grs:
    if grs not in results or grs not in decomp_piv.columns: continue
    ed, eb = [], []
    for m in WINTER:
        p  = results[grs][m]
        fd = decomp_piv.loc[m, grs] if m in decomp_piv.index else np.nan
        fb = billing_piv.loc[m, grs] if grs in billing_piv.columns and m in billing_piv.index else np.nan
        if not np.isnan(fd) and fd != 0: ed.append(abs(p-fd)/abs(fd)*100)
        if not np.isnan(fb) and fb != 0: eb.append(abs(p-fb)/abs(fb)*100)
    if ed: mapes_win_d.append(np.mean(ed))
    if eb: mapes_win_b.append(np.mean(eb))

print(f"\n  Winter MAPE vs decompose: {np.nanmean(mapes_win_d):.1f}%")
print(f"  Winter MAPE vs billing:   {np.nanmean(mapes_win_b):.1f}%")

# ── 7. Прогноз Jan 2026 ────────────────────────────────────────────────────────
print("\n=== 6. Прогноз Jan 2026 ===")
# Jan 2026 modem actuals
heat_jan26 = float(heat_s['2026-01'].sum())
cook_jan26 = float(cook_s['2026-01'].sum())
all_jan26  = heat_jan26 + cook_jan26
print(f"  Modem Jan 2026: heat={heat_jan26/1e3:.1f}тис  cook={cook_jan26/1e3:.1f}тис  total={all_jan26/1e3:.1f}тис")
print(f"  Modem Jan 2025: heat={heat_m25.iloc[0]/1e3:.1f}тис  cook={cook_m25.iloc[0]/1e3:.1f}тис")
print(f"  Зміна Jan26/Jan25: heat={heat_jan26/heat_m25.iloc[0]*100:.1f}%  cook={cook_jan26/cook_m25.iloc[0]*100:.1f}%")

# Скейл Jan26 відносно Jan25 → корегуємо annual_decomp для прогнозу
heat_scale_26 = heat_jan26 / heat_m25.iloc[0] if heat_m25.iloc[0] > 0 else 1.0
cook_scale_26 = cook_jan26 / cook_m25.iloc[0] if cook_m25.iloc[0] > 0 else 1.0

jan26_preds = {}
print(f"\n  {'ГРС':<35} {'Jan26 тис':>10} {'Jan25 тис':>10} {'Δ%':>6}")
print("  " + "-"*65)
for grs in target_grs:
    ann = ann_d.get(grs, 0)
    if ann <= 0: continue
    h = hs.get(grs, 1.0)
    # annual_decomp_26 = annual_decomp_25 * weighted_scale
    ann26 = ann * (h * heat_scale_26 + (1-h) * cook_scale_26)
    p_jan26 = ann26 * (h * w_heat[0] + (1-h) * w_cook[0]) * k_dec[1]
    # normalize
    total26 = sum(ann26 * (h * w_heat[m-1] + (1-h) * w_cook[m-1]) * k_dec[m] for m in MONTHS)
    if total26 > 0:
        p_jan26 = p_jan26 * ann26 / total26
    jan26_preds[grs] = p_jan26
    jan25 = billing_piv.loc[1, grs] if grs in billing_piv.columns else np.nan
    delta = (p_jan26 - jan25) / jan25 * 100 if not np.isnan(jan25) else np.nan
    d_str = f'{delta:>+6.1f}%' if not np.isnan(delta) else '      -'
    print(f"  {grs[:34]:<35} {p_jan26/1e3:>10.2f} {jan25/1e3 if not np.isnan(jan25) else 0:>10.2f} {d_str}")

total_jan26 = sum(jan26_preds.values())
total_jan25 = sum(billing_piv.loc[1, g] for g in target_grs if g in billing_piv.columns)
print(f"  {'ВСЬОГО':35} {total_jan26/1e3:>10.2f} {total_jan25/1e3:>10.2f} {(total_jan26-total_jan25)/total_jan25*100:>+6.1f}%")

# ── 8. Excel ───────────────────────────────────────────────────────────────────
print(f"\nЗберігаю {OUT_XLSX}...")
wb = Workbook()

def color_bias(val):
    a = abs(val)
    if a < 2:   return 'C6EFCE'   # зелений
    if a < 5:   return 'FFEB9C'   # жовтий
    if a < 10:  return 'FFCC99'   # помаранчевий
    return 'FFC7CE'                 # червоний

# ── Sheet 1: Зведена (основна) ────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Зведена'

# Рядок 1: назви місяців (merged over 3 cols: Факт | Прогноз | Bias%)
ws1.cell(1, 1, 'ГРС').font = Font(bold=True)
ws1.column_dimensions['A'].width = 36
for i, m in enumerate(MONTHS):
    c1 = 2 + i*3
    ws1.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c1+2)
    cell = ws1.cell(1, c1, UA[m])
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
ws1.merge_cells(start_row=1, start_column=2+12*3, end_row=1, end_column=2+12*3+1)
ws1.cell(1, 2+12*3, 'Jan 2026').font = Font(bold=True)

# Рядок 2: підзаголовки
ws1.cell(2, 1, '').font = Font(bold=True, italic=True)
sub_fill = PatternFill('solid', fgColor='D9E1F2')
for i, m in enumerate(MONTHS):
    c1 = 2 + i*3
    for ci, h in enumerate(['Декомп', 'Прогноз', 'Bias%'], c1):
        cell = ws1.cell(2, ci, h)
        cell.font = Font(bold=True, italic=True, size=8)
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal='center')
ws1.cell(2, 2+12*3, 'тис м³').font = Font(bold=True, italic=True, size=8)

# Дані
for ri, grs in enumerate(target_grs, 3):
    if grs not in results or grs not in decomp_piv.columns: continue
    ws1.cell(ri, 1, grs)
    for i, m in enumerate(MONTHS):
        c1 = 2 + i*3
        p  = results[grs][m]
        fd = decomp_piv.loc[m, grs] if m in decomp_piv.index else 0
        bias = (p - fd) / abs(fd) * 100 if fd != 0 else 0
        ws1.cell(ri, c1,   round(fd/1e3, 1))
        ws1.cell(ri, c1+1, round(p/1e3, 1))
        bc = ws1.cell(ri, c1+2, round(bias, 1))
        bc.fill = PatternFill('solid', fgColor=color_bias(bias))
    if grs in jan26_preds:
        ws1.cell(ri, 2+12*3, round(jan26_preds[grs]/1e3, 2))

ws1.freeze_panes = 'B3'

# ── Sheet 2: vs Billing (для порівняння) ─────────────────────────────────────
ws2 = wb.create_sheet('vs_Billing')
ws2.cell(1, 1, 'ГРС').font = Font(bold=True)
ws2.column_dimensions['A'].width = 36
for i, m in enumerate(MONTHS):
    c1 = 2 + i*3
    ws2.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c1+2)
    cell = ws2.cell(1, c1, UA[m])
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
for i, m in enumerate(MONTHS):
    c1 = 2 + i*3
    for ci, h in enumerate(['Білінг', 'Прогноз', 'Bias%'], c1):
        cell = ws2.cell(2, ci, h)
        cell.font = Font(bold=True, italic=True, size=8)
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal='center')

for ri, grs in enumerate(target_grs, 3):
    if grs not in results: continue
    ws2.cell(ri, 1, grs)
    for i, m in enumerate(MONTHS):
        c1 = 2 + i*3
        p  = results[grs][m]
        fb = billing_piv.loc[m, grs] if grs in billing_piv.columns and m in billing_piv.index else 0
        bias = (p - fb) / abs(fb) * 100 if fb != 0 else 0
        ws2.cell(ri, c1,   round(fb/1e3, 1))
        ws2.cell(ri, c1+1, round(p/1e3, 1))
        bc = ws2.cell(ri, c1+2, round(bias, 1))
        bc.fill = PatternFill('solid', fgColor=color_bias(bias))
ws2.freeze_panes = 'B3'

# ── Sheet 3: Профілі та k[m] ──────────────────────────────────────────────────
ws3 = wb.create_sheet('Профілі_та_k')
ws3.cell(1, 1, 'Профіль').font = Font(bold=True)
for ci, m in enumerate(MONTHS, 2):
    ws3.cell(1, ci, UA[m]).font = Font(bold=True)
ws3.cell(1, 14, 'Сума').font = Font(bold=True)

rows_p = [
    ('w_heat (modem heating %)', w_heat * 100),
    ('w_cook (modem cooking %)', w_cook * 100),
    ('k[m] decompose', [k_dec[m] for m in MONTHS]),
    ('w_heat × k[m]', [w_heat[m-1]*k_dec[m]*100 for m in MONTHS]),
]
for ri, (label, vals) in enumerate(rows_p, 2):
    ws3.cell(ri, 1, label)
    for ci, v in enumerate(vals, 2):
        ws3.cell(ri, ci, round(float(v), 3))
    ws3.cell(ri, 14, round(sum(vals), 3))

# Heating share per GRS
ws3.cell(7, 1, '--- hs_appl per GRS ---').font = Font(bold=True)
ws3.cell(8, 1, 'ГРС').font = Font(bold=True)
ws3.cell(8, 2, 'hs_appl').font = Font(bold=True)
ws3.cell(8, 3, 'Annual_decomp тис м³').font = Font(bold=True)
for ri, grs in enumerate(sorted(target_grs), 9):
    ws3.cell(ri, 1, grs)
    ws3.cell(ri, 2, round(hs.get(grs, 1.0), 3))
    ws3.cell(ri, 3, round(ann_d.get(grs, 0)/1e3, 1))

# ── Sheet 4: Jan 2026 прогноз ─────────────────────────────────────────────────
ws4 = wb.create_sheet('Jan_2026')
for ci, h in enumerate(['ГРС', 'Прогноз тис м³', 'Jan 2025 тис м³', 'Δ%',
                         'Annual_decomp 2025 тис'], 1):
    ws4.cell(1, ci, h).font = Font(bold=True)
ws4.column_dimensions['A'].width = 36

for ri, (grs, p26) in enumerate(sorted(jan26_preds.items(), key=lambda x: -x[1]), 2):
    jan25 = billing_piv.loc[1, grs] if grs in billing_piv.columns else np.nan
    delta = (p26 - jan25) / jan25 * 100 if not np.isnan(jan25) else np.nan
    ws4.cell(ri, 1, grs)
    ws4.cell(ri, 2, round(p26/1e3, 2))
    if not np.isnan(jan25):
        ws4.cell(ri, 3, round(jan25/1e3, 2))
        ws4.cell(ri, 4, round(delta, 1))
    ws4.cell(ri, 5, round(ann_d.get(grs, 0)/1e3, 1))

# Підсумок
ri_tot = len(jan26_preds) + 3
ws4.cell(ri_tot, 1, 'ВСЬОГО').font = Font(bold=True)
ws4.cell(ri_tot, 2, round(total_jan26/1e3, 2)).font = Font(bold=True)
ws4.cell(ri_tot, 3, round(total_jan25/1e3, 2)).font = Font(bold=True)
ws4.cell(ri_tot, 4, round((total_jan26-total_jan25)/total_jan25*100, 1)).font = Font(bold=True)

wb.save(OUT_XLSX)
print(f"Збережено: {OUT_XLSX}")
print("\nГотово!")
