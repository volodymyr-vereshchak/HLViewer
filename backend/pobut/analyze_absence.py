"""
analyze_absence.py

1. Розподіл ВСІХ не-модемних абонентів (чистих+брудних) по 2 кластерах
   (опалення / без опалення) та перевірка обсягів після перекладу на
   чисті профілі.

2. Виявлення "відсутніх" абонентів:
   - "Виїхали"    : платили в Q1, зупинились ≤ травень
   - "Нові"       : почали платити з серпня і пізніше
   - "Дача"       : є в опалювальний сезон, відсутні влітку
   - "Нерегулярні": вибіркові платежі без явного патерну
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from pobut_predictor import GROUP_MERGE, META_COLS, CT_PRIV

MONTH_ENG = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
             'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
MONTHS = list(range(1, 13))
UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}
WINTER = [1, 2, 3, 4, 10, 11, 12]
SUMMER = [6, 7, 8]

OUT_XLSX = 'data/analyze_absence.xlsx'

# ── 0. Завантаження ────────────────────────────────────────────────────────────
md = pd.read_csv('data/input/profile_pobut_daily_result.csv', sep=';', low_memory=False)
md.columns = META_COLS + list(md.columns[len(META_COLS):])
md = md[md['gas_off'].isna() & md['alternative'].isna() & md['dacha'].isna()]
modem_ids = set(pd.to_numeric(md.iloc[:,0], errors='coerce').dropna().astype(int))

ap = pd.read_csv('data/input/all_pobut_enriched.csv', low_memory=False)
ap['account_id']      = pd.to_numeric(ap['account_id'], errors='coerce')
ap['appliance_group'] = ap['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
ap['consumer_type']   = ap['consumer_type'].fillna(CT_PRIV)

bill_cols = {}
for eng, m in MONTH_ENG.items():
    col = f'{eng}_2025'
    if col in ap.columns:
        ap[col] = pd.to_numeric(ap[col], errors='coerce').fillna(0)
        bill_cols[m] = col
mc = [bill_cols[m] for m in MONTHS]
ap['annual'] = ap[mc].sum(axis=1)

# Тільки не-модемні з хоч якимось обсягом
nm = ap[
    ~ap['account_id'].isin(modem_ids)
    & ap['grs'].notna()
    & (ap[mc].abs().sum(axis=1) > 0)
].copy()
print(f"Не-модемних з обсягом: {len(nm):,}")

# ── 1. Базові метрики для класифікації ─────────────────────────────────────────
for m in MONTHS:
    nm[f'p_{m}'] = (nm[bill_cols[m]] / nm['annual'].replace(0, np.nan)).fillna(0)

nm['n_nonzero']  = (nm[mc].abs() > 0).sum(axis=1)
nm['n_negative'] = (nm[mc] < 0).sum(axis=1)
nm['max_frac']   = nm[[f'p_{m}' for m in MONTHS]].max(axis=1)

# Перший та останній місяць з платежем
def first_month(row):
    for m in MONTHS:
        if row[bill_cols[m]] != 0:
            return m
    return None

def last_month(row):
    for m in reversed(MONTHS):
        if row[bill_cols[m]] != 0:
            return m
    return None

nm['first_m'] = nm.apply(first_month, axis=1)
nm['last_m']  = nm.apply(last_month,  axis=1)
nm['span']    = nm['last_m'] - nm['first_m'] + 1  # від першого до останнього
nm['fill_rate'] = nm['n_nonzero'] / nm['span'].replace(0, 1)

# Літня пауза: всі місяці червень-серпень = 0, але є платежі до і після
nm['summer_zero'] = (nm[[bill_cols[m] for m in SUMMER]] == 0).all(axis=1)
nm['has_before_summer'] = (nm[[bill_cols[m] for m in [4,5]]].abs() > 0).any(axis=1)
nm['has_after_summer']  = (nm[[bill_cols[m] for m in [9,10]]].abs() > 0).any(axis=1)
nm['dacha_signal']      = nm['summer_zero'] & nm['has_before_summer'] & nm['has_after_summer']

# ── 2. Кластер "опалення" / "без опалення" ─────────────────────────────────────
nm['has_op'] = nm['appliance_group'].str.contains('ОП', na=False)
nm['cluster'] = np.where(nm['has_op'], 'heating', 'cooking')

# ── 3. "Чистий" критерій ──────────────────────────────────────────────────────
clean_mask = (
    (nm['n_nonzero'] >= 10)
    & (nm['max_frac'] < 0.25)
    & (nm['n_negative'] == 0)
)
nm['is_clean'] = clean_mask

# ── 4. Патерн відсутності (тільки для брудних/неповних) ──────────────────────
def absence_pattern(row):
    if row['is_clean']:
        return 'чистий'
    # Від'ємні перерахунки
    if row['n_negative'] > 0 and row['max_frac'] < 0.25 and row['n_nonzero'] >= 10:
        return 'від_ємні'
    # Виїхали: закінчили платити по травень включно (і починали з початку року)
    if row['last_m'] is not None and row['last_m'] <= 5 and row['first_m'] is not None and row['first_m'] <= 3:
        return 'виїхали'
    # Нові абоненти: почали з серпня і пізніше
    if row['first_m'] is not None and row['first_m'] >= 8:
        return 'нові_абоненти'
    # Дача: літня пауза з платежами до і після
    if row['dacha_signal'] and row['n_nonzero'] >= 4:
        return 'дача_/сезонні'
    # Квартальні / стрибок (max_frac >= 0.25 і достатньо місяців)
    if row['max_frac'] >= 0.25 and row['n_nonzero'] >= 6:
        return 'квартальні'
    # Мало місяців (загальна)
    if row['n_nonzero'] < 10:
        return 'мало_місяців'
    return 'інше'

print("Обчислення патернів відсутності...")
nm['pattern'] = nm.apply(absence_pattern, axis=1)

# ── 5. Профілі чистих по кластерах ────────────────────────────────────────────
clean = nm[nm['is_clean']].copy()
W = {}
print("\n=== Профілі чистих абонентів по кластерах (%) ===")
print(f"  {'Кластер':20s}  " + '  '.join(f'{UA[m]:>5}' for m in MONTHS) + '  Зима%')
for cl in ['heating', 'cooking']:
    sub = clean[clean['cluster'] == cl]
    w = np.zeros(12)
    for i, m in enumerate(MONTHS):
        w[i] = sub[bill_cols[m]].sum()
    w = w / w.sum() if w.sum() > 0 else w
    W[cl] = w
    heat = w[[m-1 for m in WINTER]].sum() * 100
    print(f"  {cl:20s}  " + '  '.join(f'{v*100:>5.1f}' for v in w) + f'  {heat:.1f}%')

# ── 6. Розподіл обсягів за патернами ──────────────────────────────────────────
print("\n=== Розподіл обсягів за патернами ===")
patterns_order = ['чистий', 'квартальні', 'мало_місяців', 'виїхали',
                  'нові_абоненти', 'дача_/сезонні', 'від_ємні', 'інше']

total_annual = nm['annual'].sum()
rows_summary = []
print(f"  {'Патерн':20s} {'Абон':>8} {'%абон':>7} {'Обсяг млн':>11} {'%обсяг':>7} {'mid_ann':>9}")
for pat in patterns_order:
    sub = nm[nm['pattern'] == pat]
    if len(sub) == 0:
        continue
    vol = sub['annual'].sum() / 1e6
    pct_sub = len(sub) / len(nm) * 100
    pct_vol = sub['annual'].sum() / total_annual * 100
    mid_ann = sub['annual'].median()
    print(f"  {pat:20s} {len(sub):>8,} {pct_sub:>7.1f}% {vol:>11.2f} {pct_vol:>7.1f}%  mid={mid_ann:>7.0f}")
    rows_summary.append({
        'pattern': pat, 'n': len(sub), 'pct_n': round(pct_sub,1),
        'vol_mln': round(vol,2), 'pct_vol': round(pct_vol,1),
        'median_annual': round(mid_ann,0)
    })

# По кластерах
print("\n  --- По кластерах ---")
for cl in ['heating', 'cooking']:
    sub = nm[nm['cluster'] == cl]
    vol = sub['annual'].sum() / 1e6
    clean_sub = sub[sub['is_clean']]
    print(f"  {cl:12s}: {len(sub):>7,} абон  vol={vol:.2f}млн  "
          f"чистих={len(clean_sub):,} ({len(clean_sub)/len(sub)*100:.1f}%)")

# ── 7. "Виїхали" — детальніше ──────────────────────────────────────────────────
left = nm[nm['pattern'] == 'виїхали'].copy()
print(f"\n=== 'Виїхали' (n={len(left):,}) ===")
print(f"  Загальний обсяг: {left['annual'].sum()/1e6:.3f} млн м³")
print(f"  Медіана annual: {left['annual'].median():.0f} м³")
print(f"  Середня annual: {left['annual'].mean():.0f} м³\n")
print(f"  Розподіл за last_m (місяць останнього платежу):")
for m in range(1, 6):
    n = (left['last_m'] == m).sum()
    if n > 0:
        bar = '#' * (n // 50)
        print(f"    {UA[m]}: {n:>5,}  {bar}")

print(f"\n  По ГРС (топ-10):")
for grs, grp in left.groupby('grs'):
    if len(grp) >= 10:
        print(f"    {grs:<25} n={len(grp):>4,}  vol={grp['annual'].sum()/1e3:.1f}тис")

print(f"\n  По appliance_group:")
for grp, cnt in left['appliance_group'].value_counts().items():
    print(f"    {grp:<25} {cnt:>5,}  ({cnt/len(left)*100:.1f}%)")

print(f"\n  Розподіл обсягів (annual м³):")
for lo, hi in [(0,100),(100,300),(300,1000),(1000,3000),(3000,1e9)]:
    n = ((left['annual'] >= lo) & (left['annual'] < hi)).sum()
    if n > 0:
        label = f'{lo}-{hi}' if hi < 1e9 else f'{lo}+'
        print(f"    {label:>12} м³: {n:>5,}  ({n/len(left)*100:.1f}%)")

# ── 8. "Нові абоненти" — детальніше ───────────────────────────────────────────
new_subs = nm[nm['pattern'] == 'нові_абоненти'].copy()
print(f"\n=== 'Нові абоненти' (n={len(new_subs):,}) ===")
print(f"  Загальний обсяг: {new_subs['annual'].sum()/1e6:.3f} млн м³")
print(f"  Розподіл за first_m:")
for m in range(8, 13):
    n = (new_subs['first_m'] == m).sum()
    if n > 0:
        print(f"    {UA[m]}: {n:>5,}")

# ── 9. "Дача/сезонні" — детальніше ────────────────────────────────────────────
dacha = nm[nm['pattern'] == 'дача_/сезонні'].copy()
print(f"\n=== 'Дача/сезонні' (n={len(dacha):,}) ===")
print(f"  Загальний обсяг: {dacha['annual'].sum()/1e6:.3f} млн м³")
print(f"  n_nonzero розподіл:")
for v in range(4, 10):
    n = (dacha['n_nonzero'] == v).sum()
    if n > 0:
        print(f"    {v} міс: {n:>5,}")

# ── 10. Порівняння: поточне "брудне" vs перекладене на профіль ────────────────
print("\n=== Порівняння: raw_billing vs cluster_profile для брудних ===")
dirty = nm[~nm['is_clean']].copy()
print(f"  Брудних: {len(dirty):,}  (annual={dirty['annual'].sum()/1e6:.2f} млн м³)")

# Перекласти annual на профіль кластера → redistributed monthly
dirty_pred = np.zeros(12)
for _, row in dirty.iterrows():
    cl = row['cluster']
    ann = row['annual']
    if ann > 0:
        dirty_pred += W[cl] * ann

dirty_raw = np.array([dirty[bill_cols[m]].sum() for m in MONTHS])

print(f"\n  {'Місяць':6} {'Raw billing':>12} {'Cluster profile':>16} {'Δ%':>7}")
for i, m in enumerate(MONTHS):
    raw = dirty_raw[i]
    pred = dirty_pred[i]
    delta = (pred - raw) / raw * 100 if raw != 0 else float('nan')
    print(f"  {UA[m]:6} {raw/1e3:>12.1f}тис {pred/1e3:>13.1f}тис  {delta:>+7.1f}%")

# Порівняння для "виїхали" окремо
print(f"\n  === Специфіка для 'виїхали' ===")
print(f"  Їх raw billing по місяцях vs 'правильний' профіль:")
print(f"  (якщо вони виїхали навесні, то Q3-Q4 обнуляються →")
print(f"   profile redistributes в квітні-травні їх обсяг на весь рік → завищення)")
left_raw  = np.array([left[bill_cols[m]].sum() for m in MONTHS])
left_pred = np.zeros(12)
for _, row in left.iterrows():
    cl = row['cluster']
    ann = row['annual']
    if ann > 0:
        left_pred += W[cl] * ann
print(f"  {'Місяць':6} {'Факт':>10} {'Профіль':>10}")
for i, m in enumerate(MONTHS):
    raw  = left_raw[i] / 1e3
    pred = left_pred[i] / 1e3
    flag = ' ← РЕАЛЬНЕ' if raw > 1 else (' ← ШТУЧНЕ' if pred > 1 else '')
    print(f"  {UA[m]:6} {raw:>10.1f}тис {pred:>10.1f}тис{flag}")

# ── 11. Рекомендація що робити з кожним патерном ─────────────────────────────
print("\n=== Рекомендації по кожному патерну ===")
recs = [
    ('чистий',         'Використати як є → apply W(cluster, m) × annual'),
    ('квартальні',     'Можна apply W(cluster, m) × annual (обсяг є, патерн неправильний)'),
    ('від_ємні',       'Apply W(cluster, m) × annual (обсяг є, знак ненадійний)'),
    ('нові_абоненти',  'Apply W(cluster, m) × actual_months (scaling від першого місяця)'),
    ('дача_/сезонні',  'Apply W(cluster, m) × annual, але обсяг зменшений (немає літа)'),
    ('мало_місяців',   'Apply W(cluster, m) × annual (можливо частковий рік)'),
    ('виїхали',        '⚠ НЕ перекладати! Їх annual = Q1 фактично. Вилучити з декомпозиту.'),
    ('інше',           'Apply W(cluster, m) × annual за замовчуванням'),
]
for pat, rec in recs:
    sub = nm[nm['pattern'] == pat]
    if len(sub) == 0:
        continue
    print(f"\n  [{pat}] n={len(sub):,}")
    print(f"    Рекомендація: {rec}")

# ── 12. Excel ──────────────────────────────────────────────────────────────────
print(f"\nЗберігаю {OUT_XLSX}...")
wb = Workbook()

# --- Sheet 1: Зведена по патернах ---
ws1 = wb.active
ws1.title = 'Патерни_зведена'
hdr = ['Патерн', 'Абонентів', '%абон', 'Обсяг млн м³', '%обсягу', 'Медіана м³/рік',
       'Кластер heating %', 'Кластер cooking %']
for ci, h in enumerate(hdr, 1):
    ws1.cell(1, ci, h).font = Font(bold=True)
ws1.row_dimensions[1].height = 20

ri = 2
for pat in patterns_order:
    sub = nm[nm['pattern'] == pat]
    if len(sub) == 0:
        continue
    vol  = sub['annual'].sum() / 1e6
    pct_n  = len(sub) / len(nm) * 100
    pct_v  = sub['annual'].sum() / total_annual * 100
    h_pct  = (sub['cluster'] == 'heating').mean() * 100
    c_pct  = (sub['cluster'] == 'cooking').mean() * 100
    row = [pat, len(sub), round(pct_n,1), round(vol,3), round(pct_v,1),
           round(sub['annual'].median()), round(h_pct,1), round(c_pct,1)]
    for ci, v in enumerate(row, 1):
        ws1.cell(ri, ci, v)
    ri += 1

# --- Sheet 2: Виїхали по ГРС ---
ws2 = wb.create_sheet('Виїхали')
hdr2 = ['ГРС', 'N абонентів', 'Обсяг м³', 'Серед. annual', 'Jan', 'Feb', 'Mar', 'Apr', 'May']
for ci, h in enumerate(hdr2, 1):
    ws2.cell(1, ci, h).font = Font(bold=True)

grs_left = left.groupby('grs').agg(
    n=('account_id', 'count'),
    vol=('annual', 'sum'),
    mean_ann=('annual', 'mean'),
    **{f'v_{m}': (bill_cols[m], 'sum') for m in [1,2,3,4,5]}
).reset_index().sort_values('vol', ascending=False)

for ri, (_, row) in enumerate(grs_left.iterrows(), 2):
    ws2.cell(ri, 1, row['grs'])
    ws2.cell(ri, 2, int(row['n']))
    ws2.cell(ri, 3, round(float(row['vol'])))
    ws2.cell(ri, 4, round(float(row['mean_ann'])))
    for ci, m in enumerate([1,2,3,4,5], 5):
        ws2.cell(ri, ci, round(float(row[f'v_{m}'])))

# --- Sheet 3: Порівняння redistributed vs raw по місяцях ---
ws3 = wb.create_sheet('Redistribution')
ws3.cell(1, 1, 'Категорія').font = Font(bold=True)
for ci, m in enumerate(MONTHS, 2):
    ws3.cell(1, ci, UA[m]).font = Font(bold=True)
ws3.cell(1, 14, 'Разом').font = Font(bold=True)

cats = [
    ('Факт raw billing (всі)',     nm,    'raw'),
    ('Profile predicted (чисті)',  clean, 'pred'),
    ('Profile predicted (брудні)', dirty, 'pred'),
    ('Факт (виїхали)',             left,  'raw'),
    ('Profile (виїхали)',          left,  'pred'),
]

for ri, (label, sub, mode) in enumerate(cats, 2):
    ws3.cell(ri, 1, label)
    for ci, m in enumerate(MONTHS, 2):
        if mode == 'raw':
            v = sub[bill_cols[m]].sum() / 1e3
        else:
            cl_h = sub[sub['cluster']=='heating']['annual'].sum()
            cl_c = sub[sub['cluster']=='cooking']['annual'].sum()
            v = (cl_h * W['heating'][m-1] + cl_c * W['cooking'][m-1]) / 1e3
        ws3.cell(ri, ci, round(float(v), 1))
    # Total
    ws3.cell(ri, 14, round(sub['annual'].sum()/1e3, 1))

# --- Sheet 4: Повний список "виїхали" ---
ws4 = wb.create_sheet('Список_виїхали')
cols4 = ['account_id', 'grs', 'appliance_group', 'consumer_type',
         'annual', 'n_nonzero', 'first_m', 'last_m'] + mc
for ci, c in enumerate(cols4, 1):
    ws4.cell(1, ci, c).font = Font(bold=True)
for ri, (_, row) in enumerate(left[cols4].iterrows(), 2):
    for ci, c in enumerate(cols4, 1):
        v = row[c]
        if hasattr(v, 'item'):
            v = v.item()
        ws4.cell(ri, ci, v)

wb.save(OUT_XLSX)
print(f"Збережено: {OUT_XLSX}")
print("\nГотово!")
