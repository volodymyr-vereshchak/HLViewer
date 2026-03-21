"""
annual_decompose.py

Підхід: річний білінг кожного споживача × нормалізований денний профіль групи
→ місячний прогноз без артефактів білінгового циклу.

Логіка:
  annual_i = Σ billing_i(m) за 2025  ← реальне споживання (з урахуванням перерозрахунків)
  profile(d, grp) = медіанне денне питоме споживання модемів (нормалізоване: Σ=1 за рік)
  pred_daily(i, d) = annual_i × profile_norm(d, grp)
  pred_month(i, m) = Σ pred_daily(i, d) for d in month m
  pred_grs(grs, m) = Σ pred_month(i, m) for non-modem + actual_modem(grs, m)
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from pobut_predictor import GROUP_MERGE, CT_PRIV, META_COLS

MODEM_FILE = 'data/input/profile_pobut_daily_result.csv'
SUBS_FILE  = 'data/input/all_pobut_enriched.csv'
OUT_XLSX   = 'data/annual_decompose_result.xlsx'

MONTH_ENG = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
             'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}
LARGE_GRS = 5_000_000

# ── 1. Денні профілі з модемів ───────────────────────────────────────────────
print("=== 1. Денні профілі з модемів (2025) ===")
md = pd.read_csv(MODEM_FILE, sep=';', low_memory=False)
# Перейменовуємо перші 12 колонок як в PobUtPredictor
md.columns = META_COLS + list(md.columns[len(META_COLS):])
md['appliance_group'] = md['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
md['consumer_type'] = md['consumer_type'].fillna(CT_PRIV)
md['heated_area']   = pd.to_numeric(md['heated_area'], errors='coerce').fillna(55).clip(lower=1)
md['has_OP']        = md['appliance_group'].str.contains('ОП', na=False)

# Фільтр активних (не дача, не альтернативні, не відключені)
md = md[md['gas_off'].isna() & md['alternative'].isna() & md['dacha'].isna()]

modem_ids = set(pd.to_numeric(md.iloc[:, 0], errors='coerce').dropna().astype(int))

# Зібрати всі денні колонки 2025
date_cols_2025 = []
for c in md.columns:
    if isinstance(c, str) and c.count('.') == 2 and len(c) == 10:
        try:
            d_s, m_s, y_s = c.split('.')
            dt = pd.Timestamp(f'{y_s}-{m_s}-{d_s}')
            if dt.year == 2025:
                date_cols_2025.append((dt, c))
        except:
            pass
date_cols_2025.sort()
print(f"  Денних колонок 2025: {len(date_cols_2025)}")

# Групи для яких будемо будувати профіль
GROUPS = [
    ('ОП,ПГ',     'Приватний сектор'),
    ('ОП,ПГ',     'Багатоквартирний сектор'),
    ('ОП,ПГ,ВПГ', 'Приватний сектор'),
    ('ОП,ПГ,ВПГ', 'Багатоквартирний сектор'),
    ('ПГ',        'Багатоквартирний сектор'),
    ('ПГ,ВПГ',    'Багатоквартирний сектор'),
]

# Групи без ОП: використовуємо модемний (споживчий) профіль замість білінгового.
# Ці абоненти платять нерегулярно, але споживають рівномірно протягом року.
# Модемний профіль точніше відображає реальне споживання.
MODEM_PROFILE_GROUPS = {
    ('ПГ',    'Багатоквартирний сектор'),
    ('ПГ,ВПГ','Багатоквартирний сектор'),
    ('ПГ',    'Приватний сектор'),
    ('ПГ,ВПГ','Приватний сектор'),
}

# Максимальний правдоподібний місячний обсяг для ПГ-абонента (без ОП).
# Вище → аномалія: перерахунок за кілька місяців або комерційний об'єкт.
MAX_PG_MONTHLY_M3 = 150.0

# profile_raw[key] = {date: median_daily_fraction}
# Підхід: w_i(d) = cons_i(d) / annual_i — питома частка дня в річному обсязі
# profile(d) = median( w_i(d) ) по всіх модемних абонентах групи
profile_raw = {}
MIN_CONSUMERS = 5
MIN_ANNUAL    = 10.0   # мінімальний річний обсяг (м³) для включення абонента

day_col_names = [col for _, col in date_cols_2025]

for grp, ct in GROUPS:
    mask = (md['appliance_group'] == grp) & (md['consumer_type'] == ct)
    sub  = md[mask]
    if len(sub) < MIN_CONSUMERS:
        print(f"  [{grp}, {ct[:20]}]: {len(sub)} модемів — пропускаємо")
        continue

    # Матриця: (n_consumers × n_days), заміна NaN→0, clip(0)
    data = (sub[day_col_names]
            .apply(pd.to_numeric, errors='coerce')
            .fillna(0)
            .clip(lower=0)
            .values.astype(np.float64))  # shape (n, 365)

    annual = data.sum(axis=1)            # річний обсяг кожного абонента
    valid  = annual >= MIN_ANNUAL        # лише ті, що мають достатньо даних

    if valid.sum() < MIN_CONSUMERS:
        print(f"  [{grp}, {ct[:20]}]: {valid.sum()} валідних — пропускаємо")
        continue

    # Матриця питомих часток: w_i(d) = cons_i(d) / annual_i
    fracs = data[valid] / annual[valid, np.newaxis]  # shape (n_valid, 365)

    # По кожному дню: медіана часток (зрізаємо верхній 1%)
    daily = {}
    for j, (dt, _) in enumerate(date_cols_2025):
        f = fracs[:, j]
        p99 = np.percentile(f, 99)
        f_clean = f[f <= p99]
        if len(f_clean) < MIN_CONSUMERS:
            continue
        daily[dt] = float(np.median(f_clean))

    total = sum(daily.values())
    if total <= 0:
        continue

    # Нормалізуємо щоб сума профілю за рік = 1.0
    profile_norm = {dt: v / total for dt, v in daily.items()}
    profile_raw[(grp, ct)] = profile_norm

    # Місячні ваги для відображення
    monthly_w = {}
    for dt, w in profile_norm.items():
        monthly_w[dt.month] = monthly_w.get(dt.month, 0.0) + w

    vals_str = '  '.join(f"{UA[m]}:{monthly_w.get(m,0)*100:.1f}%" for m in range(1,13))
    print(f"  [{grp}, {ct[:15]}] n={valid.sum()}  {vals_str}")

print(f"\n  Профілів побудовано: {len(profile_raw)}")

# Місячні ваги для кожного профілю
def monthly_weights(profile_norm):
    """Повертає {month: sum_of_daily_weights}"""
    mw = {}
    for dt, w in profile_norm.items():
        mw[dt.month] = mw.get(dt.month, 0.0) + w
    return mw

# Будуємо місячні ваги з модемного профілю (буде fallback для білінгового)
monthly_w_cache = {}
for key, pnorm in profile_raw.items():
    monthly_w_cache[key] = monthly_weights(pnorm)

# ── 2. Завантаження білінгу ──────────────────────────────────────────────────
print("\n=== 2. Завантаження білінгу ===")
ap = pd.read_csv(SUBS_FILE, low_memory=False)
ap['account_id']      = pd.to_numeric(ap['account_id'], errors='coerce')
ap['appliance_group'] = ap['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
ap['consumer_type']   = ap['consumer_type'].fillna(CT_PRIV)
ap['heated_area']     = pd.to_numeric(ap['heated_area'], errors='coerce').fillna(55).clip(lower=1)
ap['gas_off']         = ap['gas_off'].notna()

bill_cols = {}
for eng, m in MONTH_ENG.items():
    col = f'{eng}_2025'
    if col in ap.columns:
        ap[col] = pd.to_numeric(ap[col], errors='coerce').fillna(0)
        bill_cols[m] = col

# Річний білінг = сума 12 місяців
ap['annual_billing'] = ap[[bill_cols[m] for m in range(1, 13) if m in bill_cols]].sum(axis=1)

# Не-модемні з ненульовим місячним білінгом (хоча б один місяць ≠ 0).
# Включаємо gas_off (відключені але мали споживання) та annual_billing ≤ 0
# (перерахунки/повернення — вони мають від'ємний внесок у billing_grs,
#  тому мають бути декомпозовані з відповідним від'ємним знаком).
_mc = [bill_cols[m] for m in range(1, 13) if m in bill_cols]
nm = ap[
    ~ap['account_id'].isin(modem_ids) &
    ap['grs'].notna() &
    (ap[_mc].abs().sum(axis=1) > 0)   # хоча б один ненульовий місяць
].copy()
print(f"  Не-модемних з ненульовим білінгом: {len(nm):,}  "
      f"(з них annual>0: {(nm['annual_billing']>0).sum():,}  "
      f"annual≤0: {(nm['annual_billing']<=0).sum():,})")

# Modемні — факт по місяцях
md_grs = ap[
    ap['account_id'].isin(modem_ids) &
    ap['grs'].notna()
].copy()
print(f"  Модемних з ГРС: {len(md_grs):,}")

# ── 2b. Білінгові ваги з "чистих" споживачів ────────────────────────────────
print("\n=== 2b. Білінгові ваги з 'чистих' споживачів ===")

_mc = [bill_cols[m] for m in range(1, 13) if m in bill_cols]
ap['max_month_bill']   = ap[_mc].max(axis=1)
ap['n_months_nonzero'] = (ap[_mc] > 0).sum(axis=1)
ap['max_month_frac']   = ap['max_month_bill'] / ap['annual_billing'].clip(lower=1)

# "Чисті" споживачі: жоден місяць не перевищує 50% річного білінгу,
# і мінімум 8 місяців з ненульовим значенням → щомісячне нарахування
clean_mask = (
    ~ap['account_id'].isin(modem_ids)
    & ~ap['gas_off']
    & ap['grs'].notna()
    & (ap['annual_billing'] > 0)
    & (ap['max_month_frac'] < 0.5)
    & (ap['n_months_nonzero'] >= 8)
)
clean_all = ap[clean_mask].copy()
print(f"  'Чистих' споживачів: {len(clean_all):,}  (з {len(ap):,} всього)")

billing_monthly_w = {}   # (grp, ct) → {month: weight}  (білінговий профіль)
MIN_CLEAN = 100

for grp, ct in GROUPS:
    key = (grp, ct)

    # ПГ-групи: використовуємо модемний (споживчий) профіль — пропускаємо білінг
    if key in MODEM_PROFILE_GROUPS:
        print(f"  [{grp}, {ct[:15]}]: модемний профіль (consumption-based, без білінгу)")
        continue

    sub_c = clean_all[
        (clean_all['appliance_group'] == grp)
        & (clean_all['consumer_type'] == ct)
    ]
    if len(sub_c) < MIN_CLEAN:
        print(f"  [{grp}, {ct[:15]}]: {len(sub_c)} чистих — fallback до модемного профілю")
        continue

    fracs = {}
    for m, col in bill_cols.items():
        fracs[m] = float((sub_c[col] / sub_c['annual_billing'].clip(lower=1)).mean())

    total = sum(fracs.values())
    if total <= 0:
        continue

    billing_monthly_w[key] = {m: v / total for m, v in fracs.items()}

    vals_str = '  '.join(f"{UA[m]}:{fracs[m]/total*100:.1f}%" for m in range(1, 13))
    print(f"  [{grp}, {ct[:15]}] n={len(sub_c):,}  {vals_str}")

print(f"\n  Білінгових профілів побудовано: {len(billing_monthly_w)}")

# Порівняємо білінговий vs модемний профіль для кожної групи
print("\n  Порівняння профілів (Білінг − Модем, pp):")
for key in billing_monthly_w:
    if key not in monthly_w_cache:
        continue
    grp, ct = key
    modem_mw = monthly_w_cache[key]
    bill_mw  = billing_monthly_w[key]
    diffs = {m: (bill_mw.get(m, 0) - modem_mw.get(m, 0)) * 100 for m in range(1, 13)}
    diff_str = '  '.join(f"{UA[m]}:{diffs[m]:+.1f}" for m in range(1, 13))
    print(f"  [{grp}, {ct[:15]}]  {diff_str}")

# ── 2c. Гібридний денний профіль ─────────────────────────────────────────────
# Місячна сума = білінговий профіль (точний)
# Розподіл усередині місяця = модемний профіль (денна гранулярність)
#
# hybrid_daily(d) = billing_w(month(d)) × modem_w(d) / Σ modem_w(d') for d' ∈ month(d)
print("\n=== 2c. Гібридний денний профіль ===")

# Для кожного місяця — список дат з modемного профілю
month_days = {}  # month → [date, ...]
for dt, _ in date_cols_2025:
    month_days.setdefault(dt.month, []).append(dt)

hybrid_daily_w  = {}  # (grp, ct) → {date: weight}   (денний, сума=1)
hybrid_monthly_w = {}  # (grp, ct) → {month: weight}  (місячний, сума=1; має = billing)

for grp, ct in GROUPS:
    key = (grp, ct)
    bill_mw  = billing_monthly_w.get(key)
    modem_pn = profile_raw.get(key)      # денний модемний профіль {date: norm_weight}

    if bill_mw is None:
        continue  # немає білінгового профілю → пропускаємо

    daily = {}
    for m, b_w in bill_mw.items():
        days = month_days.get(m, [])
        if not days:
            continue
        if modem_pn is not None:
            # Сума модемних ваг за цей місяць
            month_modem_sum = sum(modem_pn.get(dt, 0.0) for dt in days)
            for dt in days:
                if month_modem_sum > 0:
                    within = modem_pn.get(dt, 0.0) / month_modem_sum
                else:
                    within = 1.0 / len(days)
                daily[dt] = b_w * within
        else:
            # Немає модемних даних — рівномірно по днях місяця
            for dt in days:
                daily[dt] = b_w / len(days)

    hybrid_daily_w[key]   = daily
    # Місячна агрегація (перевірка: має бути = billing_monthly_w)
    mw = {}
    for dt, w in daily.items():
        mw[dt.month] = mw.get(dt.month, 0.0) + w
    hybrid_monthly_w[key] = mw

print(f"  Гібридних профілів: {len(hybrid_daily_w)}")
# Верифікація: місячна сума має збігатись з білінговим профілем
max_diff = 0.0
for key in hybrid_monthly_w:
    for m in range(1, 13):
        diff = abs(hybrid_monthly_w[key].get(m, 0) - billing_monthly_w[key].get(m, 0))
        max_diff = max(max_diff, diff)
print(f"  Макс. відхилення hybrid_monthly vs billing_monthly: {max_diff:.2e}  (має бути ~0)")

# Зберігаємо hybrid_daily_w → використовується у daily_decompose_grs.py
_hdw_rows = []
for (grp, ct), dw in hybrid_daily_w.items():
    for dt, w in dw.items():
        _hdw_rows.append({'grp': grp, 'ct': ct, 'date': dt.strftime('%Y-%m-%d'), 'weight': w})
pd.DataFrame(_hdw_rows).to_csv('data/profiles/hybrid_daily_w.csv', index=False)
print(f"  Збережено hybrid_daily_w → data/profiles/hybrid_daily_w.csv  ({len(_hdw_rows)} рядків)")

# ── 3. Прогноз для не-модемних ───────────────────────────────────────────────
print("\n=== 3. Прогноз: annual × hybrid_monthly_weight ===")
print("  (Місячні ваги = білінговий профіль; денна деталізація = модемна форма)")

def get_monthly_w(grp, ct):
    """Місячні ваги: гібридний (білінг + модем) → модемний → білінговий → рівномірний.
    ПГ-групи ніколи не мають hybrid/billing → одразу йдуть у моdemний профіль.
    """
    key = (grp, ct)
    if key in hybrid_monthly_w:
        return hybrid_monthly_w[key], 'hybrid'
    if key in billing_monthly_w:
        return billing_monthly_w[key], 'billing'
    # Модемний профіль (споживчий) — використовується для ПГ-груп і як fallback
    if key in monthly_w_cache:
        return monthly_w_cache[key], 'modem'
    for (g, c), mw in hybrid_monthly_w.items():
        if g == grp:
            return mw, 'hybrid_fallback'
    for (g, c), mw in billing_monthly_w.items():
        if g == grp:
            return mw, 'billing_fallback'
    for (g, c), mw in monthly_w_cache.items():
        if g == grp:
            return mw, 'modem_fallback'
    return {m: 1/12 for m in range(1, 13)}, 'uniform'

# Інформаційний звіт по ПГ-аномаліях (залишаємо в nm — їх обсяг розкладається
# по модемному профілю, тому загальна сума декомпозиту = загальному білінгу).
_mc_list = [bill_cols[m] for m in range(1, 13) if m in bill_cols]
_pg_mask    = nm['appliance_group'].isin(['ПГ', 'ПГ,ВПГ'])
_pg_max     = nm[_mc_list].max(axis=1)
_pg_anomaly = _pg_mask & (_pg_max > MAX_PG_MONTHLY_M3)
n_pg_anom   = _pg_anomaly.sum()
if n_pg_anom > 0:
    anom_vol = nm.loc[_pg_anomaly, [bill_cols[m] for m in range(1, 13) if m in bill_cols]].sum().sum()
    print(f"\n  ПГ-аномалії (макс.місяць > {MAX_PG_MONTHLY_M3:.0f} м³): {n_pg_anom:,} абонентів, "
          f"{anom_vol:,.0f} м³ — залишаємо в nm (моdemний профіль)")

# Векторне обчислення
grps  = nm['appliance_group'].values
cts   = nm['consumer_type'].values
anns  = nm['annual_billing'].values
grs_v = nm['grs'].values

pred_records = []
source_counts = {}
missing_profile = set()

for i in range(len(nm)):
    mw, src = get_monthly_w(grps[i], cts[i])
    source_counts[src] = source_counts.get(src, 0) + 1
    if src == 'uniform':
        missing_profile.add((grps[i], cts[i]))

    for m, w in mw.items():
        pred_records.append({
            'account_id': nm.iloc[i]['account_id'],
            'grs':        grs_v[i],
            'month':      m,
            'pred':       float(anns[i]) * w,
        })

print(f"  Джерела ваг: {source_counts}")
if missing_profile:
    print(f"  УВАГА: рівномірний fallback для {len(missing_profile)} груп:")
    for k in missing_profile:
        print(f"    {k}")

pred_df = pd.DataFrame(pred_records)
pred_grs_nm = pred_df.groupby(['grs', 'month'])['pred'].sum().reset_index()
print(f"  Прогноз побудовано: {len(pred_df):,} рядків")

# ── 4. Фактичний модемний білінг по ГРС ─────────────────────────────────────
modem_parts = []
for m, col in bill_cols.items():
    g = md_grs.groupby('grs')[col].sum().reset_index()
    g.columns = ['grs', 'modem_fact']
    g['month'] = m
    modem_parts.append(g)
modem_grs = pd.concat(modem_parts, ignore_index=True)

# ── 5. Фактичний загальний білінг по ГРС ────────────────────────────────────
bill_parts = []
for m, col in bill_cols.items():
    g = ap[ap['grs'].notna()].groupby('grs')[col].sum().reset_index()
    g.columns = ['grs', 'billing']
    g['month'] = m
    bill_parts.append(g)
billing_grs = pd.concat(bill_parts, ignore_index=True)

# ── 6. Збираємо разом ────────────────────────────────────────────────────────
result = billing_grs.merge(pred_grs_nm, on=['grs','month'], how='left')
result = result.merge(modem_grs, on=['grs','month'], how='left')
result['modem_fact'] = result['modem_fact'].fillna(0)
result['pred']       = result['pred'].fillna(0)
result['total_pred'] = result['pred'] + result['modem_fact']
result['error']      = result['total_pred'] - result['billing']
result['bias_pct']   = result['error'] / result['billing'].clip(lower=1) * 100

# Річний білінг ГРС для класифікації
ann_grs = billing_grs.groupby('grs')['billing'].sum()
result['is_large'] = result['grs'].map(ann_grs) >= LARGE_GRS

# ── 7. Підсумки по місяцях ────────────────────────────────────────────────────
print("\n=== 7. Bias по місяцях (всі ГРС) ===")
print(f"  {'Міс':5s}  {'bias%':>7s}  {'MAE':>10s}  {'GRS':>5s}")
print("  " + "-"*35)
monthly_summary = []
for m in range(1, 13):
    sub = result[result['month'] == m]
    wb  = sub['billing'] > 0
    if wb.sum() == 0:
        continue
    bias = sub.loc[wb, 'error'].sum() / sub.loc[wb, 'billing'].sum() * 100
    mae  = sub.loc[wb, 'error'].abs().mean()
    n    = wb.sum()
    monthly_summary.append({'month': m, 'bias': bias, 'mae': mae, 'n_grs': n})
    print(f"  {UA[m]:5s}  {bias:>+7.2f}%  {mae:>10,.0f}  {n:>5}")

ms = pd.DataFrame(monthly_summary)

# Об'ємно-зважений річний bias: Σ error / Σ billing (правильний спосіб)
wb_all = result['billing'] > 0
annual_bias_vol = result.loc[wb_all, 'error'].sum() / result.loc[wb_all, 'billing'].sum() * 100
annual_total_pred = result['total_pred'].sum()
annual_total_bill = result.loc[wb_all, 'billing'].sum()

print(f"\n  Річний (об'ємний):   bias={annual_bias_vol:+.2f}%  "
      f"pred={annual_total_pred:,.0f}  billing={annual_total_bill:,.0f}  "
      f"Δ={annual_total_pred - annual_total_bill:+,.0f} м³")
print(f"  (Просте серед. міс): bias={ms['bias'].mean():+.2f}%  ← не зважений, спотворений серпневим +14% на малому обсязі)")

# ── 8. Грудень детально по ГРС ───────────────────────────────────────────────
print("\n=== 8. Грудень по ГРС ===")
dec = result[result['month'] == 12].copy()
dec = dec.sort_values('billing', ascending=False)
print(f"  {'GRS':45s}  {'bias%':>7s}  {'billing':>12s}  {'pred':>12s}")
print("  " + "-"*80)
for _, r in dec.iterrows():
    if r['billing'] <= 0:
        continue
    lg = ' [L]' if r['is_large'] else ''
    thr = 3 if r['is_large'] else 6
    fl  = ' !' if abs(r['bias_pct']) > thr else '  '
    print(f"  {fl}{r['grs'][:43]:43s}  {r['bias_pct']:>+7.1f}%  "
          f"{r['billing']:>12,.0f}  {r['total_pred']:>12,.0f}{lg}")

wb = dec['billing'] > 0
bias_d = dec.loc[wb,'error'].sum() / dec.loc[wb,'billing'].sum() * 100
mae_d  = dec.loc[wb,'error'].abs().mean()
l_ok   = (dec.loc[dec['is_large'] & wb, 'bias_pct'].abs() <= 3).mean() * 100
s_ok   = (dec.loc[~dec['is_large'] & wb, 'bias_pct'].abs() <= 6).mean() * 100
print(f"\n  Грудень: bias={bias_d:+.2f}%  MAE={mae_d:,.0f}")
print(f"  Вел. ≤3%: {l_ok:.0f}%   Мал. ≤6%: {s_ok:.0f}%")

# ── 9. Порівняння K_v2 vs Annual-decompose ────────────────────────────────────
print("\n=== 9. Порівняння по проблемних місяцях ===")
print(f"  {'Міс':5s}  {'K_v2 bias':>10s}  {'AnnDecomp':>10s}")
kv2 = {1:-3.8, 2:-0.1, 3:+1.3, 4:-7.2, 5:-16.5, 6:-3.2,
       7:-0.1, 8:+3.9, 9:-10.25, 10:-10.5, 11:+1.0, 12:-1.0}
for row in monthly_summary:
    m = row['month']
    print(f"  {UA[m]:5s}  {kv2.get(m, 0):>+10.2f}%  {row['bias']:>+10.2f}%")

# ── 10. Збереження ────────────────────────────────────────────────────────────
print(f"\n=== 10. Збереження → {OUT_XLSX} ===")
with pd.ExcelWriter(OUT_XLSX, engine='openpyxl') as xw:
    result.to_excel(xw, sheet_name='All_GRS_monthly', index=False)
    ms.to_excel(xw, sheet_name='Monthly_summary', index=False)
    dec.to_excel(xw, sheet_name='December_GRS', index=False)

    # Профілі місячних ваг: модемний / білінговий / гібридний
    pw_rows = []
    all_keys = set(monthly_w_cache) | set(billing_monthly_w) | set(hybrid_monthly_w)
    for (grp, ct) in sorted(all_keys):
        for src_label, src_dict in [
            ('modem',   monthly_w_cache),
            ('billing', billing_monthly_w),
            ('hybrid',  hybrid_monthly_w),
        ]:
            mw = src_dict.get((grp, ct))
            if mw is None:
                continue
            row = {'appliance_group': grp, 'consumer_type': ct, 'source': src_label}
            for m in range(1, 13):
                row[UA[m]] = round(mw.get(m, 0) * 100, 2)
            pw_rows.append(row)
    pd.DataFrame(pw_rows).to_excel(xw, sheet_name='Profile_weights_%', index=False)

    # Денний гібридний профіль (для майбутнього RT-прогнозу)
    daily_rows = []
    for (grp, ct), dw in hybrid_daily_w.items():
        for dt, w in sorted(dw.items()):
            daily_rows.append({
                'appliance_group': grp,
                'consumer_type':   ct,
                'date':            dt.strftime('%d.%m.%Y'),
                'month':           dt.month,
                'day_weight':      round(w * 100, 5),  # % від річного
            })
    pd.DataFrame(daily_rows).to_excel(xw, sheet_name='Daily_hybrid_profile', index=False)

print("  DONE")

# ── 11. Красива таблиця bias по ГРС × місяць ────────────────────────────────
print("\n=== 11. Таблиця bias → data/annual_decompose_table.xlsx ===")
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TABLE_XLSX = 'data/annual_decompose_table.xlsx'

# Кольори bias
def bias_fill(val):
    if val is None:
        return None
    v = float(val)
    if v > 20:   return PatternFill('solid', fgColor='FF4444')  # темно-червоний
    if v > 10:   return PatternFill('solid', fgColor='FF8888')
    if v > 5:    return PatternFill('solid', fgColor='FFBBBB')
    if v < -20:  return PatternFill('solid', fgColor='4444FF')  # темно-синій (перевищення)
    if v < -10:  return PatternFill('solid', fgColor='8888FF')
    if v < -5:   return PatternFill('solid', fgColor='BBBBFF')
    return None

wb2 = Workbook()
ws  = wb2.active
ws.title = 'Bias_by_GRS'

MONTHS = list(range(1, 13))
thin = Side(style='thin')
thick = Side(style='medium')

# ── Заголовки ──
ws.cell(1, 1, 'ГРС').font = Font(bold=True)
ws.cell(1, 2, 'Рядок').font = Font(bold=True)
for col_i, m in enumerate(MONTHS, start=3):
    c = ws.cell(1, col_i, UA[m])
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')
ws.cell(1, 15, 'Річний').font = Font(bold=True)
ws.freeze_panes = 'C2'

# Ширини
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 10
for col_i in range(3, 16):
    ws.column_dimensions[get_column_letter(col_i)].width = 10

# Pivot: grs → month → {billing, pred, bias}
grs_list = result[result['billing'] > 0]['grs'].unique()
# Сортуємо по річному білінгу
ann_sort = result.groupby('grs')['billing'].sum().sort_values(ascending=False)
grs_list = [g for g in ann_sort.index if g in set(grs_list)]

row_i = 2
for grs in grs_list:
    sub = result[result['grs'] == grs].set_index('month')

    bill_row  = {}
    pred_row  = {}
    bias_row  = {}
    for m in MONTHS:
        if m in sub.index and sub.loc[m, 'billing'] > 0:
            bill_row[m] = sub.loc[m, 'billing']
            pred_row[m] = sub.loc[m, 'total_pred']
            bias_row[m] = sub.loc[m, 'bias_pct']
        else:
            bill_row[m] = pred_row[m] = bias_row[m] = None

    ann_bill = sum(v for v in bill_row.values() if v)
    ann_pred = sum(v for v in pred_row.values() if v)
    ann_bias = (ann_pred - ann_bill) / ann_bill * 100 if ann_bill else None

    # Рядок 1: Білінг
    ws.cell(row_i, 1, grs)
    ws.cell(row_i, 2, 'Білінг').font = Font(italic=True, color='555555')
    for col_i, m in enumerate(MONTHS, start=3):
        c = ws.cell(row_i, col_i, round(bill_row[m]) if bill_row[m] else '')
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right')
    ws.cell(row_i, 15, round(ann_bill) if ann_bill else '').number_format = '#,##0'
    row_i += 1

    # Рядок 2: Прогноз
    ws.cell(row_i, 1, '')
    ws.cell(row_i, 2, 'Прогноз').font = Font(italic=True, color='555555')
    for col_i, m in enumerate(MONTHS, start=3):
        c = ws.cell(row_i, col_i, round(pred_row[m]) if pred_row[m] else '')
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right')
    ws.cell(row_i, 15, round(ann_pred) if ann_pred else '').number_format = '#,##0'
    row_i += 1

    # Рядок 3: Bias%
    ws.cell(row_i, 1, '')
    ws.cell(row_i, 2, 'Bias%').font = Font(bold=True)
    for col_i, m in enumerate(MONTHS, start=3):
        v = bias_row[m]
        c = ws.cell(row_i, col_i, round(v, 1) if v is not None else '')
        c.number_format = '+0.0;-0.0;0.0'
        c.alignment = Alignment(horizontal='center')
        c.font = Font(bold=True)
        fill = bias_fill(v)
        if fill:
            c.fill = fill
    # Річний bias
    c = ws.cell(row_i, 15, round(ann_bias, 1) if ann_bias is not None else '')
    c.number_format = '+0.0;-0.0;0.0'
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')
    fill = bias_fill(ann_bias)
    if fill:
        c.fill = fill

    # Лінія між ГРС
    for col_i in range(1, 16):
        ws.cell(row_i, col_i).border = Border(bottom=thin)

    row_i += 1

# ── Підсумковий рядок ──
row_i += 1
ws.cell(row_i, 1, 'ВСЬОГО').font = Font(bold=True)
ws.cell(row_i, 2, 'Bias%').font = Font(bold=True)
for col_i, m in enumerate(MONTHS, start=3):
    sub_m = result[(result['month'] == m) & (result['billing'] > 0)]
    if len(sub_m) == 0:
        continue
    b = sub_m['error'].sum() / sub_m['billing'].sum() * 100
    c = ws.cell(row_i, col_i, round(b, 1))
    c.number_format = '+0.0;-0.0;0.0'
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')
    fill = bias_fill(b)
    if fill:
        c.fill = fill
# Річний підсумок
all_b = result[result['billing'] > 0]
ann_b_total = all_b['error'].sum() / all_b['billing'].sum() * 100
c = ws.cell(row_i, 15, round(ann_b_total, 1))
c.number_format = '+0.0;-0.0;0.0'
c.font = Font(bold=True)
c.alignment = Alignment(horizontal='center')
fill = bias_fill(ann_b_total)
if fill:
    c.fill = fill

wb2.save(TABLE_XLSX)
print(f"  Збережено → {TABLE_XLSX}")
print("  Легенда: червоний = перевищення прогнозу, синій = недопрогноз")

# ── 12. Аналіз відхилень по ГРС (травень–листопад) ──────────────────────────
print("\n=== 12. Аналіз відхилень по ГРС (трав–лис) ===")
SUMMER_MONTHS = [5, 6, 7, 8, 9, 10, 11]

grs_rows = []
for grs in ann_sort.index:
    sub = result[result['grs'] == grs]
    if sub.empty:
        continue

    # Річні підсумки
    ann_bill = sub['billing'].sum()
    ann_pred = sub['total_pred'].sum()
    ann_bias = (ann_pred - ann_bill) / ann_bill * 100 if ann_bill > 0 else None

    # Літні місяці
    sub_s = sub[sub['month'].isin(SUMMER_MONTHS)]
    s_bill = sub_s['billing'].sum()
    s_pred = sub_s['total_pred'].sum()
    s_bias = (s_pred - s_bill) / s_bill * 100 if s_bill > 0 else None

    # Зимові місяці (1,2,3,4,12)
    sub_w = sub[sub['month'].isin([1,2,3,4,12])]
    w_bill = sub_w['billing'].sum()
    w_pred = sub_w['total_pred'].sum()
    w_bias = (w_pred - w_bill) / w_bill * 100 if w_bill > 0 else None

    # Модемна частка за рік (modem_fact / billing)
    modem_frac = sub['modem_fact'].sum() / ann_bill * 100 if ann_bill > 0 else 0

    # Варіативність bias по літніх місяцях (std → «шум»)
    monthly_biases = []
    for m in SUMMER_MONTHS:
        r_m = sub[sub['month'] == m]
        if len(r_m) == 1 and r_m['billing'].values[0] > 0:
            monthly_biases.append(float(r_m['bias_pct'].values[0]))
    bias_std = float(np.std(monthly_biases)) if len(monthly_biases) >= 3 else None

    row = {
        'grs': grs,
        'ann_bill_M': round(ann_bill / 1e6, 3),
        'ann_bias%':  round(ann_bias, 1) if ann_bias is not None else None,
        'sum_bias%':  round(s_bias, 1)   if s_bias is not None else None,
        'win_bias%':  round(w_bias, 1)   if w_bias is not None else None,
        'bias_std':   round(bias_std, 1) if bias_std is not None else None,
        'modem_%':    round(modem_frac, 1),
        'is_large':   bool(ann_bill >= LARGE_GRS),
    }
    for m in SUMMER_MONTHS:
        r_m = sub[sub['month'] == m]
        if len(r_m) == 1:
            row[UA[m]] = round(float(r_m['bias_pct'].values[0]), 1)
        else:
            row[UA[m]] = None
    grs_rows.append(row)

gdf = pd.DataFrame(grs_rows).sort_values('ann_bill_M', ascending=False)

# Виводимо ті, де |sum_bias%| > 5
print(f"\n  {'ГРС':42s}  {'Річн':>6s}  {'Лiто':>6s}  {'Зима':>6s}  {'Std':>5s}  {'Mod%':>5s}  "
      + '  '.join(f"{UA[m]:4s}" for m in SUMMER_MONTHS))
print("  " + "─" * 120)
for _, r in gdf.iterrows():
    s = r['sum_bias%']
    if s is None or (abs(s) <= 3 and r['ann_bill_M'] < 0.1):
        continue
    flag = ' !' if s is not None and abs(s) > 10 else '  '
    vals = '  '.join(
        f"{r[UA[m]]:+5.1f}" if r[UA[m]] is not None else "    —"
        for m in SUMMER_MONTHS
    )
    print(f"{flag}{r['grs'][:42]:42s}  {r['ann_bias%']:>+6.1f}  {r['sum_bias%']:>+6.1f}  "
          f"{r['win_bias%']:>+6.1f}  {r['bias_std']:>5.1f}  {r['modem_%']:>5.1f}  {vals}")

# Кореляція modem_% з |sum_bias%|
valid = gdf.dropna(subset=['sum_bias%', 'modem_%'])
corr = np.corrcoef(valid['modem_%'], valid['sum_bias%'].abs())[0, 1]
print(f"\n  Корел. (modem_% vs |sum_bias%|): {corr:+.3f}")
print(f"  Середній |sum_bias%| для великих ГРС: "
      f"{gdf[gdf['is_large']]['sum_bias%'].abs().mean():.1f}%")
print(f"  Середній |sum_bias%| для малих ГРС:   "
      f"{gdf[~gdf['is_large']]['sum_bias%'].abs().mean():.1f}%")

# Зберегти в Excel
GRS_XLSX = 'data/grs_summer_bias.xlsx'
with pd.ExcelWriter(GRS_XLSX, engine='openpyxl') as xw:
    gdf.to_excel(xw, sheet_name='GRS_summary', index=False)

    # Зведена таблиця: рядки = ГРС, колонки = місяці 1-12, значення = bias%
    pivot = result.pivot_table(
        index='grs', columns='month', values='bias_pct', aggfunc='first'
    ).reset_index()
    pivot.columns = ['grs'] + [UA[m] for m in range(1, 13) if m in pivot.columns[1:].tolist() or True]
    # перейменуємо числові стовпці
    pivot = result.pivot_table(
        index='grs', columns='month', values='bias_pct', aggfunc='first'
    )
    pivot.columns = [UA[c] for c in pivot.columns]
    pivot = pivot.reset_index()
    ann_b = result.groupby('grs').apply(
        lambda x: (x['total_pred'].sum() - x['billing'].sum()) / x['billing'].sum() * 100
        if x['billing'].sum() > 0 else None
    ).rename('Річний')
    pivot = pivot.merge(ann_b, on='grs', how='left')
    pivot['ann_bill_M'] = pivot['grs'].map(ann_sort / 1e6).round(3)
    pivot = pivot.sort_values('ann_bill_M', ascending=False)
    pivot.to_excel(xw, sheet_name='Bias_heatmap', index=False)

    # Форматування heatmap
    ws3 = xw.book['Bias_heatmap']
    for row in ws3.iter_rows(min_row=2, min_col=2, max_col=ws3.max_column):
        for cell in row:
            try:
                v = float(cell.value)
                cell.number_format = '+0.0;-0.0;0.0'
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
                fill = bias_fill(v)
                if fill:
                    cell.fill = fill
            except (TypeError, ValueError):
                pass

print(f"\n  Збережено → {GRS_XLSX}")

# ── 13. Порівняння декомпозиту vs білінгу ────────────────────────────────────
print("\n=== 13. Декомпозит vs Білінг → data/decompose_vs_billing.xlsx ===")
DVB_XLSX = 'data/decompose_vs_billing.xlsx'

from openpyxl.utils import get_column_letter

def num_fill(v, threshold=5.0):
    """Кольорова заливка для числових відхилень."""
    if v is None:
        return None
    try:
        v = float(v)
    except (TypeError, ValueError):
        return None
    if v > threshold:
        return PatternFill('solid', fgColor='FFCCCC')   # червоний — прогноз > білінг
    elif v < -threshold:
        return PatternFill('solid', fgColor='CCE5FF')   # синій — прогноз < білінг
    return None

with pd.ExcelWriter(DVB_XLSX, engine='openpyxl') as xw:

    # ── Лист 1: Місячні підсумки ──────────────────────────────────────────────
    rows_m = []
    for m in range(1, 13):
        sub = result[result['month'] == m]
        bill  = sub['billing'].sum()
        pred  = sub['total_pred'].sum()
        pred_nm  = sub['pred'].sum()
        pred_mod = sub['modem_fact'].sum()
        bias  = (pred - bill) / bill * 100 if bill > 0 else None
        rows_m.append({
            'Місяць':        UA[m],
            'Білінг, м³':    round(bill),
            'Декомп, м³':    round(pred),
            '  НМ, м³':      round(pred_nm),
            '  Модем, м³':   round(pred_mod),
            'Bias%':         round(bias, 2) if bias is not None else None,
            'Δ, м³':         round(pred - bill),
        })
    df_m = pd.DataFrame(rows_m)
    # Підсумковий рядок
    total_bill = result['billing'].sum()
    total_pred = result['total_pred'].sum()
    df_m.loc[len(df_m)] = {
        'Місяць':       'РАЗОМ',
        'Білінг, м³':   round(total_bill),
        'Декомп, м³':   round(total_pred),
        '  НМ, м³':     round(result['pred'].sum()),
        '  Модем, м³':  round(result['modem_fact'].sum()),
        'Bias%':        round((total_pred - total_bill) / total_bill * 100, 2) if total_bill > 0 else None,
        'Δ, м³':        round(total_pred - total_bill),
    }
    df_m.to_excel(xw, sheet_name='Місяці_всього', index=False)

    ws_m = xw.book['Місяці_всього']
    for row in ws_m.iter_rows(min_row=2, min_col=1, max_row=ws_m.max_row, max_col=ws_m.max_column):
        for cell in row:
            if cell.column_letter == 'F':   # Bias%
                try:
                    v = float(cell.value)
                    cell.number_format = '+0.00;-0.00;0.00'
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
                    fill = num_fill(v, 2.0)
                    if fill:
                        cell.fill = fill
                except (TypeError, ValueError):
                    pass
            elif cell.column_letter in ('B', 'C', 'D', 'E', 'G'):
                try:
                    float(cell.value)
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                except (TypeError, ValueError):
                    pass

    # ── Лист 2: ГРС × Місяць (Білінг) ───────────────────────────────────────
    piv_bill = result.pivot_table(
        index='grs', columns='month', values='billing', aggfunc='sum'
    ).fillna(0).reset_index()
    piv_bill.columns = ['ГРС'] + [UA[c] for c in piv_bill.columns[1:]]
    piv_bill['Річний'] = piv_bill[[UA[m] for m in range(1, 13) if UA[m] in piv_bill.columns]].sum(axis=1)
    piv_bill = piv_bill.sort_values('Річний', ascending=False)
    piv_bill.to_excel(xw, sheet_name='ГРС_Білінг', index=False)

    ws_b = xw.book['ГРС_Білінг']
    for row in ws_b.iter_rows(min_row=2, min_col=2, max_row=ws_b.max_row, max_col=ws_b.max_column):
        for cell in row:
            try:
                float(cell.value)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            except (TypeError, ValueError):
                pass

    # ── Лист 3: ГРС × Місяць (Декомпозит) ──────────────────────────────────
    piv_pred = result.pivot_table(
        index='grs', columns='month', values='total_pred', aggfunc='sum'
    ).fillna(0).reset_index()
    piv_pred.columns = ['ГРС'] + [UA[c] for c in piv_pred.columns[1:]]
    piv_pred['Річний'] = piv_pred[[UA[m] for m in range(1, 13) if UA[m] in piv_pred.columns]].sum(axis=1)
    piv_pred = piv_pred.sort_values('Річний', ascending=False)
    piv_pred.to_excel(xw, sheet_name='ГРС_Декомпозит', index=False)

    ws_p = xw.book['ГРС_Декомпозит']
    for row in ws_p.iter_rows(min_row=2, min_col=2, max_row=ws_p.max_row, max_col=ws_p.max_column):
        for cell in row:
            try:
                float(cell.value)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            except (TypeError, ValueError):
                pass

    # ── Лист 4: ГРС × Місяць (Bias%) ───────────────────────────────────────
    piv_bias = result.pivot_table(
        index='grs', columns='month', values='bias_pct', aggfunc='first'
    ).reset_index()
    piv_bias.columns = ['ГРС'] + [UA[c] for c in piv_bias.columns[1:]]
    # Річний bias
    ann_bias_ser = result.groupby('grs').apply(
        lambda x: (x['total_pred'].sum() - x['billing'].sum()) / x['billing'].sum() * 100
        if x['billing'].sum() > 0 else None, include_groups=False
    ).rename('Річний').reset_index()
    piv_bias = piv_bias.merge(ann_bias_ser, left_on='ГРС', right_on='grs', how='left').drop(columns=['grs'])
    # Річний обсяг (для сортування)
    piv_bias['_sort'] = piv_bias['ГРС'].map(ann_sort)
    piv_bias = piv_bias.sort_values('_sort', ascending=False).drop(columns=['_sort'])
    piv_bias.to_excel(xw, sheet_name='ГРС_Bias%', index=False)

    ws_bias = xw.book['ГРС_Bias%']
    for row in ws_bias.iter_rows(min_row=2, min_col=2, max_row=ws_bias.max_row, max_col=ws_bias.max_column):
        for cell in row:
            try:
                v = float(cell.value)
                cell.number_format = '+0.0;-0.0;0.0'
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
                fill = num_fill(v, 5.0)
                if fill:
                    cell.fill = fill
            except (TypeError, ValueError):
                pass

print(f"  Збережено → {DVB_XLSX}")
print("  Листи: Місяці_всього | ГРС_Білінг | ГРС_Декомпозит | ГРС_Bias%")
