"""
grs_modem_model.py

Модель прогнозування місячного споживання по ГРС.

Ідея: модеми (2507 шт, ~99% з ГРС-1,3 та ГРС-2) — «зразкові» споживачі.
Їх реальне денне/місячне споживання = головні фічі моделі.

X = [modem_heat(m), modem_cook(m), hdh_18c(m), avg_temp(m)]
y = billing_total(GRS, m)

Ridge per GRS, навчання на 2025 (12 місяців), перевірка Jan 2026.
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
import numpy as np
from sklearn.linear_model import Ridge, RidgeCV
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import r2_score, mean_absolute_error
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from pobut_predictor import GROUP_MERGE, META_COLS

UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}
MONTHS = list(range(1, 13))
OUT_XLSX = 'data/grs_modem_model.xlsx'

# Дві ГРС з найбруднішим білінгом (виключаємо з тренування як target,
# але їх модеми ВСЕ ОДНО використовуємо як X-фічі)
SKIP_BILLING_GRS = {'С-1(Запоріж) ГРС-1,3 кільце', 'ГРС (2) м.Запоріжжя'}

# ── 1. Модемні дані → місячні суми по кластерах ───────────────────────────────
print("=== 1. Завантаження модемних даних ===")
md = pd.read_csv('data/input/profile_pobut_daily_result.csv', sep=';', low_memory=False)
md.columns = META_COLS + list(md.columns[len(META_COLS):])
md = md[md['gas_off'].isna() & md['alternative'].isna() & md['dacha'].isna()]
md['appliance_group'] = md['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
md['has_op'] = md['appliance_group'].str.contains('ОП', na=False)

day_cols = [c for c in md.columns if '.' in c and len(c) == 10]
dates = pd.to_datetime([c for c in day_cols], format='%d.%m.%Y')

heat_daily = pd.Series(md[md['has_op']][day_cols].sum().values, index=dates, name='heat')
cook_daily = pd.Series(md[~md['has_op']][day_cols].sum().values, index=dates, name='cook')

# Місячні агрегати
heat_monthly = heat_daily.resample('MS').sum()
cook_monthly = cook_daily.resample('MS').sum()

modem_monthly = pd.DataFrame({'heat': heat_monthly, 'cook': cook_monthly})
modem_monthly['total'] = modem_monthly['heat'] + modem_monthly['cook']

print(f"  Modems: {len(md):,}  (heating={md['has_op'].sum()}, cooking={(~md['has_op']).sum()})")
print(f"  Данi: {dates.min().date()} → {dates.max().date()}")
print(f"  Місяців: {len(modem_monthly)}")

# ── 2. Погодні дані → місячні агрегати ────────────────────────────────────────
print("\n=== 2. Погодні дані ===")
wd = pd.read_csv('../data/weather_2025_daily_contractual.csv')
wd['date'] = pd.to_datetime(wd['date']).dt.normalize()
wd = wd.set_index('date')

# Дедуплікація (може бути дублікат першого рядку)
wd = wd[~wd.index.duplicated(keep='first')]

wd_monthly = wd.resample('MS').agg({
    'temperature': 'mean',
    'hdh_18c': 'sum',
    'hdh_15c': 'sum',
    'pressure': 'mean',
    'hours_below_0c': 'sum',
    'temperature_min': 'min',
})
print(f"  Погода: {wd.index.min().date()} → {wd.index.max().date()}")

# ── 3. Білінг по ГРС (місячний) — з annual_decompose_result.xlsx ──────────────
print("\n=== 3. Білінг по ГРС ===")
dc = pd.read_excel('data/annual_decompose_result.xlsx', sheet_name='All_GRS_monthly')
dc['grs'] = dc['grs'].astype(str)
dc['month'] = dc['month'].astype(int)

# Додамо дату (2025)
dc['date'] = pd.to_datetime(dc['month'].astype(str) + '-2025', format='%m-%Y')

all_grs = sorted(dc['grs'].unique())
target_grs = [g for g in all_grs if g not in SKIP_BILLING_GRS]
print(f"  Всього ГРС: {len(all_grs)}  |  Тренувальних (без SKIP): {len(target_grs)}")

# ── 4. Об'єднання фічей ────────────────────────────────────────────────────────
print("\n=== 4. Побудова матриці фічей ===")

# Pivot billing: rows=date, cols=grs
billing_pivot = dc.pivot(index='date', columns='grs', values='billing')
billing_pivot.index = pd.to_datetime(billing_pivot.index)

# Загальна таблиця фічей (по місяцях)
features = modem_monthly.copy()
features = features.join(wd_monthly[['temperature', 'hdh_18c', 'pressure', 'hours_below_0c']])
features = features.join(billing_pivot, how='left')
features = features.dropna(subset=['heat', 'cook'])

# Тренувальний набір: тільки 2025 місяці де є білінг
train_mask = (features.index.year == 2025)
train_df   = features[train_mask].copy()
print(f"  Train місяців (2025): {train_mask.sum()}")

# Jan 2026 — є модемні дані, немає білінгу → це наш predict
pred_mask = (features.index.year == 2026)
pred_df   = features[pred_mask].copy()
print(f"  Predict місяців (2026): {pred_mask.sum()}")

# ── 5. Тренування Ridge per GRS ───────────────────────────────────────────────
print("\n=== 5. Тренування моделей ===")

FEATURE_COLS = ['heat', 'cook', 'hdh_18c', 'temperature']

X_train = train_df[FEATURE_COLS].values
X_pred  = pred_df[FEATURE_COLS].values if len(pred_df) > 0 else None

scaler = StandardScaler()
X_train_s = scaler.fit_transform(X_train)
X_pred_s  = scaler.transform(X_pred) if X_pred is not None else None

models = {}
results = []

print(f"\n  {'ГРС':<35} {'R²':>6} {'MAE тис':>9} {'MAPE%':>8}  Bias%")
print("  " + "-"*75)

for grs in target_grs:
    if grs not in train_df.columns:
        continue
    y = train_df[grs].values
    if np.all(y == 0) or np.sum(~np.isnan(y)) < 6:
        continue

    valid = ~np.isnan(y)
    if valid.sum() < 6:
        continue

    Xv = X_train_s[valid]
    yv = y[valid]

    model = RidgeCV(alphas=[0.01, 0.1, 1, 10, 100], cv=min(5, len(yv)))
    model.fit(Xv, yv)
    models[grs] = model

    y_pred_is = model.predict(Xv)
    r2   = r2_score(yv, y_pred_is)
    mae  = mean_absolute_error(yv, y_pred_is)
    mape = np.mean(np.abs((y_pred_is - yv) / np.where(yv==0, 1, yv))) * 100
    bias = (y_pred_is.sum() - yv.sum()) / yv.sum() * 100

    results.append({'grs': grs, 'r2': r2, 'mae': mae, 'mape': mape, 'bias': bias,
                    'y_train': yv, 'y_pred_is': y_pred_is,
                    'months_train': train_df.index[valid].tolist()})

    label = grs[:34]
    print(f"  {label:<35} {r2:>6.3f} {mae/1e3:>9.2f} {mape:>7.1f}%  {bias:>+.1f}%")

r2s   = [r['r2'] for r in results]
maes  = [r['mae'] for r in results]
mapes = [r['mape'] for r in results]
print("  " + "-"*75)
print(f"  {'Середнє':35} {np.mean(r2s):>6.3f} {np.mean(maes)/1e3:>9.2f} {np.mean(mapes):>7.1f}%")

# ── 6. Прогноз Jan 2026 ────────────────────────────────────────────────────────
if X_pred_s is not None and len(pred_df) > 0:
    print(f"\n=== 6. Прогноз January 2026 ===")
    print(f"  Modem heat Jan2026: {pred_df['heat'].values[0]/1e3:.1f} тис м³")
    print(f"  Modem cook Jan2026: {pred_df['cook'].values[0]/1e3:.1f} тис м³")
    print(f"  HDH-18 Jan2026:     {pred_df['hdh_18c'].values[0]:.0f}")
    print(f"\n  {'ГРС':<35} {'Прогноз тис м³':>15}  (порівн. Jan 2025)")
    print("  " + "-"*60)
    jan26_preds = {}
    for r in results:
        grs = r['grs']
        pred_val = models[grs].predict(X_pred_s)[0]
        jan26_preds[grs] = pred_val
        # Jan 2025 billing для порівняння
        jan25_row = dc[(dc['grs']==grs) & (dc['month']==1)]
        jan25 = jan25_row['billing'].values[0] if len(jan25_row) > 0 else np.nan
        comp = f"  Jan2025: {jan25/1e3:.1f}" if not np.isnan(jan25) else ""
        print(f"  {grs[:34]:<35} {pred_val/1e3:>15.2f}{comp}")

# ── 7. Місячні передбачення (in-sample) ───────────────────────────────────────
print(f"\n=== 7. Місячний in-sample: Billing vs Predict ===")
is_results = {}
for r in results:
    grs = r['grs']
    for i, dt in enumerate(r['months_train']):
        m = dt.month
        if grs not in is_results:
            is_results[grs] = {}
        is_results[grs][m] = {'billing': r['y_train'][i], 'pred': r['y_pred_is'][i]}

print(f"\n  {'ГРС':<35}  " + '  '.join(f'{UA[m]:>5}' for m in MONTHS))
print("  " + "-"*100)
for grs in target_grs:
    if grs not in is_results:
        continue
    row_parts = []
    for m in MONTHS:
        if m in is_results[grs]:
            b = is_results[grs][m]['billing']
            p = is_results[grs][m]['pred']
            bias = (p - b) / b * 100 if b != 0 else 0
            row_parts.append(f"{bias:>+5.0f}%")
        else:
            row_parts.append("     ")
    print(f"  {grs[:34]:<35}  " + '  '.join(row_parts))

# ── 8. Збереження Excel ────────────────────────────────────────────────────────
print(f"\nЗберігаю {OUT_XLSX}...")
wb = Workbook()

# --- Sheet 1: Модельні показники ---
ws1 = wb.active
ws1.title = 'Якість_моделей'
hdr = ['ГРС', 'R²', 'MAE тис м³', 'MAPE%', 'Bias%', 'Alpha']
for ci, h in enumerate(hdr, 1):
    ws1.cell(1, ci, h).font = Font(bold=True)
for ri, r in enumerate(results, 2):
    ws1.cell(ri, 1, r['grs'])
    ws1.cell(ri, 2, round(r['r2'], 4))
    ws1.cell(ri, 3, round(r['mae']/1e3, 2))
    ws1.cell(ri, 4, round(r['mape'], 1))
    ws1.cell(ri, 5, round(r['bias'], 1))
    ws1.cell(ri, 6, round(models[r['grs']].alpha_, 3))

# --- Sheet 2: Зведена billing vs predict ---
ws2 = wb.create_sheet('Billing_vs_Predict')
# Заголовки: ГРС | Січ_bill | Січ_pred | Січ_bias% | Лют_bill | ...
hdr2 = ['ГРС']
for m in MONTHS:
    hdr2 += [f'{UA[m]}_факт', f'{UA[m]}_прогноз', f'{UA[m]}_bias%']
hdr2 += ['Jan2026_прогноз']
for ci, h in enumerate(hdr2, 1):
    cell = ws2.cell(1, ci, h)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

def bias_color(bias_pct):
    ab = abs(bias_pct)
    if ab < 5:   return 'C6EFCE'
    if ab < 10:  return 'FFEB9C'
    if ab < 20:  return 'FFCC99'
    return 'FFC7CE'

for ri, r in enumerate(results, 2):
    grs = r['grs']
    ws2.cell(ri, 1, grs)
    for i, m in enumerate(MONTHS):
        ci_base = 2 + (m-1)*3
        if m in is_results.get(grs, {}):
            b = is_results[grs][m]['billing']
            p = is_results[grs][m]['pred']
            bias = (p-b)/b*100 if b != 0 else 0
            ws2.cell(ri, ci_base,   round(b/1e3, 1))
            ws2.cell(ri, ci_base+1, round(p/1e3, 1))
            bc = ws2.cell(ri, ci_base+2, round(bias, 1))
            bc.fill = PatternFill('solid', fgColor=bias_color(bias))
    # Jan 2026
    if grs in jan26_preds:
        ws2.cell(ri, 2+12*3, round(jan26_preds[grs]/1e3, 1))

# --- Sheet 3: Modem сигнал ---
ws3 = wb.create_sheet('Modem_signal')
hdr3 = ['Місяць', 'Heat тис м³', 'Cook тис м³', 'Total тис м³', 'HDH-18', 'Avg Temp']
for ci, h in enumerate(hdr3, 1):
    ws3.cell(1, ci, h).font = Font(bold=True)
for ri, (dt, row) in enumerate(features.iterrows(), 2):
    ws3.cell(ri, 1, dt.strftime('%Y-%m'))
    ws3.cell(ri, 2, round(row['heat']/1e3, 1))
    ws3.cell(ri, 3, round(row['cook']/1e3, 1))
    ws3.cell(ri, 4, round((row['heat']+row['cook'])/1e3, 1))
    if not pd.isna(row.get('hdh_18c', np.nan)):
        ws3.cell(ri, 5, round(float(row['hdh_18c']), 0))
    if not pd.isna(row.get('temperature', np.nan)):
        ws3.cell(ri, 6, round(float(row['temperature']), 1))

# --- Sheet 4: Jan 2026 прогноз ---
ws4 = wb.create_sheet('Jan2026_прогноз')
for ci, h in enumerate(['ГРС', 'Прогноз тис м³', 'Jan 2025 факт тис м³', 'Δ%'], 1):
    ws4.cell(1, ci, h).font = Font(bold=True)
for ri, (grs, pred_val) in enumerate(sorted(jan26_preds.items(), key=lambda x: -x[1]), 2):
    jan25_row = dc[(dc['grs']==grs) & (dc['month']==1)]
    jan25 = jan25_row['billing'].values[0] if len(jan25_row) > 0 else np.nan
    ws4.cell(ri, 1, grs)
    ws4.cell(ri, 2, round(pred_val/1e3, 2))
    if not np.isnan(jan25):
        ws4.cell(ri, 3, round(jan25/1e3, 2))
        delta = (pred_val - jan25) / jan25 * 100
        ws4.cell(ri, 4, round(delta, 1))

wb.save(OUT_XLSX)
print(f"Збережено: {OUT_XLSX}")
print("\nГотово!")
