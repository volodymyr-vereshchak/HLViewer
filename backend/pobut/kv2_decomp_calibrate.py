"""
kv2_decomp_calibrate.py

Перекалібровка P_HEAT та P_nonheat K_v2 під annual_decompose.

Підхід: місячні калібрувальні множники
  C(grp, ct, m) = Σ_i annual_billing_i × W_hybrid(m) / Σ_i kv2(i, m)

Результат:
  P_HEAT_new(m)           = P_HEAT_v2(m) × C_op_heat(m)   ← один множник для всіх ОП
  P_nonheat_new(grp,ct,m) = P_nonheat(grp,ct,m) × C(grp,ct,m)  ← для решти
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter

from pobut_predictor import PobUtPredictor, GROUP_MERGE, META_COLS, CT_PRIV, HEATING_MONTHS, SUMMER_MONTHS


# ── K_v2 класи (скопійовано з kv2_vs_decompose.py, без top-level side effects) ─
class _PobUtPredictorConfigK(PobUtPredictor):
    KI_HEAT_CSV    = 'data/profiles/ki_heating_consumers.csv'
    KI_NONHEAT_CSV = 'data/profiles/ki_nonheat_consumers.csv'
    P_NONHEAT_CSV  = 'data/profiles/p_modem_nonheat.csv'
    P_HEAT = {1: 4.3167, 2: 4.3717, 3: 2.7025, 4: 1.3619,
              10: 1.1902, 11: 2.7532, 12: 3.9867}

    def _prepare_non_modem(self):
        super()._prepare_non_modem()
        self._load_k_profiles()

    def _load_k_profiles(self):
        p_df = pd.read_csv(self.P_NONHEAT_CSV)
        self._p_nonheat = {
            (r['appliance_group'], r['consumer_type'], int(r['month'])): float(r['P_modem_median'])
            for _, r in p_df.iterrows()
        }
        ki_h = pd.read_csv(self.KI_HEAT_CSV,
                           usecols=['account_id', 'grs', 'appliance_group', 'consumer_type', 'k_i_clipped'])
        ki_h['account_id'] = pd.to_numeric(ki_h['account_id'], errors='coerce')
        ki_h = ki_h[ki_h['k_i_clipped'].notna() & (ki_h['k_i_clipped'] > 0)]
        self._ki_heat_aid    = dict(zip(ki_h['account_id'].astype(int), ki_h['k_i_clipped'].astype(float)))
        self._ki_heat_grs    = ki_h.groupby(['grs', 'appliance_group', 'consumer_type'])['k_i_clipped'].median().to_dict()
        self._ki_heat_global = ki_h.groupby(['appliance_group', 'consumer_type'])['k_i_clipped'].median().to_dict()
        ki_nh = pd.read_csv(self.KI_NONHEAT_CSV,
                            usecols=['account_id', 'grs', 'appliance_group', 'consumer_type', 'k_i_clipped'])
        ki_nh['account_id'] = pd.to_numeric(ki_nh['account_id'], errors='coerce')
        ki_nh = ki_nh[ki_nh['k_i_clipped'].notna() & (ki_nh['k_i_clipped'] > 0)]
        self._ki_nh_aid    = dict(zip(ki_nh['account_id'].astype(int), ki_nh['k_i_clipped'].astype(float)))
        self._ki_nh_grs    = ki_nh.groupby(['grs', 'appliance_group', 'consumer_type'])['k_i_clipped'].median().to_dict()
        self._ki_nh_global = ki_nh.groupby(['appliance_group', 'consumer_type'])['k_i_clipped'].median().to_dict()
        self._build_nm_arrays()

    def _build_nm_arrays(self):
        nm   = self._non_modem
        n    = len(nm)
        aids = nm['account_id'].values.astype(int)
        grps = nm['appliance_group'].values
        cts  = nm['consumer_type'].values
        grs  = nm['grs'].values
        ki_h  = np.full(n, 0.85, dtype=np.float32)
        ki_nh = np.full(n, 1.0,  dtype=np.float32)
        for i in range(n):
            aid, grp, ct, g = int(aids[i]), grps[i], cts[i], grs[i]
            v = (self._ki_heat_aid.get(aid)
                 or self._ki_heat_grs.get((g, grp, ct))
                 or self._ki_heat_grs.get((g, grp, CT_PRIV))
                 or self._ki_heat_global.get((grp, ct))
                 or self._ki_heat_global.get((grp, CT_PRIV), 0.85))
            ki_h[i] = float(v or 0.85)
            v = (self._ki_nh_aid.get(aid)
                 or self._ki_nh_grs.get((g, grp, ct))
                 or self._ki_nh_grs.get((g, grp, CT_PRIV))
                 or self._ki_nh_global.get((grp, ct))
                 or self._ki_nh_global.get((grp, CT_PRIV), 1.0))
            ki_nh[i] = float(v or 1.0)
        nm['_ki_heat']    = ki_h
        nm['_ki_nonheat'] = ki_nh
        p_mat = np.zeros((n, 12), dtype=np.float32)
        for m in range(1, 13):
            for i in range(n):
                p_mat[i, m - 1] = self._p_nonheat.get((grps[i], cts[i], m), 0.0)
        self._p_nonheat_matrix = p_mat

    def _predict_non_modem(self, periods):
        nm      = self._non_modem
        aids    = nm['account_id'].values
        grps    = nm['appliance_group'].values
        grs_v   = nm['grs'].values
        areas   = nm['heated_area'].values.astype(np.float64)
        has_op  = nm['has_OP'].values
        is_dach = nm['is_dacha'].values
        ki_h    = nm['_ki_heat'].values.astype(np.float64)
        ki_nh   = nm['_ki_nonheat'].values.astype(np.float64)
        p_mat   = self._p_nonheat_matrix.astype(np.float64)
        records = []
        for (y, m) in periods:
            is_heat = m in HEATING_MONTHS
            active  = ~is_dach | np.isin(m, list(SUMMER_MONTHS))
            p_col   = p_mat[:, m - 1]
            vols    = np.zeros(len(nm))
            if is_heat:
                p_h = self.P_HEAT.get(m, 0.0)
                op  = active & has_op
                vols[op] = ki_h[op] * areas[op] * p_h
            else:
                op = active & has_op
                vols[op] = ki_nh[op] * p_col[op]
            non_op = active & ~has_op
            vols[non_op] = ki_nh[non_op] * p_col[non_op]
            vols = np.maximum(vols, 0.0)
            mask = (vols > 0) & active
            if not mask.any():
                continue
            records.append(pd.DataFrame({
                'account_id':      aids[mask],
                'appliance_group': grps[mask],
                'year': y, 'month': m,
                'volume':          vols[mask],
                'grs':             grs_v[mask],
                'type':            'pred',
            }))
        return pd.concat(records, ignore_index=True) if records else pd.DataFrame()


class PobUtPredictorConfigK_v2(_PobUtPredictorConfigK):
    """K_v2: P_HEAT з фактичних модемних даних 2025."""
    P_HEAT = {1: 3.8374, 2: 4.3598, 3: 2.7025, 4: 1.3619,
              10: 1.4048, 11: 2.3105, 12: 3.7630}

MODEM_FILE = 'data/input/profile_pobut_daily_result.csv'
SUBS_FILE  = 'data/input/all_pobut_enriched.csv'
OUT_XLSX   = 'data/kv2_decomp_calibrate.xlsx'

MONTH_ENG = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
             'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}
MONTHS = list(range(1, 13))
LARGE  = 5_000_000

GROUPS = [
    ('ОП,ПГ',     'Приватний сектор'),
    ('ОП,ПГ',     'Багатоквартирний сектор'),
    ('ОП,ПГ,ВПГ', 'Приватний сектор'),
    ('ОП,ПГ,ВПГ', 'Багатоквартирний сектор'),
    ('ПГ',        'Багатоквартирний сектор'),
    ('ПГ,ВПГ',    'Багатоквартирний сектор'),
]

# Групи без ОП: декомпозит використовує модемний (споживчий) профіль,
# тому калібрування K_v2 також повинно цілитись у модемний профіль.
MODEM_PROFILE_GROUPS = {
    ('ПГ',    'Багатоквартирний сектор'),
    ('ПГ,ВПГ','Багатоквартирний сектор'),
    ('ПГ',    'Приватний сектор'),
    ('ПГ,ВПГ','Приватний сектор'),
}

# ── Excel helpers ──────────────────────────────────────────────────────────────
thin  = Side(style='thin')
thick = Side(style='medium')
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF')

def bias_fill(v):
    if v is None: return None
    av = abs(v)
    if v > 0:
        if av > 20: return PatternFill('solid', fgColor='FF2222')
        if av > 10: return PatternFill('solid', fgColor='FF7777')
        if av >  5: return PatternFill('solid', fgColor='FFBBBB')
    else:
        if av > 20: return PatternFill('solid', fgColor='2244FF')
        if av > 10: return PatternFill('solid', fgColor='7799FF')
        if av >  5: return PatternFill('solid', fgColor='BBCCFF')
    return None

# ═══════════════════════════════════════════════════════════════════════════════
# БЛОК 1: K_v2 оригінальний прогноз
# ═══════════════════════════════════════════════════════════════════════════════
print("=== 1. K_v2 прогноз (оригінальний) ===")
p_orig  = PobUtPredictorConfigK_v2(MODEM_FILE, SUBS_FILE, mode='offline', summer_op_vpg=False)
df_orig = p_orig.predict(years=[2025])
df_pred_nm = df_orig[df_orig['type'] == 'pred'].copy()
print(f"  Non-modem прогноз: {len(df_pred_nm):,} рядків")

# ═══════════════════════════════════════════════════════════════════════════════
# БЛОК 2: Hybrid monthly weights (репліка annual_decompose, блоки 2a–2c)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== 2. Hybrid monthly weights ===")

# 2a. Денні профілі модемів
md = pd.read_csv(MODEM_FILE, sep=';', low_memory=False)
md.columns = META_COLS + list(md.columns[len(META_COLS):])
md['appliance_group'] = md['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
md['consumer_type']   = md['consumer_type'].fillna(CT_PRIV)
md = md[md['gas_off'].isna() & md['alternative'].isna() & md['dacha'].isna()]
modem_ids = set(pd.to_numeric(md.iloc[:, 0], errors='coerce').dropna().astype(int))

date_cols_2025 = []
for c in md.columns:
    if isinstance(c, str) and c.count('.') == 2 and len(c) == 10:
        try:
            d_s, m_s, y_s = c.split('.')
            dt = pd.Timestamp(f'{y_s}-{m_s}-{d_s}')
            if dt.year == 2025:
                date_cols_2025.append((dt, c))
        except:
            pass
date_cols_2025.sort()
day_col_names = [col for _, col in date_cols_2025]
print(f"  Денних колонок 2025: {len(date_cols_2025)}")

MIN_CONSUMERS = 5
MIN_ANNUAL    = 10.0
profile_raw   = {}

for grp, ct in GROUPS:
    mask = (md['appliance_group'] == grp) & (md['consumer_type'] == ct)
    sub  = md[mask]
    if len(sub) < MIN_CONSUMERS:
        continue
    data = (sub[day_col_names]
            .apply(pd.to_numeric, errors='coerce')
            .fillna(0).clip(lower=0).values.astype(np.float64))
    annual = data.sum(axis=1)
    valid  = annual >= MIN_ANNUAL
    if valid.sum() < MIN_CONSUMERS:
        continue
    fracs = data[valid] / annual[valid, np.newaxis]
    daily = {}
    for j, (dt, _) in enumerate(date_cols_2025):
        f = fracs[:, j]
        p99 = np.percentile(f, 99)
        f_clean = f[f <= p99]
        if len(f_clean) < MIN_CONSUMERS:
            continue
        daily[dt] = float(np.median(f_clean))
    total = sum(daily.values())
    if total > 0:
        profile_raw[(grp, ct)] = {dt: v / total for dt, v in daily.items()}

# 2b. Завантаження білінгу та білінгові ваги з "чистих" споживачів
ap = pd.read_csv(SUBS_FILE, low_memory=False)
ap['account_id']      = pd.to_numeric(ap['account_id'], errors='coerce')
ap['appliance_group'] = ap['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
ap['consumer_type']   = ap['consumer_type'].fillna(CT_PRIV)
ap['heated_area']     = pd.to_numeric(ap['heated_area'], errors='coerce').fillna(55).clip(lower=1)
ap['gas_off']         = ap['gas_off'].notna()

bill_cols = {}
for eng, m_num in MONTH_ENG.items():
    col = f'{eng}_2025'
    if col in ap.columns:
        ap[col] = pd.to_numeric(ap[col], errors='coerce').fillna(0)
        bill_cols[m_num] = col
ap['annual_billing'] = ap[[bill_cols[m] for m in range(1, 13) if m in bill_cols]].sum(axis=1)

_mc = [bill_cols[m] for m in range(1, 13) if m in bill_cols]
ap['max_month_bill']   = ap[_mc].max(axis=1)
ap['n_months_nonzero'] = (ap[_mc] > 0).sum(axis=1)
ap['max_month_frac']   = ap['max_month_bill'] / ap['annual_billing'].clip(lower=1)

clean_mask = (
    ~ap['account_id'].isin(modem_ids) & ~ap['gas_off'] & ap['grs'].notna()
    & (ap['annual_billing'] > 0)
    & (ap['max_month_frac'] < 0.5) & (ap['n_months_nonzero'] >= 8)
)
clean_all = ap[clean_mask].copy()
print(f"  'Чистих' споживачів для білінг. профілю: {len(clean_all):,}")

billing_monthly_w = {}
MIN_CLEAN = 100
for grp, ct in GROUPS:
    key = (grp, ct)
    # ПГ-групи: калібруємось під модемний профіль (як і annual_decompose)
    if key in MODEM_PROFILE_GROUPS:
        print(f"  [{grp}, {ct[:15]}]: модемний профіль (consumption-based, без білінгу)")
        continue
    sub_c = clean_all[
        (clean_all['appliance_group'] == grp) & (clean_all['consumer_type'] == ct)
    ]
    if len(sub_c) < MIN_CLEAN:
        print(f"  [{grp}, {ct[:15]}]: {len(sub_c)} чистих — пропускаємо білінговий профіль")
        continue
    fracs = {}
    for m, col in bill_cols.items():
        fracs[m] = float((sub_c[col] / sub_c['annual_billing'].clip(lower=1)).mean())
    total = sum(fracs.values())
    if total > 0:
        billing_monthly_w[(grp, ct)] = {m: v / total for m, v in fracs.items()}

# 2c. Гібридний денний профіль
month_days = {}
for dt, _ in date_cols_2025:
    month_days.setdefault(dt.month, []).append(dt)

hybrid_monthly_w = {}
for grp, ct in GROUPS:
    key     = (grp, ct)
    bill_mw = billing_monthly_w.get(key)
    modem_pn = profile_raw.get(key)
    if bill_mw is None:
        continue
    daily = {}
    for m, b_w in bill_mw.items():
        days = month_days.get(m, [])
        if not days:
            continue
        if modem_pn is not None:
            month_modem_sum = sum(modem_pn.get(dt2, 0.0) for dt2 in days)
            for dt2 in days:
                within = modem_pn.get(dt2, 0.0) / month_modem_sum if month_modem_sum > 0 else 1.0 / len(days)
                daily[dt2] = b_w * within
        else:
            for dt2 in days:
                daily[dt2] = b_w / len(days)
    mw = {}
    for dt2, w in daily.items():
        mw[dt2.month] = mw.get(dt2.month, 0.0) + w
    hybrid_monthly_w[key] = mw

print(f"  Гібридних профілів побудовано: {len(hybrid_monthly_w)}")

# Місячні ваги з модемного профілю (для ПГ-груп як ціль калібрації)
modem_monthly_w = {}
for key, pnorm in profile_raw.items():
    mw = {}
    for dt, w in pnorm.items():
        mw[dt.month] = mw.get(dt.month, 0.0) + w
    modem_monthly_w[key] = mw

def get_target_mw(grp, ct):
    """Цільовий місячний профіль для калібрації:
    ПГ-групи → модемний (consumption), решта → гібридний (billing-shaped)."""
    key = (grp, ct)
    if key in MODEM_PROFILE_GROUPS:
        return modem_monthly_w.get(key)
    return hybrid_monthly_w.get(key)

# ═══════════════════════════════════════════════════════════════════════════════
# БЛОК 3: Не-модемні з annual_billing
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== 3. Не-модемні споживачі ===")
nm = ap[
    ~ap['account_id'].isin(modem_ids) & ~ap['gas_off'] & ap['grs'].notna()
    & (ap['annual_billing'] > 0)
].copy()
print(f"  Не-модемних: {len(nm):,}")

# Річний білінг по (grp, ct)
nm_group_annual = nm.groupby(['appliance_group', 'consumer_type'])['annual_billing'].sum()
print("\n  Річний білінг не-модемних по групах:")
for (grp, ct), v in nm_group_annual.items():
    print(f"    [{grp}, {ct[:20]}]: {v:,.0f} м³")

# ═══════════════════════════════════════════════════════════════════════════════
# БЛОК 4: Калібрувальні множники
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== 4. Калібрувальні множники ===")

# Додаємо consumer_type до df_pred_nm через join з ap
ap_ct = (ap[['account_id', 'consumer_type']]
         .dropna(subset=['account_id'])
         .drop_duplicates('account_id'))
df_pred_nm = df_pred_nm.merge(ap_ct, on='account_id', how='left')
df_pred_nm['consumer_type'] = df_pred_nm['consumer_type'].fillna(CT_PRIV)

# Агрегований K_v2 прогноз по (grp, ct, month)
kv2_grp = (df_pred_nm
           .groupby(['appliance_group', 'consumer_type', 'month'])['volume']
           .sum()
           .reset_index()
           .rename(columns={'volume': 'kv2_sum'}))

def kv2_for(grp, ct, m):
    s = kv2_grp[
        (kv2_grp['appliance_group'] == grp) &
        (kv2_grp['consumer_type'] == ct) &
        (kv2_grp['month'] == m)
    ]
    return float(s['kv2_sum'].sum()) if not s.empty else 0.0

# --- C_op_heat(m): один множник для P_HEAT, агрегат по всіх ОП-групах ---
print("\n  C_op_heat (для P_HEAT; агрегат всіх ОП-груп по опалювальних місяцях):")
print(f"  {'Міс':5s}  {'decompose':>12s}  {'kv2':>12s}  {'C':>8s}")
C_op_heat = {}
for m in sorted(HEATING_MONTHS):
    dec_total = 0.0
    kv2_total = 0.0
    for grp, ct in GROUPS:
        if 'ОП' not in grp:
            continue
        mw = get_target_mw(grp, ct) or {}
        ann_bill = float(nm_group_annual.get((grp, ct), 0.0))
        dec_total += ann_bill * mw.get(m, 0.0)
        kv2_total += kv2_for(grp, ct, m)
    C_op_heat[m] = dec_total / kv2_total if kv2_total > 0 else 1.0
    print(f"  {UA[m]:5s}  {dec_total:>12,.0f}  {kv2_total:>12,.0f}  {C_op_heat[m]:>8.4f}")

# --- C(grp, ct, m): per-group множники для P_nonheat ---
# Застосовуємо тільки до: не-ОП груп (всі місяці) + ОП груп (літні місяці)
print("\n  C(grp, ct, m) для P_nonheat:")
C_nonheat = {}  # {(grp, ct, m): multiplier}
for grp, ct in GROUPS:
    mw = get_target_mw(grp, ct)
    if mw is None:
        for m in MONTHS:
            C_nonheat[(grp, ct, m)] = 1.0
        continue
    ann_bill = float(nm_group_annual.get((grp, ct), 0.0))
    for m in MONTHS:
        is_op      = 'ОП' in grp
        is_heating = m in HEATING_MONTHS
        if is_op and is_heating:
            # P_nonheat для ОП в опалювальних місяцях не використовується → без зміни
            C_nonheat[(grp, ct, m)] = 1.0
            continue
        dec_sum = ann_bill * mw.get(m, 0.0)
        kv2_sum = kv2_for(grp, ct, m)
        if kv2_sum > 0 and dec_sum > 0:
            C_nonheat[(grp, ct, m)] = dec_sum / kv2_sum
        else:
            C_nonheat[(grp, ct, m)] = 1.0

# Відображення суттєвих відхилень
print(f"  {'Група':25s}  {'ТС':20s}  {'Міс':5s}  {'C':>8s}")
for (grp, ct, m), c in sorted(C_nonheat.items()):
    if abs(c - 1.0) > 0.05:
        print(f"  {grp:25s}  {ct[:20]:20s}  {UA[m]:5s}  {c:>8.4f}")

# ═══════════════════════════════════════════════════════════════════════════════
# БЛОК 5: Нові P-профілі
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== 5. Нові P-профілі ===")

P_HEAT_v2  = PobUtPredictorConfigK_v2.P_HEAT
P_HEAT_new = {m: P_HEAT_v2[m] * C_op_heat.get(m, 1.0) for m in HEATING_MONTHS}

print("\n  P_HEAT оновлені (м³/м²):")
print(f"  {'Міс':5s}  {'P_HEAT_v2':>12s}  {'P_HEAT_new':>12s}  {'Δ%':>8s}")
for m in sorted(HEATING_MONTHS):
    old_v = P_HEAT_v2[m]
    new_v = P_HEAT_new[m]
    delta = (new_v - old_v) / old_v * 100
    print(f"  {UA[m]:5s}  {old_v:>12.4f}  {new_v:>12.4f}  {delta:>+8.2f}%")

# Нові P_nonheat (зчитуємо з CSV, потім множимо на C)
p_df = pd.read_csv(_PobUtPredictorConfigK.P_NONHEAT_CSV)
p_nonheat_orig = {
    (r['appliance_group'], r['consumer_type'], int(r['month'])): float(r['P_modem_median'])
    for _, r in p_df.iterrows()
}

p_nonheat_new = {}
for k, v in p_nonheat_orig.items():
    grp, ct, m = k
    c = C_nonheat.get((grp, ct, m), 1.0)
    p_nonheat_new[k] = v * c

print("\n  P_nonheat оновлені (зміни > 3%):")
print(f"  {'Група':25s}  {'ТС':20s}  {'Міс':5s}  {'Orig':>8s}  {'New':>8s}  {'Δ%':>8s}")
for (grp, ct, m), new_v in sorted(p_nonheat_new.items()):
    old_v = p_nonheat_orig[(grp, ct, m)]
    if abs(new_v - old_v) / max(abs(old_v), 0.001) > 0.03:
        delta = (new_v - old_v) / old_v * 100
        print(f"  {grp:25s}  {ct[:20]:20s}  {UA[m]:5s}  {old_v:>8.3f}  {new_v:>8.3f}  {delta:>+8.2f}%")

# ═══════════════════════════════════════════════════════════════════════════════
# БЛОК 6: Re-run K_v2 з новими P-профілями
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== 6. K_v2 калібрований прогноз ===")

# Зберігаємо в module-level щоб клас міг їх використати
_p_heat_calib    = dict(P_HEAT_new)
_p_nonheat_calib = dict(p_nonheat_new)


class PobUtPredictorConfigK_v2_Decomp(PobUtPredictorConfigK_v2):
    """K_v2 перекалібрований під annual_decompose.

    Перевизначає P_HEAT та P_nonheat з калібрувальних множників.
    """
    P_HEAT = _p_heat_calib

    def _load_k_profiles(self):
        # Завантажуємо оригінальні профілі (будує _p_nonheat та _p_nonheat_matrix)
        super()._load_k_profiles()
        # Замінюємо значення каліброваними
        for k, v in _p_nonheat_calib.items():
            self._p_nonheat[k] = v
        # Перебудовуємо матрицю з новими значеннями
        self._build_nm_arrays()


p_calib  = PobUtPredictorConfigK_v2_Decomp(MODEM_FILE, SUBS_FILE, mode='offline', summer_op_vpg=False)
df_calib = p_calib.predict(years=[2025])
print("  Прогноз калібрований готовий")

# ═══════════════════════════════════════════════════════════════════════════════
# Агрегація по ГРС × місяць (fact + pred)
# ═══════════════════════════════════════════════════════════════════════════════
def agg_to_grs(df_p):
    t = (df_p
         .groupby(['grs', 'year', 'month', 'type'])['volume']
         .sum().reset_index()
         .pivot_table(index=['grs', 'year', 'month'], columns='type',
                      values='volume', aggfunc='sum', fill_value=0)
         .reset_index())
    t.columns.name = None
    t['fact'] = t.get('fact', pd.Series(0, index=t.index))
    t['pred'] = t.get('pred', pd.Series(0, index=t.index))
    t['total'] = t['fact'] + t['pred']
    return t[['grs', 'month', 'total']].copy()


grs_orig  = agg_to_grs(df_orig)
grs_calib = agg_to_grs(df_calib)

# Завантаження decompose як еталону
dc = pd.read_excel('data/annual_decompose_result.xlsx', sheet_name='All_GRS_monthly')
dc = dc[['grs', 'month', 'total_pred']].rename(columns={'total_pred': 'decompose'})

# Об'єднання трьох наборів
m_all = dc.merge(grs_orig.rename(columns={'total': 'kv2_orig'}),   on=['grs', 'month'], how='inner')
m_all = m_all.merge(grs_calib.rename(columns={'total': 'kv2_cal'}), on=['grs', 'month'], how='left')
m_all = m_all[m_all['decompose'] > 0].copy()
m_all['kv2_cal']   = m_all['kv2_cal'].fillna(0)
m_all['err_orig']  = m_all['kv2_orig'] - m_all['decompose']
m_all['err_cal']   = m_all['kv2_cal']  - m_all['decompose']
m_all['bias_orig'] = m_all['err_orig'] / m_all['decompose'] * 100
m_all['bias_cal']  = m_all['err_cal']  / m_all['decompose'] * 100

ann_grs = m_all.groupby('grs')['decompose'].sum().sort_values(ascending=False)

# Підсумок по місяцях
print("\n  Результати після калібрації:")
print(f"  {'Міс':5s}  {'Bias orig%':>11s}  {'Bias calib%':>12s}  {'Покращ.':>10s}")
for mo in MONTHS:
    s = m_all[m_all['month'] == mo]
    if s.empty:
        continue
    b_o = s['err_orig'].sum() / s['decompose'].sum() * 100
    b_c = s['err_cal'].sum()  / s['decompose'].sum() * 100
    imp = abs(b_o) - abs(b_c)
    flag = '✓' if imp > 0 else '✗'
    print(f"  {UA[mo]:5s}  {b_o:>+11.2f}%  {b_c:>+12.2f}%  {imp:>+10.2f}pp {flag}")

mae_orig  = m_all['err_orig'].abs().mean()
mae_calib = m_all['err_cal'].abs().mean()
bias_orig = m_all['err_orig'].sum() / m_all['decompose'].sum() * 100
bias_cal  = m_all['err_cal'].sum()  / m_all['decompose'].sum() * 100
print(f"\n  MAE:  {mae_orig:,.0f} → {mae_calib:,.0f}  (Δ {mae_calib - mae_orig:+,.0f})")
print(f"  Bias: {bias_orig:+.2f}% → {bias_cal:+.2f}%")

# ═══════════════════════════════════════════════════════════════════════════════
# БЛОК 7: Excel
# ═══════════════════════════════════════════════════════════════════════════════
print(f"\n=== 7. Збереження → {OUT_XLSX} ===")
wb = Workbook()

# ════════ Лист 1: По_місяцях ══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'По_місяцях'

headers1 = ['Показник'] + [UA[mo] for mo in MONTHS] + ['Річний']
for ci, h in enumerate(headers1, 1):
    c = ws1.cell(1, ci, h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = Alignment(horizontal='center')
ws1.column_dimensions['A'].width = 20
for ci in range(2, 15):
    ws1.column_dimensions[get_column_letter(ci)].width = 11
ws1.freeze_panes = 'B2'

dec_row   = {mo: m_all[m_all['month'] == mo]['decompose'].sum() for mo in MONTHS}
orig_row  = {mo: m_all[m_all['month'] == mo]['kv2_orig'].sum()  for mo in MONTHS}
calib_row = {mo: m_all[m_all['month'] == mo]['kv2_cal'].sum()   for mo in MONTHS}
bo_row    = {mo: (orig_row[mo]  - dec_row[mo]) / dec_row[mo] * 100 if dec_row[mo] else 0 for mo in MONTHS}
bc_row    = {mo: (calib_row[mo] - dec_row[mo]) / dec_row[mo] * 100 if dec_row[mo] else 0 for mo in MONTHS}

ann_dec   = sum(dec_row.values())
ann_orig  = sum(orig_row.values())
ann_calib = sum(calib_row.values())
ann_bo    = (ann_orig  - ann_dec) / ann_dec * 100 if ann_dec else 0
ann_bc    = (ann_calib - ann_dec) / ann_dec * 100 if ann_dec else 0

rows_def = [
    ('Декомпозит (факт)',  dec_row,   '#,##0',         False, ann_dec),
    ('K_v2 оригінал',     orig_row,  '#,##0',         False, ann_orig),
    ('Bias orig %',       bo_row,    '+0.0;-0.0;0.0', True,  ann_bo),
    ('K_v2 калібрований', calib_row, '#,##0',         False, ann_calib),
    ('Bias calib %',      bc_row,    '+0.0;-0.0;0.0', True,  ann_bc),
]
for ri, (label, data, fmt, is_bias, ann_v) in enumerate(rows_def, 2):
    ws1.cell(ri, 1, label).font = Font(bold=is_bias, italic=not is_bias)
    ws1.cell(ri, 1).alignment = Alignment(horizontal='left')
    for ci, mo in enumerate(MONTHS, 2):
        c = ws1.cell(ri, ci, round(data[mo], 1 if is_bias else 0))
        c.number_format = fmt; c.alignment = Alignment(horizontal='right')
        if is_bias:
            c.font = Font(bold=True)
            f = bias_fill(data[mo])
            if f: c.fill = f
    c = ws1.cell(ri, 14, round(ann_v, 1 if is_bias else 0))
    c.number_format = fmt; c.alignment = Alignment(horizontal='right')
    if is_bias:
        c.font = Font(bold=True)
        f = bias_fill(ann_v)
        if f: c.fill = f
    for ci in range(1, 15):
        ws1.cell(ri, ci).border = Border(bottom=thin)

# ════════ Лист 2: По_ГРС ══════════════════════════════════════════════════════
ws2 = wb.create_sheet('По_ГРС')

for ci, h in enumerate(['ГРС', ''] + [UA[mo] for mo in MONTHS] + ['Річний'], 1):
    c = ws2.cell(1, ci, h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = Alignment(horizontal='center')
ws2.column_dimensions['A'].width = 42
ws2.column_dimensions['B'].width = 14
for ci in range(3, 16):
    ws2.column_dimensions[get_column_letter(ci)].width = 11
ws2.freeze_panes = 'C2'

row_i = 2
for grs in ann_grs.index:
    sub = m_all[m_all['grs'] == grs].set_index('month')
    if sub.empty:
        continue
    is_large = ann_grs[grs] >= LARGE
    row_fill = PatternFill('solid', fgColor='D6E4F0' if is_large else 'F5F5F5')

    d_data  = {mo: float(sub.loc[mo, 'decompose']) if mo in sub.index else 0.0 for mo in MONTHS}
    c_data  = {mo: float(sub.loc[mo, 'kv2_cal'])   if mo in sub.index else 0.0 for mo in MONTHS}
    bc_data = {mo: (c_data[mo] - d_data[mo]) / d_data[mo] * 100 if d_data[mo] > 0 else None for mo in MONTHS}
    ann_d   = sum(d_data.values())
    ann_c   = sum(c_data.values())
    ann_bc_grs = (ann_c - ann_d) / ann_d * 100 if ann_d else None

    for li, (label, data, fmt, is_bias) in enumerate([
        ('Декомпозит',   d_data,  '#,##0',         False),
        ('K_v2 calib',   c_data,  '#,##0',         False),
        ('Bias calib %', bc_data, '+0.0;-0.0;0.0', True),
    ]):
        c1 = ws2.cell(row_i, 1, grs if li == 0 else '')
        c1.font = Font(bold=is_large)
        if li == 0:
            c1.fill = row_fill
        c2 = ws2.cell(row_i, 2, label)
        c2.font = Font(bold=is_bias, italic=not is_bias, color='444444')
        if li == 0:
            c2.fill = row_fill

        for ci, mo in enumerate(MONTHS, 3):
            v = data[mo]
            if v is None:
                ws2.cell(row_i, ci, '-').alignment = Alignment(horizontal='center')
            else:
                c = ws2.cell(row_i, ci, round(v, 1 if is_bias else 0))
                c.number_format = fmt
                c.alignment = Alignment(horizontal='right')
                if is_bias:
                    c.font = Font(bold=True)
                    f = bias_fill(v)
                    if f: c.fill = f

        ann_label_v = ann_d if li == 0 else ann_c if li == 1 else ann_bc_grs
        c = ws2.cell(row_i, 15,
                     round(ann_label_v, 1 if is_bias else 0) if ann_label_v is not None else '-')
        c.number_format = fmt if is_bias else '#,##0'
        c.alignment = Alignment(horizontal='right')
        if is_bias and ann_label_v is not None:
            c.font = Font(bold=True)
            f = bias_fill(ann_label_v)
            if f: c.fill = f

        border_side = thick if (is_bias and is_large) else thin if is_bias else Side()
        for ci in range(1, 16):
            ws2.cell(row_i, ci).border = Border(bottom=border_side)
        row_i += 1

# Підсумковий рядок
row_i += 1
ws2.cell(row_i, 1, 'ВСЬОГО').font = Font(bold=True)
ws2.cell(row_i, 2, 'Bias calib %').font = Font(bold=True)
for ci, mo in enumerate(MONTHS, 3):
    sm = m_all[m_all['month'] == mo]
    if sm.empty:
        continue
    b = (sm['kv2_cal'].sum() - sm['decompose'].sum()) / sm['decompose'].sum() * 100
    c = ws2.cell(row_i, ci, round(b, 1))
    c.number_format = '+0.0;-0.0;0.0'; c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='right')
    f = bias_fill(b)
    if f: c.fill = f
all_b = (m_all['kv2_cal'].sum() - m_all['decompose'].sum()) / m_all['decompose'].sum() * 100
c = ws2.cell(row_i, 15, round(all_b, 1))
c.number_format = '+0.0;-0.0;0.0'; c.font = Font(bold=True)
c.alignment = Alignment(horizontal='right')
f = bias_fill(all_b)
if f: c.fill = f

wb.save(OUT_XLSX)
print(f"  DONE → {OUT_XLSX}")
print(f"  Листи: По_місяцях | По_ГРС ({row_i} рядків)")

# ═══════════════════════════════════════════════════════════════════════════════
# БЛОК 8: Друк нових P-профілів (підсумкова таблиця)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n=== 8. Підсумок нових P-профілів ===")

print("\n  P_HEAT оновлені (м³/м²):")
print(f"  {'Міс':5s}  {'P_HEAT_v2':>12s}  {'P_HEAT_new':>12s}  {'Δ%':>8s}")
for m in sorted(HEATING_MONTHS):
    old_v = P_HEAT_v2[m]
    new_v = P_HEAT_new[m]
    delta = (new_v - old_v) / old_v * 100
    print(f"  {UA[m]:5s}  {old_v:>12.4f}  {new_v:>12.4f}  {delta:>+8.2f}%")

print("\n  P_nonheat оновлені (всі зміни):")
print(f"  {'Група':25s}  {'ТС':20s}  {'Міс':5s}  {'Orig':>8s}  {'New':>8s}  {'Δ%':>8s}")
for (grp, ct, m), new_v in sorted(p_nonheat_new.items(), key=lambda x: (x[0][0], x[0][1], x[0][2])):
    old_v = p_nonheat_orig[(grp, ct, m)]
    delta = (new_v - old_v) / old_v * 100 if old_v else 0
    mark = ' !' if abs(delta) > 10 else ''
    print(f"  {grp:25s}  {ct[:20]:20s}  {UA[m]:5s}  {old_v:>8.3f}  {new_v:>8.3f}  {delta:>+8.2f}%{mark}")
