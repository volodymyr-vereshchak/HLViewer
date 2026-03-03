"""
rt_predict_month.py

Оперативний прогноз місячного споживання по ГРС.
Запускається щомісяця після оновлення вхідних файлів.

══════════════════════════════════════════════════════════════════════════
ВХІДНІ ДАНІ
══════════════════════════════════════════════════════════════════════════

1. SUBS_FILE  — CSV абонентів з місячним білінгом
   Формат  : стандартний all_pobut_enriched.csv
   Кодування: UTF-8 або latin-1, роздільник ','
   Обов'язкові колонки:
     account_id       — ID абонента
     grs              — назва ГРС
     appliance_group  — прилади (визначає ОП)
     heated_area      — опалювальна площа (м²)
     gas_off          — якщо не NaN → абонент відключений (виключається)
     alternative      — якщо не NaN → альт. опалення (виключається)
     dacha            — якщо не NaN → дача (виключається)
     jan_2025 ... dec_2025 ... jan_2026 ...
                      — місячний білінг (м³). Назва колонки: {mon}_{рік},
                        де mon = jan/feb/mar/apr/may/jun/jul/aug/sep/oct/nov/dec
   Оновлення: щомісяця додати нову колонку (напр. feb_2026)

2. MODEM_FILE — CSV щоденних показань лічильників-модемів
   Формат  : роздільник ';', стандартний profile_pobut_daily_result.csv
   Рядки   : кожен рядок — один лічильник-модем
   META-колонки (перші 15): account_id, grs, consumer_type, appliance_group,
     alternative, dacha, gas_off, address, gas_contract, contract_date,
     heated_area, floors, rooms, floor_area, people  (порядок з META_COLS)
   Дата-колонки: 'DD.MM.YYYY' — щоденний обсяг (м³/день)
     Підтримуються будь-які роки: 2025, 2026, ...
     Файл може містити різну кількість модемів між запусками
   Оновлення: вивантажити поточні дані (включно з місяцем прогнозу)

3. WEATHER_FILE (опціонально) — добова температура
   Формат: CSV з колонками 'date' (YYYY-MM-DD) та 'temperature'
   Якщо файл не знайдено — сезони визначаються тільки по ref_heat (KMeans)

══════════════════════════════════════════════════════════════════════════
ЛОГІКА
══════════════════════════════════════════════════════════════════════════

  Місяці білінгу  → тренувальний набір (Y)
  Місяці модемів  → ref_heat та ref_all (X)
  Місяць прогнозу = перший місяць у модемах, якого немає в білінгу

  Модель (per GRS, два сезони):
    Опал.:  q_m = a_heat + b_heat × ref_heat_m
    Літо:   q_m = a_sum  + b_sum × ref_all_m + c_sum × ref_heat_m
  Де q_m — місячний обсяг ГРС (м³), ref_* — середнє за добовими сигналами

  Валідація: LOO-month CV на тренувальних місяцях
  Прогноз  : застосовується до місяць прогнозу

══════════════════════════════════════════════════════════════════════════
ВИВІД
══════════════════════════════════════════════════════════════════════════

  rt_predict_{YYYYMM}.xlsx:
    Прогноз_ГРС      — прогнози по ГРС з метриками
    Деталі_моделей   — коефіцієнти та LOO-bias
    Дані_тренування  — Y vs Ŷ по тренувальних місяцях
    Ref_сигнали      — ref_heat/ref_all по місяцях + сезон

══════════════════════════════════════════════════════════════════════════
"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score, mean_absolute_error
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from pobut_predictor import GROUP_MERGE, META_COLS

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
SUBS_FILE    = "data/input/all_pobut_enriched.csv"
MODEM_FILE   = "data/input/profile_pobut_daily_result.csv"
WEATHER_FILE = "../data/weather_2025_daily_contractual.csv"   # опційно
OUT_DIR      = "data"

N_SKIP_TOP   = 3    # виключаємо N найбільших ГРС за річним об'ємом
MIN_TRAIN_M  = 5    # мінімум тренувальних місяців для моделі ГРС
MIN_HEAT_M   = 3    # мінімум опалювальних місяців для heating-моделі

MON_TO_NUM = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
              'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
NUM_TO_MON = {v: k for k, v in MON_TO_NUM.items()}
UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}

HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HEAT_FILL = PatternFill('solid', fgColor='FCE4D6')
SUM_FILL  = PatternFill('solid', fgColor='E2EFDA')
PRED_FILL = PatternFill('solid', fgColor='D9E1F2')

def color_bias(v):
    a = abs(v)
    if a <= 3:  return None
    if a <= 7:  return 'FFEB9C'
    if a <= 15: return 'FFC7CE'
    return 'FF4C4C'


# ── 1. Абоненти та місяці білінгу ─────────────────────────────────────────────
print("=== 1. Завантаження абонентів ===")
subs = pd.read_csv(SUBS_FILE, low_memory=False)
subs['appliance_group'] = subs['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
subs['heated_area'] = pd.to_numeric(subs['heated_area'], errors='coerce').fillna(55.0)
subs['heated_area'] = subs['heated_area'].where(subs['heated_area'] > 0, 55.0)

# Визначаємо всі місяці білінгу: колонки вигляду jan_2025, feb_2026, …
bill_cols = [c for c in subs.columns
             if re.match(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)_\d{4}$', c)]

def col_to_ts(c):
    mon, yr = c.split('_')
    return pd.Timestamp(f"{yr}-{MON_TO_NUM[mon]:02d}-01")

bill_ts_map = {c: col_to_ts(c) for c in bill_cols}
bill_months = sorted(bill_ts_map.values())

for c in bill_cols:
    subs[c] = pd.to_numeric(subs[c], errors='coerce').fillna(0)

# Тільки активні (без відключень, альт., дач)
subs_act = subs[
    subs['gas_off'].isna() &
    subs['alternative'].isna() &
    subs['dacha'].isna() &
    subs['grs'].notna()
].copy()

print(f"  Активних абонентів: {len(subs_act):,}")
print(f"  ГРС: {subs_act['grs'].nunique()}")
print(f"  Місяці білінгу: {[str(m.date()) for m in bill_months]}")

# Y_train: місячне GRS-споживання (сума всіх абонентів)
# Індекс = GRS, колонки = Timestamps
ts_cols = {c: bill_ts_map[c] for c in bill_cols}
Y_df = (subs_act.groupby('grs')[bill_cols]
               .sum()
               .rename(columns=ts_cols))
Y_df.columns = pd.to_datetime(Y_df.columns)

grs_annual = Y_df.sum(axis=1).sort_values(ascending=False)
print(f"\n  Топ-5 ГРС за річним об'ємом:")
for g, v in grs_annual.head(5).items():
    print(f"    {g[:50]:50} {v/1e6:.3f} млн м³")


# ── 2. Модемні дані → ref-сигнали по місяцях ──────────────────────────────────
print("\n=== 2. Завантаження модемних даних ===")
md = pd.read_csv(MODEM_FILE, sep=';', low_memory=False)
md.columns = META_COLS + list(md.columns[len(META_COLS):])
md['appliance_group'] = md['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
md = md[md['gas_off'].isna() & md['alternative'].isna() & md['dacha'].isna()]
md['has_OP'] = md['appliance_group'].str.contains('ОП', na=False)
md['heated_area'] = pd.to_numeric(md['heated_area'], errors='coerce').fillna(55.0)
md['heated_area'] = md['heated_area'].where(md['heated_area'] > 0, 55.0)

# Всі дата-колонки (будь-який рік: DD.MM.YYYY)
date_cols = sorted(
    [c for c in md.columns
     if isinstance(c, str) and c.count('.') == 2 and len(c) == 10
     and c[:2].isdigit() and c[3:5].isdigit() and c[6:].isdigit()],
    key=lambda c: (int(c[6:]), int(c[3:5]), int(c[:2]))
)

# Парсинг дат і групування по місяцях
dates_all = [pd.Timestamp(f"{c[6:]}-{c[3:5]}-{c[:2]}") for c in date_cols]
month_to_dcols = {}
for dc, dt in zip(date_cols, dates_all):
    mts = dt.replace(day=1)
    month_to_dcols.setdefault(mts, []).append(dc)

modem_months = sorted(month_to_dcols)
print(f"  Рядків у файлі модемів: {len(md):,}")
print(f"  Місяців у модемному файлі: {[str(m.date()) for m in modem_months]}")

# Визначаємо місяць прогнозу
predict_months = [m for m in modem_months if m not in set(bill_months)]
if not predict_months:
    print("\n  ❌ Усі місяці модемів вже є в білінгу.")
    print("     Оновіть MODEM_FILE — він повинен містити місяць прогнозу.")
    sys.exit(1)

predict_month = max(predict_months)  # найновіший місяць без білінгу
train_months  = [m for m in modem_months if m in set(bill_months)]

print(f"\n  Місяць прогнозу:    {predict_month.strftime('%B %Y')}")
print(f"  Тренувальних міс.:  {len(train_months)}")

for c in date_cols:
    md[c] = pd.to_numeric(md[c], errors='coerce').fillna(0)

md_vals = md[date_cols].values.astype(np.float32)   # (N_modems, N_days)

# Активні модеми: хоча б якийсь обсяг за весь доступний період
annual_total = md_vals.sum(axis=1)
active_mask  = annual_total > 0

md_act    = md[active_mask].reset_index(drop=True)
vals_act  = md_vals[active_mask]
md_op     = md_act[md_act['has_OP']].copy()
vals_op   = vals_act[md_act['has_OP'].values]

print(f"\n  Активних модемів (всього): {len(md_act):,}")
print(f"  З них ОП:                  {len(md_op):,}")

# ref_heat та ref_all по місяцях
def monthly_ref(md_sub, vals_sub, months_map):
    """Обчислює ref-сигнал (обсяг/площа) для кожного місяця."""
    area = md_sub['heated_area'].values.astype(np.float32)
    result = {}
    for mts, dcols in sorted(months_map.items()):
        idx_d = [date_cols.index(c) for c in dcols]
        vol_m = vals_sub[:, idx_d].sum(axis=1)  # місячний обсяг per modem
        a_mask = vol_m > 0
        if a_mask.sum() == 0:
            result[mts] = 0.0
        else:
            result[mts] = float(vol_m[a_mask].sum()) / float(area[a_mask].sum())
    return result

ref_heat_m = monthly_ref(md_op, vals_op, month_to_dcols)
ref_all_m  = monthly_ref(md_act, vals_act, month_to_dcols)

print("\n  Місячні ref-сигнали [м³/(м²·міс)]:")
print(f"  {'Місяць':12}  {'ref_heat':>10}  {'ref_all':>10}  {'Тип':6}")
for mts in sorted(month_to_dcols):
    rh = ref_heat_m.get(mts, 0)
    ra = ref_all_m.get(mts, 0)
    tp = 'ПРОГН' if mts == predict_month else 'трен.'
    print(f"  {str(mts.date()):12}  {rh:10.4f}  {ra:10.4f}  {tp}")


# ── 3. Сезон (KMeans на ref_heat тренувальних місяців) ────────────────────────
print("\n=== 3. Визначення сезонів ===")
rh_train = np.array([ref_heat_m[m] for m in train_months])

if len(train_months) < 4:
    print("  Мало місяців для KMeans, використовуємо поріг ref_heat > median")
    threshold = float(np.median(rh_train))
else:
    km = KMeans(n_clusters=2, random_state=42, n_init=10)
    km.fit(rh_train.reshape(-1, 1))
    c0, c1 = sorted(km.cluster_centers_.flatten())
    threshold = (c0 + c1) / 2.0
    print(f"  Центроїди KMeans: літо={c0:.4f}  зима={c1:.4f}")

print(f"  Поріг: ref_heat > {threshold:.4f} → опалювальний місяць")

is_heat_train = {m: ref_heat_m[m] > threshold for m in train_months}
heating_months_set = {m for m, h in is_heat_train.items() if h}
summer_months_set  = {m for m, h in is_heat_train.items() if not h}

# Сезон для місяця прогнозу
is_predict_heating = ref_heat_m[predict_month] > threshold
predict_season = 'опал' if is_predict_heating else 'літо'

print(f"\n  Опалювальних місяців: {len(heating_months_set)}")
print(f"  Літніх місяців:       {len(summer_months_set)}")
print(f"  Сезон прогнозу ({predict_month.strftime('%b %Y')}): {predict_season}")

for mts in sorted(train_months):
    seas = 'опал ▓' if is_heat_train[mts] else 'літо ░'
    print(f"    {str(mts.date()):12} ref_heat={ref_heat_m[mts]:.4f}  [{seas}]")


# ── 4. Двосезонна регресія per GRS ────────────────────────────────────────────
print("\n=== 4. Двосезонні моделі per GRS ===")

skip_grs = set(grs_annual.head(N_SKIP_TOP).index)
print(f"  Виключаємо топ-{N_SKIP_TOP} ГРС (найбільші):")
for g in sorted(skip_grs, key=lambda x: -grs_annual[x]):
    print(f"    {g}")

target_grs = [g for g in Y_df.index if g not in skip_grs]
print(f"  Моделюємо ГРС: {len(target_grs)}")
print()

results   = []
loo_preds = {}   # grs → {month: pred_m3}
models    = {}   # grs → {'heat': lr_h, 'sum': lr_s, 'threshold': threshold}

for grs in target_grs:
    # Вибираємо тільки ті тренувальні місяці, для яких є і Y, і X
    avail = [m for m in train_months if m in Y_df.columns]
    if len(avail) < MIN_TRAIN_M:
        continue

    Y_vec  = np.array([Y_df.loc[grs, m] for m in avail])
    RH_vec = np.array([ref_heat_m[m] for m in avail])
    RA_vec = np.array([ref_all_m[m]  for m in avail])
    is_h   = np.array([is_heat_train[m] for m in avail])

    if Y_vec.sum() < 50:
        continue

    # Вектори сезонів
    h_idx = np.where(is_h)[0]
    s_idx = np.where(~is_h)[0]

    # LOO-CV по місяцях
    y_loo = np.full(len(avail), np.nan)

    # --- Heating LOO ---
    if len(h_idx) >= MIN_HEAT_M:
        for ti, m_test in enumerate(np.array(avail)[h_idx]):
            te = (np.array(avail)[h_idx] == m_test)
            tr = ~te
            if tr.sum() < 2:
                continue
            Xtr = RH_vec[h_idx][tr].reshape(-1, 1)
            Ytr = Y_vec[h_idx][tr]
            Xte = RH_vec[h_idx][te].reshape(-1, 1)
            lr = LinearRegression().fit(Xtr, Ytr)
            y_loo[h_idx[te]] = np.maximum(lr.predict(Xte), 0.0)
    else:
        # Недостатньо опалювальних місяців → середнє
        if len(h_idx) > 0:
            y_loo[h_idx] = Y_vec[h_idx].mean()

    # --- Summer LOO ---
    if len(s_idx) >= 3:
        for ti, m_test in enumerate(np.array(avail)[s_idx]):
            te = (np.array(avail)[s_idx] == m_test)
            tr = ~te
            if tr.sum() < 2:
                continue
            Xtr = np.column_stack([RA_vec[s_idx][tr], RH_vec[s_idx][tr]])
            Ytr = Y_vec[s_idx][tr]
            Xte = np.column_stack([RA_vec[s_idx][te], RH_vec[s_idx][te]])
            lr = LinearRegression().fit(Xtr, Ytr)
            y_loo[s_idx[te]] = np.maximum(lr.predict(Xte), 0.0)
    elif len(s_idx) > 0:
        y_loo[s_idx] = Y_vec[s_idx].mean()

    # Заповнити NaN (недостатньо даних для LOO)
    nan_m = np.isnan(y_loo)
    if nan_m.any() and (~nan_m).sum() >= 2:
        lr_fb = LinearRegression().fit(RH_vec[~nan_m].reshape(-1, 1), Y_vec[~nan_m])
        y_loo[nan_m] = np.maximum(lr_fb.predict(RH_vec[nan_m].reshape(-1, 1)), 0.0)

    valid = ~np.isnan(y_loo)
    if valid.sum() < 3:
        continue

    r2  = r2_score(Y_vec[valid], y_loo[valid]) if Y_vec[valid].std() > 0 else 0.0
    mae = mean_absolute_error(Y_vec[valid], y_loo[valid])

    # Повні моделі (для прогнозу)
    lr_h, lr_s = None, None
    if len(h_idx) >= 2:
        lr_h = LinearRegression().fit(RH_vec[h_idx].reshape(-1, 1), Y_vec[h_idx])
    if len(s_idx) >= 2:
        Xs = np.column_stack([RA_vec[s_idx], RH_vec[s_idx]])
        lr_s = LinearRegression().fit(Xs, Y_vec[s_idx])

    models[grs] = {'heat': lr_h, 'sum': lr_s, 'threshold': threshold}
    loo_preds[grs] = {m: y_loo[i] for i, m in enumerate(avail) if not np.isnan(y_loo[i])}

    # LOO bias по місяцях
    bias_by_m = {}
    for i, m in enumerate(avail):
        if not np.isnan(y_loo[i]) and Y_vec[i] > 0:
            bias_by_m[m] = round((y_loo[i] - Y_vec[i]) / Y_vec[i] * 100, 1)

    a_h = float(lr_h.intercept_) if lr_h else np.nan
    b_h = float(lr_h.coef_[0])   if lr_h else np.nan
    a_s = float(lr_s.intercept_) if lr_s else np.nan
    b_s = float(lr_s.coef_[0])   if lr_s else np.nan
    c_s = float(lr_s.coef_[1])   if lr_s else np.nan

    results.append({
        'grs': grs, 'n_train': len(avail), 'annual_m3': round(Y_vec.sum()),
        'a_heat': round(a_h, 0) if not np.isnan(a_h) else '',
        'b_heat': round(b_h, 0) if not np.isnan(b_h) else '',
        'a_sum':  round(a_s, 0) if not np.isnan(a_s) else '',
        'b_sum':  round(b_s, 0) if not np.isnan(b_s) else '',
        'c_sum':  round(c_s, 0) if not np.isnan(c_s) else '',
        'r2_loo': round(r2, 3), 'mae_month': round(mae, 0),
        **{str(m.date()): bias_by_m.get(m) for m in avail},
    })

    print(f"  {grs[:38]:38}  R²={r2:+.3f}  MAE={mae/1e3:5.1f}тис  "
          f"ah={a_h:7.0f} bh={b_h:7.0f}"
          + (f"  as={a_s:6.0f} bs={b_s:6.0f} cs={c_s:5.0f}" if lr_s else ""))

res_df = pd.DataFrame(results)
print(f"\n  Змодельовано ГРС: {len(res_df)}")
print(f"  Медіана R² (LOO): {res_df['r2_loo'].median():.3f}")
print(f"  Медіана MAE/міс:  {res_df['mae_month'].median()/1e3:.1f} тис м³")


# ── 5. Прогноз цільового місяця ────────────────────────────────────────────────
print(f"\n=== 5. Прогноз {predict_month.strftime('%B %Y')} ({predict_season}) ===")

preds = {}
rh_pred = ref_heat_m[predict_month]
ra_pred = ref_all_m[predict_month]

for grs, m_models in models.items():
    lr_h = m_models['heat']
    lr_s = m_models['sum']
    thr  = m_models['threshold']

    if is_predict_heating:
        if lr_h is not None:
            pred = float(lr_h.predict([[rh_pred]])[0])
        elif lr_s is not None:
            pred = float(lr_s.predict([[ra_pred, rh_pred]])[0])
        else:
            continue
    else:
        if lr_s is not None:
            pred = float(lr_s.predict([[ra_pred, rh_pred]])[0])
        elif lr_h is not None:
            pred = float(lr_h.predict([[rh_pred]])[0])
        else:
            continue

    preds[grs] = max(pred, 0.0)

total_pred = sum(preds.values())
print(f"\n  Прогноз по ГРС ({predict_month.strftime('%b %Y')}):")
for grs in sorted(preds, key=lambda g: -preds[g]):
    r2_v = res_df.loc[res_df['grs'] == grs, 'r2_loo'].values
    r2_s = f"R²={r2_v[0]:.3f}" if len(r2_v) else ""
    print(f"    {grs[:45]:45}  {preds[grs]/1e3:8.1f} тис м³  {r2_s}")
print(f"\n  РАЗОМ: {total_pred/1e6:.3f} млн м³")


# ── 6. Збереження Excel ────────────────────────────────────────────────────────
out_file = os.path.join(OUT_DIR, f"rt_predict_{predict_month.strftime('%Y%m')}.xlsx")
print(f"\n=== 6. Збереження {out_file} ===")

wb = Workbook()

# ─ Лист 1: Прогноз ГРС ────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = f"Прогноз_{predict_month.strftime('%Y%m')}"
ws1.column_dimensions['A'].width = 42

# Знаходимо попередній рік — для порівняння
prev_year_month = predict_month.replace(year=predict_month.year - 1)
has_prev = prev_year_month in Y_df.columns

hdr1 = ['ГРС', f"Прогноз {predict_month.strftime('%b %Y')} (м³)",
        'Сезон', 'R² (LOO)']
if has_prev:
    hdr1.append(f"Факт {prev_year_month.strftime('%b %Y')} (м³)")
    hdr1.append('YoY %')
hdr1 += ['MAE/міс (м³)', 'Кіл. тренув.']

for ci, h in enumerate(hdr1, 1):
    c = ws1.cell(1, ci, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
for ci in range(2, len(hdr1) + 1):
    ws1.column_dimensions[get_column_letter(ci)].width = 16

for ri, grs in enumerate(sorted(preds, key=lambda g: -preds[g]), 2):
    row_r = res_df[res_df['grs'] == grs].iloc[0] if (res_df['grs'] == grs).any() else None
    r2_v  = float(row_r['r2_loo'])  if row_r is not None else 0.0
    mae_v = float(row_r['mae_month']) if row_r is not None else 0.0
    n_tr  = int(row_r['n_train'])   if row_r is not None else 0

    ws1.cell(ri, 1, grs)
    pc = ws1.cell(ri, 2, round(preds[grs]))
    pc.number_format = '#,##0'
    pc.fill = PRED_FILL
    ws1.cell(ri, 3, predict_season).alignment = Alignment(horizontal='center')

    r2c = ws1.cell(ri, 4, round(r2_v, 3))
    r2c.number_format = '0.000'
    r2c.alignment = Alignment(horizontal='center')
    r2_color = ('00B050' if r2_v >= 0.90 else '92D050' if r2_v >= 0.75
                else 'FFEB9C' if r2_v >= 0.50 else 'FF7C80')
    r2c.font = Font(bold=True, color=r2_color)

    ci_next = 5
    if has_prev:
        prev_val = float(Y_df.loc[grs, prev_year_month]) if grs in Y_df.index else 0.0
        ws1.cell(ri, ci_next, round(prev_val)).number_format = '#,##0'
        ci_next += 1
        if prev_val > 0:
            yoy = (preds[grs] - prev_val) / prev_val * 100
            yc = ws1.cell(ri, ci_next, round(yoy, 1))
            yc.number_format = '+0.0;-0.0;0.0'
            f = color_bias(yoy)
            if f: yc.fill = PatternFill('solid', fgColor=f)
        ci_next += 1

    ws1.cell(ri, ci_next,     round(mae_v)).number_format = '#,##0'
    ws1.cell(ri, ci_next + 1, n_tr).alignment = Alignment(horizontal='center')

# Підсумок
ri_tot = len(preds) + 3
ws1.cell(ri_tot, 1, 'РАЗОМ').font = Font(bold=True)
tot_c = ws1.cell(ri_tot, 2, round(total_pred))
tot_c.number_format = '#,##0'
tot_c.font = Font(bold=True)
if has_prev:
    tot_prev = sum(float(Y_df.loc[g, prev_year_month])
                   for g in preds if g in Y_df.index)
    ws1.cell(ri_tot, 5, round(tot_prev)).number_format = '#,##0'
    if tot_prev > 0:
        yoy_tot = (total_pred - tot_prev) / tot_prev * 100
        tc = ws1.cell(ri_tot, 6, round(yoy_tot, 1))
        tc.number_format = '+0.0;-0.0;0.0'
        tc.font = Font(bold=True)

ws1.freeze_panes = 'B2'


# ─ Лист 2: Деталі моделей ─────────────────────────────────────────────────────
ws2 = wb.create_sheet('Деталі_моделей')
ws2.column_dimensions['A'].width = 42
hdr2 = ['ГРС', 'R² LOO', 'MAE/міс', 'N тренув.',
        'a_heat', 'b_heat', 'a_sum', 'b_sum(ref_all)', 'c_sum(ref_heat)']
# Додаємо LOO bias по місяцях
for m in train_months:
    hdr2.append(m.strftime('%b%y'))

for ci, h in enumerate(hdr2, 1):
    c = ws2.cell(1, ci, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = HDR_FILL if ci <= 9 else (HEAT_FILL if is_heat_train.get(train_months[ci-10], False) else SUM_FILL)
    c.alignment = Alignment(horizontal='center')
for ci in range(2, len(hdr2) + 1):
    ws2.column_dimensions[get_column_letter(ci)].width = 11

for ri, row in enumerate(res_df.sort_values('r2_loo', ascending=False).to_dict('records'), 2):
    ws2.cell(ri, 1, row['grs'])
    r2c = ws2.cell(ri, 2, row['r2_loo'])
    r2c.number_format = '0.000'
    r2_color = ('00B050' if row['r2_loo'] >= 0.90 else '92D050' if row['r2_loo'] >= 0.75
                else 'FFEB9C' if row['r2_loo'] >= 0.50 else 'FF7C80')
    r2c.font = Font(bold=True, color=r2_color)
    ws2.cell(ri, 3, row['mae_month']).number_format = '#,##0'
    ws2.cell(ri, 4, row['n_train']).alignment = Alignment(horizontal='center')
    for ci, k in enumerate(['a_heat','b_heat','a_sum','b_sum','c_sum'], 5):
        if row[k] != '':
            ws2.cell(ri, ci, row[k]).number_format = '#,##0'
    for ci, m in enumerate(train_months, 10):
        v = row.get(str(m.date()))
        if v is not None:
            cell = ws2.cell(ri, ci, v)
            cell.number_format = '+0.0;-0.0;0.0'
            cell.alignment = Alignment(horizontal='center')
            f = color_bias(v)
            if f: cell.fill = PatternFill('solid', fgColor=f)

ws2.freeze_panes = 'B2'


# ─ Лист 3: Дані тренування ────────────────────────────────────────────────────
ws3 = wb.create_sheet('Дані_тренування')
ws3.column_dimensions['A'].width = 42
hdr3 = ['ГРС'] + [m.strftime('%b %Y') + '\nФакт' for m in train_months] \
                + [m.strftime('%b %Y') + '\nПрогн' for m in train_months] \
                + [m.strftime('%b %Y') + '\nBias%' for m in train_months]
for ci, h in enumerate(hdr3, 1):
    c = ws3.cell(1, ci, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)
ws3.row_dimensions[1].height = 28

n_m = len(train_months)
for ri, grs in enumerate(sorted(loo_preds, key=lambda g: -grs_annual.get(g, 0)), 2):
    ws3.cell(ri, 1, grs)
    for ci, m in enumerate(train_months, 2):
        fact = float(Y_df.loc[grs, m]) if grs in Y_df.index and m in Y_df.columns else 0.0
        pred = loo_preds[grs].get(m, np.nan)
        ws3.cell(ri, ci,       round(fact)).number_format = '#,##0'
        if not np.isnan(pred):
            ws3.cell(ri, ci + n_m, round(pred)).number_format = '#,##0'
            if fact > 0:
                bias = (pred - fact) / fact * 100
                bc = ws3.cell(ri, ci + 2 * n_m, round(bias, 1))
                bc.number_format = '+0.0;-0.0;0.0'
                bc.alignment = Alignment(horizontal='center')
                f = color_bias(bias)
                if f: bc.fill = PatternFill('solid', fgColor=f)

ws3.freeze_panes = 'B2'


# ─ Лист 4: Ref-сигнали ────────────────────────────────────────────────────────
ws4 = wb.create_sheet('Ref_сигнали')
for ci, h in enumerate(['Місяць', 'ref_heat', 'ref_all', 'Сезон', 'Тип'], 1):
    c = ws4.cell(1, ci, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = HDR_FILL
for ri, mts in enumerate(sorted(month_to_dcols), 2):
    rh = ref_heat_m.get(mts, 0)
    ra = ref_all_m.get(mts, 0)
    seas = 'опал' if rh > threshold else 'літо'
    tp = 'ПРОГНОЗ' if mts == predict_month else 'тренування'
    ws4.cell(ri, 1, mts.strftime('%Y-%m'))
    ws4.cell(ri, 2, round(rh, 5)).number_format = '0.00000'
    ws4.cell(ri, 3, round(ra, 5)).number_format = '0.00000'
    sc = ws4.cell(ri, 4, seas)
    sc.fill = HEAT_FILL if seas == 'опал' else SUM_FILL
    tc = ws4.cell(ri, 5, tp)
    if tp == 'ПРОГНОЗ':
        tc.font = Font(bold=True)
        tc.fill = PRED_FILL

ws4.column_dimensions['A'].width = 12
for ci in range(2, 6):
    ws4.column_dimensions[get_column_letter(ci)].width = 14

wb.save(out_file)
print(f"  Збережено: {out_file}")
print(f"  Листи: {ws1.title} | Деталі_моделей | Дані_тренування | Ref_сигнали")

print(f"""
=== Підсумок ===
  Місяць прогнозу : {predict_month.strftime('%B %Y')}  ({predict_season})
  Тренувальних міс: {len(train_months)}
  Змодельовано ГРС: {len(preds)}
  Прогноз разом  : {total_pred/1e6:.3f} млн м³
  Медіана R² (LOO): {res_df['r2_loo'].median():.3f}

  Якість: {'✓ добра (R²≥0.85)' if res_df['r2_loo'].median() >= 0.85 else
           '⚠ прийнятна' if res_df['r2_loo'].median() >= 0.65 else
           '✗ погана — перевірте вхідні дані'}
""")
print("=== Готово ===")
