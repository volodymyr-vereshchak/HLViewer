"""
grs_rt_models.py

Двосезонні моделі прогнозу добового споживання по ГРС.

Сезони визначаються автоматично по ref_heat (KMeans k=2 на ОП-модемах):
  - Опалювальний:   q_grs(d) = a_heat + b_heat * ref_heat(d)
  - Неопалювальний: q_grs(d) = a_sum  + b_sum  * ref_all(d) + c_sum * ref_heat(d)

ref_heat(d) = sum_OP_active(v_i(d)) / sum_OP_area    [м³/(м²·день)]
ref_all(d)  = sum_ALL_active(v_i(d)) / sum_ALL_area   [м³/(м²·день)]

Влітку ОП-абоненти теж споживають (ПГ/ВПГ) → ref_all r=0.99 з Y.
ref_heat в літній моделі: у вересні починає зростати до порогу → вловлює
наближення опалювального сезону, знижує LOO-bias Лип/Сер з -9/-14% до +4/-3%.

Тренування:  Y = daily_decompose_grs.xlsx (GRS_daily sheet)
Валідація:   LOO-month CV для кожного сезону окремо
Виключення:  топ-3 ГРС за річним об'ємом

Результат:   data/grs_rt_models.xlsx
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score, mean_absolute_error
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pobut_predictor import GROUP_MERGE, META_COLS

MODEM_FILE   = "data/input/profile_pobut_daily_result.csv"
DECOMP_FILE  = "data/daily_decompose_grs.xlsx"
WEATHER_CSV  = "../data/weather_2025_daily_contractual.csv"
OUT_XLSX     = "data/grs_rt_models.xlsx"
N_SKIP_TOP   = 3   # виключаємо N найбільших ГРС

UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}


# ── 1. Модемні дані → ref_signal ─────────────────────────────────────────────
print("=== 1. ref_signal з ОП-модемів ===")
md = pd.read_csv(MODEM_FILE, sep=';', low_memory=False)
md.columns = META_COLS + list(md.columns[len(META_COLS):])
md['appliance_group'] = md['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
md = md[md['gas_off'].isna() & md['alternative'].isna() & md['dacha'].isna()]
md['has_OP'] = md['appliance_group'].str.contains('\u041e\u041f', na=False)
md['heated_area'] = pd.to_numeric(md['heated_area'], errors='coerce').fillna(55.0)
md['heated_area'] = md['heated_area'].where(md['heated_area'] > 0, 55.0)

md_op = md[md['has_OP']].copy()

date_cols = sorted(
    [c for c in md_op.columns
     if isinstance(c, str) and c.count('.') == 2 and len(c) == 10 and c.endswith('.2025')],
    key=lambda c: (int(c[6:]), int(c[3:5]), int(c[:2]))
)
dates_2025 = [pd.Timestamp(f"2025-{c[3:5]}-{c[:2]}") for c in date_cols]

for c in date_cols:
    md_op[c] = pd.to_numeric(md_op[c], errors='coerce').fillna(0)

md_vals = md_op[date_cols].values.astype(np.float32)

# Фільтр активних (як у daily_decompose_grs.py)
days_by_month = {m: [i for i, c in enumerate(date_cols) if int(c[3:5]) == m]
                 for m in range(1, 13)}
monthly = np.array([md_vals[:, days_by_month[m]].sum(axis=1) for m in range(1, 13)]).T
annual_modem = monthly.sum(axis=1)
max_frac = monthly.max(axis=1) / np.maximum(annual_modem, 1)
active_mask = (annual_modem > 0) & (max_frac < 0.60)

md_act   = md_op[active_mask].reset_index(drop=True)
vals_act = md_vals[active_mask]
area_op  = float(md_act['heated_area'].sum())

# ref_heat — тільки ОП-модеми (для визначення сезону та опал. моделі)
ref_heat = vals_act.sum(axis=0).astype(np.float64) / area_op

# ref_all — всі активні модеми (r≈0.99 влітку, ОП-абоненти теж споживають ПГ/ВПГ)
for c in date_cols:
    md[c] = pd.to_numeric(md[c], errors='coerce').fillna(0)
vals_all_raw = md[date_cols].values.astype(np.float32)
monthly_all  = np.array([vals_all_raw[:, days_by_month[m]].sum(axis=1)
                          for m in range(1, 13)]).T
act_all_mask = monthly_all.sum(axis=1) > 0
area_all     = float(md[act_all_mask]['heated_area'].sum())
ref_all      = vals_all_raw[act_all_mask].sum(axis=0).astype(np.float64) / area_all

date_to_heat = {d: ref_heat[i] for i, d in enumerate(dates_2025)}
date_to_all  = {d: ref_all[i]  for i, d in enumerate(dates_2025)}

print(f"  Активних ОП-модемів:  {len(md_act):,}  площа: {area_op:,.0f} м²")
print(f"  Активних всіх мод.:   {act_all_mask.sum():,}  площа: {area_all:,.0f} м²")
print(f"  ref_heat: min={ref_heat.min():.5f}  max={ref_heat.max():.5f}")
print(f"  ref_all:  min={ref_all.min():.5f}  max={ref_all.max():.5f}")


# ── 2. Визначення сезонів через KMeans(k=2) ───────────────────────────────────
print("\n=== 2. Визначення сезонів (KMeans на ref_heat) ===")

km = KMeans(n_clusters=2, random_state=42, n_init=10)
km.fit(ref_heat.reshape(-1, 1))
c0, c1 = sorted(km.cluster_centers_.flatten())   # c0=літо, c1=зима
threshold = (c0 + c1) / 2.0

# Масив міток для кожного дня
is_heating_full = ref_heat > threshold   # bool, (n_days,)
season_label    = np.where(is_heating_full, 'опал', 'літо')

print(f"  Центроїди: літо={c0:.5f}  зима={c1:.5f}")
print(f"  Поріг:     {threshold:.5f} м³/(м²·д)")
print(f"  Опалювальних днів: {is_heating_full.sum()} ({is_heating_full.mean()*100:.1f}%)")
print(f"  Неопалювальних:    {(~is_heating_full).sum()} ({(~is_heating_full).mean()*100:.1f}%)")

# Показати розподіл сезонів по місяцях
months_full = np.array([d.month for d in dates_2025])
print("\n  Склад сезонів по місяцях (опал. днів / всього):")
for m in range(1, 13):
    idx_m = months_full == m
    n_h = is_heating_full[idx_m].sum()
    n_t = idx_m.sum()
    bar = '█' * n_h + '░' * (n_t - n_h)
    print(f"    {UA[m]:4}: {n_h:2}/{n_t}  {bar}")


# ── 3. Погодні дані ───────────────────────────────────────────────────────────
print("\n=== 3. Погодні дані ===")
wdf = pd.read_csv(WEATHER_CSV)
wdf['date'] = pd.to_datetime(wdf['date']).dt.normalize()
wdf = (wdf[wdf['date'].dt.year == 2025]
       .drop_duplicates('date')
       .sort_values('date')
       .reset_index(drop=True))
wdf['temp'] = pd.to_numeric(wdf['temperature'], errors='coerce').fillna(wdf['temperature'].median())
# 7-денне ковзне середнє — знижує шум, доступне в реальному часі (rolling back)
wdf['temp_roll7'] = wdf['temp'].rolling(7, min_periods=1).mean()
date_to_temp      = dict(zip(wdf['date'], wdf['temp']))
date_to_temp7     = dict(zip(wdf['date'], wdf['temp_roll7']))
print(f"  Погодних днів 2025: {len(wdf)}")
print(f"  Температура: min={wdf['temp'].min():.1f}°C  max={wdf['temp'].max():.1f}°C")


# ── 4. Завантаження добового декомпозиту по ГРС ───────────────────────────────
print("\n=== 3. Завантаження daily_decompose_grs.xlsx ===")
df_raw = pd.read_excel(DECOMP_FILE, sheet_name='GRS_daily', index_col=0)
df_raw.columns = pd.to_datetime(df_raw.columns)

all_decomp_dates = sorted([d for d in df_raw.columns if d.year == 2025])
df_2025 = df_raw[all_decomp_dates].copy()
grs_all = df_2025.index.tolist()

annual_by_grs = df_2025.sum(axis=1).sort_values(ascending=False)
print(f"  ГРС всього: {len(grs_all)}")
print("  Топ-5 за річним об'ємом:")
for g, v in annual_by_grs.head(5).items():
    print(f"    {g[:55]:55}  {v/1e6:.3f} млн м³")


# ── 5. Виключення топ-N ГРС ───────────────────────────────────────────────────
skip_grs = set(annual_by_grs.head(N_SKIP_TOP).index)
print(f"\n  Виключаємо топ-{N_SKIP_TOP}:")
for g in sorted(skip_grs, key=lambda x: -annual_by_grs[x]):
    print(f"    {g}")
target_grs = [g for g in grs_all if g not in skip_grs]
print(f"  ГРС для моделювання: {len(target_grs)}")


# ── 6. Вирівнювання дат та побудова матриць фічів ────────────────────────────
common_dates = sorted(set(all_decomp_dates) & set(dates_2025))
print(f"\n  Спільних дат: {len(common_dates)}")

ref_heat_c = np.array([date_to_heat[d]  for d in common_dates])
ref_all_c  = np.array([date_to_all[d]   for d in common_dates])
temp_c     = np.array([date_to_temp.get(d, np.nan)  for d in common_dates])
temp7_c    = np.array([date_to_temp7.get(d, np.nan) for d in common_dates])
# fallback для відсутніх дат погоди
temp_c  = np.where(np.isnan(temp_c),  np.nanmedian(temp_c),  temp_c)
temp7_c = np.where(np.isnan(temp7_c), np.nanmedian(temp7_c), temp7_c)

heat_arr = np.array([is_heating_full[dates_2025.index(d)] for d in common_dates])
mon_arr  = np.array([d.month for d in common_dates])

# Опалювальна модель: ref_heat
X_heat = ref_heat_c.reshape(-1, 1)
# Літня модель: ref_all + ref_heat
# ref_all:  r≈0.99 in-sample (ОП-абоненти теж споживають ПГ/ВПГ влітку)
# ref_heat: в не-опалювальних днях < threshold, але НЕ нуль — у вересні
#   перші ОП-модеми починають вмикатись → ref_heat повільно зростає
#   → у LOO-Вер: жовтневі не-опал. дні (вищий ref_heat) навчають що
#     ref_heat↑ → споживання↑ → менший недоприогноз вересня.
# Температура НЕ додається: у LOO знак нестабільний.
X_sum  = np.column_stack([ref_all_c, ref_heat_c])

heating_months = sorted({mon_arr[i] for i in range(len(mon_arr)) if heat_arr[i]})
nonheat_months = sorted({mon_arr[i] for i in range(len(mon_arr)) if not heat_arr[i]})
print(f"  Опалювальні місяці (мають опал. дні): {[UA[m] for m in heating_months]}")
print(f"  Неопалювальні місяці:                  {[UA[m] for m in nonheat_months]}")


# ── 7. Двосезонна регресія per GRS ────────────────────────────────────────────
print("\n=== 7. Двосезонна модель [LOO-month CV] ===")
print(f"  Опал.: q = a + b * ref_heat  (дні з ref_heat > {threshold:.5f})")
print(f"  Літо:  q = a + b*ref_all + c*ref_heat  (всі модеми + ОП-сигнал)")
print()

results  = []
pred_cv  = {}

for grs in target_grs:
    Y = df_2025.loc[grs, common_dates].values.astype(np.float64)

    if Y.sum() < 50.0:
        continue

    y_pred = np.full(len(Y), np.nan)

    # ── Опалювальні дні: LOO по місяцях (ref_heat) ──────────────────────────
    h_idx    = np.where(heat_arr)[0]
    h_months = mon_arr[h_idx]
    Xh = X_heat[h_idx]
    Yh = Y[h_idx]

    y_pred_heat = np.zeros(len(h_idx))
    for test_m in np.unique(h_months):
        tr = h_months != test_m
        te = h_months == test_m
        if tr.sum() < 20 or te.sum() == 0:
            continue
        lr = LinearRegression()
        lr.fit(Xh[tr], Yh[tr])
        y_pred_heat[te] = lr.predict(Xh[te])

    y_pred[h_idx] = np.maximum(y_pred_heat, 0.0)

    # ── Неопалювальні дні: LOO по місяцях (ref_all) ─────────────────────────
    # ref_all = всі модеми (ОП+не-ОП), r≈0.99 in-sample.
    # Температура не покращує LOO через нестабільний знак коефіцієнту влітку.
    s_idx    = np.where(~heat_arr)[0]
    s_months = mon_arr[s_idx]
    Xs = X_sum[s_idx]     # [ref_all, temp7]
    Ys = Y[s_idx]

    y_pred_summer = np.zeros(len(s_idx))
    for test_m in np.unique(s_months):
        tr = s_months != test_m
        te = s_months == test_m
        if tr.sum() < 20 or te.sum() == 0:
            y_pred_summer[te] = np.mean(Ys)
            continue
        lr = LinearRegression()
        lr.fit(Xs[tr], Ys[tr])
        y_pred_summer[te] = lr.predict(Xs[te])

    y_pred[s_idx] = np.maximum(y_pred_summer, 0.0)

    # Залишки NaN (якщо якийсь опал. місяць не мав пари) — заповнити лінійною моделлю
    nan_mask = np.isnan(y_pred)
    if nan_mask.any():
        lr_fb = LinearRegression().fit(X_heat[~nan_mask], Y[~nan_mask])
        y_pred[nan_mask] = np.maximum(lr_fb.predict(X_heat[nan_mask]), 0.0)

    r2  = r2_score(Y, y_pred) if Y.std() > 0 else 0.0
    mae = mean_absolute_error(Y, y_pred)

    # Bias по місяцях
    monthly_bias = {}
    for m in range(1, 13):
        idx_m = mon_arr == m
        if idx_m.sum() == 0:
            continue
        ys = Y[idx_m].sum()
        yp = y_pred[idx_m].sum()
        monthly_bias[m] = round((yp - ys) / max(abs(ys), 1) * 100, 1)

    # Повні моделі (для збереження коефіцієнтів)
    lr_h = LinearRegression().fit(Xh, Yh)
    a_heat = float(lr_h.intercept_);  b_heat = float(lr_h.coef_[0])
    lr_s = LinearRegression().fit(Xs, Ys)
    a_sum = float(lr_s.intercept_)
    b_sum = float(lr_s.coef_[0])   # коеф при ref_all
    c_sum = float(lr_s.coef_[1])   # коеф при ref_heat

    results.append({
        'grs':      grs,
        'n_days':   len(Y),
        'annual_m3': round(Y.sum()),
        'a_heat':   round(a_heat, 2),
        'b_heat':   round(b_heat, 2),
        'a_sum':    round(a_sum,  2),
        'b_sum':    round(b_sum,  2),
        'c_sum':    round(c_sum,  2),
        'r2':       round(r2, 3),
        'mae_day':  round(mae, 1),
        **{UA[m]: monthly_bias.get(m) for m in range(1, 13)},
    })
    pred_cv[grs] = y_pred

    print(f"  {grs[:38]:38s}  R²={r2:+.3f}  MAE={mae:7,.1f}  "
          f"ah={a_heat:7.1f} bh={b_heat:8.0f}  as={a_sum:6.1f} bs={b_sum:7.0f} cs={c_sum:7.1f}")

res_df = pd.DataFrame(results)
print(f"\n  Змодельовано ГРС: {len(res_df)}")
print(f"  Медіана R²:       {res_df['r2'].median():.3f}")
print(f"  Медіана MAE/day:  {res_df['mae_day'].median():,.1f} м³/день")

print("\n  Медіанний місячний bias%:")
for m in range(1, 13):
    col = UA[m]
    if col in res_df.columns:
        v = res_df[col].median()
        mark = '← опал' if m in heating_months else '← літо'
        bar  = '█' * min(int(abs(v) / 2), 20)
        print(f"    {col:4}  {v:+6.1f}%  {bar}  {mark}")


# ── 7. Місячна агрегація ──────────────────────────────────────────────────────
print("\n=== 7. Місячна агрегація (факт vs прогноз CV) ===")
month_summary = []
for m in range(1, 13):
    idx_m = mon_arr == m
    fact_tot = sum(df_2025.loc[g, common_dates].values[idx_m].sum()
                   for g in pred_cv)
    pred_tot = sum(pred_cv[g][idx_m].sum() for g in pred_cv)
    bias = (pred_tot - fact_tot) / max(fact_tot, 1) * 100
    season = 'опал' if m in heating_months else 'літо'
    month_summary.append({'month': m, 'season': season,
                          'fact_m3': fact_tot, 'pred_m3': pred_tot,
                          'bias_pct': round(bias, 2)})
    print(f"  {UA[m]:4} [{season:4}]  факт={fact_tot/1e3:8.1f}тис  "
          f"прогноз={pred_tot/1e3:8.1f}тис  bias={bias:+.2f}%")

ms_df = pd.DataFrame(month_summary)
tot_f = ms_df['fact_m3'].sum()
tot_p = ms_df['pred_m3'].sum()
print(f"\n  Рік  факт={tot_f/1e6:.3f} млн  прогноз={tot_p/1e6:.3f} млн  "
      f"bias={(tot_p-tot_f)/tot_f*100:+.2f}%")

# Окремо по сезонах
for seas in ['опал', 'літо']:
    sf = ms_df[ms_df['season'] == seas]['fact_m3'].sum()
    sp = ms_df[ms_df['season'] == seas]['pred_m3'].sum()
    print(f"  {seas:5} факт={sf/1e6:.3f} млн  прогноз={sp/1e6:.3f} млн  "
          f"bias={(sp-sf)/max(sf,1)*100:+.2f}%")


# ── 8. Збереження Excel ───────────────────────────────────────────────────────
print(f"\n=== 8. Збереження {OUT_XLSX} ===")

def color_r2(v):
    if v >= 0.95: return '00B050'
    if v >= 0.85: return '92D050'
    if v >= 0.70: return 'FFEB9C'
    return 'FF7C80'

def color_bias(v):
    a = abs(v)
    if a <= 2:  return None
    if a <= 5:  return 'FFEB9C'
    if a <= 10: return 'FFC7CE'
    return 'FF4C4C'

HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
HEAT_FILL  = PatternFill('solid', fgColor='FCE4D6')   # опал. місяць — персиковий
SUM_FILL   = PatternFill('solid', fgColor='E2EFDA')   # літній — зелений

wb = Workbook()

# ─ Sheet 1: Коефіцієнти та метрики ────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Моделі_GRS'

hdr = ['ГРС', 'N_днів', 'Річний м³',
       'a_heat', 'b_heat', 'a_sum', 'b_sum(ref_all)', 'c_sum(ref_heat)',
       'R²_CV', 'MAE_день'] + [UA[m] + '_%' for m in range(1, 13)]
for ci, h in enumerate(hdr, 1):
    c = ws1.cell(1, ci, h)
    c.font      = Font(bold=True, color='FFFFFF')
    c.fill      = HDR_FILL
    c.alignment = Alignment(horizontal='center')
ws1.column_dimensions['A'].width = 40
for ci in range(2, len(hdr) + 1):
    ws1.column_dimensions[get_column_letter(ci)].width = 12

for ri, row in enumerate(
        res_df.sort_values('r2', ascending=False).to_dict('records'), 2):
    ws1.cell(ri, 1, row['grs'])
    ws1.cell(ri, 2, row['n_days'])
    ws1.cell(ri, 3, row['annual_m3']).number_format = '#,##0'
    ws1.cell(ri, 4, row['a_heat']).number_format = '#,##0.00'
    ws1.cell(ri, 5, row['b_heat']).number_format = '#,##0.00'
    ws1.cell(ri, 6, row['a_sum']).number_format  = '#,##0.00'
    ws1.cell(ri, 7, row['b_sum']).number_format  = '#,##0.00'
    ws1.cell(ri, 8, row['c_sum']).number_format  = '#,##0.00'

    r2_c = ws1.cell(ri, 9, row['r2'])
    r2_c.number_format = '0.000'
    r2_c.font = Font(bold=True, color=color_r2(row['r2']))

    ws1.cell(ri, 10, row['mae_day']).number_format = '#,##0.0'

    for ci, m in enumerate(range(1, 13), 11):
        v = row.get(UA[m])
        if v is None:
            continue
        cell = ws1.cell(ri, ci, v)
        cell.number_format = '+0.0;-0.0;0.0'
        cell.alignment     = Alignment(horizontal='center')
        fill_c = color_bias(v)
        if fill_c:
            cell.fill = PatternFill('solid', fgColor=fill_c)

# Підфарбувати заголовки місяців за сезоном
for ci, m in enumerate(range(1, 13), 11):
    h_cell = ws1.cell(1, ci)
    h_cell.fill = HEAT_FILL if m in heating_months else SUM_FILL
    h_cell.font = Font(bold=True)

# Рядок медіан
ri_med = len(res_df) + 3
ws1.cell(ri_med, 1, 'Медіана').font = Font(bold=True)
ws1.cell(ri_med, 9,  round(res_df['r2'].median(), 3)).number_format = '0.000'
ws1.cell(ri_med, 10, round(res_df['mae_day'].median(), 1)).number_format = '#,##0.0'
for ci, m in enumerate(range(1, 13), 11):
    col_ua = UA[m]
    if col_ua in res_df.columns:
        v = round(float(res_df[col_ua].median()), 1)
        cell = ws1.cell(ri_med, ci, v)
        cell.number_format = '+0.0;-0.0;0.0'
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        fill_c = color_bias(v)
        if fill_c:
            cell.fill = PatternFill('solid', fgColor=fill_c)

ws1.freeze_panes = 'B2'

# ─ Sheet 2: Місячний підсумок ─────────────────────────────────────────────────
ws2 = wb.create_sheet('Місячний_підсумок')
for ci, h in enumerate(['Місяць', 'Сезон', 'Факт м³', 'Прогноз м³', 'Bias%'], 1):
    ws2.cell(1, ci, h).font = Font(bold=True, color='FFFFFF')
    ws2.cell(1, ci).fill   = HDR_FILL

for ri, row in enumerate(ms_df.to_dict('records'), 2):
    ws2.cell(ri, 1, UA[row['month']])
    ws2.cell(ri, 2, row['season'])
    ws2.cell(ri, 3, round(row['fact_m3'])).number_format = '#,##0'
    ws2.cell(ri, 4, round(row['pred_m3'])).number_format = '#,##0'
    bc = ws2.cell(ri, 4 + 1, row['bias_pct'])
    bc.number_format = '+0.00;-0.00;0.00'
    fill_c = color_bias(row['bias_pct'])
    if fill_c:
        bc.fill = PatternFill('solid', fgColor=fill_c)
    s_fill = HEAT_FILL if row['season'] == 'опал' else SUM_FILL
    for ci2 in range(1, 6):
        ws2.cell(ri, ci2).fill = s_fill

ri_tot = len(ms_df) + 3
ws2.cell(ri_tot, 1, 'РІК').font = Font(bold=True)
ws2.cell(ri_tot, 3, round(tot_f)).number_format = '#,##0'
ws2.cell(ri_tot, 4, round(tot_p)).number_format = '#,##0'
tot_bias = round((tot_p - tot_f) / max(tot_f, 1) * 100, 2)
ws2.cell(ri_tot, 5, tot_bias).number_format = '+0.00;-0.00;0.00'

# ─ Sheet 3: GRS × місяць ─────────────────────────────────────────────────────
ws3 = wb.create_sheet('GRS_місяць')
ws3.cell(1, 1, 'ГРС').font = Font(bold=True)
ws3.column_dimensions['A'].width = 40
for mi, m in enumerate(range(1, 13)):
    c0   = 2 + mi * 3
    s_f  = HEAT_FILL if m in heating_months else SUM_FILL
    ws3.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + 2)
    hc   = ws3.cell(1, c0, UA[m])
    hc.font      = Font(bold=True)
    hc.fill      = s_f
    hc.alignment = Alignment(horizontal='center')
    for ci2, sub in enumerate(['Факт', 'Прогн', 'Bias%'], c0):
        c = ws3.cell(2, ci2, sub)
        c.font      = Font(bold=True, italic=True, size=8)
        c.fill      = s_f
        c.alignment = Alignment(horizontal='center')
ws3.freeze_panes = 'B3'

ri3 = 3
for grs in target_grs:
    if grs not in pred_cv:
        continue
    ws3.cell(ri3, 1, grs)
    Y_g  = df_2025.loc[grs, common_dates].values.astype(np.float64)
    yp_g = pred_cv[grs]
    for mi, m in enumerate(range(1, 13)):
        idx_m = mon_arr == m
        ys    = Y_g[idx_m].sum()
        yp    = yp_g[idx_m].sum()
        bias  = (yp - ys) / max(abs(ys), 1) * 100
        c0    = 2 + mi * 3
        ws3.cell(ri3, c0,     round(ys / 1e3, 1))
        ws3.cell(ri3, c0 + 1, round(yp / 1e3, 1))
        bc = ws3.cell(ri3, c0 + 2, round(bias, 1))
        bc.number_format = '+0.0;-0.0;0.0'
        fill_c = color_bias(bias)
        if fill_c:
            bc.fill = PatternFill('solid', fgColor=fill_c)
    ri3 += 1

# ─ Sheet 4: Сезонна маска (день → сезон) ─────────────────────────────────────
ws4 = wb.create_sheet('Сезонна_маска')
for ci, h in enumerate(['Дата', 'ref_heat', 'ref_all', 'Сезон', 'Місяць'], 1):
    ws4.cell(1, ci, h).font = Font(bold=True)
ws4.column_dimensions['A'].width = 14
for ri, d in enumerate(common_dates, 2):
    ref_h  = date_to_heat[d]
    ref_a  = date_to_all[d]
    is_h   = ref_h > threshold
    ws4.cell(ri, 1, d.strftime('%Y-%m-%d'))
    ws4.cell(ri, 2, round(ref_h, 6))
    ws4.cell(ri, 3, round(ref_a, 6))
    sc = ws4.cell(ri, 4, 'опал' if is_h else 'літо')
    sc.fill = HEAT_FILL if is_h else SUM_FILL
    ws4.cell(ri, 5, d.month)

# ─ Sheet 5: Виключені ГРС ────────────────────────────────────────────────────
ws5 = wb.create_sheet('Виключені_ГРС')
for ci, h in enumerate(['ГРС', 'Річний м³', 'Причина'], 1):
    ws5.cell(1, ci, h).font = Font(bold=True)
for ri2, g in enumerate(sorted(skip_grs, key=lambda x: -annual_by_grs[x]), 2):
    ws5.cell(ri2, 1, g)
    ws5.cell(ri2, 2, round(annual_by_grs[g])).number_format = '#,##0'
    ws5.cell(ri2, 3, "Топ-3 ГРС за об'ємом")

# ─ Sheet 6: Bias-хітмап (GRS × Місяць) ──────────────────────────────────────
ws6 = wb.create_sheet('Bias_хітмап')
ws6.column_dimensions['A'].width = 40

# Заголовок рядок 1: назви місяців + Річний
hdr6 = ['ГРС'] + [UA[m] for m in range(1, 13)] + ['Річний']
for ci, h in enumerate(hdr6, 1):
    c = ws6.cell(1, ci, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
# Підфарбувати заголовки за сезоном
for ci, m in enumerate(range(1, 13), 2):
    ws6.cell(1, ci).fill = HEAT_FILL if m in heating_months else SUM_FILL
    ws6.cell(1, ci).font = Font(bold=True)

# Рядок 2: підпис Факт тис / Прогн тис / Bias%
ws6.cell(2, 1, 'Bias % (LOO-CV)').font = Font(italic=True, size=9, color='595959')

ri6 = 3
for grs in sorted(pred_cv, key=lambda g: -df_2025.loc[g, common_dates].sum()):
    ws6.cell(ri6, 1, grs)
    Y_g  = df_2025.loc[grs, common_dates].values.astype(np.float64)
    yp_g = pred_cv[grs]
    ann_f = Y_g.sum(); ann_p = yp_g.sum()
    for ci, m in enumerate(range(1, 13), 2):
        idx_m = mon_arr == m
        ys = Y_g[idx_m].sum(); yp = yp_g[idx_m].sum()
        if ys == 0: continue
        bias = (yp - ys) / abs(ys) * 100
        cell = ws6.cell(ri6, ci, round(bias, 1))
        cell.number_format = '+0.0;-0.0;0.0'
        cell.alignment = Alignment(horizontal='center')
        fill_c = color_bias(bias)
        if fill_c:
            cell.fill = PatternFill('solid', fgColor=fill_c)
    ann_bias = (ann_p - ann_f) / max(abs(ann_f), 1) * 100
    ac = ws6.cell(ri6, 14, round(ann_bias, 1))
    ac.number_format = '+0.0;-0.0;0.0'
    ac.alignment = Alignment(horizontal='center')
    ac.font = Font(bold=True)
    fill_c = color_bias(ann_bias)
    if fill_c:
        ac.fill = PatternFill('solid', fgColor=fill_c)
    ri6 += 1

# Медіана рядок
ri6 += 1
ws6.cell(ri6, 1, 'Медіана').font = Font(bold=True)
for ci, m in enumerate(range(1, 13), 2):
    col_ua = UA[m]
    if col_ua in res_df.columns:
        v = round(float(res_df[col_ua].median()), 1)
        cell = ws6.cell(ri6, ci, v)
        cell.number_format = '+0.0;-0.0;0.0'
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        fill_c = color_bias(v)
        if fill_c:
            cell.fill = PatternFill('solid', fgColor=fill_c)

ws6.freeze_panes = 'B3'

# ─ Sheet 7: Об'єми (Факт vs Прогноз, тис м³) ─────────────────────────────────
ws7 = wb.create_sheet('Обсяги_тис_м3')
ws7.column_dimensions['A'].width = 40

# Подвійний заголовок: місяці (merged) + Рядок 2: Ф/П для кожного місяця
for ci, m in enumerate(range(1, 13)):
    c0 = 2 + ci * 2
    sf = HEAT_FILL if m in heating_months else SUM_FILL
    ws7.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + 1)
    hc = ws7.cell(1, c0, UA[m])
    hc.font = Font(bold=True); hc.fill = sf
    hc.alignment = Alignment(horizontal='center')
    ws7.cell(2, c0,     'Факт').font  = Font(bold=True, italic=True, size=8)
    ws7.cell(2, c0 + 1, 'Прогн').font = Font(bold=True, italic=True, size=8)
    ws7.cell(2, c0).fill     = sf
    ws7.cell(2, c0 + 1).fill = sf
# Річна колонка
ws7.merge_cells(start_row=1, start_column=26, end_row=1, end_column=27)
ws7.cell(1, 26, 'РІК').font = Font(bold=True); ws7.cell(1, 26).fill = HDR_FILL
ws7.cell(1, 26).alignment = Alignment(horizontal='center')
ws7.cell(2, 26, 'Факт').font  = Font(bold=True, italic=True, size=8)
ws7.cell(2, 27, 'Прогн').font = Font(bold=True, italic=True, size=8)
ws7.cell(1, 1, 'ГРС').font = Font(bold=True, color='FFFFFF'); ws7.cell(1, 1).fill = HDR_FILL

ri7 = 3
for grs in sorted(pred_cv, key=lambda g: -df_2025.loc[g, common_dates].sum()):
    ws7.cell(ri7, 1, grs)
    Y_g  = df_2025.loc[grs, common_dates].values.astype(np.float64)
    yp_g = pred_cv[grs]
    for ci, m in enumerate(range(1, 13)):
        c0 = 2 + ci * 2
        ys = Y_g[mon_arr == m].sum()
        yp = yp_g[mon_arr == m].sum()
        ws7.cell(ri7, c0,     round(ys / 1e3, 1)).number_format = '#,##0.0'
        ws7.cell(ri7, c0 + 1, round(yp / 1e3, 1)).number_format = '#,##0.0'
    ws7.cell(ri7, 26, round(Y_g.sum()  / 1e3, 1)).number_format = '#,##0.0'
    ws7.cell(ri7, 27, round(yp_g.sum() / 1e3, 1)).number_format = '#,##0.0'
    ri7 += 1

ws7.freeze_panes = 'B3'

wb.save(OUT_XLSX)
print(f"  Збережено: {OUT_XLSX}")
print(f"  Листи: Моделі_GRS | Місячний_підсумок | GRS_місяць | Сезонна_маска | Виключені_ГРС | Bias_хітмап | Обсяги_тис_м3")
print(f"\n  Поріг сезону: ref_signal > {threshold:.5f} = опалювальний день")
print("\n=== Готово ===")
