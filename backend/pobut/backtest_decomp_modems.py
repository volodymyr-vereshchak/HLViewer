"""
backtest_decomp_modems.py

Для кожної ГРС: Ridge-модель із системних модемних фіч по групах.

Фічі (system-wide, щоденні суми по всіх модемах):
  m_opg_pr    — ОП,ПГ Приватний сектор
  m_opg_mk    — ОП,ПГ Багатоквартирний сектор
  m_opgv_pr   — ОП,ПГ,ВПГ Приватний сектор
  m_opgv_mk   — ОП,ПГ,ВПГ Багатоквартирний сектор
  m_pg_mk     — ПГ Багатоквартирний сектор
  m_pgv_mk    — ПГ,ВПГ Багатоквартирний сектор
  m_total     — усі групи разом
  m_op_all    — усі ОП-групи
  m_pg_only   — тільки ПГ/ПГ,ВПГ
  sin_doy, cos_doy, sin_m, cos_m  — сезонні

Таргет: monthly_pred(grs, m) / days_in_month(m)
Бектест: train Jan–Sep 2025, test Oct–Dec 2025
Per-GRS Ridge з CV на alpha. Усі 35 ГРС.

Примітка: погодні фічі (температура, hdh_18c) тестувались, але погіршують
  результат через мультиколінеарність з модемними фічами → CV вибирає α→0
  (near-OLS), що призводить до перенавчання і росту жовтневого bias до -19%.
  Структурна причина: жовтневий декомпозит є billing-based (рахунки спалахують
  на початку сезону), тоді як модемне споживання в жовтні ще тільки зростає.
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from sklearn.linear_model import Ridge
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import cross_val_score
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pobut_predictor import GROUP_MERGE, META_COLS, CT_PRIV

MODEM_FILE  = 'data/input/profile_pobut_daily_result.csv'
DECOMP_XLSX = 'data/annual_decompose_result.xlsx'
# Режим таргету: 'decompose' — total_pred (профільний декомпозит)
#                'billing'   — сирий місячний білінг по ГРС
TARGET_MODE = 'billing'

OUT_XLSX = f'data/backtest_decomp_modems_{TARGET_MODE}.xlsx'

TRAIN_MONTHS = list(range(1, 13))   # Jan–Dec (всі місяці)
TEST_MONTHS  = list(range(1, 13))   # Jan–Dec (in-sample)
ALL_MONTHS   = list(range(1, 13))
UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}
LARGE_GRS = 5_000_000
ALPHAS    = [0.001, 0.01, 0.1, 1, 10, 100, 1000]

# Групи модемів → короткі коди фіч
GROUPS_CODE = [
    (('ОП,ПГ',     'Приватний сектор'),       'm_opg_pr'),
    (('ОП,ПГ',     'Багатоквартирний сектор'), 'm_opg_mk'),
    (('ОП,ПГ,ВПГ', 'Приватний сектор'),       'm_opgv_pr'),
    (('ОП,ПГ,ВПГ', 'Багатоквартирний сектор'), 'm_opgv_mk'),
    (('ПГ',        'Багатоквартирний сектор'), 'm_pg_mk'),
    (('ПГ,ВПГ',    'Багатоквартирний сектор'), 'm_pgv_mk'),
]
GRP_FEAT_NAMES = [code for _, code in GROUPS_CODE]

# ── 1. Модемні дані ───────────────────────────────────────────────────────────
print("=== 1. Завантаження модемних даних ===")
md = pd.read_csv(MODEM_FILE, sep=';', low_memory=False)
md.columns = META_COLS + list(md.columns[len(META_COLS):])
md['appliance_group'] = md['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
md['consumer_type']   = md['consumer_type'].fillna(CT_PRIV)
md['account_id']      = pd.to_numeric(md['account_id'], errors='coerce')
md = md[md['gas_off'].isna() & md['alternative'].isna() & md['dacha'].isna()]
print(f"  Активних модемів: {len(md):,}")

# Знаходимо денні колонки 2025
date_cols_2025 = []
for c in md.columns:
    if isinstance(c, str) and c.count('.') == 2 and len(c) == 10:
        try:
            dd, mm, yy = c.split('.')
            dt = pd.Timestamp(f'{yy}-{mm}-{dd}')
            if dt.year == 2025:
                date_cols_2025.append((dt, c))
        except Exception:
            pass
date_cols_2025.sort()
day_col_names = [col for _, col in date_cols_2025]
print(f"  Денних колонок 2025: {len(date_cols_2025)}")

# Конвертуємо числові значення
md_num = (md[day_col_names]
          .apply(pd.to_numeric, errors='coerce')
          .fillna(0).clip(lower=0))

# ── 2. Системні фічі по групах (всі модеми) ──────────────────────────────────
print("\n=== 2. Системні фічі по групах ===")
feat_rows = {}   # feat_name → array shape (n_days,)

for (grp, ct), fname in GROUPS_CODE:
    mask = ((md['appliance_group'] == grp) & (md['consumer_type'] == ct)).values
    n    = mask.sum()
    if n > 0:
        daily_sum = md_num.values[mask].sum(axis=0)
    else:
        daily_sum = np.zeros(len(date_cols_2025))
    feat_rows[fname] = daily_sum
    print(f"  {fname:12s}  n={n:4d}  "
          f"mean={daily_sum.mean():7.1f}  max={daily_sum.max():7.1f}  "
          f"total={daily_sum.sum():,.0f} м³")

# Агрегати
feat_rows['m_total']  = sum(feat_rows[c] for c in GRP_FEAT_NAMES)
feat_rows['m_op_all'] = (feat_rows['m_opg_pr']  + feat_rows['m_opg_mk'] +
                         feat_rows['m_opgv_pr'] + feat_rows['m_opgv_mk'])
feat_rows['m_pg_only'] = feat_rows['m_pg_mk'] + feat_rows['m_pgv_mk']

# Датафрейм фіч індексований по даті
dates_arr = [dt for dt, _ in date_cols_2025]
df_sys = pd.DataFrame(feat_rows, index=dates_arr)
df_sys.index.name = 'date'
df_sys['month']   = [dt.month      for dt in dates_arr]
df_sys['doy']     = [dt.day_of_year for dt in dates_arr]
df_sys['sin_doy'] = np.sin(2 * np.pi * df_sys['doy'] / 365)
df_sys['cos_doy'] = np.cos(2 * np.pi * df_sys['doy'] / 365)
df_sys['sin_m']   = np.sin(2 * np.pi * df_sys['month'] / 12)
df_sys['cos_m']   = np.cos(2 * np.pi * df_sys['month'] / 12)

df_sys = df_sys.reset_index()

FEAT_COLS = GRP_FEAT_NAMES + ['m_total', 'm_op_all', 'm_pg_only',
                               'sin_doy', 'cos_doy', 'sin_m', 'cos_m']
print(f"\n  Фіч: {len(FEAT_COLS)}  ({FEAT_COLS})")

# ── 3a. Денні групові форми ───────────────────────────────────────────────────
# shape(grp, d) = modem_grp(d) / Σ_{d'∈month} modem_grp(d')  → сума по місяцю = 1.0
print("\n=== 3a. Денні групові форми ===")
dates_arr  = [dt for dt, _ in date_cols_2025]
months_arr = np.array([dt.month for dt in dates_arr])

SHAPE_KEYS = GRP_FEAT_NAMES + ['m_total']   # + system total as fallback
grp_shapes = {}
for fname in SHAPE_KEYS:
    daily  = feat_rows[fname].astype(float)
    shape  = np.zeros_like(daily)
    for m in range(1, 13):
        mm = months_arr == m
        s  = daily[mm].sum()
        shape[mm] = daily[mm] / s if s > 0 else 1.0 / mm.sum()
    grp_shapes[fname] = shape
    sums = {m: grp_shapes[fname][months_arr == m].sum() for m in [1, 7, 10]}
    print(f"  {fname:12s}  сума_по_міс(Січ={sums[1]:.3f}  Лип={sums[7]:.3f}  Жов={sums[10]:.3f})")

# ── 3b. Груповий мікс по ГРС (не-модемні споживачі) ──────────────────────────
print("\n=== 3b. Груповий мікс по ГРС (all_pobut_enriched) ===")
ALL_POBUT_CSV = 'data/input/all_pobut_enriched.csv'
BILL_COLS     = ['jan_2025','feb_2025','mar_2025','apr_2025','may_2025','jun_2025',
                 'jul_2025','aug_2025','sep_2025','oct_2025','nov_2025','dec_2025']

ap = pd.read_csv(ALL_POBUT_CSV)
ap['account_id']      = pd.to_numeric(ap['account_id'], errors='coerce')
ap['appliance_group'] = ap['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
ap['consumer_type']   = ap['consumer_type'].fillna(CT_PRIV)
ap = ap[ap['gas_off'].isna() & ap['alternative'].isna() & ap['dacha'].isna()]

modem_ids = set(md['account_id'].dropna().astype(int))
nm = ap[~ap['account_id'].isin(modem_ids)].copy()
nm['annual_billing'] = nm[BILL_COLS].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1)
nm = nm[nm['annual_billing'] > 0]

GRP_KEY_TO_FNAME = {(grp, ct): fname for (grp, ct), fname in GROUPS_CODE}
nm['fname'] = nm.apply(
    lambda r: GRP_KEY_TO_FNAME.get((r['appliance_group'], r['consumer_type']), 'm_total'),
    axis=1)

grs_mix = (nm.groupby(['grs', 'fname'])['annual_billing'].sum()
             .reset_index()
             .rename(columns={'annual_billing': 'grp_billing'}))
grs_total = nm.groupby('grs')['annual_billing'].sum().rename('total_billing')
grs_mix   = grs_mix.merge(grs_total, on='grs')
grs_mix['mix'] = grs_mix['grp_billing'] / grs_mix['total_billing']
grs_mix_dict = (grs_mix.groupby('grs')
                        .apply(lambda g: dict(zip(g['fname'], g['mix'])),
                               include_groups=False)
                        .to_dict())
print(f"  Не-модемних споживачів: {len(nm):,}   ГРС з міксом: {len(grs_mix_dict)}")

# ── 3c. Місячний декомпозит + денний таргет (справжній варіант) ──────────────
print(f"\n=== 3c. Місячний {'білінг' if TARGET_MODE=='billing' else 'декомпозит'} → денний таргет через форми ===")
dc_raw = pd.read_excel(DECOMP_XLSX, sheet_name='All_GRS_monthly')
dc_raw['grs'] = dc_raw['grs'].astype(str)
if TARGET_MODE == 'billing':
    dc = dc_raw[['grs', 'month', 'billing']].rename(columns={'billing': 'total_pred'}).copy()
    print("  Режим: СИРИЙ БІЛІНГ (billing) → дневний таргет")
else:
    dc = dc_raw[['grs', 'month', 'total_pred']].copy()
    print("  Режим: ДЕКОМПОЗИТ (total_pred) → дневний таргет")

days_in_month = {m: pd.Period(f'2025-{m:02d}', 'M').days_in_month for m in range(1, 13)}
ann_decomp = dc.groupby('grs')['total_pred'].sum().sort_values(ascending=False)

# Для кожної ГРС: eff_shape(d) = Σ_grp mix(grp) × grp_shapes(grp, d)
# daily_target(d) = monthly_decomp(m_of_d) × eff_shape(d)
# Перевірка: Σ_{d∈m} daily_target(d) = monthly_decomp(m) × Σ_{d∈m} eff_shape(d) = monthly_decomp(m) ✓
grs_daily_tgt = {}   # grs → np.array len(dates)

for grs in ann_decomp.index:
    dc_grs = dc[dc['grs'] == grs].set_index('month')['total_pred'].to_dict()
    if not dc_grs:
        continue

    mix = grs_mix_dict.get(grs, {'m_total': 1.0})
    total_mix = sum(mix.values())

    eff_shape = np.zeros(len(dates_arr))
    for fname, w in mix.items():
        eff_shape += (w / total_mix) * grp_shapes.get(fname, grp_shapes['m_total'])

    monthly_vals = np.array([dc_grs.get(m, 0.0) for m in months_arr])
    grs_daily_tgt[grs] = monthly_vals * eff_shape

print(f"  ГРС з денними таргетами: {len(grs_daily_tgt)}")
# Санітарна перевірка: місячні суми мають збігатися з original
sample_grs = list(ann_decomp.index)[0]
for m in [1, 10, 12]:
    mm = months_arr == m
    orig = dc[(dc['grs'] == sample_grs) & (dc['month'] == m)]['total_pred'].sum()
    recon = grs_daily_tgt[sample_grs][mm].sum()
    print(f"  [{sample_grs[:30]}] міс={m:2d}  orig={orig:,.0f}  реконстр={recon:,.0f}  "
          f"delta={recon-orig:+,.0f}")

# ── 4. Per-GRS Ridge моделі ───────────────────────────────────────────────────
te_lbl  = '+'.join(UA[m] for m in TEST_MONTHS)
tr_lbl  = f"Jan–{UA[TRAIN_MONTHS[-1]]}"
print(f"\n=== 4. Per-GRS Ridge (train {tr_lbl}, test {te_lbl}) ===")
te_hdrs = '  '.join(f"{UA[mo]:>7s}" for mo in TEST_MONTHS)
print(f"  {'ГРС':42s}  {'млн':>5s}  {'R²tr':>6s}  {'R²te':>6s}  "
      f"{'MAEte_d':>9s}  {te_hdrs}  α")
print("  " + "─" * (88 + len(TEST_MONTHS) * 10))

all_preds   = []
grs_metrics = []

for grs in ann_decomp.index:
    if grs not in grs_daily_tgt:
        continue
    tgt_arr = grs_daily_tgt[grs]
    if tgt_arr.sum() == 0:
        continue

    # Приєднуємо денний таргет до системних фіч (по позиції — обидва 365 елементів)
    df_grs = df_sys.copy()
    df_grs['daily_target'] = tgt_arr
    # total_pred по місяцях (для bias-розрахунку)
    dc_grs_monthly = dc[dc['grs'] == grs].set_index('month')['total_pred'].to_dict()
    df_grs['total_pred'] = df_grs['month'].map(dc_grs_monthly).fillna(0)
    df_grs = df_grs.sort_values('date').reset_index(drop=True)

    is_large = float(ann_decomp[grs]) >= LARGE_GRS
    ann_M    = round(float(ann_decomp[grs]) / 1e6, 2)

    tr = df_grs[df_grs['month'].isin(TRAIN_MONTHS)]
    te = df_grs[df_grs['month'].isin(TEST_MONTHS)]

    if len(tr) < 30 or len(te) < 10:
        continue

    X_tr = tr[FEAT_COLS].values.astype(np.float64)
    y_tr = tr['daily_target'].values.astype(np.float64)
    X_te = te[FEAT_COLS].values.astype(np.float64)
    y_te = te['daily_target'].values.astype(np.float64)

    sc = StandardScaler()
    X_tr_s = sc.fit_transform(X_tr)
    X_te_s = sc.transform(X_te)

    # CV на alpha
    best_alpha, best_cv = 1.0, -np.inf
    for alpha in ALPHAS:
        try:
            s = cross_val_score(Ridge(alpha=alpha), X_tr_s, y_tr, cv=5, scoring='r2')
            if s.mean() > best_cv:
                best_cv, best_alpha = s.mean(), alpha
        except Exception:
            pass

    mdl = Ridge(alpha=best_alpha)
    mdl.fit(X_tr_s, y_tr)
    y_tr_pred = mdl.predict(X_tr_s)
    y_te_pred = mdl.predict(X_te_s)

    def r2(yt, yp):
        ss = np.sum((yt - yt.mean()) ** 2)
        return float(1 - np.sum((yt - yp) ** 2) / ss) if ss > 0 else 0.0

    r2_tr = r2(y_tr, y_tr_pred)
    r2_te = r2(y_te, y_te_pred)
    mae_te = float(np.abs(y_te - y_te_pred).mean())

    # Monthly bias на test
    te_df  = te.copy()
    te_df['pred_day'] = y_te_pred
    mb = {}
    for mo in TEST_MONTHS:
        s = te_df[te_df['month'] == mo]
        if len(s) == 0: continue
        b = (s['pred_day'].sum() - s['daily_target'].sum()) / s['daily_target'].sum() * 100
        mb[mo] = round(b, 1)

    flag  = '[L]' if is_large else '   '
    mb_str = '  '.join(f"{mb.get(mo, 0):>+7.1f}%" for mo in TEST_MONTHS)
    print(f"  {flag}{grs[:40]:40s}  {ann_M:>5.2f}  {r2_tr:>6.3f}  {r2_te:>6.3f}  "
          f"{mae_te:>9,.0f}  {mb_str}  {best_alpha}")

    # Зберігаємо
    for i, row in tr.iterrows():
        all_preds.append({'grs': grs, 'date': row['date'], 'month': row['month'],
                          'target_day': float(y_tr[i - tr.index[0]]),
                          'pred_day':   float(y_tr_pred[i - tr.index[0]]),
                          'split': 'train'})
    for i, row in te.iterrows():
        all_preds.append({'grs': grs, 'date': row['date'], 'month': row['month'],
                          'target_day': float(y_te[i - te.index[0]]),
                          'pred_day':   float(y_te_pred[i - te.index[0]]),
                          'split': 'test'})

    grs_metrics.append({
        'grs':        grs,
        'is_large':   is_large,
        'ann_M':      ann_M,
        'r2_train':   round(r2_tr, 3),
        'r2_test':    round(r2_te, 3),
        'mae_test_d': round(mae_te),
        'alpha':      best_alpha,
        **{f'mb_{mo}': mb.get(mo) for mo in TEST_MONTHS},
    })

df_preds   = pd.DataFrame(all_preds)
df_metrics = pd.DataFrame(grs_metrics).sort_values('ann_M', ascending=False)

# ── 5. Підсумки по місяцях ────────────────────────────────────────────────────
print("\n=== 5. Підсумки по місяцях ===")
# При train=test=12 кожен запис є і в train, і в test — беремо лише 'test'
df_te = df_preds[df_preds['split'] == 'test'].copy()
print(f"  {'Міс':5s}  {'Bias%':>9s}  {'MAE_d':>9s}  {'R²':>7s}  {'Декомп_м³':>14s}  {'Прогноз_м³':>14s}")
print("  " + "─" * 75)
for mo in ALL_MONTHS:
    s = df_te[df_te['month'] == mo]
    if s.empty: continue
    tt = s['target_day'].sum()
    if tt == 0: continue
    bias = (s['pred_day'].sum() - tt) / tt * 100
    mae  = (s['pred_day'] - s['target_day']).abs().mean()
    ss_t = ((s['target_day'] - s['target_day'].mean()) ** 2).sum()
    ss_r = ((s['target_day'] - s['pred_day']) ** 2).sum()
    r2v  = float(1 - ss_r / ss_t) if ss_t > 0 else 0.0
    # target_day вже в м³ (не м³/день) — сума по місяцю = місячний декомпозит
    dec_m  = s['target_day'].sum()
    pred_m = s['pred_day'].sum()
    print(f"  {UA[mo]:5s}  {bias:>+9.2f}%  {mae:>9,.0f}  {r2v:>7.3f}  {dec_m:>14,.0f}  {pred_m:>14,.0f}")

mae_all  = (df_te['pred_day'] - df_te['target_day']).abs().mean()
bias_all = (df_te['pred_day'].sum() - df_te['target_day'].sum()) / df_te['target_day'].sum() * 100
print(f"\n  Річний (in-sample): MAE={mae_all:,.0f} м³/день  Bias={bias_all:+.2f}%")

mb_cols = [f'mb_{mo}' for mo in TEST_MONTHS]
ok3  = df_metrics[df_metrics['is_large']][mb_cols].abs().max(axis=1).lt(3).sum()
ok5  = df_metrics[mb_cols].abs().max(axis=1).lt(5).sum()
ok10 = df_metrics[mb_cols].abs().max(axis=1).lt(10).sum()
print(f"  Великих ГРС bias <3%: {ok3}/{df_metrics['is_large'].sum()}")
print(f"  Усіх ГРС: <5%={ok5}  <10%={ok10}  з {len(df_metrics)}")

# ── 6. Excel ──────────────────────────────────────────────────────────────────
print(f"\n=== 6. Збереження → {OUT_XLSX} ===")

thin  = Side(style='thin')
thick = Side(style='medium')
HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_FONT  = Font(bold=True, color='FFFFFF', size=9)
GRS_L_FILL = PatternFill('solid', fgColor='D6E4F0')   # великі ГРС
GRS_S_FILL = PatternFill('solid', fgColor='F5F5F5')   # малі ГРС

def bias_fill(v, thr=(3, 7, 15)):
    if v is None: return None
    try: v = float(v)
    except: return None
    av = abs(v)
    t1, t2, t3 = thr
    if v > 0:
        if av > t3: return PatternFill('solid', fgColor='FF4444')
        if av > t2: return PatternFill('solid', fgColor='FF9999')
        if av > t1: return PatternFill('solid', fgColor='FFCCCC')
    else:
        if av > t3: return PatternFill('solid', fgColor='4466FF')
        if av > t2: return PatternFill('solid', fgColor='99AAFF')
        if av > t1: return PatternFill('solid', fgColor='CCDDFF')
    return None

def r2_fill(v):
    try: v = float(v)
    except: return None
    if v >= 0.99: return PatternFill('solid', fgColor='92D050')
    if v >= 0.95: return PatternFill('solid', fgColor='C6EFCE')
    if v >= 0.85: return PatternFill('solid', fgColor='FFEB9C')
    return PatternFill('solid', fgColor='FFC7CE')

def fmt_hdr(ws, headers, col_widths=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

# При train=test=12 кожен рядок є і train, і test — беремо лише 'test'
df_te = df_preds[df_preds['split'] == 'test'].copy()
# target_day вже в м³ (Σ_{d∈month} = monthly_decompose), НЕ множимо на дні
df_te['target_m'] = df_te['target_day']
df_te['pred_m']   = df_te['pred_day']

# Агрегати: місяць (всі ГРС разом)
monthly_all = (df_te.groupby('month')
               .agg(target_m=('target_m','sum'), pred_m=('pred_m','sum'))
               .reset_index())
monthly_all['bias_pct'] = ((monthly_all['pred_m'] - monthly_all['target_m'])
                           / monthly_all['target_m'] * 100)

# Агрегати: ГРС × місяць
monthly_grs = (df_te.groupby(['grs','month'])
               .agg(target_m=('target_m','sum'), pred_m=('pred_m','sum'))
               .reset_index())
monthly_grs['bias_pct'] = ((monthly_grs['pred_m'] - monthly_grs['target_m'])
                           / monthly_grs['target_m'] * 100)

grs_order = df_metrics.sort_values('ann_M', ascending=False)['grs'].tolist()

wb = Workbook()

# ════ Лист 1: Місяці_всього ════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Місяці_всього'
MO_COLS = [UA[m] for m in ALL_MONTHS]
hdr1 = ['Показник'] + MO_COLS + ['Річний']
fmt_hdr(ws1, hdr1, [22] + [10]*12 + [12])
ws1.freeze_panes = 'B2'

def write_row(ws, ri, label, vals_by_mo, annual, is_bias, bold=False):
    c = ws.cell(ri, 1, label)
    c.font = Font(bold=bold or is_bias, italic=not is_bias and not bold)
    for ci, mo in enumerate(ALL_MONTHS, 2):
        v = vals_by_mo.get(mo)
        if v is None:
            ws.cell(ri, ci, '—').alignment = Alignment(horizontal='center')
            continue
        cell = ws.cell(ri, ci, round(v, 1) if is_bias else round(v))
        cell.number_format = '+0.0;-0.0;0.0' if is_bias else '#,##0'
        cell.alignment = Alignment(horizontal='right')
        if is_bias:
            cell.font = Font(bold=True)
            f = bias_fill(v); f and setattr(cell, 'fill', f)
    c_ann = ws.cell(ri, len(hdr1), round(annual, 1) if is_bias else round(annual))
    c_ann.number_format = '+0.0;-0.0;0.0' if is_bias else '#,##0'
    c_ann.alignment = Alignment(horizontal='right')
    if is_bias:
        c_ann.font = Font(bold=True)
        f = bias_fill(annual); f and setattr(c_ann, 'fill', f)
    for ci in range(1, len(hdr1)+1):
        ws.cell(ri, ci).border = Border(bottom=thin)

dec_by_mo   = monthly_all.set_index('month')['target_m'].to_dict()
pred_by_mo  = monthly_all.set_index('month')['pred_m'].to_dict()
bias_by_mo  = monthly_all.set_index('month')['bias_pct'].to_dict()
ann_dec  = sum(dec_by_mo.values())
ann_pred = sum(pred_by_mo.values())
ann_bias = (ann_pred - ann_dec) / ann_dec * 100 if ann_dec else 0

write_row(ws1, 2, 'Декомпозит (таргет), м³', dec_by_mo,  ann_dec,  False, bold=True)
write_row(ws1, 3, 'Прогноз Ridge, м³',        pred_by_mo, ann_pred, False)
write_row(ws1, 4, 'Bias %',                   bias_by_mo, ann_bias, True)
for ci in range(1, len(hdr1)+1):
    ws1.cell(4, ci).border = Border(bottom=thick)

# ════ Лист 2: Декомпозит_по_ГРС ══════════════════════════════════════════════
ws2 = wb.create_sheet('Декомпозит_по_ГРС')
hdr2 = ['ГРС', 'Річний'] + MO_COLS
fmt_hdr(ws2, hdr2, [44, 12] + [10]*12)
ws2.freeze_panes = 'B2'

for ri, grs in enumerate(grs_order, 2):
    sub = monthly_grs[monthly_grs['grs'] == grs].set_index('month')['target_m']
    ann = sub.sum()
    is_l = ann >= LARGE_GRS
    rfill = GRS_L_FILL if is_l else GRS_S_FILL
    ws2.cell(ri, 1, grs).font = Font(bold=is_l); ws2.cell(ri, 1).fill = rfill
    c = ws2.cell(ri, 2, round(ann)); c.number_format = '#,##0'; c.fill = rfill
    c.alignment = Alignment(horizontal='right'); c.font = Font(bold=is_l)
    for ci, mo in enumerate(ALL_MONTHS, 3):
        v = sub.get(mo, 0)
        c = ws2.cell(ri, ci, round(v)); c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right'); c.fill = rfill

# Підсумок
ri_tot = len(grs_order) + 2
ws2.cell(ri_tot, 1, 'ВСЬОГО').font = Font(bold=True)
ws2.cell(ri_tot, 2, round(monthly_all['target_m'].sum())).number_format = '#,##0'
ws2.cell(ri_tot, 2).font = Font(bold=True)
for ci, mo in enumerate(ALL_MONTHS, 3):
    v = monthly_all[monthly_all['month']==mo]['target_m'].sum()
    c = ws2.cell(ri_tot, ci, round(v)); c.number_format='#,##0'; c.font=Font(bold=True)
    c.alignment = Alignment(horizontal='right')

# ════ Лист 3: Прогноз_по_ГРС ════════════════════════════════════════════════
ws3 = wb.create_sheet('Прогноз_по_ГРС')
hdr3 = ['ГРС', 'Річний'] + MO_COLS
fmt_hdr(ws3, hdr3, [44, 12] + [10]*12)
ws3.freeze_panes = 'B2'

for ri, grs in enumerate(grs_order, 2):
    sub = monthly_grs[monthly_grs['grs'] == grs].set_index('month')['pred_m']
    ann = sub.sum()
    is_l = ann >= LARGE_GRS
    rfill = GRS_L_FILL if is_l else GRS_S_FILL
    ws3.cell(ri, 1, grs).font = Font(bold=is_l); ws3.cell(ri, 1).fill = rfill
    c = ws3.cell(ri, 2, round(ann)); c.number_format = '#,##0'; c.fill = rfill
    c.alignment = Alignment(horizontal='right'); c.font = Font(bold=is_l)
    for ci, mo in enumerate(ALL_MONTHS, 3):
        v = sub.get(mo, 0)
        c = ws3.cell(ri, ci, round(v)); c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right'); c.fill = rfill

ri_tot3 = len(grs_order) + 2
ws3.cell(ri_tot3, 1, 'ВСЬОГО').font = Font(bold=True)
ws3.cell(ri_tot3, 2, round(monthly_all['pred_m'].sum())).number_format = '#,##0'
ws3.cell(ri_tot3, 2).font = Font(bold=True)
for ci, mo in enumerate(ALL_MONTHS, 3):
    v = monthly_all[monthly_all['month']==mo]['pred_m'].sum()
    c = ws3.cell(ri_tot3, ci, round(v)); c.number_format='#,##0'; c.font=Font(bold=True)
    c.alignment = Alignment(horizontal='right')

# ════ Лист 4: Bias_по_ГРС (теплокарта) ══════════════════════════════════════
ws4 = wb.create_sheet('Bias_по_ГРС')
hdr4 = ['ГРС', 'Річний%'] + MO_COLS
fmt_hdr(ws4, hdr4, [44, 10] + [9]*12)
ws4.freeze_panes = 'B2'

for ri, grs in enumerate(grs_order, 2):
    sub_b  = monthly_grs[monthly_grs['grs']==grs].set_index('month')
    sub_dc = sub_b['target_m']
    sub_pr = sub_b['pred_m']
    ann_dc = sub_dc.sum(); ann_pr = sub_pr.sum()
    ann_b  = (ann_pr - ann_dc) / ann_dc * 100 if ann_dc else 0
    is_l   = ann_dc >= LARGE_GRS

    ws4.cell(ri, 1, grs).font = Font(bold=is_l)
    c = ws4.cell(ri, 2, round(ann_b, 1))
    c.number_format = '+0.0;-0.0;0.0'; c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')
    f = bias_fill(ann_b, (3,7,15)); f and setattr(c, 'fill', f)

    for ci, mo in enumerate(ALL_MONTHS, 3):
        v = sub_b['bias_pct'].get(mo) if mo in sub_b.index else None
        if v is None:
            ws4.cell(ri, ci, '—').alignment = Alignment(horizontal='center'); continue
        v = float(v)
        c = ws4.cell(ri, ci, round(v, 1))
        c.number_format = '+0.0;-0.0;0.0'; c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
        thr = (3,7,15) if is_l else (5,10,20)
        f = bias_fill(v, thr); f and setattr(c, 'fill', f)
    ws4.cell(ri, ri).border   # placeholder
    for ci in range(1, len(hdr4)+1):
        ws4.cell(ri, ci).border = Border(bottom=thin)

# Підсумок всього
ri_tot4 = len(grs_order) + 2
ws4.cell(ri_tot4, 1, 'ВСЬОГО').font = Font(bold=True)
all_dc = monthly_all['target_m'].sum(); all_pr = monthly_all['pred_m'].sum()
ann_b_all = (all_pr - all_dc) / all_dc * 100 if all_dc else 0
c = ws4.cell(ri_tot4, 2, round(ann_b_all, 1))
c.number_format='+0.0;-0.0;0.0'; c.font=Font(bold=True)
c.alignment=Alignment(horizontal='center')
f=bias_fill(ann_b_all,(3,7,15)); f and setattr(c,'fill',f)
for ci, mo in enumerate(ALL_MONTHS, 3):
    r = monthly_all[monthly_all['month']==mo]
    if r.empty or r['target_m'].sum()==0: continue
    b = float(r['bias_pct'].iloc[0])
    c = ws4.cell(ri_tot4, ci, round(b,1))
    c.number_format='+0.0;-0.0;0.0'; c.font=Font(bold=True)
    c.alignment=Alignment(horizontal='center')
    f=bias_fill(b,(3,7,15)); f and setattr(c,'fill',f)

# ════ Лист 5: Метрики_по_ГРС ═════════════════════════════════════════════════
ws5 = wb.create_sheet('Метрики_по_ГРС')
hdr5 = ['ГРС', 'Млн.м³', 'R²_трен', 'R²_тест', 'MAE_д', 'α'] + \
       [f'{UA[m]}%' for m in ALL_MONTHS]
fmt_hdr(ws5, hdr5, [44, 9, 9, 9, 9, 7] + [8]*12)
ws5.freeze_panes = 'B2'

for ri, row in enumerate(df_metrics.itertuples(), 2):
    is_l  = row.is_large
    rfill = GRS_L_FILL if is_l else GRS_S_FILL
    ws5.cell(ri, 1, row.grs).font = Font(bold=is_l); ws5.cell(ri, 1).fill = rfill
    ws5.cell(ri, 2, row.ann_M).number_format = '0.00'
    ws5.cell(ri, 2).alignment = Alignment(horizontal='right')
    for ci, attr in [(3,'r2_train'),(4,'r2_test')]:
        v = getattr(row, attr)
        c = ws5.cell(ri, ci, v); c.number_format='0.000'
        c.alignment=Alignment(horizontal='center')
        f=r2_fill(v); f and setattr(c,'fill',f)
    ws5.cell(ri, 5, row.mae_test_d).number_format='#,##0'
    ws5.cell(ri, 5).alignment=Alignment(horizontal='right')
    ws5.cell(ri, 6, row.alpha).alignment=Alignment(horizontal='center')
    for ci, mo in enumerate(ALL_MONTHS, 7):
        v = getattr(row, f'mb_{mo}', None)
        if v is None: continue
        c = ws5.cell(ri, ci, round(v,1))
        c.number_format='+0.0;-0.0;0.0'; c.font=Font(bold=True)
        c.alignment=Alignment(horizontal='center')
        thr=(3,7,15) if is_l else (5,10,20)
        f=bias_fill(v,thr); f and setattr(c,'fill',f)
    for ci in range(1, len(hdr5)+1):
        ws5.cell(ri, ci).border=Border(bottom=thin)

# ════ Лист 6: Зведена (Таргет | Прогноз | Bias% по кожному місяцю) ══════════
ws6 = wb.create_sheet('Зведена')
ws6.freeze_panes = 'B3'

# Заголовок рядок 1: назви місяців (об'єднані по 3 стовпці)
ws6.cell(1, 1, 'ГРС').font = HDR_FONT
ws6.cell(1, 1).fill = HDR_FILL
ws6.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
ws6.column_dimensions['A'].width = 44
ws6.row_dimensions[1].height = 22
ws6.row_dimensions[2].height = 22

col_start = 2
for mo in ALL_MONTHS:
    # Merge 3 cells for month name in row 1
    end_col = col_start + 2
    ws6.merge_cells(start_row=1, start_column=col_start,
                    end_row=1, end_column=end_col)
    c = ws6.cell(1, col_start, UA[mo])
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    # Sub-headers in row 2
    for ci, lbl in enumerate(['Таргет', 'Прогноз', 'Bias%'], col_start):
        c2 = ws6.cell(2, ci, lbl)
        c2.font = HDR_FONT
        c2.fill = PatternFill('solid', fgColor='2E75B6')
        c2.alignment = Alignment(horizontal='center')
        ws6.column_dimensions[get_column_letter(ci)].width = 11
    col_start += 3

# Рядки ГРС
for ri, grs in enumerate(grs_order, 3):
    sub   = monthly_grs[monthly_grs['grs'] == grs].set_index('month')
    ann_t = sub['target_m'].sum()
    is_l  = ann_t >= LARGE_GRS
    rfill = GRS_L_FILL if is_l else GRS_S_FILL

    c = ws6.cell(ri, 1, grs)
    c.font = Font(bold=is_l); c.fill = rfill

    col_start = 2
    for mo in ALL_MONTHS:
        if mo in sub.index:
            tgt = float(sub.loc[mo, 'target_m'])
            prd = float(sub.loc[mo, 'pred_m'])
            bia = float(sub.loc[mo, 'bias_pct'])
        else:
            tgt = prd = bia = 0.0

        c1 = ws6.cell(ri, col_start,   round(tgt))
        c2 = ws6.cell(ri, col_start+1, round(prd))
        c3 = ws6.cell(ri, col_start+2, round(bia, 1))
        for c_ in (c1, c2):
            c_.number_format = '#,##0'
            c_.alignment = Alignment(horizontal='right')
            c_.fill = rfill
        c3.number_format = '+0.0;-0.0;0.0'
        c3.alignment = Alignment(horizontal='center')
        c3.font = Font(bold=True)
        thr = (3,7,15) if is_l else (5,10,20)
        f = bias_fill(bia, thr); f and setattr(c3, 'fill', f)
        col_start += 3

    for ci in range(1, 1 + len(ALL_MONTHS)*3 + 1):
        ws6.cell(ri, ci).border = Border(bottom=thin)

# Підсумок ВСЬОГО
ri_tot6 = len(grs_order) + 3
ws6.cell(ri_tot6, 1, 'ВСЬОГО').font = Font(bold=True)
col_start = 2
for mo in ALL_MONTHS:
    r = monthly_all[monthly_all['month'] == mo]
    if r.empty:
        col_start += 3; continue
    tgt = float(r['target_m'].sum())
    prd = float(r['pred_m'].sum())
    bia = float(r['bias_pct'].iloc[0]) if tgt > 0 else 0.0

    c1 = ws6.cell(ri_tot6, col_start,   round(tgt))
    c2 = ws6.cell(ri_tot6, col_start+1, round(prd))
    c3 = ws6.cell(ri_tot6, col_start+2, round(bia, 1))
    for c_ in (c1, c2):
        c_.number_format = '#,##0'; c_.font = Font(bold=True)
        c_.alignment = Alignment(horizontal='right')
    c3.number_format = '+0.0;-0.0;0.0'; c3.font = Font(bold=True)
    c3.alignment = Alignment(horizontal='center')
    f = bias_fill(bia, (3,7,15)); f and setattr(c3, 'fill', f)
    col_start += 3

ws6.row_dimensions[ri_tot6].height = 18

wb.save(OUT_XLSX)
print(f"  DONE → {OUT_XLSX}")
print(f"  Листи: Місяці_всього | Декомпозит_по_ГРС | Прогноз_по_ГРС | Bias_по_ГРС | Метрики_по_ГРС | Зведена")
