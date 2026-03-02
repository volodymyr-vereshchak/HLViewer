"""
find_theft_heating.py

Пошук підозрілих абонентів з ОП (потенційне викрадення газу).
Аналіз окремо по ГРС.

Індикатори підозрілості:
  A) Питоме споживання < X% медіани по ГРС + consumer_type
  B) Нульовий білінг у пікові зимові місяці (Гру/Січ/Лют)
  C) Частка опалювальних місяців у річному обсязі < норми
  D) Аномально низьке абсолютне питоме (< 5 м³/м²/рік)

Оцінка (0–10):  score = A_pts + B_pts + C_pts + D_pts
Рівень:  Критичний ≥7 | Високий ≥4 | Середній ≥2 | Низький ≥1
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from pobut_predictor import GROUP_MERGE

SUBS_FILE  = "data/input/all_pobut_enriched.csv"
MODEM_FILE = "data/input/profile_pobut_daily_result.csv"
OUT_XLSX   = "data/theft_suspects_heating.xlsx"

MONTH_COLS = ['jan_2025','feb_2025','mar_2025','apr_2025','may_2025','jun_2025',
              'jul_2025','aug_2025','sep_2025','oct_2025','nov_2025','dec_2025']
MONTH_UA   = ['Січ','Лют','Бер','Кві','Тра','Чер','Лип','Сер','Вер','Жов','Лис','Гру']
WINTER_IDX = [11, 0, 1]   # Гру=11, Січ=0, Лют=1  (індекси в MONTH_COLS)
HEAT_IDX   = [9, 10, 11, 0, 1, 2]   # Жов–Бер

# Пороги оцінки
RATIO_PTS  = [(0.20, 4), (0.30, 3), (0.40, 2), (0.50, 1)]  # ratio < threshold → pts
ABS_LOW    = 5.0   # м³/м²/рік — критично низьке абсолютне споживання
HEAT_SHARE_LOW = 0.45  # частка опал. місяців < цього → підозріло

UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}

HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
CRIT_FILL  = PatternFill('solid', fgColor='FF4C4C')
HIGH_FILL  = PatternFill('solid', fgColor='FFC7CE')
MED_FILL   = PatternFill('solid', fgColor='FFEB9C')
LOW_FILL   = PatternFill('solid', fgColor='FFFFCC')
MOD_FILL   = PatternFill('solid', fgColor='D9E1F2')


# ── 1. Завантаження ────────────────────────────────────────────────────────────
print("=== 1. Завантаження абонентів ===")
ap = pd.read_csv(SUBS_FILE, low_memory=False)
ap['appliance_group'] = ap['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
ap['has_OP']     = ap['appliance_group'].str.contains('ОП', na=False)
ap['has_VPG']    = ap['appliance_group'].str.contains('ВПГ', na=False)
ap['heated_area'] = pd.to_numeric(ap['heated_area'], errors='coerce').fillna(55.0)
ap['heated_area'] = ap['heated_area'].where(ap['heated_area'] > 0, 55.0)

for c in MONTH_COLS:
    ap[c] = pd.to_numeric(ap[c], errors='coerce').fillna(0)

ap['annual'] = ap[MONTH_COLS].sum(axis=1)

# Тільки ОП-абоненти, активні, з ненульовим білінгом
op = ap[
    ap['has_OP'] &
    ap['gas_off'].isna() &
    ap['alternative'].isna() &
    ap['dacha'].isna() &
    ap['grs'].notna() &
    (ap['annual'] > 0)
].copy().reset_index(drop=True)

print(f"  ОП-абонентів з білінгом: {len(op):,}")
print(f"  ГРС: {op['grs'].nunique()}")

# Modem IDs — для позначки (мають реал-тайм дані)
md_raw = pd.read_csv(MODEM_FILE, sep=';', low_memory=False, usecols=[0])
md_raw.columns = ['account_id']
md_raw['account_id'] = pd.to_numeric(md_raw['account_id'], errors='coerce')
modem_ids = set(md_raw['account_id'].dropna().astype(int))

op['is_modem'] = op['account_id'].isin(modem_ids)
print(f"  З них модемних: {op['is_modem'].sum():,}")


# ── 2. Питомі метрики ─────────────────────────────────────────────────────────
print("\n=== 2. Розрахунок метрик ===")

# Матриця місячних значень
m_vals = op[MONTH_COLS].values.astype(np.float64)  # (N, 12)

# Питоме річне споживання
op['specific'] = op['annual'] / op['heated_area']

# Зимові нулі (Гру/Січ/Лют)
winter_vals = m_vals[:, WINTER_IDX]   # (N, 3)
op['winter_zero_months'] = (winter_vals == 0).sum(axis=1).astype(int)

# Частка опалювальних місяців у річному об'ємі
heat_vals = m_vals[:, HEAT_IDX]       # (N, 6)
op['heat_share'] = heat_vals.sum(axis=1) / np.maximum(op['annual'].values, 1)

# Максимальна частка одного місяця
op['max_month_share'] = m_vals.max(axis=1) / np.maximum(op['annual'].values, 1)
op['max_month_idx']   = m_vals.argmax(axis=1).astype(int)

# Кількість ненульових місяців
op['nonzero_months'] = (m_vals > 0).sum(axis=1).astype(int)


# ── 3. Еталонні медіани по ГРС × consumer_type ────────────────────────────────
print("\n=== 3. Еталонні медіани ===")
ref = (op.groupby(['grs', 'consumer_type'])['specific']
         .median().rename('grs_ct_median').reset_index())
op = op.merge(ref, on=['grs', 'consumer_type'], how='left')

# Якщо мало абонентів у групі — fallback на GRS-медіану
grs_ref = op.groupby('grs')['specific'].median().rename('grs_median').reset_index()
op = op.merge(grs_ref, on='grs', how='left')
op['ref_median'] = op['grs_ct_median'].fillna(op['grs_median'])

op['ratio_to_ref'] = op['specific'] / np.maximum(op['ref_median'], 1)

print("  GRS × consumer_type медіана питомого (м³/м²):")
for _, row in ref.iterrows():
    ct_short = 'Прив' if 'Прив' in row['consumer_type'] else 'БКС'
    print(f"    {row['grs'][:40]:40} [{ct_short}] → {row['grs_ct_median']:.1f}")


# ── 4. Оцінка підозрілості ────────────────────────────────────────────────────
print("\n=== 4. Розрахунок оцінки підозрілості ===")

def compute_score(row):
    score = 0

    # A) ratio до медіани
    r = row['ratio_to_ref']
    for thr, pts in RATIO_PTS:
        if r < thr:
            score += pts
            break

    # B) зимові нулі
    wz = int(row['winter_zero_months'])
    if wz == 3:   score += 3
    elif wz == 2: score += 2
    elif wz == 1: score += 1

    # C) низька частка опалювальних місяців
    hs = row['heat_share']
    if hs < HEAT_SHARE_LOW:
        score += 2
    elif hs < 0.60:
        score += 1

    # D) абсолютно низьке питоме
    if row['specific'] < ABS_LOW:
        score += 2

    return score

op['score'] = op.apply(compute_score, axis=1)

def score_level(s):
    if s >= 7: return 'Критичний'
    if s >= 4: return 'Високий'
    if s >= 2: return 'Середній'
    if s >= 1: return 'Низький'
    return 'Норма'

op['level'] = op['score'].apply(score_level)

suspects = op[op['score'] >= 1].copy()

print(f"\n  Підозрілих до виключень (score≥1): {len(suspects):,}")

# ── 4б. Виключення хибних позитивів ──────────────────────────────────────────
# Колонка dacha вже фільтрує офіційно відмічених («не мешкають», «ЗСУ», «дача»...).
# Але є схожі без позначки — визначаємо по паттерну платежів.

WINTER_MC = ['oct_2025','nov_2025','dec_2025','jan_2025','feb_2025','mar_2025']
SUMMER_MC = ['apr_2025','may_2025','jun_2025','jul_2025','aug_2025','sep_2025']
LATE_MC   = ['jun_2025','jul_2025','aug_2025','sep_2025','oct_2025','nov_2025','dec_2025']

winter_bill = suspects[WINTER_MC].clip(lower=0).sum(axis=1)
summer_bill = suspects[SUMMER_MC].clip(lower=0).sum(axis=1)
late_bill   = suspects[LATE_MC].clip(lower=0).sum(axis=1)

# 1. Сезонники / дачники: нуль у всі опалювальні місяці (Жов–Бер), є влітку
#    → живуть/відвідують тільки в тепло, зима нульова — не крадіжка
is_seasonal = (winter_bill == 0) & (summer_bill > 0)

# 2. Відключились / виїхали протягом року:
#    — мали платіж у Січ-Тра (початок року, опалювальний сезон)
#    — нічого з Чер до кінця (повне припинення)
#    — мало місяців загалом (<=5) → типово: евакуйовані, відключені без позначки
early_nonzero = (
    (suspects['jan_2025'].clip(lower=0) > 0) |
    (suspects['feb_2025'].clip(lower=0) > 0) |
    (suspects['mar_2025'].clip(lower=0) > 0) |
    (suspects['apr_2025'].clip(lower=0) > 0) |
    (suspects['may_2025'].clip(lower=0) > 0)
)
is_departed = (
    early_nonzero &
    (late_bill <= 0) &
    (suspects['nonzero_months'] <= 5) &
    ~is_seasonal
)

# 3. Мінімальний / технічний рахунок: 1–2 місяці, < 30 м³
#    → коригувальні платежі, тестове підключення, або рахунок що майже закрито
is_tiny = (suspects['nonzero_months'] <= 2) & (suspects['annual'] < 30)

# Призначення причини (пріоритет: seasonal > departed > tiny)
excl_reason = pd.Series('', index=suspects.index)
excl_reason[is_tiny]     = 'Мінімальний рахунок'
excl_reason[is_departed] = 'Відключились/виїхали'
excl_reason[is_seasonal] = 'Сезонне використання'

suspects['excl_reason'] = excl_reason
real_suspects = suspects[suspects['excl_reason'] == ''].copy()
excluded      = suspects[suspects['excl_reason'] != ''].copy()

print(f"\n=== 4б. Виключення хибних позитивів ===")
for reason, cnt in excluded['excl_reason'].value_counts().items():
    print(f"  {reason:28}: {cnt:,}")
print(f"  Разом виключено:            {len(excluded):,}")
print(f"\n  Справжніх підозрілих (score≥1): {len(real_suspects):,}")
for lvl in ['Критичний','Високий','Середній','Низький']:
    n = (real_suspects['level'] == lvl).sum()
    print(f"    {lvl:12}: {n:5,}")

print(f"\n  По ГРС (справжні підозрілі):")
grs_cnts = real_suspects.groupby('grs')['score'].agg(['count','mean']).sort_values('count', ascending=False)
for grs, row in grs_cnts.iterrows():
    print(f"    {grs[:45]:45}  {int(row['count']):4} абон.  avg_score={row['mean']:.1f}")


# ── 5. Місяць максимального платежу → назва ────────────────────────────────────
real_suspects = real_suspects.copy()
real_suspects['max_month_name'] = real_suspects['max_month_idx'].apply(lambda i: MONTH_UA[i])
excluded = excluded.copy()
excluded['max_month_name'] = excluded['max_month_idx'].apply(lambda i: MONTH_UA[i])


# ── 6. Збереження Excel ────────────────────────────────────────────────────────
print(f"\n=== 5. Збереження {OUT_XLSX} ===")

def level_fill(lvl):
    if lvl == 'Критичний': return CRIT_FILL
    if lvl == 'Високий':   return HIGH_FILL
    if lvl == 'Середній':  return MED_FILL
    if lvl == 'Низький':   return LOW_FILL
    return None

HDR = ['ГРС', 'Тип', 'Прилади', 'account_id', 'Площа м²',
       'Річний м³', 'Питоме м³/м²', 'Еталон м³/м²', 'Ratio',
       'Зим.нулів', 'Опал.частка', 'Ненульових міс.',
       'Макс.місяць', *MONTH_UA,
       'Оцінка', 'Рівень', 'Модем']

wb = Workbook()

# ─ Лист 1: Всі підозрілі (сортування: GRS → score desc) ─────────────────────
ws1 = wb.active
ws1.title = 'Підозрілі_всі'
ws1.column_dimensions['A'].width = 38
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 12
for ci in range(5, len(HDR)+1):
    ws1.column_dimensions[get_column_letter(ci)].width = 11

for ci, h in enumerate(HDR, 1):
    c = ws1.cell(1, ci, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')

out = real_suspects.sort_values(['grs', 'score'], ascending=[True, False]).reset_index(drop=True)

for ri, row in enumerate(out.itertuples(index=False), 2):
    lvl = row.level
    fill = level_fill(lvl)

    ws1.cell(ri,  1, row.grs)
    ws1.cell(ri,  2, 'Прив' if 'Прив' in str(row.consumer_type) else 'БКС')
    ws1.cell(ri,  3, row.appliance_group)
    ws1.cell(ri,  4, int(row.account_id))
    ws1.cell(ri,  5, round(row.heated_area, 1)).number_format = '#,##0.0'
    ws1.cell(ri,  6, round(row.annual, 1)).number_format     = '#,##0.0'
    ws1.cell(ri,  7, round(row.specific, 2)).number_format   = '#,##0.00'
    ws1.cell(ri,  8, round(row.ref_median, 2)).number_format = '#,##0.00'

    ratio_c = ws1.cell(ri, 9, round(row.ratio_to_ref, 3))
    ratio_c.number_format = '0.000'
    if row.ratio_to_ref < 0.30:
        ratio_c.font = Font(bold=True, color='CC0000')
    elif row.ratio_to_ref < 0.50:
        ratio_c.font = Font(color='C65911')

    ws1.cell(ri, 10, int(row.winter_zero_months)).alignment = Alignment(horizontal='center')
    ws1.cell(ri, 11, round(row.heat_share, 2)).number_format = '0.00'
    ws1.cell(ri, 12, int(row.nonzero_months)).alignment = Alignment(horizontal='center')
    ws1.cell(ri, 13, row.max_month_name).alignment = Alignment(horizontal='center')

    for mi, mc in enumerate(MONTH_COLS, 14):
        v = getattr(row, mc)
        ws1.cell(ri, mi, round(v, 1) if v > 0 else '').number_format = '#,##0.0'

    sc = ws1.cell(ri, 26, int(row.score))
    sc.font = Font(bold=True)
    sc.alignment = Alignment(horizontal='center')

    lc = ws1.cell(ri, 27, lvl)
    lc.alignment = Alignment(horizontal='center')

    ws1.cell(ri, 28, '✓' if row.is_modem else '').alignment = Alignment(horizontal='center')

    if fill:
        for ci in range(1, 29):
            ws1.cell(ri, ci).fill = fill

ws1.freeze_panes = 'E2'
ws1.auto_filter.ref = f'A1:{get_column_letter(len(HDR))}1'

print(f"  Лист 'Підозрілі_всі': {len(out):,} рядків")


# ─ Лист 2: Статистика по ГРС ─────────────────────────────────────────────────
ws2 = wb.create_sheet('Статистика_ГРС')
ws2.column_dimensions['A'].width = 40
for ci in range(2, 15):
    ws2.column_dimensions[get_column_letter(ci)].width = 13

hdr2 = ['ГРС', 'Всього ОП', 'Підозрілих', 'Крит.', 'Вис.', 'Сер.', 'Низ.',
        'Медіана pitch.м³/м²', 'Медіана ratio підоз.',
        'Зим.нулів ≥2', 'Без опал.місяців',
        '% підозр від ОП', 'Сер.score підоз.']
for ci, h in enumerate(hdr2, 1):
    c = ws2.cell(1, ci, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')

for ri, grs in enumerate(sorted(op['grs'].unique()), 2):
    grs_op   = op[op['grs'] == grs]
    grs_susp = real_suspects[real_suspects['grs'] == grs]
    n_tot    = len(grs_op)
    n_susp   = len(grs_susp)
    n_crit   = (grs_susp['level'] == 'Критичний').sum()
    n_high   = (grs_susp['level'] == 'Високий').sum()
    n_med    = (grs_susp['level'] == 'Середній').sum()
    n_low    = (grs_susp['level'] == 'Низький').sum()
    med_spec = grs_susp['specific'].median() if n_susp > 0 else np.nan
    med_rat  = grs_susp['ratio_to_ref'].median() if n_susp > 0 else np.nan
    win2     = (grs_susp['winter_zero_months'] >= 2).sum()
    no_heat  = (grs_susp['heat_share'] < 0.30).sum()
    pct      = n_susp / max(n_tot, 1) * 100
    avg_sc   = grs_susp['score'].mean() if n_susp > 0 else 0

    ws2.cell(ri,  1, grs)
    ws2.cell(ri,  2, n_tot).number_format = '#,##0'
    ws2.cell(ri,  3, n_susp)
    ws2.cell(ri,  4, n_crit).alignment = Alignment(horizontal='center')
    ws2.cell(ri,  5, n_high).alignment = Alignment(horizontal='center')
    ws2.cell(ri,  6, n_med).alignment  = Alignment(horizontal='center')
    ws2.cell(ri,  7, n_low).alignment  = Alignment(horizontal='center')
    if not np.isnan(med_spec):
        ws2.cell(ri,  8, round(med_spec, 1)).number_format = '#,##0.0'
    if not np.isnan(med_rat):
        ws2.cell(ri,  9, round(med_rat, 3)).number_format = '0.000'
    ws2.cell(ri, 10, int(win2))
    ws2.cell(ri, 11, int(no_heat))
    ws2.cell(ri, 12, round(pct, 1)).number_format = '#,##0.0'
    ws2.cell(ri, 13, round(avg_sc, 1)).number_format = '#,##0.0'

    if n_crit > 0:
        ws2.cell(ri, 4).fill = CRIT_FILL
    if n_high > 0:
        ws2.cell(ri, 5).fill = HIGH_FILL

ws2.freeze_panes = 'B2'


# ─ Листи 3+: Окремо по ГРС (тільки Критичний+Високий) ───────────────────────
GRS_LEVELS = ['Критичний', 'Високий']
created_sheets = 0
for grs in sorted(real_suspects['grs'].unique()):
    g_df = real_suspects[
        (real_suspects['grs'] == grs) &
        (real_suspects['level'].isin(GRS_LEVELS))
    ].sort_values('score', ascending=False)

    if len(g_df) == 0:
        continue

    # Скорочена назва ГРС для назви листа (до 28 символів)
    sh_name = grs[:28].strip().replace('/', '-').replace('\\', '-').replace('?','').replace('*','').replace('[','').replace(']','').replace(':','')
    ws = wb.create_sheet(sh_name)
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    for ci in range(4, 20):
        ws.column_dimensions[get_column_letter(ci)].width = 10

    hdr_g = ['account_id', 'Площа м²', 'Прилади',
              'Річний м³', 'Питоме', 'Еталон', 'Ratio',
              'Зим.нулів', 'Опал.ч-ка',
              *MONTH_UA,
              'Оцінка', 'Рівень', 'Модем']
    for ci, h in enumerate(hdr_g, 1):
        c = ws.cell(1, ci, h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center')

    for ri2, row in enumerate(g_df.itertuples(index=False), 2):
        fill = level_fill(row.level)
        ws.cell(ri2, 1, int(row.account_id))
        ws.cell(ri2, 2, round(row.heated_area, 0))
        ws.cell(ri2, 3, row.appliance_group)
        ws.cell(ri2, 4, round(row.annual, 1)).number_format = '#,##0.0'
        ws.cell(ri2, 5, round(row.specific, 2)).number_format = '#,##0.00'
        ws.cell(ri2, 6, round(row.ref_median, 2)).number_format = '#,##0.00'
        rc = ws.cell(ri2, 7, round(row.ratio_to_ref, 3))
        rc.number_format = '0.000'
        if row.ratio_to_ref < 0.30: rc.font = Font(bold=True, color='CC0000')
        ws.cell(ri2, 8, int(row.winter_zero_months)).alignment = Alignment(horizontal='center')
        ws.cell(ri2, 9, round(row.heat_share, 2)).number_format = '0.00'
        for mi, mc in enumerate(MONTH_COLS, 10):
            v = getattr(row, mc)
            ws.cell(ri2, mi, round(v, 1) if v > 0 else '')
        sc2 = ws.cell(ri2, 22, int(row.score))
        sc2.font = Font(bold=True); sc2.alignment = Alignment(horizontal='center')
        lc2 = ws.cell(ri2, 23, row.level); lc2.alignment = Alignment(horizontal='center')
        ws.cell(ri2, 24, '✓' if row.is_modem else '').alignment = Alignment(horizontal='center')
        if fill:
            for ci in range(1, 25):
                ws.cell(ri2, ci).fill = fill

    ws.freeze_panes = 'D2'
    created_sheets += 1

print(f"  Лист 'Статистика_ГРС': {op['grs'].nunique()} ГРС")
print(f"  Листи по ГРС (Крит+Вис): {created_sheets} ГРС")

# ─ Лист "Виключені" (хибні позитиви з поясненням) ───────────────────────────
EXC_FILL = PatternFill('solid', fgColor='D9D9D9')
ws_exc = wb.create_sheet('Виключені')
ws_exc.column_dimensions['A'].width = 38
ws_exc.column_dimensions['B'].width = 10
ws_exc.column_dimensions['C'].width = 26
ws_exc.column_dimensions['D'].width = 12
for ci in range(5, 30):
    ws_exc.column_dimensions[get_column_letter(ci)].width = 11

hdr_exc = ['ГРС', 'Тип', 'Причина виключення', 'account_id', 'Площа м²',
           'Річний м³', 'Питоме м³/м²', 'Еталон м³/м²', 'Ratio',
           'Зим.нулів', 'Опал.частка', 'Ненульових міс.',
           'Макс.місяць', *MONTH_UA, 'Оцінка', 'Рівень', 'Модем']
for ci, h in enumerate(hdr_exc, 1):
    c = ws_exc.cell(1, ci, h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')

exc_out = excluded.sort_values(['excl_reason','grs','score'], ascending=[True,True,False]).reset_index(drop=True)
for ri, row in enumerate(exc_out.itertuples(index=False), 2):
    ws_exc.cell(ri,  1, row.grs)
    ws_exc.cell(ri,  2, 'Прив' if 'Прив' in str(row.consumer_type) else 'БКС')
    ws_exc.cell(ri,  3, row.excl_reason).font = Font(italic=True, color='595959')
    ws_exc.cell(ri,  4, int(row.account_id))
    ws_exc.cell(ri,  5, round(row.heated_area, 1))
    ws_exc.cell(ri,  6, round(row.annual, 1)).number_format = '#,##0.0'
    ws_exc.cell(ri,  7, round(row.specific, 2)).number_format = '#,##0.00'
    ws_exc.cell(ri,  8, round(row.ref_median, 2)).number_format = '#,##0.00'
    ws_exc.cell(ri,  9, round(row.ratio_to_ref, 3)).number_format = '0.000'
    ws_exc.cell(ri, 10, int(row.winter_zero_months))
    ws_exc.cell(ri, 11, round(row.heat_share, 2)).number_format = '0.00'
    ws_exc.cell(ri, 12, int(row.nonzero_months))
    ws_exc.cell(ri, 13, row.max_month_name)
    for mi, mc in enumerate(MONTH_COLS, 14):
        v = getattr(row, mc)
        ws_exc.cell(ri, mi, round(v, 1) if v > 0 else '')
    ws_exc.cell(ri, 26, int(row.score))
    ws_exc.cell(ri, 27, row.level)
    ws_exc.cell(ri, 28, '✓' if row.is_modem else '')
    for ci2 in range(1, 29):
        ws_exc.cell(ri, ci2).fill = EXC_FILL

ws_exc.freeze_panes = 'E2'
ws_exc.auto_filter.ref = 'A1:AB1'
print(f"  Лист 'Виключені': {len(exc_out):,} рядків")

wb.save(OUT_XLSX)
print(f"\n  Збережено: {OUT_XLSX}")

# ── Підсумок у консоль ─────────────────────────────────────────────────────────
print("\n=== Підсумок ===")
print(f"  Всього ОП-абонентів:  {len(op):,}")
print(f"  Підозрілих до виключ: {len(suspects):,}  ({len(suspects)/len(op)*100:.1f}%)")
print(f"  Виключено (хиб.пол.): {len(excluded):,}  (з них сезонних:{(excluded['excl_reason']=='Сезонне використання').sum()}, виїхали:{(excluded['excl_reason']=='Відключились/виїхали').sum()}, мінім:{(excluded['excl_reason']=='Мінімальний рахунок').sum()})")
print(f"  Справжніх підозрілих: {len(real_suspects):,}  ({len(real_suspects)/len(op)*100:.1f}%)")
for lvl in ['Критичний','Високий','Середній','Низький']:
    n = (real_suspects['level'] == lvl).sum()
    bar = '█' * min(n // 20, 30)
    print(f"    {lvl:12}: {n:5,}  {bar}")
print()
print(f"  Топ-10 підозрілих по ratio (після виключень):")
top10 = real_suspects.nsmallest(10, 'ratio_to_ref')[
    ['grs','account_id','heated_area','annual','specific','ref_median','ratio_to_ref','level','score']
]
for _, r in top10.iterrows():
    print(f"    {r['grs'][:35]:35} acc={int(r['account_id'])}  "
          f"spec={r['specific']:.1f}  ref={r['ref_median']:.1f}  "
          f"ratio={r['ratio_to_ref']:.3f}  [{r['level']}]")

print("\n=== Готово ===")
