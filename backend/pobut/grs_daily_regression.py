"""
grs_daily_regression.py

Лінійна регресія денного споживання по ГРС.

Еталон (Y): clean_daily_grs(d) = Σ_consumer [ annual_i × hybrid_daily_w(group_i, d) ]
             (annual_decompose — очищений від білінгового шуму)

Фічі (X):
  - modem_grp(d): сума добового споживання ВСІХ модемів по групі (не per-GRS)
      → ОП_ПГ, ОП_ПГ_ВПГ, ПГ, ПГ_ВПГ
  - temperature, HDD_18, wind_speed  (погода)
  - month_sin, month_cos, dow_sin, dow_cos  (циклічний календар)
  - is_heating  (0/1)

Модель: Ridge regression per GRS  (або єдина глобальна з GRS-dummy)
Оцінка: leave-one-month-out CV → R², MAE per GRS per month
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from pathlib import Path
from sklearn.linear_model import Ridge, RidgeCV
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import r2_score, mean_absolute_error

from pobut_predictor import GROUP_MERGE, META_COLS, CT_PRIV

# ── константи ─────────────────────────────────────────────────────────────────
MODEM_FILE  = 'data/input/profile_pobut_daily_result.csv'
SUBS_FILE   = 'data/input/all_pobut_enriched.csv'
WEATHER_CSV = '../data/weather_2025_daily_contractual.csv'
OUT_XLSX    = 'data/grs_daily_regression.xlsx'

GROUPS = [
    ('ОП,ПГ',     'Приватний сектор'),
    ('ОП,ПГ',     'Багатоквартирний сектор'),
    ('ОП,ПГ,ВПГ', 'Приватний сектор'),
    ('ОП,ПГ,ВПГ', 'Багатоквартирний сектор'),
    ('ПГ',        'Багатоквартирний сектор'),
    ('ПГ,ВПГ',    'Багатоквартирний сектор'),
]
HEATING_MONTHS = {1, 2, 3, 4, 10, 11, 12}
MIN_ANNUAL     = 10.0
UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}

MONTH_ENG = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
             'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

# ── 1. Модемні дані: агрегація по групі × день ───────────────────────────────
print("=== 1. Модемні денні суми по групах ===")
md = pd.read_csv(MODEM_FILE, sep=';', low_memory=False)
md.columns = META_COLS + list(md.columns[len(META_COLS):])
md['appliance_group'] = md['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
md['consumer_type']   = md['consumer_type'].fillna(CT_PRIV)
md = md[md['gas_off'].isna() & md['alternative'].isna() & md['dacha'].isna()]
modem_ids = set(pd.to_numeric(md['account_id'], errors='coerce').dropna().astype(int))

# Збираємо денні колонки 2025
date_cols = []
for c in md.columns:
    if isinstance(c, str) and c.count('.') == 2 and len(c) == 10:
        try:
            d_s, m_s, y_s = c.split('.')
            dt = pd.Timestamp(f'{y_s}-{m_s}-{d_s}')
            if dt.year == 2025:
                date_cols.append((dt, c))
        except Exception:
            continue
date_cols.sort()
print(f"  Денних колонок 2025: {len(date_cols)}")

# Головні групи для фічей (об'єднуємо МКД+Прив в одну фічу)
FEAT_GROUPS = ['ОП,ПГ', 'ОП,ПГ,ВПГ', 'ПГ', 'ПГ,ВПГ']
feat_col_names = {g: f"modem_{g.replace(',','_').replace(' ','')}" for g in FEAT_GROUPS}

modem_daily = {}   # group -> pd.Series(date -> sum)
for grp in FEAT_GROUPS:
    sub = md[md['appliance_group'] == grp]
    if len(sub) == 0:
        modem_daily[grp] = pd.Series(dtype=float)
        continue
    day_col_list = [col for _, col in date_cols]
    data = (sub[day_col_list]
            .apply(pd.to_numeric, errors='coerce')
            .fillna(0).clip(lower=0)
            .values.astype(np.float64))
    sums = data.sum(axis=0)  # (365,)
    modem_daily[grp] = pd.Series(sums, index=[dt for dt, _ in date_cols])
    print(f"  {grp}: {len(sub)} модемів, річна сума={sums.sum():,.0f} м³")

# DataFrame фічей по модемах
modem_feat_df = pd.DataFrame({
    feat_col_names[g]: modem_daily[g]
    for g in FEAT_GROUPS
}).reset_index().rename(columns={'index': 'date'})
modem_feat_df['date'] = pd.to_datetime(modem_feat_df['date'])

# ── 2. Погода ─────────────────────────────────────────────────────────────────
print("\n=== 2. Погода ===")
wdf = pd.read_csv(WEATHER_CSV)
wdf['date'] = pd.to_datetime(wdf['date']).dt.normalize()
wdf = wdf[['date','temperature','hdh_18c','wind_speed']].rename(
    columns={'temperature':'temp','hdh_18c':'hdd'})
wdf = wdf[wdf['date'].dt.year == 2025].drop_duplicates('date')
print(f"  Погодних днів 2025: {len(wdf)}")

# ── 3. Еталон: clean daily GRS з annual_decompose ────────────────────────────
print("\n=== 3. Еталон: clean daily GRS ===")

# 3a. Білінгові місячні ваги (з "чистих" споживачів)
ap = pd.read_csv(SUBS_FILE, low_memory=False)
ap['account_id']      = pd.to_numeric(ap['account_id'], errors='coerce')
ap['appliance_group'] = ap['appliance_group'].astype(str).str.strip().replace(GROUP_MERGE)
ap['consumer_type']   = ap['consumer_type'].fillna(CT_PRIV)
ap['gas_off']         = ap['gas_off'].notna()

bill_cols = {}
for eng, m in MONTH_ENG.items():
    col = f'{eng}_2025'
    if col in ap.columns:
        ap[col] = pd.to_numeric(ap[col], errors='coerce').fillna(0)
        bill_cols[m] = col

ap['annual_billing'] = ap[[bill_cols[m] for m in range(1,13) if m in bill_cols]].sum(axis=1)

_mc = [bill_cols[m] for m in range(1,13) if m in bill_cols]
ap['max_month_frac'] = ap[_mc].max(axis=1) / ap['annual_billing'].clip(lower=1)
ap['n_months_nz']    = (ap[_mc] > 0).sum(axis=1)

clean_mask = (
    ~ap['account_id'].isin(modem_ids) & ~ap['gas_off'] & ap['grs'].notna()
    & (ap['annual_billing'] > 0)
    & (ap['max_month_frac'] < 0.5) & (ap['n_months_nz'] >= 8)
)
clean_all = ap[clean_mask]
print(f"  'Чистих' для білінгового профілю: {len(clean_all):,}")

billing_monthly_w = {}
for grp, ct in GROUPS:
    sub_c = clean_all[(clean_all['appliance_group']==grp)&(clean_all['consumer_type']==ct)]
    if len(sub_c) < 100: continue
    fracs = {m: float((sub_c[col]/sub_c['annual_billing'].clip(lower=1)).mean())
             for m, col in bill_cols.items()}
    total = sum(fracs.values())
    if total > 0:
        billing_monthly_w[(grp,ct)] = {m: v/total for m, v in fracs.items()}

# 3b. Модемний денний профіль (нормалізований по annual споживача)
day_col_names_all = [col for _, col in date_cols]
profile_raw = {}
for grp, ct in GROUPS:
    sub = md[(md['appliance_group']==grp)&(md['consumer_type']==ct)]
    if len(sub) < 5: continue
    data = (sub[day_col_names_all]
            .apply(pd.to_numeric, errors='coerce')
            .fillna(0).clip(lower=0)
            .values.astype(np.float64))
    annual = data.sum(axis=1)
    valid  = annual >= MIN_ANNUAL
    if valid.sum() < 5: continue
    fracs = data[valid] / annual[valid, np.newaxis]
    daily_med = {}
    for j, (dt, _) in enumerate(date_cols):
        f = fracs[:, j]
        p99 = np.percentile(f, 99)
        f_c = f[f <= p99]
        daily_med[dt] = float(np.median(f_c)) if len(f_c) >= 5 else float(np.median(f))
    total = sum(daily_med.values())
    if total > 0:
        profile_raw[(grp,ct)] = {dt: v/total for dt, v in daily_med.items()}

# 3c. Гібридний денний профіль (billing monthly × modem within-month shape)
month_days = {}
for dt, _ in date_cols:
    month_days.setdefault(dt.month, []).append(dt)

hybrid_daily_w = {}
for grp, ct in GROUPS:
    bill_mw  = billing_monthly_w.get((grp,ct))
    modem_pn = profile_raw.get((grp,ct))
    if bill_mw is None: continue
    daily = {}
    for m, b_w in bill_mw.items():
        days = month_days.get(m, [])
        if not days: continue
        if modem_pn:
            msum = sum(modem_pn.get(dt, 0) for dt in days)
            for dt in days:
                daily[dt] = b_w * (modem_pn.get(dt,0)/msum if msum>0 else 1/len(days))
        else:
            for dt in days:
                daily[dt] = b_w / len(days)
    hybrid_daily_w[(grp,ct)] = daily

# 3d. Денна сума по ГРС: annual_sum(grs, grp, ct) × w(grp, ct, d)
nm = ap[
    ~ap['account_id'].isin(modem_ids) & ~ap['gas_off'] & ap['grs'].notna()
    & (ap['annual_billing'] > 0)
].copy()
modem_grs_df = ap[ap['account_id'].isin(modem_ids) & ap['grs'].notna()].copy()

# Сума річного білінгу по (grs, grp, ct)
grp_annual = nm.groupby(['grs','appliance_group','consumer_type'])['annual_billing'].sum()

# Денна матриця: date × grs
all_dates = [dt for dt, _ in date_cols]
grs_list  = nm['grs'].unique().tolist()
grs_idx   = {g: i for i, g in enumerate(grs_list)}
n_grs     = len(grs_list)
n_days    = len(all_dates)

daily_matrix = np.zeros((n_days, n_grs), dtype=np.float64)

for (grs, grp, ct), ann_sum in grp_annual.items():
    if grs not in grs_idx: continue
    hw = hybrid_daily_w.get((grp,ct))
    if hw is None:
        # fallback: по групі без CT
        for (g,c), h in hybrid_daily_w.items():
            if g == grp: hw = h; break
    if hw is None: continue
    gi = grs_idx[grs]
    for j, dt in enumerate(all_dates):
        daily_matrix[j, gi] += ann_sum * hw.get(dt, 0)

# Додаємо модемний факт по ГРС
md_ann = pd.read_csv(SUBS_FILE, low_memory=False)
md_ann['account_id'] = pd.to_numeric(md_ann['account_id'], errors='coerce')
md_grs_map = md_ann[md_ann['account_id'].isin(modem_ids)][['account_id','grs']].dropna()
md_grs_map = dict(zip(md_grs_map['account_id'].astype(int), md_grs_map['grs']))

md['account_id_int'] = pd.to_numeric(md['account_id'], errors='coerce')
for j, (dt, col) in enumerate(date_cols):
    vals = pd.to_numeric(md[col], errors='coerce').fillna(0).clip(lower=0).values
    for i_m, acc_id in enumerate(md['account_id_int'].values):
        if np.isnan(acc_id): continue
        grs_name = md_grs_map.get(int(acc_id))
        if grs_name and grs_name in grs_idx:
            daily_matrix[j, grs_idx[grs_name]] += vals[i_m]

# Перетворення в DataFrame
clean_daily = pd.DataFrame(daily_matrix, index=all_dates, columns=grs_list)
clean_daily.index.name = 'date'
clean_daily = clean_daily.reset_index()
clean_daily = clean_daily.melt(id_vars='date', var_name='grs', value_name='clean_cons')
print(f"  Денних записів: {len(clean_daily):,}  ({n_grs} ГРС × {n_days} днів)")

# ── 4. Збираємо фічі ─────────────────────────────────────────────────────────
print("\n=== 4. Збираємо фічі ===")
feat = clean_daily.merge(modem_feat_df, on='date', how='left')
feat = feat.merge(wdf, on='date', how='left')
feat['month']      = feat['date'].dt.month
feat['dow']        = feat['date'].dt.dayofweek
feat['is_heating'] = feat['month'].isin(HEATING_MONTHS).astype(int)
feat['month_sin']  = np.sin(2*np.pi*feat['month']/12)
feat['month_cos']  = np.cos(2*np.pi*feat['month']/12)
feat['dow_sin']    = np.sin(2*np.pi*feat['dow']/7)
feat['dow_cos']    = np.cos(2*np.pi*feat['dow']/7)
feat['hdd']        = feat['hdd'].fillna(0)
feat['temp']       = feat['temp'].fillna(feat['temp'].median())
feat['wind_speed'] = feat['wind_speed'].fillna(feat['wind_speed'].median())

MODEM_FEATS = list(feat_col_names.values())
WEATHER_FEATS = ['temp','hdd','wind_speed']
CAL_FEATS     = ['is_heating','month_sin','month_cos','dow_sin','dow_cos']
ALL_FEATS     = MODEM_FEATS + WEATHER_FEATS + CAL_FEATS

feat = feat.dropna(subset=ALL_FEATS + ['clean_cons'])
print(f"  Записів після join: {len(feat):,}")

# ── 5. Регресія по кожному ГРС (leave-one-month-out CV) ──────────────────────
print("\n=== 5. Ridge regression per GRS (LOO-month CV) ===")

results = []
for grs in grs_list:
    sub = feat[feat['grs'] == grs].copy().reset_index(drop=True)
    if len(sub) < 60:
        continue

    X = sub[ALL_FEATS].values
    y = sub['clean_cons'].values
    months = sub['month'].values
    dates  = sub['date'].values

    # LOO-month CV
    y_pred_cv = np.zeros(len(y))
    for test_m in range(1, 13):
        tr = months != test_m
        te = months == test_m
        if tr.sum() < 30 or te.sum() == 0:
            continue
        sc = StandardScaler()
        Xtr = sc.fit_transform(X[tr])
        Xte = sc.transform(X[te])
        mdl = RidgeCV(alphas=[0.1, 1, 10, 100]).fit(Xtr, y[tr])
        y_pred_cv[te] = mdl.predict(Xte)

    r2  = r2_score(y, y_pred_cv)
    mae = mean_absolute_error(y, y_pred_cv)

    # Bias по місяцях
    monthly_bias = {}
    for m in range(1, 13):
        idx = months == m
        if idx.sum() == 0: continue
        b = (y_pred_cv[idx].sum() - y[idx].sum()) / max(y[idx].sum(), 1) * 100
        monthly_bias[m] = round(b, 1)

    results.append({
        'grs': grs, 'n_days': len(sub),
        'r2': round(r2, 3), 'mae': round(mae, 0),
        **{UA[m]: monthly_bias.get(m) for m in range(1,13)}
    })

    print(f"  {grs[:40]:40s}  R²={r2:+.3f}  MAE={mae:7,.0f}")

res_df = pd.DataFrame(results)

# Глобальна статистика
print(f"\n  Медіана R²:  {res_df['r2'].median():.3f}")
print(f"  Медіана MAE: {res_df['mae'].median():,.0f}")
print(f"  GRS з R²>0.95: {(res_df['r2']>0.95).sum()}")
print(f"  GRS з R²>0.90: {(res_df['r2']>0.90).sum()}")

# ── 6. Збереження ─────────────────────────────────────────────────────────────
print(f"\n=== 6. Збереження → {OUT_XLSX} ===")
from openpyxl import Workbook as OXL_WB
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter

def bias_fill(v):
    if v is None: return None
    av = abs(v)
    if v > 0:
        if av > 20: return PatternFill('solid', fgColor='FF2222')
        if av > 10: return PatternFill('solid', fgColor='FF7777')
        if av >  5: return PatternFill('solid', fgColor='FFBBBB')
    else:
        if av > 20: return PatternFill('solid', fgColor='2244FF')
        if av > 10: return PatternFill('solid', fgColor='7799FF')
        if av >  5: return PatternFill('solid', fgColor='BBCCFF')
    return None

wb = OXL_WB()
ws = wb.active; ws.title = 'GRS_results'
thin = Side(style='thin')
HDR = PatternFill('solid', fgColor='1F4E79')

headers = ['ГРС','Днів','R²','MAE'] + [UA[m] for m in range(1,13)]
for ci, h in enumerate(headers, 1):
    c = ws.cell(1, ci, h); c.font = Font(bold=True, color='FFFFFF'); c.fill = HDR
    c.alignment = Alignment(horizontal='center')
ws.column_dimensions['A'].width = 42
for ci in range(2, len(headers)+1): ws.column_dimensions[get_column_letter(ci)].width = 9
ws.freeze_panes = 'B2'

for ri, r in enumerate(res_df.sort_values('r2', ascending=False).to_dict('records'), 2):
    ws.cell(ri, 1, r['grs'])
    ws.cell(ri, 2, r['n_days']).number_format = '#,##0'
    c = ws.cell(ri, 3, r['r2']); c.number_format = '0.000'
    c.font = Font(bold=True, color='006400' if r['r2']>0.9 else 'CC4400' if r['r2']<0.7 else '000000')
    ws.cell(ri, 4, r['mae']).number_format = '#,##0'
    for ci, m in enumerate(range(1,13), 5):
        v = r.get(UA[m])
        if v is None: continue
        c = ws.cell(ri, ci, v); c.number_format = '+0.0;-0.0;0.0'
        c.alignment = Alignment(horizontal='center'); c.font = Font(bold=(abs(v)>5))
        f = bias_fill(v)
        if f: c.fill = f
    for ci in range(1, len(headers)+1):
        ws.cell(ri, ci).border = Border(bottom=thin)

# Глобальна статистика
ri = len(results) + 3
ws.cell(ri, 1, 'Медіана').font = Font(bold=True)
ws.cell(ri, 3, round(res_df['r2'].median(),3)).number_format = '0.000'
ws.cell(ri, 4, round(res_df['mae'].median())).number_format = '#,##0'

wb.save(OUT_XLSX)
print(f"  DONE → {OUT_XLSX}")
