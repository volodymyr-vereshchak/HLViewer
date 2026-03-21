"""
kv2_vs_decompose.py

Порівняння K_v2 прогнозу з "фактом" = annual_decompose (очищені дані).

  Факт    = annual_decompose total_pred (annual_billing × hybrid_profile)
  Прогноз = K_v2 (PobUtPredictorConfigK_v2)

Excel-таблиця: По_місяцях | По_ГРС  (колірне кодування bias)
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter

from pobut_predictor import PobUtPredictor, GROUP_MERGE, META_COLS, CT_PRIV, HEATING_MONTHS, SUMMER_MONTHS

# ── K_v2 клас (P_HEAT з фактичних modemних даних 2025) ───────────────────────
import re as _re

class _PobUtPredictorConfigK(PobUtPredictor):
    KI_HEAT_CSV    = 'data/profiles/ki_heating_consumers.csv'
    KI_NONHEAT_CSV = 'data/profiles/ki_nonheat_consumers.csv'
    P_NONHEAT_CSV  = 'data/profiles/p_modem_nonheat.csv'
    P_HEAT = {1: 4.3167, 2: 4.3717, 3: 2.7025, 4: 1.3619,
              10: 1.1902, 11: 2.7532, 12: 3.9867}

    def _prepare_non_modem(self):
        super()._prepare_non_modem()
        self._load_k_profiles()

    def _load_k_profiles(self):
        p_df = pd.read_csv(self.P_NONHEAT_CSV)
        self._p_nonheat = {
            (r['appliance_group'], r['consumer_type'], int(r['month'])): float(r['P_modem_median'])
            for _, r in p_df.iterrows()
        }
        ki_h = pd.read_csv(self.KI_HEAT_CSV,
                           usecols=['account_id','grs','appliance_group','consumer_type','k_i_clipped'])
        ki_h['account_id'] = pd.to_numeric(ki_h['account_id'], errors='coerce')
        ki_h = ki_h[ki_h['k_i_clipped'].notna() & (ki_h['k_i_clipped'] > 0)]
        self._ki_heat_aid    = dict(zip(ki_h['account_id'].astype(int), ki_h['k_i_clipped'].astype(float)))
        self._ki_heat_grs    = ki_h.groupby(['grs','appliance_group','consumer_type'])['k_i_clipped'].median().to_dict()
        self._ki_heat_global = ki_h.groupby(['appliance_group','consumer_type'])['k_i_clipped'].median().to_dict()
        ki_nh = pd.read_csv(self.KI_NONHEAT_CSV,
                            usecols=['account_id','grs','appliance_group','consumer_type','k_i_clipped'])
        ki_nh['account_id'] = pd.to_numeric(ki_nh['account_id'], errors='coerce')
        ki_nh = ki_nh[ki_nh['k_i_clipped'].notna() & (ki_nh['k_i_clipped'] > 0)]
        self._ki_nh_aid    = dict(zip(ki_nh['account_id'].astype(int), ki_nh['k_i_clipped'].astype(float)))
        self._ki_nh_grs    = ki_nh.groupby(['grs','appliance_group','consumer_type'])['k_i_clipped'].median().to_dict()
        self._ki_nh_global = ki_nh.groupby(['appliance_group','consumer_type'])['k_i_clipped'].median().to_dict()
        self._build_nm_arrays()

    def _build_nm_arrays(self):
        nm   = self._non_modem
        n    = len(nm)
        aids = nm['account_id'].values.astype(int)
        grps = nm['appliance_group'].values
        cts  = nm['consumer_type'].values
        grs  = nm['grs'].values
        ki_h  = np.full(n, 0.85, dtype=np.float32)
        ki_nh = np.full(n, 1.0,  dtype=np.float32)
        for i in range(n):
            aid, grp, ct, g = int(aids[i]), grps[i], cts[i], grs[i]
            v = (self._ki_heat_aid.get(aid)
                 or self._ki_heat_grs.get((g, grp, ct))
                 or self._ki_heat_grs.get((g, grp, CT_PRIV))
                 or self._ki_heat_global.get((grp, ct))
                 or self._ki_heat_global.get((grp, CT_PRIV), 0.85))
            ki_h[i] = float(v or 0.85)
            v = (self._ki_nh_aid.get(aid)
                 or self._ki_nh_grs.get((g, grp, ct))
                 or self._ki_nh_grs.get((g, grp, CT_PRIV))
                 or self._ki_nh_global.get((grp, ct))
                 or self._ki_nh_global.get((grp, CT_PRIV), 1.0))
            ki_nh[i] = float(v or 1.0)
        nm['_ki_heat']    = ki_h
        nm['_ki_nonheat'] = ki_nh
        p_mat = np.zeros((n, 12), dtype=np.float32)
        for m in range(1, 13):
            for i in range(n):
                p_mat[i, m-1] = self._p_nonheat.get((grps[i], cts[i], m), 0.0)
        self._p_nonheat_matrix = p_mat

    def _predict_non_modem(self, periods):
        nm      = self._non_modem
        aids    = nm['account_id'].values
        grps    = nm['appliance_group'].values
        grs_v   = nm['grs'].values
        areas   = nm['heated_area'].values.astype(np.float64)
        has_op  = nm['has_OP'].values
        is_dach = nm['is_dacha'].values
        ki_h    = nm['_ki_heat'].values.astype(np.float64)
        ki_nh   = nm['_ki_nonheat'].values.astype(np.float64)
        p_mat   = self._p_nonheat_matrix.astype(np.float64)
        records = []
        for (y, m) in periods:
            is_heat = m in HEATING_MONTHS
            active  = ~is_dach | np.isin(m, list(SUMMER_MONTHS))
            p_col   = p_mat[:, m - 1]
            vols    = np.zeros(len(nm))
            if is_heat:
                p_h = self.P_HEAT.get(m, 0.0)
                op  = active & has_op
                vols[op] = ki_h[op] * areas[op] * p_h
            else:
                op = active & has_op
                vols[op] = ki_nh[op] * p_col[op]
            non_op = active & ~has_op
            vols[non_op] = ki_nh[non_op] * p_col[non_op]
            vols = np.maximum(vols, 0.0)
            mask = (vols > 0) & active
            if not mask.any(): continue
            records.append(pd.DataFrame({
                'account_id':      aids[mask],
                'appliance_group': grps[mask],
                'year': y, 'month': m,
                'volume':          vols[mask],
                'grs':             grs_v[mask],
                'type':            'pred',
            }))
        return pd.concat(records, ignore_index=True) if records else pd.DataFrame()


class PobUtPredictorConfigK_v2(_PobUtPredictorConfigK):
    """K_v2: P_HEAT з фактичних modemних даних 2025."""
    P_HEAT = {1: 3.8374, 2: 4.3598, 3: 2.7025, 4: 1.3619,
              10: 1.4048, 11: 2.3105, 12: 3.7630}

MODEM_FILE = 'data/input/profile_pobut_daily_result.csv'
SUBS_FILE  = 'data/input/all_pobut_enriched.csv'
OUT_XLSX   = 'data/kv2_vs_decompose.xlsx'

UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}
MONTHS = list(range(1, 13))
LARGE  = 5_000_000

thin  = Side(style='thin')
thick = Side(style='medium')
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF')

def bias_fill(v):
    if v is None: return None
    av = abs(v)
    if v > 0:
        if av > 20: return PatternFill('solid', fgColor='FF2222')
        if av > 10: return PatternFill('solid', fgColor='FF7777')
        if av >  5: return PatternFill('solid', fgColor='FFBBBB')
    else:
        if av > 20: return PatternFill('solid', fgColor='2244FF')
        if av > 10: return PatternFill('solid', fgColor='7799FF')
        if av >  5: return PatternFill('solid', fgColor='BBCCFF')
    return None

# ── 1. Запуск K_v2 ────────────────────────────────────────────────────────────
print("=== 1. K_v2 прогноз ===")
p = PobUtPredictorConfigK_v2(MODEM_FILE, SUBS_FILE, mode='offline', summer_op_vpg=False)
df_pred = p.predict(years=[2025])

kv2 = (df_pred
       .groupby(['grs','year','month','type'])['volume']
       .sum().reset_index()
       .pivot_table(index=['grs','year','month'], columns='type',
                    values='volume', aggfunc='sum', fill_value=0)
       .reset_index())
kv2.columns.name = None
kv2['fact']  = kv2.get('fact',  pd.Series(0, index=kv2.index))
kv2['pred']  = kv2.get('pred',  pd.Series(0, index=kv2.index))
kv2['kv2']   = kv2['fact'] + kv2['pred']   # повний K_v2 прогноз (modем_факт + предикт)
kv2 = kv2[['grs','month','kv2']].copy()
print(f"  ГРС: {kv2['grs'].nunique()}, місяців: {kv2['month'].nunique()}")

# ── 2. Завантаження annual_decompose як "факт" ────────────────────────────────
print("\n=== 2. annual_decompose факт ===")
dc = pd.read_excel('data/annual_decompose_result.xlsx', sheet_name='All_GRS_monthly')
dc = dc[['grs','month','total_pred']].rename(columns={'total_pred': 'decompose'})
print(f"  ГРС: {dc['grs'].nunique()}, місяців: {dc['month'].nunique()}")

# ── 3. Об'єднання ─────────────────────────────────────────────────────────────
m = dc.merge(kv2, on=['grs','month'], how='inner')
m = m[m['decompose'] > 0].copy()
m['error']    = m['kv2'] - m['decompose']
m['bias_pct'] = m['error'] / m['decompose'] * 100

# Річний обсяг по ГРС для сортування
ann_grs = m.groupby('grs')['decompose'].sum().sort_values(ascending=False)
print(f"\n  Записів: {len(m)}")
print(f"\n  MAE    = {m['error'].abs().mean():,.0f}")
print(f"  MAPE   = {(m['error'].abs()/m['decompose']).mean()*100:.2f}%")
print(f"  Bias   = {m['error'].sum()/m['decompose'].sum()*100:+.2f}%")

print("\n  Медіанний bias% по місяцях:")
for mo in MONTHS:
    s = m[m['month']==mo]
    if s.empty: continue
    b = s['error'].sum() / s['decompose'].sum() * 100
    bar = '█' * min(int(abs(b)/2), 25)
    print(f"    {UA[mo]:4s}: {b:+6.1f}%  {bar}")

# ── 4. Excel ──────────────────────────────────────────────────────────────────
print(f"\n=== 4. Збереження → {OUT_XLSX} ===")
wb = Workbook()

# ════════════════ Лист 1: По_місяцях ═════════════════════════════════════════
ws1 = wb.active
ws1.title = 'По_місяцях'

headers1 = ['Показник'] + [UA[mo] for mo in MONTHS] + ['Річний']
for ci, h in enumerate(headers1, 1):
    c = ws1.cell(1, ci, h); c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
ws1.column_dimensions['A'].width = 16
for ci in range(2, 15): ws1.column_dimensions[get_column_letter(ci)].width = 11
ws1.freeze_panes = 'B2'

dec_row  = {mo: m[m['month']==mo]['decompose'].sum() for mo in MONTHS}
kv2_row  = {mo: m[m['month']==mo]['kv2'].sum()       for mo in MONTHS}
bias_row = {mo: (kv2_row[mo]-dec_row[mo])/dec_row[mo]*100 if dec_row[mo] else 0
            for mo in MONTHS}
ann_dec  = sum(dec_row.values())
ann_kv2  = sum(kv2_row.values())
ann_bias = (ann_kv2 - ann_dec) / ann_dec * 100 if ann_dec else 0

rows_def = [
    ('Декомпозит (факт)', dec_row,  '#,##0',         False),
    ('K_v2 (прогноз)',    kv2_row,  '#,##0',         False),
    ('Bias %',            bias_row, '+0.0;-0.0;0.0', True),
]
for ri, (label, data, fmt, is_bias) in enumerate(rows_def, 2):
    ws1.cell(ri, 1, label).font = Font(bold=is_bias, italic=not is_bias)
    ws1.cell(ri, 1).alignment = Alignment(horizontal='left')
    for ci, mo in enumerate(MONTHS, 2):
        c = ws1.cell(ri, ci, round(data[mo], 1 if is_bias else 0))
        c.number_format = fmt; c.alignment = Alignment(horizontal='right')
        if is_bias:
            c.font = Font(bold=True)
            f = bias_fill(data[mo])
            if f: c.fill = f
    ann_v = ann_dec if label.startswith('Декомп') else ann_kv2 if label.startswith('K_v2') else ann_bias
    c = ws1.cell(ri, 14, round(ann_v, 1 if is_bias else 0))
    c.number_format = fmt; c.alignment = Alignment(horizontal='right')
    if is_bias:
        c.font = Font(bold=True)
        f = bias_fill(ann_v)
        if f: c.fill = f
    for ci in range(1, 15):
        ws1.cell(ri, ci).border = Border(bottom=thin)

# ════════════════ Лист 2: По_ГРС ═════════════════════════════════════════════
ws2 = wb.create_sheet('По_ГРС')

ws2.cell(1, 1, 'ГРС').font = HDR_FONT; ws2.cell(1, 1).fill = HDR_FILL
ws2.cell(1, 2, '').fill = HDR_FILL
ws2.column_dimensions['A'].width = 42
ws2.column_dimensions['B'].width = 12
for ci, mo in enumerate(MONTHS, 3):
    c = ws2.cell(1, ci, UA[mo]); c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
    ws2.column_dimensions[get_column_letter(ci)].width = 11
c = ws2.cell(1, 15, 'Річний'); c.font = HDR_FONT; c.fill = HDR_FILL
c.alignment = Alignment(horizontal='center')
ws2.column_dimensions['O'].width = 11
ws2.freeze_panes = 'C2'

row_i = 2
for grs in ann_grs.index:
    sub = m[m['grs'] == grs].set_index('month')
    if sub.empty: continue
    is_large  = ann_grs[grs] >= LARGE
    row_fill  = PatternFill('solid', fgColor='D6E4F0' if is_large else 'F5F5F5')

    d_data  = {mo: float(sub.loc[mo,'decompose']) if mo in sub.index else 0.0 for mo in MONTHS}
    k_data  = {mo: float(sub.loc[mo,'kv2'])       if mo in sub.index else 0.0 for mo in MONTHS}
    bi_data = {mo: (k_data[mo]-d_data[mo])/d_data[mo]*100 if d_data[mo]>0 else None for mo in MONTHS}
    ann_d   = sum(d_data.values())
    ann_k   = sum(k_data.values())
    ann_bi  = (ann_k - ann_d) / ann_d * 100 if ann_d else None

    for label, data, fmt, is_bias in [
        ('Декомпозит', d_data,  '#,##0',         False),
        ('K_v2',       k_data,  '#,##0',         False),
        ('Bias %',     bi_data, '+0.0;-0.0;0.0', True),
    ]:
        c1 = ws2.cell(row_i, 1, grs if label == 'Декомпозит' else '')
        c1.font = Font(bold=is_large)
        if label == 'Декомпозит': c1.fill = row_fill
        c2 = ws2.cell(row_i, 2, label)
        c2.font = Font(bold=is_bias, italic=not is_bias, color='444444')
        if label == 'Декомпозит': c2.fill = row_fill

        for ci, mo in enumerate(MONTHS, 3):
            v = data[mo]
            if v is None:
                ws2.cell(row_i, ci, '-').alignment = Alignment(horizontal='center')
            else:
                c = ws2.cell(row_i, ci, round(v, 1 if is_bias else 0))
                c.number_format = fmt
                c.alignment = Alignment(horizontal='right')
                if is_bias:
                    c.font = Font(bold=True)
                    f = bias_fill(v)
                    if f: c.fill = f

        ann_v = ann_d if label == 'Декомпозит' else ann_k if label == 'K_v2' else ann_bi
        c = ws2.cell(row_i, 15,
                     round(ann_v, 1 if is_bias else 0) if ann_v is not None else '-')
        c.number_format = fmt if is_bias else '#,##0'
        c.alignment = Alignment(horizontal='right')
        if is_bias and ann_v is not None:
            c.font = Font(bold=True)
            f = bias_fill(ann_v)
            if f: c.fill = f

        border_side = thick if (is_bias and is_large) else thin if is_bias else Side()
        for ci in range(1, 16):
            ws2.cell(row_i, ci).border = Border(bottom=border_side)
        row_i += 1

# Підсумковий рядок
row_i += 1
ws2.cell(row_i, 1, 'ВСЬОГО').font = Font(bold=True)
ws2.cell(row_i, 2, 'Bias %').font = Font(bold=True)
for ci, mo in enumerate(MONTHS, 3):
    sm = m[m['month'] == mo]
    if sm.empty: continue
    b = (sm['kv2'].sum() - sm['decompose'].sum()) / sm['decompose'].sum() * 100
    c = ws2.cell(row_i, ci, round(b, 1))
    c.number_format = '+0.0;-0.0;0.0'; c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='right')
    f = bias_fill(b)
    if f: c.fill = f
all_b = (m['kv2'].sum() - m['decompose'].sum()) / m['decompose'].sum() * 100
c = ws2.cell(row_i, 15, round(all_b, 1))
c.number_format = '+0.0;-0.0;0.0'; c.font = Font(bold=True)
c.alignment = Alignment(horizontal='right')
f = bias_fill(all_b)
if f: c.fill = f

wb.save(OUT_XLSX)
print(f"  DONE → {OUT_XLSX}")
print(f"  Листи: По_місяцях | По_ГРС ({row_i} рядків)")
