import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import os; os.chdir(os.path.dirname(os.path.abspath(__file__)))
import numpy as np, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter

UA = {1:'Січ',2:'Лют',3:'Бер',4:'Кві',5:'Тра',6:'Чер',
      7:'Лип',8:'Сер',9:'Вер',10:'Жов',11:'Лис',12:'Гру'}
MONTHS = list(range(1, 13))
LARGE  = 5_000_000
thin   = Side(style='thin')
thick  = Side(style='medium')
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF')

def bias_fill(v):
    if v is None: return None
    av = abs(v)
    if v > 0:
        if av > 20: return PatternFill('solid', fgColor='FF2222')
        if av > 10: return PatternFill('solid', fgColor='FF7777')
        if av >  5: return PatternFill('solid', fgColor='FFBBBB')
    else:
        if av > 20: return PatternFill('solid', fgColor='2244FF')
        if av > 10: return PatternFill('solid', fgColor='7799FF')
        if av >  5: return PatternFill('solid', fgColor='BBCCFF')
    return None

# ── дані ──────────────────────────────────────────────────────────────────────
result  = pd.read_excel('data/annual_decompose_result.xlsx', sheet_name='All_GRS_monthly')
ann_grs = result.groupby('grs')['billing'].sum().sort_values(ascending=False)

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# Лист 1: підсумок по місяцях
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'По_місяцях'

headers = ['Показник'] + [UA[m] for m in MONTHS] + ['Річний']
for ci, h in enumerate(headers, 1):
    c = ws.cell(1, ci, h); c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
ws.column_dimensions['A'].width = 14
for ci in range(2, 15): ws.column_dimensions[get_column_letter(ci)].width = 11
ws.freeze_panes = 'B2'

bill_row = {m: result[result['month']==m]['billing'].sum()    for m in MONTHS}
pred_row = {m: result[result['month']==m]['total_pred'].sum() for m in MONTHS}
bias_row = {m: (pred_row[m]-bill_row[m])/bill_row[m]*100 if bill_row[m] else 0 for m in MONTHS}
ann_bill = sum(bill_row.values())
ann_pred = sum(pred_row.values())
ann_bias = (ann_pred - ann_bill) / ann_bill * 100

rows_def = [
    ('Білінг',  bill_row, '#,##0',         False),
    ('Прогноз', pred_row, '#,##0',         False),
    ('Bias %',  bias_row, '+0.0;-0.0;0.0', True),
]
for ri, (label, data, fmt, bold) in enumerate(rows_def, 2):
    ws.cell(ri, 1, label).font = Font(bold=bold, italic=not bold)
    ws.cell(ri, 1).alignment = Alignment(horizontal='left')
    for ci, m in enumerate(MONTHS, 2):
        c = ws.cell(ri, ci, round(data[m], 1 if bold else 0))
        c.number_format = fmt; c.alignment = Alignment(horizontal='right')
        if bold: c.font = Font(bold=True)
        if bold:
            f = bias_fill(data[m])
            if f: c.fill = f
    ann_v = ann_bill if label == 'Білінг' else ann_pred if label == 'Прогноз' else ann_bias
    c = ws.cell(ri, 14, round(ann_v, 1 if bold else 0))
    c.number_format = fmt; c.alignment = Alignment(horizontal='right')
    if bold: c.font = Font(bold=True)
    if bold:
        f = bias_fill(ann_v)
        if f: c.fill = f
    for ci in range(1, 15):
        ws.cell(ri, ci).border = Border(bottom=thin)

# ══════════════════════════════════════════════════════════════════════════════
# Лист 2: по ГРС × місяць
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('По_ГРС')

ws2.cell(1, 1, 'ГРС').font = HDR_FONT; ws2.cell(1, 1).fill = HDR_FILL
ws2.cell(1, 2, '').fill = HDR_FILL
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 10
for ci, m in enumerate(MONTHS, 3):
    c = ws2.cell(1, ci, UA[m]); c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
    ws2.column_dimensions[get_column_letter(ci)].width = 11
c = ws2.cell(1, 15, 'Річний'); c.font = HDR_FONT; c.fill = HDR_FILL
c.alignment = Alignment(horizontal='center')
ws2.column_dimensions['O'].width = 11
ws2.freeze_panes = 'C2'

row_i = 2
for grs in ann_grs.index:
    sub = result[result['grs'] == grs].set_index('month')
    if sub.empty: continue
    is_large = ann_grs[grs] >= LARGE
    row_fill = PatternFill('solid', fgColor='D6E4F0' if is_large else 'F5F5F5')

    b_data  = {m: float(sub.loc[m,'billing'])    if m in sub.index else 0.0 for m in MONTHS}
    p_data  = {m: float(sub.loc[m,'total_pred']) if m in sub.index else 0.0 for m in MONTHS}
    bi_data = {m: (p_data[m]-b_data[m])/b_data[m]*100 if b_data[m]>0 else None for m in MONTHS}
    ann_b  = sum(b_data.values())
    ann_p  = sum(p_data.values())
    ann_bi = (ann_p - ann_b) / ann_b * 100 if ann_b else None

    for label, data, fmt, bold in [
        ('Білінг',  b_data,  '#,##0',         False),
        ('Прогноз', p_data,  '#,##0',         False),
        ('Bias %',  bi_data, '+0.0;-0.0;0.0', True),
    ]:
        c1 = ws2.cell(row_i, 1, grs if label == 'Білінг' else '')
        c1.font = Font(bold=is_large)
        if label == 'Білінг': c1.fill = row_fill
        c2 = ws2.cell(row_i, 2, label)
        c2.font = Font(bold=bold, italic=not bold, color='444444')
        if label == 'Білінг': c2.fill = row_fill

        for ci, m in enumerate(MONTHS, 3):
            v = data[m]
            if v is None:
                ws2.cell(row_i, ci, '-').alignment = Alignment(horizontal='center')
            else:
                c = ws2.cell(row_i, ci, round(v, 1 if bold else 0))
                c.number_format = fmt
                c.alignment = Alignment(horizontal='right')
                if bold:
                    c.font = Font(bold=True)
                    f = bias_fill(v)
                    if f: c.fill = f

        ann_v = ann_b if label == 'Білінг' else ann_p if label == 'Прогноз' else ann_bi
        c = ws2.cell(row_i, 15, round(ann_v, 1 if bold else 0) if ann_v is not None else '-')
        c.number_format = fmt if bold else '#,##0'
        c.alignment = Alignment(horizontal='right')
        if bold:
            c.font = Font(bold=True)
            f = bias_fill(ann_v)
            if f: c.fill = f

        border_side = thick if (label == 'Bias %' and is_large) else thin if label == 'Bias %' else Side()
        for ci in range(1, 16):
            ws2.cell(row_i, ci).border = Border(bottom=border_side)
        row_i += 1

# підсумковий рядок
row_i += 1
ws2.cell(row_i, 1, 'ВСЬОГО').font = Font(bold=True)
ws2.cell(row_i, 2, 'Bias %').font = Font(bold=True)
for ci, m in enumerate(MONTHS, 3):
    sm = result[result['month'] == m]
    b  = (sm['total_pred'].sum() - sm['billing'].sum()) / sm['billing'].sum() * 100
    c  = ws2.cell(row_i, ci, round(b, 1))
    c.number_format = '+0.0;-0.0;0.0'; c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='right')
    f = bias_fill(b)
    if f: c.fill = f
all_bi = (result['total_pred'].sum() - result['billing'].sum()) / result['billing'].sum() * 100
c = ws2.cell(row_i, 15, round(all_bi, 1))
c.number_format = '+0.0;-0.0;0.0'; c.font = Font(bold=True)
c.alignment = Alignment(horizontal='right')
f = bias_fill(all_bi)
if f: c.fill = f

OUT = 'data/annual_decompose_table.xlsx'
wb.save(OUT)
print(f'Збережено -> {OUT}')
print(f'Листи: По_місяцях | По_ГРС ({row_i} рядків)')
