"""
Convert existing enterprise.xlsx + line_id.xlsx → new universal Excel format.

The new format uses:
  - Виробник   → short_name from `manufacturer` table
  - Модель     → model_name from `corector_type` table  (first match by type_dev)
  - Лінія      → name from `gas_volume_line` table
  - All other columns remain the same

Output: backend/data/enterprise_import.xlsx

Usage:
    python -m backend.db.preload_db.export_enterprise_to_new_excel
"""
import asyncio
import io
import logging
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlmodel import select

from backend.db.engine import async_session_factory
from backend.db.models.device_catalog_model import Manufacturer, CorectorType
from backend.db.models.line_model import Line
from backend.services.enterprise_mappings import load_mappings

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

OUTPUT_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "enterprise_import.xlsx"
)

_COLUMNS  = ["Підприємство", "Серійний номер", "Виробник", "Модель коректора",
             "Канал (0-based)", "Активний", "Увімкнений", "Лінія (назва)"]
_COL_WIDTHS = [40, 18, 16, 24, 16, 12, 14, 30]


async def build_lookups():
    async with async_session_factory() as session:
        mfrs    = (await session.execute(select(Manufacturer))).scalars().all()
        ctypes  = (await session.execute(select(CorectorType))).scalars().all()
        lines   = (await session.execute(select(Line))).scalars().all()

    # mf_dev → short_name
    mf_dev_to_short: dict[int, str] = {m.mf_dev: m.short_name for m in mfrs}

    # (manufacturer_id, type_dev) → model_name  (first match wins)
    mfr_id_by_mfdev = {m.mf_dev: m.id for m in mfrs}
    type_to_model: dict[tuple[int, int], str] = {}
    for ct in ctypes:
        key = (ct.manufacturer_id, ct.type_dev)
        if key not in type_to_model:          # keep first match for duplicates
            type_to_model[key] = ct.model_name

    # line_id → line name
    line_id_to_name: dict[int, str] = {ln.id: ln.name for ln in lines}

    return mf_dev_to_short, mfr_id_by_mfdev, type_to_model, line_id_to_name


def write_excel(rows: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дані"

    hdr_fill  = PatternFill("solid", fgColor="2E7D32")
    hint_fill = PatternFill("solid", fgColor="1B5E20")
    hdr_font  = Font(bold=True, color="FFFFFF")
    hint_font = Font(italic=True, color="A5D6A7")

    # Row 1 — headers
    for col_idx, (col_name, width) in enumerate(zip(_COLUMNS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 2 — hints
    hints = ["Назва точки обліку", "Серійний номер", "Скорочена назва",
             "Модель з довідника", "0, 1, 2…", "Так / Ні", "Так / Ні",
             "Точна назва лінії або порожньо"]
    for col_idx, hint in enumerate(hints, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font = hint_font
        cell.fill = hint_fill

    ws.freeze_panes = "A3"

    # Data rows
    for row_idx, r in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=r["enterprise_name"])
        ws.cell(row=row_idx, column=2, value=r["ser_num"])
        ws.cell(row=row_idx, column=3, value=r["mfr_short"])
        ws.cell(row=row_idx, column=4, value=r["model_name"])
        ws.cell(row=row_idx, column=5, value=r["ch_num"])
        ws.cell(row=row_idx, column=6, value="Так" if r["active"] else "Ні")
        ws.cell(row=row_idx, column=7, value="Так" if r["enabled"] else "Ні")
        ws.cell(row=row_idx, column=8, value=r["line_name"])

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    logger.info(f"Saved {len(rows)} rows → {output_path}")


async def main() -> None:
    logger.info("Loading lookup tables from DB…")
    mf_dev_to_short, mfr_id_by_mfdev, type_to_model, line_id_to_name = await build_lookups()

    logger.info("Reading enterprise.xlsx + line_id.xlsx…")
    df = load_mappings()

    if df is None or df.empty:
        logger.error("No data in enterprise.xlsx — exiting")
        return

    rows = []
    skipped = 0
    import pandas as pd

    for _, row in df.iterrows():
        mf_dev   = int(row["mfDev"])
        type_dev = int(row["typeDev"])

        mfr_short = mf_dev_to_short.get(mf_dev)
        if not mfr_short:
            logger.warning(f"Unknown mf_dev={mf_dev} for {row['enterprise_name']} — skipped")
            skipped += 1
            continue

        mfr_id    = mfr_id_by_mfdev.get(mf_dev)
        model_name = type_to_model.get((mfr_id, type_dev))
        if not model_name:
            logger.warning(
                f"No model for mf_dev={mf_dev} type_dev={type_dev} "
                f"({row['enterprise_name']}) — model left blank"
            )
            model_name = ""

        line_id_val = None if pd.isna(row["line_id"]) else int(row["line_id"])
        line_name   = line_id_to_name.get(line_id_val, "") if line_id_val else ""

        rows.append({
            "enterprise_name": str(row["enterprise_name"]).strip(),
            "ser_num":   int(row["serNum"]),
            "mfr_short": mfr_short,
            "model_name": model_name,
            "ch_num":    int(row["chNum"]),
            "active":    bool(row["active"]),
            "enabled":   bool(row["enabled"]),
            "line_name": line_name,
        })

    logger.info(f"Converted {len(rows)} rows ({skipped} skipped)")
    output = os.path.abspath(OUTPUT_PATH)
    write_excel(rows, output)
    logger.info(f"Done! File: {output}")


if __name__ == "__main__":
    asyncio.run(main())
