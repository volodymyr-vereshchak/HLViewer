"""
Enterprise Volumes API Endpoint

Provides endpoints for fetching enterprise volume data from DPD API,
aggregating by line_id and date, and CRUD management of enterprise records.
"""

import io
import logging
import pandas as pd
from datetime import datetime, date
from typing import List, Optional
from collections import defaultdict
from fastapi import APIRouter, Depends, Query, status, HTTPException, UploadFile, File
from backend.api.endpoints.auth_ep import get_branch_filter
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from sqlmodel import select

from backend.db.engine import async_session_factory
from backend.db.dao.enterprise_dao import EnterpriseDao
from backend.db.models.enterprise_model import EnterpriseRead, EnterpriseCreate, EnterpriseUpdate
from backend.db.models.enterprise_models import (
    EnterpriseVolumeResponse,
    DeviceVolume,
    EnterpriseVolumeError,
    EnterpriseMapping
)
from backend.services.dpd_client import DPDClient
from backend.services.enterprise_mappings import get_devices_for_lines, get_devices_for_lines_db, load_mappings

logger = logging.getLogger(__name__)


class EnterpriseRouter:
    """Router for enterprise volume endpoints."""

    def __init__(self):
        self.router = APIRouter()
        self.router.add_api_route(
            path="/enterprise/volumes/",
            tags=["enterprise"],
            endpoint=self.get_enterprise_volumes,
            methods=["GET"],
            response_model=List[EnterpriseVolumeResponse],
            status_code=status.HTTP_200_OK,
            summary="Get enterprise volume data",
            description=(
                "Fetches volume data for enterprise calculators from DPD API, "
                "aggregated by line_id and date. Returns empty array if no "
                "mappings exist for specified lines."
            ),
            responses={
                200: {
                    "description": "Successfully retrieved enterprise volumes",
                    "model": List[EnterpriseVolumeResponse]
                },
                400: {
                    "description": "Invalid request parameters",
                    "model": EnterpriseVolumeError
                },
                500: {
                    "description": "Server error (e.g., mappings file not found)",
                    "model": EnterpriseVolumeError
                },
                503: {
                    "description": "DPD API unavailable",
                    "model": EnterpriseVolumeError
                }
            }
        )
        self.router.add_api_route(
            path="/enterprise/mappings/",
            tags=["enterprise"],
            endpoint=self.get_all_enterprises_excel,
            methods=["GET"],
            response_model=List[EnterpriseMapping],
            status_code=status.HTTP_200_OK,
            summary="Get all enterprise mappings (Excel source)",
            description=(
                "Returns list of all enterprises from Excel mappings with their device "
                "information and active status. Legacy endpoint."
            ),
        )


    async def get_enterprise_volumes(
        self,
        line_id: Optional[List[int]] = Query(default=None, description="Line IDs to fetch volumes for (optional if serNum+chNum provided)"),
        from_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
        to_date: str = Query(..., description="End date (YYYY-MM-DD)"),
        period_type: str = Query(
            default="daily",
            pattern="^(daily|hourly)$",
            description="Data granularity: 'daily' or 'hourly'"
        ),
        serNum: Optional[int] = Query(None, description="Optional: Filter by device serial number"),
        mfDev: Optional[int] = Query(None, description="Optional: Manufacturer code"),
        typeDev: Optional[int] = Query(None, description="Optional: Device type code"),
        chNum: Optional[int] = Query(None, description="Optional: Filter by device channel number"),
    ) -> List[EnterpriseVolumeResponse]:
        logger.info(
            f"Fetching enterprise volumes for lines {line_id}, "
            f"period {from_date} to {to_date}, granularity: {period_type}"
        )

        try:
            date_from = datetime.strptime(from_date.split(" ")[0].split("T")[0], "%Y-%m-%d")
            date_to = datetime.strptime(to_date.split(" ")[0].split("T")[0], "%Y-%m-%d")
            if date_from > date_to:
                raise ValueError("from_date must be <= to_date")
        except ValueError as e:
            logger.error(f"Invalid date parameters: {e}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid date format: {e}. Use YYYY-MM-DD format."
            )

        if not line_id and (serNum is None or chNum is None or mfDev is None or typeDev is None):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Provide line_id or serNum+chNum"
            )

        try:
            async with async_session_factory() as session:
                if line_id:
                    devices = await get_devices_for_lines_db(line_id, session)
                    if serNum is not None and chNum is not None:
                        devices = [d for d in devices if d["serNum"] == serNum and d["chNum"] == chNum]
                else:
                    # No line_id — lookup by full device identity
                    ent = await EnterpriseDao(session).get_by_device(serNum, mfDev, typeDev, chNum)
                    if not ent:
                        logger.warning(f"No enterprise found: serNum={serNum} mfDev={mfDev} typeDev={typeDev} chNum={chNum}")
                        return []
                    devices = [{
                        "line_id": ent.line_id,
                        "branch_id": ent.branch_id,
                        "serNum": ent.ser_num,
                        "mfDev": ent.mf_dev,
                        "typeDev": ent.type_dev,
                        "chNum": ent.ch_num,
                        "enterprise_name": ent.enterprise_name,
                    }]
        except Exception as e:
            logger.error(f"Error loading enterprise mappings from DB: {e}")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))

        if not devices:
            logger.info(f"No enterprise mappings found for lines {line_id}")
            return []

        branch_id = next((d["branch_id"] for d in devices if d.get("branch_id")), None)
        if branch_id is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Could not determine branch for requested lines"
            )

        try:
            async with async_session_factory() as cred_session:
                client = await DPDClient.for_branch(branch_id, cred_session)
            volumes_data = await client.get_volumes(
                devices, date_from, date_to, type_request=period_type
            )
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
        except Exception as e:
            logger.error(f"DPD API error: {e}")
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail=f"DPD API unavailable: {e}"
            )

        if not volumes_data:
            logger.warning("No volume data returned from DPD API")
            return []

        device_map = {
            (d["serNum"], d["mfDev"], d["typeDev"], d["chNum"]): d
            for d in devices
        }

        aggregated = defaultdict(lambda: {"total": 0.0, "devices": []})

        for record in volumes_data:
            device_key = (
                record.get("serNum"),
                record.get("mfDev"),
                record.get("typeDev"),
                record.get("chNum")
            )

            device_info = device_map.get(device_key)
            if not device_info:
                logger.warning(f"Device not found in mappings: {device_key}")
                continue

            volume = record.get("dvstAlwrk")

            record_date_str = record.get("date") or record.get("period")
            if not record_date_str:
                logger.warning(f"No date field in record: {record}")
                continue

            try:
                if period_type == 'hourly':
                    if isinstance(record_date_str, str):
                        clean_str = record_date_str.split(".")[0]
                        try:
                            record_period = datetime.strptime(clean_str, "%Y-%m-%dT%H:%M:%S")
                        except ValueError:
                            record_period = datetime.strptime(clean_str, "%Y-%m-%dT%H:%M")
                    elif isinstance(record_date_str, datetime):
                        record_period = record_date_str
                    else:
                        logger.warning(f"Invalid datetime format: {record_date_str}")
                        continue
                else:
                    if isinstance(record_date_str, str):
                        record_period = datetime.strptime(
                            record_date_str.split("T")[0], "%Y-%m-%d"
                        ).date()
                    elif isinstance(record_date_str, datetime):
                        record_period = record_date_str.date()
                    elif isinstance(record_date_str, date):
                        record_period = record_date_str
                    else:
                        logger.warning(f"Invalid date format: {record_date_str}")
                        continue
            except Exception as e:
                logger.warning(f"Error parsing period {record_date_str}: {e}")
                continue

            key = (device_info["line_id"], record_period)
            temperature = record.get("temper")
            pressure = record.get("press")
            # Pressure unit as reported by the DPD API (e.g. "kPa").
            pressure_unit = record.get("pressUnit")
            if isinstance(pressure_unit, str):
                pressure_unit = pressure_unit.strip() or None

            if volume is not None:
                aggregated[key]["total"] += volume
            aggregated[key]["devices"].append(
                DeviceVolume(
                    serNum=device_info["serNum"],
                    mfDev=device_info["mfDev"],
                    typeDev=device_info["typeDev"],
                    chNum=device_info["chNum"],
                    enterprise_name=device_info.get("enterprise_name", ""),
                    volume=volume,
                    temperature=temperature,
                    pressure=pressure,
                    pressure_unit=pressure_unit
                )
            )

        result = []
        for (line_id_val, period_val), data in aggregated.items():
            result.append(
                EnterpriseVolumeResponse(
                    line_id=line_id_val,
                    period=period_val,
                    total_volume=round(data["total"], 2),
                    device_count=len(data["devices"]),
                    devices=data["devices"]
                )
            )

        result.sort(key=lambda x: (x.line_id is None, x.line_id or 0, x.period))

        logger.info(
            f"Returning {len(result)} aggregated enterprise volume records "
            f"for {len(devices)} devices"
        )

        return result

    async def get_all_enterprises_excel(self) -> List[EnterpriseMapping]:
        """Legacy endpoint: read from Excel."""
        logger.info("Fetching all enterprise mappings (Excel)")

        try:
            df = load_mappings()
        except FileNotFoundError as e:
            logger.error(f"Enterprise mappings file not found: {e}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=str(e)
            )
        except Exception as e:
            logger.error(f"Error loading enterprise mappings: {e}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error loading enterprise mappings: {e}"
            )

        if df is None or df.empty:
            logger.warning("No enterprise mappings available")
            return []

        result = []
        for _, row in df.iterrows():
            line_id_val = None if pd.isna(row["line_id"]) else int(row["line_id"])
            result.append(
                EnterpriseMapping(
                    line_id=line_id_val,
                    serNum=int(row["serNum"]),
                    mfDev=int(row["mfDev"]),
                    typeDev=int(row["typeDev"]),
                    chNum=int(row["chNum"]),
                    enterprise_name=str(row["enterprise_name"]),
                    active=bool(row["active"]),
                    enabled=bool(row["enabled"])
                )
            )

        result.sort(key=lambda x: (x.line_id is None, x.line_id or 0, x.enterprise_name))
        logger.info(f"Returning {len(result)} enterprise mappings")
        return result


# Create router instance (class-based routes)
enterprise_router = EnterpriseRouter().router

# ─── DB CRUD routes ───────────────────────────────────────────────────────────

_crud_router = APIRouter(tags=["enterprise"])


@_crud_router.get(
    "/enterprise-mappings/",
    response_model=List[EnterpriseRead],
    summary="List all enterprises (DB)",
)
async def list_enterprises(branch_ids: list[int] | None = Depends(get_branch_filter)):
    async with async_session_factory() as session:
        dao = EnterpriseDao(session=session)
        if branch_ids is None:
            return await dao.get_all()
        return await dao.get_by_branch_ids(branch_ids)


@_crud_router.post(
    "/enterprise-mappings/",
    response_model=EnterpriseRead,
    status_code=status.HTTP_201_CREATED,
    summary="Create enterprise (DB)",
)
async def create_enterprise(data: EnterpriseCreate):
    async with async_session_factory() as session:
        dao = EnterpriseDao(session=session)
        return await dao.create(data)


@_crud_router.patch(
    "/enterprise-mappings/{enterprise_id}",
    response_model=EnterpriseRead,
    summary="Update enterprise (DB)",
)
async def update_enterprise(enterprise_id: int, data: EnterpriseUpdate):
    async with async_session_factory() as session:
        dao = EnterpriseDao(session=session)
        item = await dao.update(enterprise_id, data)
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Enterprise not found")
    return item


@_crud_router.delete(
    "/enterprise-mappings/{enterprise_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete enterprise (DB)",
)
async def delete_enterprise(enterprise_id: int):
    async with async_session_factory() as session:
        dao = EnterpriseDao(session=session)
        deleted = await dao.delete(enterprise_id)
    if not deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Enterprise not found")


# ─── Excel template & upload ─────────────────────────────────────────────────

_COLUMNS  = ["Підприємство", "Серійний номер", "Виробник", "Модель коректора",
             "Канал (0-based)", "Активний", "Увімкнений", "ID лінії", "Назва лінії (довідково)"]
_COL_WIDTHS = [40, 18, 16, 24, 16, 12, 14, 12, 32]


async def _lines_with_context(session):
    """Lines for the reference sheet: (line_id, line_name, calc_name, lumg_name).

    Line names are NOT globally unique (a line is unique only within its
    calculator), so the reference carries the ID plus calculator + LUMG context
    to disambiguate. The data sheet then references the line by ID."""
    from backend.db.models.line_model import Line
    from backend.db.models.gas_volume_calc_model import GasVolumeCalc
    from backend.db.models.lumg_model import Lumg

    stmt = (
        select(Line.id, Line.name, GasVolumeCalc.name, Lumg.name)
        .join(GasVolumeCalc, Line.gas_volume_calc_id == GasVolumeCalc.id)
        .join(Lumg, GasVolumeCalc.lumg_id == Lumg.id)
        .order_by(Lumg.name, GasVolumeCalc.name, Line.name)
    )
    return (await session.execute(stmt)).all()


async def _build_template_workbook() -> openpyxl.Workbook:
    from backend.db.dao.device_catalog_dao import ManufacturerDao, CorectorTypeDao
    from backend.db.dao.line_dao import LineDao

    async with async_session_factory() as session:
        manufacturers = await ManufacturerDao(session).get_all()
        corector_types = await CorectorTypeDao(session).get_all()
        lines_ctx = await _lines_with_context(session)  # (id, name, calc, lumg)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дані"

    hdr_fill  = PatternFill("solid", fgColor="2E7D32")
    hint_fill = PatternFill("solid", fgColor="1B5E20")
    hdr_font  = Font(bold=True, color="FFFFFF")
    hint_font = Font(italic=True, color="A5D6A7")
    ref_hdr_fill = PatternFill("solid", fgColor="1565C0")
    ref_hdr_font = Font(bold=True, color="FFFFFF")

    # ── Sheet 1: Data ────────────────────────────────────────────────────────
    for col_idx, (col_name, width) in enumerate(zip(_COLUMNS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    mfr_names  = " / ".join(m.short_name for m in manufacturers) or "РадмирТех / ГРЕМПІС / Тандем / Укргазтех"
    hints = [
        "Назва точки обліку",
        "Напр.: 123456 (без дефісів)",
        mfr_names,
        "Модель (напр. ВЕГА-1.01) — див. аркуш 'Довідник'",
        "0, 1, 2, …",
        "Так / Ні",
        "Так / Ні",
        "ID лінії з аркуша 'Довідник' (або порожньо)",
        "Заповнюється автоматично при експорті",
    ]
    for col_idx, hint in enumerate(hints, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font = hint_font
        cell.fill = hint_fill

    # example row
    ex_mfr   = manufacturers[0].short_name if manufacturers else "РадмирТех"
    ex_model = corector_types[0].model_name if corector_types else "ВЕГА-1.01"
    ex_lid   = lines_ctx[0][0] if lines_ctx else ""
    ex_lname = lines_ctx[0][1] if lines_ctx else ""
    for col_idx, val in enumerate(
        ["ТОВ Завод №1", 123456, ex_mfr, ex_model, 0, "Так", "Так", ex_lid, ex_lname], start=1
    ):
        ws.cell(row=3, column=col_idx, value=val)

    ws.freeze_panes = "A3"

    # ── Sheet 2: Reference ────────────────────────────────────────────────────
    ref = wb.create_sheet("Довідник")

    # Manufacturers + models
    ref.cell(row=1, column=1, value="Виробник (скорочено)").font = ref_hdr_font
    ref.cell(row=1, column=1).fill = ref_hdr_fill
    ref.cell(row=1, column=2, value="Модель коректора").font = ref_hdr_font
    ref.cell(row=1, column=2).fill = ref_hdr_fill
    ref.column_dimensions["A"].width = 20
    ref.column_dimensions["B"].width = 28

    mfr_map = {m.id: m.short_name for m in manufacturers}
    row = 2
    for ct in corector_types:
        ref.cell(row=row, column=1, value=mfr_map.get(ct.manufacturer_id, ""))
        ref.cell(row=row, column=2, value=ct.model_name)
        row += 1

    # Lines — ID + назва + вичислювач + ЛУМГ (look up the ID by name here, then
    # put the ID into the 'ID лінії' column on the data sheet).
    line_ref_cols = [("ID лінії", 10), ("Назва лінії", 32), ("Вичислювач", 28), ("ЛУМГ", 24)]
    for j, (title, width) in enumerate(line_ref_cols, start=4):  # columns D, E, F, G
        c = ref.cell(row=1, column=j, value=title)
        c.font = ref_hdr_font
        c.fill = ref_hdr_fill
        ref.column_dimensions[get_column_letter(j)].width = width
    for i, (lid, lname, cname, lumgname) in enumerate(lines_ctx, start=2):
        ref.cell(row=i, column=4, value=lid)
        ref.cell(row=i, column=5, value=lname)
        ref.cell(row=i, column=6, value=cname)
        ref.cell(row=i, column=7, value=lumgname)

    return wb


@_crud_router.get(
    "/enterprise-mappings/template",
    summary="Download Excel template for enterprise import",
    response_class=StreamingResponse,
)
async def download_template():
    wb = await _build_template_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enterprise_template.xlsx"},
    )


@_crud_router.get(
    "/enterprise-mappings/export",
    summary="Export current enterprises from DB to Excel",
    response_class=StreamingResponse,
)
async def export_enterprises():
    from backend.db.dao.device_catalog_dao import ManufacturerDao, CorectorTypeDao
    from backend.db.dao.line_dao import LineDao

    async with async_session_factory() as session:
        enterprises = await EnterpriseDao(session).get_all()
        manufacturers = await ManufacturerDao(session).get_all()
        corector_types = await CorectorTypeDao(session).get_all()
        lines = await LineDao(session).get_all()

    mfr_by_mfdev = {m.mf_dev: m.short_name for m in manufacturers}
    mfr_id_by_mfdev = {m.mf_dev: m.id for m in manufacturers}
    ct_model: dict[tuple, str] = {}
    for ct in corector_types:
        key = (ct.manufacturer_id, ct.type_dev)
        if key not in ct_model:
            ct_model[key] = ct.model_name
    line_by_id = {ln.id: ln.name for ln in lines}

    hdr_fill  = PatternFill("solid", fgColor="2E7D32")
    hint_fill = PatternFill("solid", fgColor="1B5E20")
    hdr_font  = Font(bold=True, color="FFFFFF")
    hint_font = Font(italic=True, color="A5D6A7")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дані"

    for col_idx, (col_name, width) in enumerate(zip(_COLUMNS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    hints = ["Назва точки обліку", "Серійний номер", "Скорочена назва",
             "Модель з довідника", "0, 1, 2…", "Так / Ні", "Так / Ні",
             "ID лінії (ключ для імпорту)", "Назва лінії (довідково)"]
    for col_idx, hint in enumerate(hints, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font = hint_font
        cell.fill = hint_fill

    ws.freeze_panes = "A3"

    for row_idx, ent in enumerate(enterprises, start=3):
        mfr_short = mfr_by_mfdev.get(ent.mf_dev, str(ent.mf_dev))
        mfr_id = mfr_id_by_mfdev.get(ent.mf_dev)
        model_name = ct_model.get((mfr_id, ent.type_dev), "") if mfr_id else ""
        line_name = line_by_id.get(ent.line_id, "") if ent.line_id else ""
        ws.cell(row=row_idx, column=1, value=ent.enterprise_name)
        ws.cell(row=row_idx, column=2, value=ent.ser_num)
        ws.cell(row=row_idx, column=3, value=mfr_short)
        ws.cell(row=row_idx, column=4, value=model_name)
        ws.cell(row=row_idx, column=5, value=ent.ch_num)
        ws.cell(row=row_idx, column=6, value="Так" if ent.active else "Ні")
        ws.cell(row=row_idx, column=7, value="Так" if ent.enabled else "Ні")
        ws.cell(row=row_idx, column=8, value=ent.line_id if ent.line_id else None)
        ws.cell(row=row_idx, column=9, value=line_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enterprise_export.xlsx"},
    )


@_crud_router.post(
    "/enterprise-mappings/upload",
    summary="Import enterprises from Excel",
)
async def upload_enterprises(file: UploadFile = File(...), branch_id: Optional[int] = None):
    from backend.db.dao.device_catalog_dao import ManufacturerDao, CorectorTypeDao
    from backend.db.dao.line_dao import LineDao
    from backend.db.models.enterprise_model import Enterprise as EnterpriseModel

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Не вдалося відкрити файл: {e}")

    # Pre-load lookup tables from DB
    async with async_session_factory() as session:
        manufacturers   = await ManufacturerDao(session).get_all()
        corector_types  = await CorectorTypeDao(session).get_all()
        all_lines       = await LineDao(session).get_all()

    mfr_by_short: dict[str, object]        = {m.short_name.strip(): m for m in manufacturers}
    ct_by_mfr_model: dict[tuple, object]   = {
        (ct.manufacturer_id, ct.model_name.strip()): ct for ct in corector_types
    }
    line_ids_set: set = {ln.id for ln in all_lines}
    # name → [line_id, …]; a name maps to >1 id when calculators reuse line names
    line_ids_by_name: dict[str, list] = defaultdict(list)
    for ln in all_lines:
        line_ids_by_name[ln.name.strip().lower()].append(ln.id)


    def parse_bool(val, default=True) -> bool:
        if val is None:
            return default
        return str(val).strip().lower() not in ("ні", "нет", "no", "false", "0", "н")

    records: list[dict] = []
    errors:  list[str]  = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or not any(row):
            continue

        row_padded = (list(row) + [None] * 9)[:9]
        (enterprise_name, ser_num, mfr_str, model_str, ch_num,
         active_str, enabled_str, line_key, _line_name_info) = row_padded

        if not enterprise_name:
            errors.append(f"Рядок {row_idx}: порожня назва — пропущено")
            continue

        # ser_num
        try:
            ser_num_int = int(str(ser_num).replace("-", "").lstrip("0") or "0")
        except (ValueError, TypeError):
            errors.append(f"Рядок {row_idx}: некоректний серійний номер '{ser_num}'")
            continue

        # manufacturer → mf_dev
        mfr_clean = str(mfr_str).strip() if mfr_str else ""
        mfr = mfr_by_short.get(mfr_clean)
        if not mfr:
            errors.append(
                f"Рядок {row_idx}: невідомий виробник '{mfr_clean}'. "
                f"Допустимі: {', '.join(mfr_by_short)}"
            )
            continue

        # model → type_dev
        model_clean = str(model_str).strip() if model_str else ""
        ct = ct_by_mfr_model.get((mfr.id, model_clean))
        if not ct:
            errors.append(
                f"Рядок {row_idx}: модель '{model_clean}' не знайдена для виробника '{mfr_clean}'"
            )
            continue

        # ch_num
        try:
            ch_num_int = int(ch_num)
        except (ValueError, TypeError):
            errors.append(f"Рядок {row_idx}: некоректний канал '{ch_num}'")
            continue

        # line (optional): column "ID лінії". Primary = numeric line ID; fall back
        # to a name match for older files — but reject ambiguous names (a name that
        # belongs to several lines across calculators), telling the user to use the ID.
        line_id = None
        if line_key is not None and str(line_key).strip():
            key = str(line_key).strip()
            if key.replace(".", "", 1).isdigit():            # numeric → line ID
                lid = int(float(key))
                if lid in line_ids_set:
                    line_id = lid
                else:
                    errors.append(f"Рядок {row_idx}: ID лінії {lid} не існує — line_id=null")
            else:                                            # text → name lookup
                ids = line_ids_by_name.get(key.lower(), [])
                if len(ids) == 1:
                    line_id = ids[0]
                elif len(ids) > 1:
                    errors.append(
                        f"Рядок {row_idx}: назва лінії '{key}' неоднозначна "
                        f"({len(ids)} збігів) — вкажіть ID лінії з аркуша 'Довідник'"
                    )
                else:
                    errors.append(f"Рядок {row_idx}: лінія '{key}' не знайдена — line_id=null")

        records.append({
            "enterprise_name": str(enterprise_name).strip(),
            "branch_id": branch_id,
            "ser_num":  ser_num_int,
            "mf_dev":   mfr.mf_dev,
            "type_dev": ct.type_dev,
            "ch_num":   ch_num_int,
            "active":   parse_bool(active_str),
            "enabled":  parse_bool(enabled_str),
            "line_id":  line_id,
        })

    if not records:
        return {"imported": 0, "warnings": len(errors), "errors": errors}

    async with async_session_factory() as session:
        from sqlalchemy.dialects.postgresql import insert as pg_insert
        from backend.db.models.enterprise_model import Enterprise as EnterpriseModel

        stmt = pg_insert(EnterpriseModel).values(records)
        stmt = stmt.on_conflict_do_update(
            constraint='uq_enterprise_device',
            set_={
                'enterprise_name': stmt.excluded.enterprise_name,
                'branch_id':       stmt.excluded.branch_id,
                'line_id':         stmt.excluded.line_id,
                'active':          stmt.excluded.active,
                'enabled':         stmt.excluded.enabled,
            }
        ).returning(EnterpriseModel.id)

        result = await session.execute(stmt)
        await session.commit()
        returned_ids = [row[0] for row in result]

    # Count inserted vs updated by comparing with existing ids
    # (simple approach: just report total)
    logger.info(f"Upserted {len(returned_ids)} enterprise records from Excel ({len(errors)} warnings)")
    return {
        "imported": len(returned_ids),
        "warnings": len(errors),
        "errors":   errors,
    }


# Attach CRUD router to the main enterprise router so main.py needs no change
enterprise_router.include_router(_crud_router)
